#!/usr/bin/env python3
"""Unified rule/config loader with Excel-first, JSON-fallback strategy.

This module centralizes rule file loading so business scripts do not need
to know whether a rule currently lives in Excel or JSON.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


SHEET_RUNTIME = "\u8fd0\u884c\u63a7\u5236"
SHEET_RECON = "\u52fe\u7a3d\u89c4\u5219"
SHEET_RATIO_ALIAS = "\u6307\u6807\u522b\u540d\u6620\u5c04"
SHEET_RATIO_RULE = "\u8d22\u52a1\u6bd4\u7387\u89c4\u5219"


def _config_paths(project_root: Path) -> Dict[str, Path]:
    cfg_dir = project_root / "config"
    return {
        "rulebook_xlsx": cfg_dir / "rulebook.xlsx",
        "runtime_json": cfg_dir / "runtime_controls.json",
        "workbook_rules_json": cfg_dir / "workbook_rules_v1.json",
        "ratio_rules_json": cfg_dir / "financial_ratio_rules_v1.json",
    }


def _get_sheet_by_loose_name(wb, candidates: List[str]):
    sheet_map = {str(name).strip(): name for name in wb.sheetnames}
    for cand in candidates:
        key = str(cand).strip()
        if key in sheet_map:
            return wb[sheet_map[key]]
    return None


def _parse_bool(v: Any, default: bool = True) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"1", "true", "yes", "y", "\u662f"}:
        return True
    if s in {"0", "false", "no", "n", "\u5426"}:
        return False
    return default


def _parse_jsonish(v: Any) -> Any:
    if v is None or isinstance(v, (int, float, bool, list, dict)):
        return v
    text = str(v).strip()
    if not text:
        return None
    low = text.lower()
    if low in {"true", "false"}:
        return low == "true"
    try:
        if "." in text:
            return float(text)
        return int(text)
    except Exception:
        pass
    if (text.startswith("{") and text.endswith("}")) or (text.startswith("[") and text.endswith("]")):
        try:
            return json.loads(text)
        except Exception:
            return text
    return text


def _split_aliases(text: str) -> List[str]:
    raw = str(text or "").strip()
    if not raw:
        return []
    for sep in ["\uff1b", ";", "\uff0c", ",", "\u3001", "|", "/", "\\", "\n", "\r"]:
        raw = raw.replace(sep, ",")
    return [x.strip() for x in raw.split(",") if x.strip()]


def _safe_read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def load_runtime_controls(project_root: Path) -> Dict[str, Any]:
    paths = _config_paths(project_root)
    rb = paths["rulebook_xlsx"]
    if rb.exists():
        try:
            wb = load_workbook(rb, data_only=False)
            ws = _get_sheet_by_loose_name(wb, [SHEET_RUNTIME, "runtime_controls"])
            if ws is not None:
                out: Dict[str, Any] = {}
                for r in range(2, ws.max_row + 1):
                    key = str(ws.cell(r, 1).value or "").strip()
                    if not key:
                        continue
                    out[key] = _parse_jsonish(ws.cell(r, 2).value)
                if out:
                    return out
        except Exception:
            pass
    return _safe_read_json(paths["runtime_json"], {})


def load_workbook_rules(project_root: Path) -> Dict[str, Any]:
    paths = _config_paths(project_root)
    rb = paths["rulebook_xlsx"]
    if rb.exists():
        try:
            wb = load_workbook(rb, data_only=False)
            recon_rules: List[Dict[str, Any]] = []
            ws_recon = _get_sheet_by_loose_name(wb, [SHEET_RECON, "reconciliation_rules"])
            if ws_recon is not None:
                for r in range(2, ws_recon.max_row + 1):
                    rid = str(ws_recon.cell(r, 1).value or "").strip()
                    desc = str(ws_recon.cell(r, 2).value or "").strip()
                    formula = str(ws_recon.cell(r, 3).value or "").strip()
                    enabled = _parse_bool(ws_recon.cell(r, 4).value, True)
                    if rid and "=" in formula:
                        recon_rules.append(
                            {"id": rid, "description": desc, "formula": formula, "enabled": enabled}
                        )

            alias_map: Dict[str, List[str]] = {}
            ws_alias = _get_sheet_by_loose_name(wb, [SHEET_RATIO_ALIAS, "ratio_aliases"])
            if ws_alias is not None:
                for r in range(2, ws_alias.max_row + 1):
                    rid = str(ws_alias.cell(r, 1).value or "").strip()
                    alias_raw = str(ws_alias.cell(r, 2).value or "").strip()
                    if not rid or not alias_raw:
                        continue
                    vals = _split_aliases(alias_raw)
                    if vals:
                        alias_map[rid] = vals

            if recon_rules or alias_map:
                return {
                    "version": "v1",
                    "reconciliation": {"rules": recon_rules},
                    "ratio_aliases": alias_map,
                }
        except Exception:
            pass
    return _safe_read_json(paths["workbook_rules_json"], {})


def load_ratio_rules(project_root: Path) -> Dict[str, Any]:
    paths = _config_paths(project_root)
    rb = paths["rulebook_xlsx"]
    if rb.exists():
        try:
            wb = load_workbook(rb, data_only=False)
            ws = _get_sheet_by_loose_name(wb, [SHEET_RATIO_RULE, "ratio_rules"])
            if ws is not None:
                rules: List[Dict[str, Any]] = []
                for r in range(2, ws.max_row + 1):
                    rid = str(ws.cell(r, 1).value or "").strip()
                    if not rid:
                        continue
                    name = str(ws.cell(r, 2).value or "").strip() or rid
                    group = str(ws.cell(r, 3).value or "").strip()
                    enabled = _parse_bool(ws.cell(r, 4).value, True)
                    desc = str(ws.cell(r, 5).value or "").strip()
                    direct_aliases = _split_aliases(str(ws.cell(r, 6).value or ""))
                    calc_formula = str(ws.cell(r, 7).value or "").strip()
                    prefer_direct = _parse_bool(ws.cell(r, 8).value, True)
                    try:
                        direct_value_divisor = float(ws.cell(r, 9).value or 1.0)
                    except Exception:
                        direct_value_divisor = 1.0
                    rules.append(
                        {
                            "id": rid,
                            "name": name,
                            "group": group,
                            "enabled": enabled,
                            "description": desc,
                            "direct_aliases": direct_aliases,
                            "calc_formula": calc_formula,
                            "prefer_direct": prefer_direct,
                            "direct_value_divisor": direct_value_divisor,
                        }
                    )
                if rules:
                    return {"rules": rules}
        except Exception:
            pass
    cfg = _safe_read_json(paths["ratio_rules_json"], {})
    return cfg if isinstance(cfg, dict) else {"rules": []}

