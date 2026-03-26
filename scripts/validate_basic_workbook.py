#!/usr/bin/env python3
"""Stage 2 validator for the single master workbook.

Behavior:
- Reads the manually edited master workbook.
- Rebuilds reconciliation sheet.
- Rebuilds issue list (missing values + reconciliation failures).
- Never rebuilds templates or auto-filled statement values.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from build_basic_data_workbook import (
    build_missing_sheet,
    build_ratio_sheet,
    build_recon_sheet_from_template_rules,
    get_code_aliases,
    get_interest_debt_items,
    get_sheet_by_name_loose,
    load_ratio_alias_map,
    load_ratio_cfg,
    load_recon_rules,
    load_runtime_cfg,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_AMOUNT_UNIT = "元"
DEFAULT_ANALYSIS_TEXT_TEMPLATES = {
    "asset_abs": "{name}，{y1}年{v1}{unit}，{y2}年{v2}{unit}，{y3}年{v3}{unit}；{y2}年较{y1}年{p21}，{y3}年较{y2}年{p32}。",
    "liability_abs": "{name}，{y1}年{v1}{unit}，{y2}年{v2}{unit}，{y3}年{v3}{unit}；{y2}年较{y1}年{p21}，{y3}年较{y2}年{p32}。",
    "asset_total_struct": "资产总计仅描述绝对量变化，不做占比分析。",
    "asset_subtotal_struct": "占总资产比例：{y1}年{r2_1}%，{y2}年{r2_2}%，{y3}年{r2_3}%；{y2}较{y1}{r2_21}，{y3}较{y2}{r2_32}。",
    "asset_item_struct": "{base1_label}：{y1}年{r1_1}%，{y2}年{r1_2}%，{y3}年{r1_3}%；{y2}较{y1}{r1_21}，{y3}较{y2}{r1_32}。占总资产比例：{y1}年{r2_1}%，{y2}年{r2_2}%，{y3}年{r2_3}%；{y2}较{y1}{r2_21}，{y3}较{y2}{r2_32}。",
    "liability_total_struct": "负债合计仅描述绝对量变化，不做占比分析。",
    "liability_subtotal_struct": "占负债总计比例：{y1}年{r2_1}%，{y2}年{r2_2}%，{y3}年{r2_3}%；{y2}较{y1}{r2_21}，{y3}较{y2}{r2_32}。",
    "liability_item_struct": "{base1_label}：{y1}年{r1_1}%，{y2}年{r1_2}%，{y3}年{r1_3}%；{y2}较{y1}{r1_21}，{y3}较{y2}{r1_32}。占负债总计比例：{y1}年{r2_1}%，{y2}年{r2_2}%，{y3}年{r2_3}%；{y2}较{y1}{r2_21}，{y3}较{y2}{r2_32}。",
    # phrase policies for absolute/relative trend wording
    "scale_phrase_up": "增加{abs_pct}%",
    "scale_phrase_down": "减少{abs_pct}%",
    "scale_phrase_stable": "保持稳定（变动{abs_pct}%）",
    "struct_phrase_up": "上升{abs_pp}个百分点",
    "struct_phrase_down": "下降{abs_pp}个百分点",
    "struct_phrase_stable": "基本稳定（变动{abs_pp}个百分点）",
}
DEFAULT_ANALYSIS_TEMPLATE_UNITS = {
    "asset_abs": DEFAULT_AMOUNT_UNIT,
    "liability_abs": DEFAULT_AMOUNT_UNIT,
}
DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT = 2.0


def normalize_project_id(project_id: str) -> str:
    text = re.sub(r"[^0-9A-Za-z._-]+", "_", (project_id or "").strip())
    return text or "default_project"


def get_runtime_text(runtime_cfg: Optional[Dict[str, object]], key: str, default: str) -> str:
    if not isinstance(runtime_cfg, dict):
        return default
    val = runtime_cfg.get(key)
    txt = str(val or "").strip()
    return txt or default


def get_amount_unit(runtime_cfg: Optional[Dict[str, object]] = None, default: str = DEFAULT_AMOUNT_UNIT) -> str:
    if isinstance(runtime_cfg, dict):
        raw = runtime_cfg.get("amount_unit")
        txt = str(raw or "").strip()
        if txt:
            return txt
    return default


def _render_template(template: str, values: Dict[str, Any]) -> str:
    txt = str(template or "")

    def repl(match: re.Match) -> str:
        key = match.group(1)
        v = values.get(key)
        return str(v) if v is not None else "待补充"

    return re.sub(r"\{([A-Za-z0-9_]+)\}", repl, txt).strip()


def _pick_tpl_by_code(tpl: Dict[str, str], base_key: str, code: str) -> str:
    """Per-subject template takes priority, then fallback to base template."""
    c = str(code or "").strip()
    if c:
        v = tpl.get(f"{base_key}_{c}")
        if isinstance(v, str) and v.strip():
            return v
    return tpl.get(base_key, DEFAULT_ANALYSIS_TEXT_TEMPLATES.get(base_key, ""))


def _pick_tpl_unit_by_code(unit_map: Dict[str, str], base_key: str, code: str, default: str = DEFAULT_AMOUNT_UNIT) -> str:
    c = str(code or "").strip()
    if c:
        u = str(unit_map.get(f"{base_key}_{c}", "") or "").strip()
        if u:
            return u
    u = str(unit_map.get(base_key, "") or "").strip()
    return u or default


def load_analysis_text_templates(runtime_cfg: Optional[Dict[str, object]]) -> Dict[str, str]:
    cfg = dict(DEFAULT_ANALYSIS_TEXT_TEMPLATES)

    # 1) project config override
    if isinstance(runtime_cfg, dict):
        raw = runtime_cfg.get("analysis_text_templates")
        if isinstance(raw, dict):
            for k, v in raw.items():
                kk = str(k or "").strip()
                vv = str(v or "").strip()
                if kk and vv and "?" not in vv and "�" not in vv:
                    cfg[kk] = vv

    # 2) rulebook sheet override
    rb = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not rb.exists():
        return cfg
    try:
        wb = load_workbook(rb, data_only=True)
    except Exception:
        return cfg
    ws = get_sheet_by_name_loose(wb, "analysis_text_templates")
    if ws is None:
        ws = get_sheet_by_name_loose(wb, "分析文本模板")
    if ws is None:
        return cfg
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_key = headers.get("template_key", 1)
    c_txt = headers.get("template_text", 2)
    c_enabled = headers.get("enabled", 3)
    for r in range(2, ws.max_row + 1):
        enabled_raw = str(ws.cell(r, c_enabled).value or "1").strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        key = str(ws.cell(r, c_key).value or "").strip()
        txt = str(ws.cell(r, c_txt).value or "").strip()
        if not key or not txt or "?" in txt or "�" in txt:
            continue
        cfg[key] = txt
    return cfg


def load_analysis_template_units(runtime_cfg: Optional[Dict[str, object]]) -> Dict[str, str]:
    amount_unit = get_amount_unit(runtime_cfg)
    cfg = {
        "asset_abs": amount_unit,
        "liability_abs": amount_unit,
        "_default": amount_unit,
    }

    if isinstance(runtime_cfg, dict):
        raw = runtime_cfg.get("analysis_template_units")
        if isinstance(raw, dict):
            for k, v in raw.items():
                kk = str(k or "").strip()
                vv = str(v or "").strip()
                if kk and vv:
                    cfg[kk] = vv

    rb = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not rb.exists():
        return cfg
    try:
        wb = load_workbook(rb, data_only=True)
    except Exception:
        return cfg
    ws = get_sheet_by_name_loose(wb, "analysis_text_templates")
    if ws is None:
        ws = get_sheet_by_name_loose(wb, "分析文本模板")
    if ws is None:
        return cfg
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_key = headers.get("template_key", 1)
    c_unit = headers.get("unit")
    c_enabled = headers.get("enabled", 3)
    if not c_unit:
        return cfg
    for r in range(2, ws.max_row + 1):
        enabled_raw = str(ws.cell(r, c_enabled).value or "1").strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        key = str(ws.cell(r, c_key).value or "").strip()
        unit = str(ws.cell(r, c_unit).value or "").strip()
        if key and unit:
            cfg[key] = unit
    return cfg


def load_analysis_thresholds(runtime_cfg: Optional[Dict[str, object]]) -> Dict[str, Any]:
    """
    Two-level threshold policy:
    1) subject threshold (optional)
    2) global threshold (required, fallback to runtime config/code default)
    """
    global_scale = DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT
    global_struct = DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT
    if isinstance(runtime_cfg, dict):
        try:
            global_scale = float(runtime_cfg.get("analysis_stable_threshold_pct", DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT))
        except Exception:
            global_scale = DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT
        try:
            global_struct = float(runtime_cfg.get("analysis_struct_stable_pp", global_scale))
        except Exception:
            global_struct = global_scale
    out: Dict[str, Any] = {
        "global_scale_pct": global_scale,
        "global_struct_pp": global_struct,
        "subject_scale_pct": {},
        "subject_struct_pp": {},
        # backward compatibility aliases
        "global_pct": global_scale,
        "subject_pct": {},
    }

    rb = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not rb.exists():
        return out
    try:
        wb = load_workbook(rb, data_only=True)
    except Exception:
        return out
    ws = get_sheet_by_name_loose(wb, "analysis_thresholds")
    if ws is None:
        ws = get_sheet_by_name_loose(wb, "分析阈值配置")
    if ws is None:
        return out

    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_scope = headers.get("scope", 1)
    c_subject = headers.get("subject_code", 2)
    c_scale = headers.get("scale_stable_pct", headers.get("stable_threshold_pct", 3))
    c_struct = headers.get("struct_stable_pp", c_scale)
    c_enabled = headers.get("enabled", 4)
    for r in range(2, ws.max_row + 1):
        enabled_raw = str(ws.cell(r, c_enabled).value or "1").strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        scope = str(ws.cell(r, c_scope).value or "").strip().lower()
        subject = str(ws.cell(r, c_subject).value or "").strip().upper()
        try:
            scale_v = float(ws.cell(r, c_scale).value)
        except Exception:
            continue
        try:
            struct_v = float(ws.cell(r, c_struct).value)
        except Exception:
            struct_v = scale_v
        if scope in {"global", "全局"}:
            out["global_scale_pct"] = scale_v
            out["global_struct_pp"] = struct_v
            out["global_pct"] = scale_v
        elif scope in {"subject", "科目"} and re.match(r"^BS\d{3}$", subject):
            out["subject_scale_pct"][subject] = scale_v
            out["subject_struct_pp"][subject] = struct_v
            out["subject_pct"][subject] = scale_v
    return out


def effective_stable_thresholds(code: str, threshold_cfg: Dict[str, Any]) -> Tuple[float, float]:
    code_u = str(code or "").strip().upper()
    by_scale = threshold_cfg.get("subject_scale_pct", {}) if isinstance(threshold_cfg, dict) else {}
    by_struct = threshold_cfg.get("subject_struct_pp", {}) if isinstance(threshold_cfg, dict) else {}
    if (not by_scale) and isinstance(threshold_cfg, dict):
        by_scale = threshold_cfg.get("subject_pct", {}) or {}
    if code_u and isinstance(by_scale, dict) and code_u in by_scale:
        try:
            s_scale = float(by_scale[code_u])
        except Exception:
            s_scale = float((threshold_cfg or {}).get("global_scale_pct", (threshold_cfg or {}).get("global_pct", DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT)))
        try:
            s_struct = float(by_struct.get(code_u, s_scale)) if isinstance(by_struct, dict) else s_scale
        except Exception:
            s_struct = s_scale
        return s_scale, s_struct
    try:
        g_scale = float((threshold_cfg or {}).get("global_scale_pct", (threshold_cfg or {}).get("global_pct", DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT)))
    except Exception:
        g_scale = DEFAULT_ANALYSIS_STABLE_THRESHOLD_PCT
    try:
        g_struct = float((threshold_cfg or {}).get("global_struct_pp", g_scale))
    except Exception:
        g_struct = g_scale
    return g_scale, g_struct


def parse_years(ws, year_row: int = 3, year_start_col: int = 3) -> List[str]:
    years: List[str] = []
    current_year = dt.date.today().year
    for c in range(max(1, year_start_col), ws.max_column + 1):
        text = str(ws.cell(max(1, year_row), c).value or "")
        m = re.search(r"(20\d{2})", text)
        if not m:
            continue
        y = int(m.group(1))
        if 2000 <= y <= current_year + 1:
            years.append(str(y))
    uniq: List[str] = []
    for y in years:
        if y not in uniq:
            uniq.append(y)
    return uniq


def normalize_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_statement_values(ws, years: List[str]) -> Dict[Tuple[str, str], Optional[float]]:
    values: Dict[Tuple[str, str], Optional[float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0] or "").strip()
        if not re.match(r"^(BS|IS|CF)\d+$", code):
            continue
        name = row[1]
        if not name:
            continue
        for i, y in enumerate(years):
            val = normalize_num(row[2 + i])
            values[(str(name), y)] = val
    return values


def collect_missing_rows(ws, years: List[str]) -> List[dict]:
    rows: List[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0] or "").strip()
        if not re.match(r"^(BS|IS|CF)\d+$", code):
            continue
        name = row[1]
        if not name:
            continue
        for i, y in enumerate(years):
            val = normalize_num(row[2 + i])
            if val is None:
                rows.append(
                    {
                        "报表": ws.title,
                        "科目编码": code,
                        "科目名称": name,
                        "期间": y,
                        "状态": "待补充",
                        "说明": "主文件该字段为空，请手工补录或修正",
                    }
                )
    return rows


def collect_recon_issues(recon_ws) -> List[dict]:
    issues: List[dict] = []
    if recon_ws is None:
        return issues
    for row in recon_ws.iter_rows(min_row=2, values_only=True):
        rule_id, desc, period, left, right, diff, result = row
        if result == "是":
            continue
        issues.append(
            {
                "报表": recon_ws.title,
                "科目编码": rule_id,
                "科目名称": desc,
                "期间": period,
                "状态": result or "待补充",
                "说明": f"左值={left}, 右值={right}, 差异={diff}",
            }
        )
    return issues


def create_interest_debt_sheet_if_missing(
    wb,
    years: List[str],
    bs_values: Dict[Tuple[str, str], Optional[float]],
    runtime_cfg: Optional[Dict[str, object]] = None,
) -> None:
    debt_sheet_name = get_runtime_text(runtime_cfg, "interest_debt_sheet_name", "有息负债明细")
    total_label = get_runtime_text(runtime_cfg, "interest_debt_total_label", "有息负债合计")

    if get_sheet_by_name_loose(wb, debt_sheet_name) is not None:
        return

    ws = wb.create_sheet(debt_sheet_name)
    amount_unit = get_amount_unit(runtime_cfg)
    ws.append(["期间", "子项", f"值({amount_unit})", "说明"])
    items = get_interest_debt_items(runtime_cfg)
    for y in years:
        total = 0.0
        has_val = False
        for item in items:
            v = bs_values.get((item, y))
            if v is not None:
                total += v
                has_val = True
            ws.append([y, item, v, "根据资产负债表自动初始化"])
        ws.append([y, total_label, round(total, 2) if has_val else None, "根据资产负债表自动初始化"])


def extract_interest_debt_values(
    wb,
    years: List[str],
    runtime_cfg: Optional[Dict[str, object]] = None,
) -> Tuple[Dict[Tuple[str, str], Optional[float]], List[dict]]:
    values: Dict[Tuple[str, str], Optional[float]] = {}
    missing_rows: List[dict] = []

    debt_sheet_name = get_runtime_text(runtime_cfg, "interest_debt_sheet_name", "有息负债明细")
    total_label = get_runtime_text(runtime_cfg, "interest_debt_total_label", "有息负债合计")

    ws = get_sheet_by_name_loose(wb, debt_sheet_name)
    if ws is None:
        return values, missing_rows

    for row in ws.iter_rows(min_row=2, values_only=True):
        period, item, val, _ = row
        if str(item or "").strip() != total_label:
            continue
        y = str(period or "").strip()
        if y not in years:
            continue
        n = normalize_num(val)
        values[(total_label, y)] = n
        if n is None:
            missing_rows.append(
                {
                    "报表": ws.title,
                    "科目编码": "DEBT999",
                    "科目名称": total_label,
                    "期间": y,
                    "状态": "待补充",
                    "说明": f"{total_label}为空，请补录或修正",
                }
            )
    return values, missing_rows


def load_ratio_candidates(project_id: str) -> List[Path]:
    manifest_path = PROJECT_ROOT / "data" / project_id / "input_discovery.json"
    if not manifest_path.exists():
        return []
    try:
        cfg = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    selected = cfg.get("selected_workpaper")
    if not selected:
        return []
    p = Path(selected)
    return [p] if p.exists() else []


def _bs_code_num(code: str) -> Optional[int]:
    m = re.match(r"^BS(\d+)$", str(code or "").strip())
    if not m:
        return None
    try:
        return int(m.group(1))
    except ValueError:
        return None


def _safe_div(numer: Optional[float], denom: Optional[float]) -> Optional[float]:
    if numer is None or denom in (None, 0):
        return None
    return numer / denom


def _pct_change(curr: Optional[float], prev: Optional[float]) -> Optional[float]:
    if curr is None or prev in (None, 0):
        return None
    return (curr - prev) / abs(prev) * 100.0


def _abs_change(curr: Optional[float], prev: Optional[float]) -> Optional[float]:
    if curr is None or prev is None:
        return None
    return curr - prev


def _change_label(pct_value: Optional[float], stable_pct: float = 2.0) -> str:
    if pct_value is None:
        return "待补充"
    if pct_value > stable_pct:
        return "增加"
    if pct_value < -stable_pct:
        return "减少"
    return "保持稳定"


def _share_label(delta_pct: Optional[float], stable_pct: float = 2.0) -> str:
    if delta_pct is None:
        return "待补充"
    if delta_pct > stable_pct:
        return "上升"
    if delta_pct < -stable_pct:
        return "下降"
    return "基本稳定"


def _fmt_num(v: Optional[float], digits: int = 2) -> str:
    if v is None:
        return "待补充"
    return f"{v:.{digits}f}"


def _scale_phrase(pct_value: Optional[float], stable_pct: float = 2.0, text_templates: Optional[Dict[str, str]] = None) -> str:
    tpl = dict(DEFAULT_ANALYSIS_TEXT_TEMPLATES)
    tpl.update(text_templates or {})
    if pct_value is None:
        return "待补充"
    if pct_value > stable_pct:
        return _render_template(tpl["scale_phrase_up"], {"abs_pct": _fmt_num(abs(pct_value))})
    if pct_value < -stable_pct:
        return _render_template(tpl["scale_phrase_down"], {"abs_pct": _fmt_num(abs(pct_value))})
    return _render_template(tpl["scale_phrase_stable"], {"abs_pct": _fmt_num(abs(pct_value))})


def _struct_phrase(delta_pct: Optional[float], stable_pct: float = 2.0, text_templates: Optional[Dict[str, str]] = None) -> str:
    tpl = dict(DEFAULT_ANALYSIS_TEXT_TEMPLATES)
    tpl.update(text_templates or {})
    if delta_pct is None:
        return "待补充"
    if delta_pct > stable_pct:
        return _render_template(tpl["struct_phrase_up"], {"abs_pp": _fmt_num(abs(delta_pct))})
    if delta_pct < -stable_pct:
        return _render_template(tpl["struct_phrase_down"], {"abs_pp": _fmt_num(abs(delta_pct))})
    return _render_template(tpl["struct_phrase_stable"], {"abs_pp": _fmt_num(abs(delta_pct))})


def _replace_sheet(wb, sheet_name: str):
    old = get_sheet_by_name_loose(wb, sheet_name)
    if old is not None:
        del wb[old.title]
    return wb.create_sheet(sheet_name)


def build_asset_analysis_sheets(
    wb,
    bs_ws,
    years: List[str],
    stable_threshold_pct: float = 2.0,
    current_total_code: str = "BS027",
    noncurrent_total_code: str = "BS056",
    asset_total_code: str = "BS057",
    text_templates: Optional[Dict[str, str]] = None,
    text_template_units: Optional[Dict[str, str]] = None,
    threshold_cfg: Optional[Dict[str, Any]] = None,
) -> None:
    if len(years) < 3:
        return

    y1, y2, y3 = years[0], years[1], years[2]
    amount_unit = str((text_template_units or {}).get("_default", DEFAULT_AMOUNT_UNIT) if isinstance(text_template_units, dict) else DEFAULT_AMOUNT_UNIT)
    tpl = dict(DEFAULT_ANALYSIS_TEXT_TEMPLATES)
    tpl.update(text_templates or {})
    tpl_units = dict(DEFAULT_ANALYSIS_TEMPLATE_UNITS)
    tpl_units.update(text_template_units or {})

    rows: List[dict] = []
    code_year_values: Dict[str, Dict[str, Optional[float]]] = {}

    for r in range(2, bs_ws.max_row + 1):
        code = str(bs_ws.cell(r, 1).value or "").strip()
        name = str(bs_ws.cell(r, 2).value or "").strip().rstrip("：:").strip()
        n = _bs_code_num(code)
        if n is None or n > 57:
            continue
        vals = {
            y1: normalize_num(bs_ws.cell(r, 3).value),
            y2: normalize_num(bs_ws.cell(r, 4).value),
            y3: normalize_num(bs_ws.cell(r, 5).value),
        }
        # Skip pure section rows.
        if (name.endswith("：") or name.endswith(":")) and all(v is None for v in vals.values()):
            continue
        code_year_values[code] = vals
        rows.append({"code": code, "code_num": n, "name": name, "values": vals})

    current_totals = code_year_values.get(current_total_code, {})
    noncurrent_totals = code_year_values.get(noncurrent_total_code, {})
    asset_totals = code_year_values.get(asset_total_code, {})

    ws_scale = _replace_sheet(wb, "分析_规模变化_资产类")
    ws_scale.append(
        [
            "科目编码",
            "科目名称",
            f"{y1}年({amount_unit})",
            f"{y2}年({amount_unit})",
            f"{y3}年({amount_unit})",
            f"{y2}较{y1}变动额({amount_unit})",
            f"{y2}较{y1}变动率(%)",
            f"{y2}较{y1}变化判断",
            f"{y3}较{y2}变动额({amount_unit})",
            f"{y3}较{y2}变动率(%)",
            f"{y3}较{y2}变化判断",
            "定量描述_绝对量",
        ]
    )

    ws_struct = _replace_sheet(wb, "分析_结构占比_资产类")
    ws_struct.append(
        [
            "科目编码",
            "科目名称",
            "结构口径1",
            f"{y1}年口径1占比(%)",
            f"{y2}年口径1占比(%)",
            f"{y3}年口径1占比(%)",
            f"{y2}较{y1}口径1变动(百分点)",
            f"{y3}较{y2}口径1变动(百分点)",
            f"{y1}年占总资产比例(%)",
            f"{y2}年占总资产比例(%)",
            f"{y3}年占总资产比例(%)",
            f"{y2}较{y1}占总资产变动(百分点)",
            f"{y3}较{y2}占总资产变动(百分点)",
            "结构变化判断",
            "定量描述_相对量",
        ]
    )

    for item in rows:
        code = item["code"]
        name = item["name"]
        n = item["code_num"]
        v1 = item["values"].get(y1)
        v2 = item["values"].get(y2)
        v3 = item["values"].get(y3)

        d21 = _abs_change(v2, v1)
        p21 = _pct_change(v2, v1)
        d32 = _abs_change(v3, v2)
        p32 = _pct_change(v3, v2)
        eff_scale, eff_struct = effective_stable_thresholds(
            code, threshold_cfg or {"global_scale_pct": stable_threshold_pct, "global_struct_pp": stable_threshold_pct}
        )
        l21 = _change_label(p21, stable_pct=eff_scale)
        l32 = _change_label(p32, stable_pct=eff_scale)

        abs_tpl = _pick_tpl_by_code(tpl, "asset_abs", code)
        abs_unit = _pick_tpl_unit_by_code(tpl_units, "asset_abs", code, default=amount_unit)
        abs_text = _render_template(
            abs_tpl,
            {
                "name": name,
                "y1": y1,
                "y2": y2,
                "y3": y3,
                "v1": _fmt_num(v1),
                "v2": _fmt_num(v2),
                "v3": _fmt_num(v3),
                "p21": _scale_phrase(p21, stable_pct=eff_scale, text_templates=tpl),
                "p32": _scale_phrase(p32, stable_pct=eff_scale, text_templates=tpl),
                "unit": abs_unit,
            },
        )

        ws_scale.append(
            [
                code,
                name,
                v1,
                v2,
                v3,
                d21,
                p21,
                l21,
                d32,
                p32,
                l32,
                abs_text,
            ]
        )

        # Relative structure: asset total row only keeps absolute changes (no ratio narrative).
        ratio_base1_name = ""
        r1_1 = r1_2 = r1_3 = None
        r1_d21 = r1_d32 = None
        rt_1 = _safe_div(v1, asset_totals.get(y1))
        rt_2 = _safe_div(v2, asset_totals.get(y2))
        rt_3 = _safe_div(v3, asset_totals.get(y3))
        rt_d21 = (rt_2 - rt_1) * 100 if rt_2 is not None and rt_1 is not None else None
        rt_d32 = (rt_3 - rt_2) * 100 if rt_3 is not None and rt_2 is not None else None
        rt_d31 = (rt_3 - rt_1) * 100 if rt_3 is not None and rt_1 is not None else None
        struct_label = _share_label(rt_d32, stable_pct=eff_struct)
        struct_text = ""

        if code == asset_total_code:
            struct_label = "不适用"
            struct_tpl = _pick_tpl_by_code(tpl, "asset_struct", code) or tpl["asset_total_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["asset_total_struct"]
            struct_text = _render_template(struct_tpl, {})
        elif code in {current_total_code, noncurrent_total_code}:
            struct_label = _share_label(rt_d32, stable_pct=eff_struct)
            struct_tpl = _pick_tpl_by_code(tpl, "asset_struct", code) or tpl["asset_subtotal_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["asset_subtotal_struct"]
            struct_text = _render_template(
                struct_tpl,
                {
                    "y1": y1,
                    "y2": y2,
                    "y3": y3,
                    "r2_1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "r2_2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "r2_3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "r2_21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r2_32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r2_31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                    # Legacy aliases (backward compatible with existing templates).
                    "rt1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "rt2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "rt3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "rt21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "rt32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "rt31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                },
            )
        else:
            if n <= 27:
                ratio_base1_name = "占流动资产比例"
                b1_1 = current_totals.get(y1)
                b1_2 = current_totals.get(y2)
                b1_3 = current_totals.get(y3)
            else:
                ratio_base1_name = "占非流动资产比例"
                b1_1 = noncurrent_totals.get(y1)
                b1_2 = noncurrent_totals.get(y2)
                b1_3 = noncurrent_totals.get(y3)

            r1_1 = _safe_div(v1, b1_1)
            r1_2 = _safe_div(v2, b1_2)
            r1_3 = _safe_div(v3, b1_3)
            r1_d21 = (r1_2 - r1_1) * 100 if r1_2 is not None and r1_1 is not None else None
            r1_d32 = (r1_3 - r1_2) * 100 if r1_3 is not None and r1_2 is not None else None
            r1_d31 = (r1_3 - r1_1) * 100 if r1_3 is not None and r1_1 is not None else None
            struct_label = _share_label(r1_d32, stable_pct=eff_struct)
            struct_tpl = _pick_tpl_by_code(tpl, "asset_struct", code) or tpl["asset_item_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["asset_item_struct"]
            struct_text = _render_template(
                struct_tpl,
                {
                    "base1_label": ratio_base1_name,
                    "y1": y1,
                    "y2": y2,
                    "y3": y3,
                    "r1_1": _fmt_num(r1_1 * 100 if r1_1 is not None else None),
                    "r1_2": _fmt_num(r1_2 * 100 if r1_2 is not None else None),
                    "r1_3": _fmt_num(r1_3 * 100 if r1_3 is not None else None),
                    "r1_21": _struct_phrase(r1_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r1_32": _struct_phrase(r1_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r1_31": _struct_phrase(r1_d31, stable_pct=eff_struct, text_templates=tpl),
                    "r2_1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "r2_2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "r2_3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "r2_21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r2_32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r2_31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                    # Legacy aliases (backward compatible with existing templates).
                    "r11": _fmt_num(r1_1 * 100 if r1_1 is not None else None),
                    "r12": _fmt_num(r1_2 * 100 if r1_2 is not None else None),
                    "r13": _fmt_num(r1_3 * 100 if r1_3 is not None else None),
                    "r121": _struct_phrase(r1_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r132": _struct_phrase(r1_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r131": _struct_phrase(r1_d31, stable_pct=eff_struct, text_templates=tpl),
                    "rt1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "rt2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "rt3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "rt21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "rt32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "rt31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                },
            )

        ws_struct.append(
            [
                code,
                name,
                ratio_base1_name,
                r1_1 * 100 if r1_1 is not None else None,
                r1_2 * 100 if r1_2 is not None else None,
                r1_3 * 100 if r1_3 is not None else None,
                r1_d21,
                r1_d32,
                rt_1 * 100 if rt_1 is not None else None,
                rt_2 * 100 if rt_2 is not None else None,
                rt_3 * 100 if rt_3 is not None else None,
                rt_d21,
                rt_d32,
                struct_label,
                struct_text,
            ]
        )


def build_liability_analysis_sheets(
    wb,
    bs_ws,
    years: List[str],
    stable_threshold_pct: float = 2.0,
    current_total_code: str = "BS089",
    noncurrent_total_code: str = "BS102",
    liability_total_code: str = "BS103",
    text_templates: Optional[Dict[str, str]] = None,
    text_template_units: Optional[Dict[str, str]] = None,
    threshold_cfg: Optional[Dict[str, Any]] = None,
) -> None:
    if len(years) < 3:
        return

    y1, y2, y3 = years[0], years[1], years[2]
    amount_unit = str((text_template_units or {}).get("_default", DEFAULT_AMOUNT_UNIT) if isinstance(text_template_units, dict) else DEFAULT_AMOUNT_UNIT)
    tpl = dict(DEFAULT_ANALYSIS_TEXT_TEMPLATES)
    tpl.update(text_templates or {})
    tpl_units = dict(DEFAULT_ANALYSIS_TEMPLATE_UNITS)
    tpl_units.update(text_template_units or {})

    rows: List[dict] = []
    code_year_values: Dict[str, Dict[str, Optional[float]]] = {}

    for r in range(2, bs_ws.max_row + 1):
        code = str(bs_ws.cell(r, 1).value or "").strip()
        name = str(bs_ws.cell(r, 2).value or "").strip().rstrip("：:").strip()
        n = _bs_code_num(code)
        if n is None or n < 58 or n > 103:
            continue
        vals = {
            y1: normalize_num(bs_ws.cell(r, 3).value),
            y2: normalize_num(bs_ws.cell(r, 4).value),
            y3: normalize_num(bs_ws.cell(r, 5).value),
        }
        if (name.endswith("：") or name.endswith(":")) and all(v is None for v in vals.values()):
            continue
        code_year_values[code] = vals
        rows.append({"code": code, "code_num": n, "name": name, "values": vals})

    current_totals = code_year_values.get(current_total_code, {})
    noncurrent_totals = code_year_values.get(noncurrent_total_code, {})
    liability_totals = code_year_values.get(liability_total_code, {})

    ws_scale = _replace_sheet(wb, "分析_规模变化_负债类")
    ws_scale.append(
        [
            "科目编码",
            "科目名称",
            f"{y1}年({amount_unit})",
            f"{y2}年({amount_unit})",
            f"{y3}年({amount_unit})",
            f"{y2}较{y1}变动额({amount_unit})",
            f"{y2}较{y1}变动率(%)",
            f"{y2}较{y1}变化判断",
            f"{y3}较{y2}变动额({amount_unit})",
            f"{y3}较{y2}变动率(%)",
            f"{y3}较{y2}变化判断",
            "定量描述_绝对量",
        ]
    )

    ws_struct = _replace_sheet(wb, "分析_结构占比_负债类")
    ws_struct.append(
        [
            "科目编码",
            "科目名称",
            "结构口径1",
            f"{y1}年口径1占比(%)",
            f"{y2}年口径1占比(%)",
            f"{y3}年口径1占比(%)",
            f"{y2}较{y1}口径1变动(百分点)",
            f"{y3}较{y2}口径1变动(百分点)",
            f"{y1}年占负债总计比例(%)",
            f"{y2}年占负债总计比例(%)",
            f"{y3}年占负债总计比例(%)",
            f"{y2}较{y1}占负债总计变动(百分点)",
            f"{y3}较{y2}占负债总计变动(百分点)",
            "结构变化判断",
            "定量描述_相对量",
        ]
    )

    for item in rows:
        code = item["code"]
        name = item["name"]
        n = item["code_num"]
        v1 = item["values"].get(y1)
        v2 = item["values"].get(y2)
        v3 = item["values"].get(y3)

        d21 = _abs_change(v2, v1)
        p21 = _pct_change(v2, v1)
        d32 = _abs_change(v3, v2)
        p32 = _pct_change(v3, v2)
        eff_scale, eff_struct = effective_stable_thresholds(
            code, threshold_cfg or {"global_scale_pct": stable_threshold_pct, "global_struct_pp": stable_threshold_pct}
        )
        l21 = _change_label(p21, stable_pct=eff_scale)
        l32 = _change_label(p32, stable_pct=eff_scale)

        abs_tpl = _pick_tpl_by_code(tpl, "liability_abs", code)
        abs_unit = _pick_tpl_unit_by_code(tpl_units, "liability_abs", code, default=amount_unit)
        abs_text = _render_template(
            abs_tpl,
            {
                "name": name,
                "y1": y1,
                "y2": y2,
                "y3": y3,
                "v1": _fmt_num(v1),
                "v2": _fmt_num(v2),
                "v3": _fmt_num(v3),
                "p21": _scale_phrase(p21, stable_pct=eff_scale, text_templates=tpl),
                "p32": _scale_phrase(p32, stable_pct=eff_scale, text_templates=tpl),
                "unit": abs_unit,
            },
        )

        ws_scale.append(
            [
                code,
                name,
                v1,
                v2,
                v3,
                d21,
                p21,
                l21,
                d32,
                p32,
                l32,
                abs_text,
            ]
        )

        ratio_base1_name = ""
        r1_1 = r1_2 = r1_3 = None
        r1_d21 = r1_d32 = None
        rt_1 = _safe_div(v1, liability_totals.get(y1))
        rt_2 = _safe_div(v2, liability_totals.get(y2))
        rt_3 = _safe_div(v3, liability_totals.get(y3))
        rt_d21 = (rt_2 - rt_1) * 100 if rt_2 is not None and rt_1 is not None else None
        rt_d32 = (rt_3 - rt_2) * 100 if rt_3 is not None and rt_2 is not None else None
        rt_d31 = (rt_3 - rt_1) * 100 if rt_3 is not None and rt_1 is not None else None
        struct_label = _share_label(rt_d32, stable_pct=eff_struct)
        struct_text = ""

        if code == liability_total_code:
            struct_label = "不适用"
            struct_tpl = _pick_tpl_by_code(tpl, "liability_struct", code) or tpl["liability_total_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["liability_total_struct"]
            struct_text = _render_template(struct_tpl, {})
        elif code in {current_total_code, noncurrent_total_code}:
            struct_tpl = _pick_tpl_by_code(tpl, "liability_struct", code) or tpl["liability_subtotal_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["liability_subtotal_struct"]
            struct_text = _render_template(
                struct_tpl,
                {
                    "y1": y1,
                    "y2": y2,
                    "y3": y3,
                    "r2_1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "r2_2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "r2_3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "r2_21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r2_32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r2_31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                    # Legacy aliases (backward compatible with existing templates).
                    "rt1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "rt2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "rt3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "rt21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "rt32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "rt31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                },
            )
        else:
            if n <= 89:
                ratio_base1_name = "占流动负债比例"
                b1_1 = current_totals.get(y1)
                b1_2 = current_totals.get(y2)
                b1_3 = current_totals.get(y3)
            else:
                ratio_base1_name = "占非流动负债比例"
                b1_1 = noncurrent_totals.get(y1)
                b1_2 = noncurrent_totals.get(y2)
                b1_3 = noncurrent_totals.get(y3)

            r1_1 = _safe_div(v1, b1_1)
            r1_2 = _safe_div(v2, b1_2)
            r1_3 = _safe_div(v3, b1_3)
            r1_d21 = (r1_2 - r1_1) * 100 if r1_2 is not None and r1_1 is not None else None
            r1_d32 = (r1_3 - r1_2) * 100 if r1_3 is not None and r1_2 is not None else None
            r1_d31 = (r1_3 - r1_1) * 100 if r1_3 is not None and r1_1 is not None else None
            struct_label = _share_label(r1_d32, stable_pct=eff_struct)
            struct_tpl = _pick_tpl_by_code(tpl, "liability_struct", code) or tpl["liability_item_struct"]
            if not struct_tpl.strip():
                struct_tpl = tpl["liability_item_struct"]
            struct_text = _render_template(
                struct_tpl,
                {
                    "base1_label": ratio_base1_name,
                    "y1": y1,
                    "y2": y2,
                    "y3": y3,
                    "r1_1": _fmt_num(r1_1 * 100 if r1_1 is not None else None),
                    "r1_2": _fmt_num(r1_2 * 100 if r1_2 is not None else None),
                    "r1_3": _fmt_num(r1_3 * 100 if r1_3 is not None else None),
                    "r1_21": _struct_phrase(r1_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r1_32": _struct_phrase(r1_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r1_31": _struct_phrase(r1_d31, stable_pct=eff_struct, text_templates=tpl),
                    "r2_1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "r2_2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "r2_3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "r2_21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r2_32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r2_31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                    # Legacy aliases (backward compatible with existing templates).
                    "r11": _fmt_num(r1_1 * 100 if r1_1 is not None else None),
                    "r12": _fmt_num(r1_2 * 100 if r1_2 is not None else None),
                    "r13": _fmt_num(r1_3 * 100 if r1_3 is not None else None),
                    "r121": _struct_phrase(r1_d21, stable_pct=eff_struct, text_templates=tpl),
                    "r132": _struct_phrase(r1_d32, stable_pct=eff_struct, text_templates=tpl),
                    "r131": _struct_phrase(r1_d31, stable_pct=eff_struct, text_templates=tpl),
                    "rt1": _fmt_num(rt_1 * 100 if rt_1 is not None else None),
                    "rt2": _fmt_num(rt_2 * 100 if rt_2 is not None else None),
                    "rt3": _fmt_num(rt_3 * 100 if rt_3 is not None else None),
                    "rt21": _struct_phrase(rt_d21, stable_pct=eff_struct, text_templates=tpl),
                    "rt32": _struct_phrase(rt_d32, stable_pct=eff_struct, text_templates=tpl),
                    "rt31": _struct_phrase(rt_d31, stable_pct=eff_struct, text_templates=tpl),
                },
            )

        ws_struct.append(
            [
                code,
                name,
                ratio_base1_name,
                r1_1 * 100 if r1_1 is not None else None,
                r1_2 * 100 if r1_2 is not None else None,
                r1_3 * 100 if r1_3 is not None else None,
                r1_d21,
                r1_d32,
                rt_1 * 100 if rt_1 is not None else None,
                rt_2 * 100 if rt_2 is not None else None,
                rt_3 * 100 if rt_3 is not None else None,
                rt_d21,
                rt_d32,
                struct_label,
                struct_text,
            ]
        )


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate single master workbook (stage 2)")
    parser.add_argument(
        "--project-id",
        default="default_project",
        help="Project identifier used for output isolation (default: default_project)",
    )
    parser.add_argument(
        "--master-workbook",
        default="",
        help="Master workbook path override; default is outputs/<project_id>/<project_id>_项目主文件.xlsx",
    )
    args = parser.parse_args()

    runtime_cfg = load_runtime_cfg()
    project_id = normalize_project_id(args.project_id)

    if args.master_workbook:
        master_workbook = Path(args.master_workbook)
        if not master_workbook.is_absolute():
            master_workbook = PROJECT_ROOT / master_workbook
    else:
        master_workbook = PROJECT_ROOT / "outputs" / project_id / f"{project_id}_项目主文件.xlsx"

    if not master_workbook.exists():
        print(f"Master workbook not found: {master_workbook}")
        print("Please run stage 1 first: scripts/init_basic_workbook.py")
        return 1

    wb = load_workbook(master_workbook)
    bs_sheet_candidates = runtime_cfg.get("bs_sheet_candidates", ["资产负债表", "资产负债表 "])
    bs_ws = None
    for sname in bs_sheet_candidates:
        bs_ws = get_sheet_by_name_loose(wb, str(sname))
        if bs_ws is not None:
            break
    if bs_ws is None:
        print("Master workbook is invalid: missing 资产负债表 sheet")
        return 1

    year_row = int(runtime_cfg.get("year_header_row", 3))
    year_start_col = int(runtime_cfg.get("year_start_col", 3))
    years = parse_years(bs_ws, year_row=year_row, year_start_col=year_start_col)
    if not years:
        print("Cannot detect years from 资产负债表 header")
        return 1

    all_values: Dict[str, Dict[Tuple[str, str], Optional[float]]] = {}
    missing_rows: List[dict] = []

    statement_sheets = runtime_cfg.get("statement_sheets", [])
    visited_statement_titles = set()
    for sname in statement_sheets:
        ws = get_sheet_by_name_loose(wb, str(sname))
        if ws is None:
            continue
        if ws.title in visited_statement_titles:
            continue
        visited_statement_titles.add(ws.title)
        canonical = ws.title.strip()
        all_values[canonical] = extract_statement_values(ws, years)
        missing_rows.extend(collect_missing_rows(ws, years))

    create_interest_debt_sheet_if_missing(wb, years, all_values.get("资产负债表", {}), runtime_cfg=runtime_cfg)
    debt_values, debt_missing = extract_interest_debt_values(wb, years, runtime_cfg=runtime_cfg)
    if debt_values:
        all_values[get_runtime_text(runtime_cfg, "interest_debt_sheet_name", "有息负债明细")] = debt_values
    missing_rows.extend(debt_missing)

    delete_sheet_names = runtime_cfg.get("rebuild_sheets", []) + runtime_cfg.get("drop_sheets", [])
    for name in delete_sheet_names:
        ws = get_sheet_by_name_loose(wb, str(name))
        if ws is not None:
            del wb[ws.title]

    recon_rules = load_recon_rules()
    ratio_alias_map = load_ratio_alias_map()
    ratio_cfg = load_ratio_cfg()
    tolerance_abs = float(runtime_cfg.get("recon_tolerance_abs", 1.0))
    code_aliases = get_code_aliases(runtime_cfg)

    build_recon_sheet_from_template_rules(wb, years, recon_rules, tolerance_abs, code_aliases=code_aliases)
    ratio_candidates = load_ratio_candidates(project_id)
    build_ratio_sheet(
        wb,
        years,
        all_values,
        ratio_cfg,
        ratio_candidates,
        ratio_alias_map,
        runtime_cfg=runtime_cfg,
    )

    recon_ws = get_sheet_by_name_loose(wb, "勾稽校验")
    missing_rows.extend(collect_recon_issues(recon_ws))
    build_missing_sheet(wb, missing_rows)
    threshold_cfg = load_analysis_thresholds(runtime_cfg)
    stable_threshold_pct = float(
        threshold_cfg.get("global_scale_pct", threshold_cfg.get("global_pct", runtime_cfg.get("analysis_stable_threshold_pct", 2.0)))
    )
    analysis_text_templates = load_analysis_text_templates(runtime_cfg)
    analysis_template_units = load_analysis_template_units(runtime_cfg)
    build_asset_analysis_sheets(
        wb,
        bs_ws,
        years,
        stable_threshold_pct=stable_threshold_pct,
        current_total_code=str(runtime_cfg.get("asset_current_total_code", "BS027")),
        noncurrent_total_code=str(runtime_cfg.get("asset_noncurrent_total_code", "BS056")),
        asset_total_code=str(runtime_cfg.get("asset_total_code", "BS057")),
        text_templates=analysis_text_templates,
        text_template_units=analysis_template_units,
        threshold_cfg=threshold_cfg,
    )
    build_liability_analysis_sheets(
        wb,
        bs_ws,
        years,
        stable_threshold_pct=stable_threshold_pct,
        current_total_code=str(runtime_cfg.get("liability_current_total_code", "BS089")),
        noncurrent_total_code=str(runtime_cfg.get("liability_noncurrent_total_code", "BS102")),
        liability_total_code=str(runtime_cfg.get("liability_total_code", "BS103")),
        text_templates=analysis_text_templates,
        text_template_units=analysis_template_units,
        threshold_cfg=threshold_cfg,
    )

    first_order = runtime_cfg.get("sheet_order", [])
    if first_order:
        wb._sheets.sort(key=lambda ws: first_order.index(ws.title) if ws.title in first_order else 100)

    wb.save(master_workbook)

    recon_ws = get_sheet_by_name_loose(wb, "勾稽校验")
    if recon_ws is None:
        print("Validation failed: missing 勾稽校验 sheet after rebuild")
        return 1

    recon_yes = recon_no = recon_pending = 0
    for row in recon_ws.iter_rows(min_row=2, values_only=True):
        result = str(row[6] or "").strip()
        if result == "是":
            recon_yes += 1
        elif result == "否":
            recon_no += 1
        else:
            recon_pending += 1

    print(f"Validated: {master_workbook}")
    print(f"Years: {', '.join(years)}")
    print(f"Recon summary: 是={recon_yes}, 否={recon_no}, 待补充={recon_pending}")
    print(f"Issue rows: {len(missing_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
