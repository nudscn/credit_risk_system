#!/usr/bin/env python3
"""Build standards-first basic financial data workbook (V1).

Output: a single xlsx containing:
- 5 statement sheets
- per-subject detail sheets
- reconciliation sheet
- ratio sheet
- missing/difference sheet

Current data adapters (structured first):
1) Workpaper Excel (分析底稿-十堰城运.xlsx)
2) Balance sheet Excel (25十运04[258597.SH]-资产负债表.xlsx)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from rule_loader import load_ratio_rules, load_runtime_controls, load_workbook_rules

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_WORKBOOK_PATH = PROJECT_ROOT / "config" / "主文件模板.xlsx"
OUTPUT_FILENAME = "项目主文件.xlsx"
RECON_RULE_CONFIG_PATH = PROJECT_ROOT / "config" / "recon_rules_workbook_v1.json"
RATIO_RULE_CONFIG_PATH = PROJECT_ROOT / "config" / "financial_ratio_rules_v1.json"
WORKBOOK_RULES_CONFIG_PATH = PROJECT_ROOT / "config" / "workbook_rules_v1.json"
MAIN_TABLE_SCHEMA_PATH = PROJECT_ROOT / "config" / "main_table_schema_v1.json"
RUNTIME_CONTROL_PATH = PROJECT_ROOT / "config" / "runtime_controls.json"
RULEBOOK_XLSX_PATH = PROJECT_ROOT / "config" / "rulebook.xlsx"

WORKPAPER_PATH = PROJECT_ROOT / "inputs" / "分析底稿-十堰城运.xlsx"
BS_PATH = PROJECT_ROOT / "inputs" / "25十运04[258597.SH]-资产负债表.xlsx"
WORKPAPER_CANDIDATES: List[Path] = []
BS_CANDIDATES: List[Path] = []

SOURCE_PRIORITY = ["workpaper", "audit_excel", "rating_report"]
DEFAULT_INTEREST_DEBT_ITEMS = ["短期借款", "一年内到期的非流动负债", "长期借款", "应付债券", "租赁负债", "长期应付款"]
DEFAULT_CODE_ALIASES = {
    "CF013": "CF037",
    "CF014": "CF038",
    "CF015": "CF039",
}
FORCED_BS_CODE_REDIRECTS = {
    "BS000": "BS027",  # 流动资产 <- 流动资产合计
    "BS028": "BS056",  # 非流动资产 <- 非流动资产合计
    "BS058": "BS089",  # 流动负债 <- 流动负债合计
    "BS090": "BS102",  # 非流动负债 <- 非流动负债合计
}


def load_main_table_schema() -> Dict[str, Any]:
    default_schema = {
        "prefix_columns": ["科目编码", "科目名称"],
        "value_col_template": "{year}年（万元）",
        "status_col_template": "{year}年状态",
    }
    if not MAIN_TABLE_SCHEMA_PATH.exists():
        return default_schema
    try:
        cfg = json.loads(MAIN_TABLE_SCHEMA_PATH.read_text(encoding="utf-8"))
    except Exception:
        return default_schema
    schema = dict(default_schema)
    schema.update({k: v for k, v in cfg.items() if v})
    return schema


def parse_bool(v: Any, default: bool = True) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in {"1", "true", "yes", "y", "\u662f"}:
        return True
    if s in {"0", "false", "no", "n", "\u5426"}:
        return False
    return default


def parse_jsonish_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, bool, list, dict)):
        return v
    text = str(v).strip()
    if text == "":
        return ""
    try:
        return json.loads(text)
    except Exception:
        return text


def get_interest_debt_items(runtime_cfg: Optional[Dict[str, Any]] = None) -> List[str]:
    cfg = runtime_cfg or {}
    raw = cfg.get("interest_debt_items")
    if isinstance(raw, list):
        vals = [str(x).strip() for x in raw if str(x).strip()]
        if vals:
            return vals
    return list(DEFAULT_INTEREST_DEBT_ITEMS)


def get_code_aliases(runtime_cfg: Optional[Dict[str, Any]] = None) -> Dict[str, str]:
    cfg = runtime_cfg or {}
    raw = cfg.get("code_aliases")
    out: Dict[str, str] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            kk = str(k or "").strip().upper()
            vv = str(v or "").strip().upper()
            if kk and vv:
                out[kk] = vv
    if out:
        return out
    return dict(DEFAULT_CODE_ALIASES)


def load_runtime_cfg_from_excel(cfg_path: Path = RULEBOOK_XLSX_PATH) -> Dict[str, Any]:
    if not cfg_path.exists():
        return {}
    try:
        wb = load_workbook(cfg_path, data_only=False)
    except Exception:
        return {}
    ws = get_sheet_by_name_loose(wb, "\u8fd0\u884c\u63a7\u5236")
    if ws is None:
        return {}
    out: Dict[str, Any] = {}
    for r in range(2, ws.max_row + 1):
        key = str(ws.cell(r, 1).value or "").strip()
        raw = ws.cell(r, 2).value
        if not key:
            continue
        out[key] = parse_jsonish_value(raw)
    return out


def load_workbook_rules_cfg_from_excel(cfg_path: Path = RULEBOOK_XLSX_PATH) -> Dict[str, Any]:
    if not cfg_path.exists():
        return {}
    try:
        wb = load_workbook(cfg_path, data_only=False)
    except Exception:
        return {}

    recon_rules: List[Dict[str, Any]] = []
    ws_recon = get_sheet_by_name_loose(wb, "\u52fe\u7a3d\u89c4\u5219")
    if ws_recon is not None:
        for r in range(2, ws_recon.max_row + 1):
            rid = str(ws_recon.cell(r, 1).value or "").strip()
            desc = str(ws_recon.cell(r, 2).value or "").strip()
            formula = str(ws_recon.cell(r, 3).value or "").strip()
            enabled = parse_bool(ws_recon.cell(r, 4).value, True)
            if not rid or "=" not in formula:
                continue
            recon_rules.append(
                {
                    "id": rid,
                    "description": desc,
                    "formula": formula,
                    "enabled": enabled,
                }
            )

    ratio_aliases: Dict[str, List[str]] = {}
    ws_alias = get_sheet_by_name_loose(wb, "\u6307\u6807\u522b\u540d\u6620\u5c04")
    if ws_alias is not None:
        for r in range(2, ws_alias.max_row + 1):
            rid = str(ws_alias.cell(r, 1).value or "").strip()
            alias_raw = str(ws_alias.cell(r, 2).value or "").strip()
            if not rid or not alias_raw:
                continue
            vals = split_aliases(alias_raw)
            if vals:
                ratio_aliases[rid] = vals

    if not recon_rules and not ratio_aliases:
        return {}
    return {
        "version": "v1",
        "reconciliation": {"rules": recon_rules},
        "ratio_aliases": ratio_aliases,
    }


def load_ratio_cfg_from_excel(cfg_path: Path = RULEBOOK_XLSX_PATH) -> Dict[str, Any]:
    if not cfg_path.exists():
        return {"rules": []}
    try:
        wb = load_workbook(cfg_path, data_only=False)
    except Exception:
        return {"rules": []}
    ws = get_sheet_by_name_loose(wb, "\u8d22\u52a1\u6bd4\u7387\u89c4\u5219")
    if ws is None:
        return {"rules": []}

    rules: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, 1).value or "").strip()
        name = str(ws.cell(r, 2).value or "").strip()
        group = str(ws.cell(r, 3).value or "").strip()
        enabled = parse_bool(ws.cell(r, 4).value, True)
        desc = str(ws.cell(r, 5).value or "").strip()
        direct_aliases = split_aliases(str(ws.cell(r, 6).value or ""))
        calc_formula = str(ws.cell(r, 7).value or "").strip()
        prefer_direct = parse_bool(ws.cell(r, 8).value, True)
        direct_value_divisor = to_float(ws.cell(r, 9).value) or 1.0
        if not rid:
            continue
        rules.append(
            {
                "id": rid,
                "name": name or rid,
                "group": group,
                "enabled": enabled,
                "description": desc,
                "direct_aliases": direct_aliases,
                "calc_formula": calc_formula,
                "prefer_direct": prefer_direct,
                "direct_value_divisor": direct_value_divisor,
            }
        )
    return {"rules": rules}


def load_runtime_cfg() -> Dict[str, Any]:
    cfg = load_runtime_controls(PROJECT_ROOT)
    if isinstance(cfg, dict):
        return cfg
    return {}


def load_workbook_rules_cfg(cfg_path: Path = WORKBOOK_RULES_CONFIG_PATH) -> Dict[str, Any]:
    cfg = load_workbook_rules(PROJECT_ROOT)
    if isinstance(cfg, dict) and cfg:
        return cfg
    if not cfg_path.exists():
        return {}
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def render_main_table_header(years: List[str], schema: Dict[str, Any]) -> List[str]:
    prefix = schema.get("prefix_columns", ["科目编码", "科目名称"])
    if not isinstance(prefix, list) or len(prefix) < 2:
        prefix = ["科目编码", "科目名称"]
    value_tpl = str(schema.get("value_col_template", "{year}年（万元）"))
    status_tpl = str(schema.get("status_col_template", "{year}年状态"))
    value_cols = [value_tpl.format(year=y) for y in years]
    status_cols = [status_tpl.format(year=y) for y in years]
    return prefix + value_cols + status_cols


def normalize_project_id(project_id: str) -> str:
    text = re.sub(r"[^0-9A-Za-z._-]+", "_", (project_id or "").strip())
    return text or "default_project"


def resolve_path(path_text: str) -> Path:
    p = Path(path_text)
    return p if p.is_absolute() else PROJECT_ROOT / p


def list_excel_files(search_dirs: List[Path]) -> List[Path]:
    files: List[Path] = []
    seen = set()
    for d in search_dirs:
        if not d.exists() or not d.is_dir():
            continue
        for ext in ("*.xlsx", "*.xls"):
            for p in d.rglob(ext):
                if p.is_file():
                    key = str(p.resolve())
                    if key not in seen:
                        seen.add(key)
                        files.append(p)
    return files


def list_pdf_files(search_dirs: List[Path]) -> List[Path]:
    files: List[Path] = []
    seen = set()
    for d in search_dirs:
        if not d.exists() or not d.is_dir():
            continue
        for p in d.rglob("*.pdf"):
            if p.is_file():
                key = str(p.resolve())
                if key not in seen:
                    seen.add(key)
                    files.append(p)
    return files


def list_project_files(search_dirs: List[Path]) -> List[Path]:
    exts = {".xlsx", ".xls", ".pdf", ".doc", ".docx"}
    files: List[Path] = []
    seen = set()
    for d in search_dirs:
        if not d.exists() or not d.is_dir():
            continue
        for p in d.rglob("*"):
            if p.is_file() and p.suffix.lower() in exts:
                key = str(p.resolve())
                if key not in seen:
                    seen.add(key)
                    files.append(p)
    return files


def score_workpaper_file(path: Path) -> int:
    name = path.name.lower()
    score = 0
    if "底稿" in name or "workpaper" in name:
        score += 5
    if "分析" in name:
        score += 2
    try:
        sheet_names = set(pd.ExcelFile(path).sheet_names)
        for sn in sheet_names:
            if ("资产负债表" in sn) or ("利润表" in sn) or ("现金流量表" in sn):
                score += 3
            if ("资产" in sn) or ("负债" in sn) or ("主营构成" in sn):
                score += 1
    except Exception:
        pass
    return score


def score_bs_file(path: Path) -> int:
    name = path.name.lower()
    score = 0
    if "资产负债表" in name or "balance" in name:
        score += 5
    if "bs" in name:
        score += 2
    try:
        excel = pd.ExcelFile(path)
        if any("资产负债表" in sn for sn in excel.sheet_names):
            score += 6
        probe_sheet = next((sn for sn in excel.sheet_names if "资产负债表" in sn), excel.sheet_names[0])
        df = pd.read_excel(path, sheet_name=probe_sheet, header=None, nrows=6)
        txt = " ".join(str(v) for v in df.fillna("").values.flatten().tolist())
        if re.search(r"20\d{2}", txt):
            score += 2
    except Exception:
        pass
    return score


def score_general_relevance(path: Path, wp_score: int, bs_score: int) -> int:
    name = path.name.lower()
    score = max(wp_score, bs_score)
    if "审计" in name or "audit" in name:
        score += 3
    if "财务" in name or "报表" in name:
        score += 2
    if "底稿" in name or "workpaper" in name:
        score += 2
    return score


def choose_best(paths: List[Path], scorer) -> Optional[Path]:
    best_path: Optional[Path] = None
    best_score = -1
    for p in paths:
        s = scorer(p)
        if s > best_score:
            best_score = s
            best_path = p
    if best_score <= 0:
        return None
    return best_path


def discover_project_sources(
    project_id: str, input_dir: str, workpaper_path: str, bs_path: str
) -> Tuple[Path, Path, List[Path], List[Path], Dict[str, Any]]:
    search_dirs: List[Path] = []
    if input_dir:
        search_dirs.append(resolve_path(input_dir))
    else:
        search_dirs.extend([PROJECT_ROOT / "inputs" / project_id, PROJECT_ROOT / "inputs"])

    excel_files = list_excel_files(search_dirs)
    pdf_files = list_pdf_files(search_dirs)
    all_files = list_project_files(search_dirs)

    wp_score_map: Dict[str, int] = {}
    bs_score_map: Dict[str, int] = {}
    relevance_rows: List[Dict[str, Any]] = []
    for p in excel_files:
        wp_s = score_workpaper_file(p)
        bs_s = score_bs_file(p)
        key = str(p)
        wp_score_map[key] = wp_s
        bs_score_map[key] = bs_s
        relevance_rows.append(
            {
                "path": key,
                "workpaper_score": wp_s,
                "balance_sheet_score": bs_s,
                "relevance_score": score_general_relevance(p, wp_s, bs_s),
            }
        )

    workpaper_scores = [{"path": row["path"], "score": row["workpaper_score"]} for row in relevance_rows]
    bs_scores = [{"path": row["path"], "score": row["balance_sheet_score"]} for row in relevance_rows]
    workpaper_scores.sort(key=lambda x: x["score"], reverse=True)
    bs_scores.sort(key=lambda x: x["score"], reverse=True)
    relevance_rows.sort(key=lambda x: x["relevance_score"], reverse=True)

    if workpaper_path:
        wp = resolve_path(workpaper_path)
    else:
        picked = choose_best(excel_files, score_workpaper_file)
        wp = picked if picked else PROJECT_ROOT / "inputs" / "__missing_workpaper__.xlsx"

    if bs_path:
        bs = resolve_path(bs_path)
    else:
        picked = choose_best(excel_files, score_bs_file)
        bs = picked if picked else PROJECT_ROOT / "inputs" / "__missing_balance_sheet__.xlsx"

    # Prefer distinct sources when both roles can be satisfied by different files.
    if (
        not workpaper_path
        and not bs_path
        and wp.exists()
        and bs.exists()
        and wp.resolve() == bs.resolve()
        and len(excel_files) > 1
    ):
        for item in bs_scores:
            cand = Path(item["path"])
            if item["score"] > 0 and cand.resolve() != wp.resolve():
                bs = cand
                break

    wp_candidates = [Path(row["path"]) for row in workpaper_scores if row["score"] > 0]
    bs_candidates = [Path(row["path"]) for row in bs_scores if row["score"] > 0]
    if not wp_candidates and wp.exists():
        wp_candidates = [wp]
    if not bs_candidates and bs.exists():
        bs_candidates = [bs]

    manifest = {
        "project_id": project_id,
        "priority_dimensions": ["报表", "审计报告", "评级报告"],
        "search_dirs": [str(p) for p in search_dirs],
        "all_file_count": len(all_files),
        "excel_file_count": len(excel_files),
        "pdf_file_count": len(pdf_files),
        "all_files": [str(p) for p in all_files],
        "excel_files": [str(p) for p in excel_files],
        "pdf_files": [str(p) for p in pdf_files],
        "financial_statement_files_top20": [
            str(p)
            for p in sorted(
                [f for f in all_files if ("报表" in f.name or "财务" in f.name or "statement" in f.name.lower())],
                key=lambda x: x.name,
            )[:20]
        ],
        "audit_report_files_top20": [
            str(p)
            for p in sorted(
                [f for f in all_files if ("审计" in f.name or "audit" in f.name.lower())],
                key=lambda x: x.name,
            )[:20]
        ],
        "rating_report_files_top20": [
            str(p)
            for p in sorted(
                [f for f in all_files if ("评级" in f.name or "rating" in f.name.lower())],
                key=lambda x: x.name,
            )[:20]
        ],
        "workpaper_candidates_top10": workpaper_scores[:10],
        "balance_sheet_candidates_top10": bs_scores[:10],
        "excel_relevance_top20": relevance_rows[:20],
        "selected_workpaper": str(wp),
        "selected_balance_sheet": str(bs),
    }
    return wp, bs, wp_candidates, bs_candidates, manifest


@dataclass
class SourceRecord:
    source_type: str
    file: str
    sheet: str
    row_idx: int
    matched_name: str
    value_yiyuan: float


@dataclass
class RatioRecord:
    source_type: str
    file: str
    sheet: str
    row_idx: int
    matched_name: str
    value: float


def normalize_name(name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip().lower()
    # Normalize common abbreviations/synonyms used in workpapers.
    text = text.replace("及现金等价物", "")
    text = text.replace("（合计）", "合计").replace("(合计)", "合计")
    for ch in [" ", "\t", "\n", "（", "）", "(", ")", "，", ",", "。", ":", "：", "-", "_"]:
        text = text.replace(ch, "")
    return text


def has_combined_subject_marker(name: str) -> bool:
    text = str(name or "")
    markers = ["及", "和", "与", "/", "、"]
    return any(m in text for m in markers)


def should_skip_combined_match(alias_raw: str, source_raw: str) -> bool:
    if not alias_raw or not source_raw:
        return False
    if normalize_name(alias_raw) == normalize_name(source_raw):
        return False
    # Prevent mapping single subjects (e.g. 应收账款) from merged rows (e.g. 应收票据及应收账款).
    return has_combined_subject_marker(source_raw) and not has_combined_subject_marker(alias_raw)


def should_skip_qualified_subject_match(alias_raw: str, source_raw: str) -> bool:
    """Avoid fuzzy mapping between base and qualified subjects, e.g. 待摊费用 <- 长期待摊费用."""
    a = str(alias_raw or "")
    s = str(source_raw or "")
    if not a or not s:
        return False
    if normalize_name(a) == normalize_name(s):
        return False

    qualifiers = ["长期", "短期", "一年内", "其中", "合计", "净额", "非流动", "流动", "清理"]
    for q in qualifiers:
        if (q in s) != (q in a):
            na = normalize_name(a)
            ns = normalize_name(s)
            # Skip both directions when one side adds a qualifier:
            # e.g. 待摊费用 <-> 长期待摊费用, 固定资产清理 <-> 固定资产
            if na and (na in ns or ns in na):
                return True
    return False


def to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    s = str(value).strip().replace(",", "").replace("，", "")
    if s == "":
        return None
    negative = False
    # Accounting style negatives: (123.45) / （123.45）
    if (s.startswith("(") and s.endswith(")")) or (s.startswith("（") and s.endswith("）")):
        negative = True
        s = s[1:-1].strip()
    if s.startswith("-"):
        negative = True
        s = s[1:].strip()
    try:
        v = float(s)
        return -v if negative else v
    except ValueError:
        return None


def parse_year_token(value: Any) -> Optional[str]:
    txt = str(value or "").strip()
    if not txt:
        return None
    m = re.search(r"(20\d{2})", txt)
    if not m:
        return None
    y = int(m.group(1))
    current_year = dt.date.today().year
    if y < 2000 or y > current_year + 1:
        return None
    return str(y)


def yiyuan_to_wanyuan(value_yiyuan: Optional[float]) -> Optional[float]:
    if value_yiyuan is None:
        return None
    # Source files are already in 万元, so keep original unit.
    return round(value_yiyuan, 2)


def detect_latest_three_years(candidate_paths: Optional[List[Path]] = None) -> List[str]:
    years = set()
    runtime_cfg = load_runtime_cfg()
    paths = candidate_paths or []
    if not paths:
        paths = [WORKPAPER_PATH, BS_PATH]

    for src_path in paths:
        if not src_path.exists():
            continue
        try:
            xls = pd.ExcelFile(src_path)
        except Exception:
            continue
        for sheet_name in xls.sheet_names[:3]:
            try:
                df = pd.read_excel(src_path, sheet_name=sheet_name, header=None, nrows=20)
            except Exception:
                continue
            if df.empty:
                continue
            for r in range(df.shape[0]):
                for c in range(df.shape[1]):
                    y = parse_year_token(df.iloc[r, c])
                    if y:
                        years.add(y)

    if not years:
        count = int(runtime_cfg.get("fallback_recent_years", 3))
        end_year_cfg = runtime_cfg.get("fallback_end_year", "current")
        end_year = dt.date.today().year if str(end_year_cfg) == "current" else int(end_year_cfg)
        start_year = end_year - count + 1
        return [str(y) for y in range(start_year, end_year + 1)]

    sorted_years = sorted(list(years))
    return sorted_years[-3:]


def load_workpaper_records(years: List[str], candidate_paths: Optional[List[Path]] = None) -> Dict[Tuple[str, str], List[SourceRecord]]:
    result: Dict[Tuple[str, str], List[SourceRecord]] = {}
    paths = candidate_paths if candidate_paths is not None else WORKPAPER_CANDIDATES
    if not paths:
        paths = [WORKPAPER_PATH]

    # report -> candidate sheet-name keywords, fallback year->col map
    report_sheet_cfg = {
        "资产负债表": (["资产负债表", "资产", "负债"], {"2022": 1, "2023": 4, "2024": 8}),
        "利润表": (["利润表", "损益", "主营构成"], {"2022": 1, "2023": 4, "2024": 8}),
        "现金流量表": (["现金流量表", "现金流"], {"2022": 1, "2023": 2, "2024": 3}),
    }
    for src_path in paths:
        if not src_path.exists():
            continue
        try:
            excel = pd.ExcelFile(src_path)
        except Exception:
            continue

        for report_name, (keywords, fallback_col_map) in report_sheet_cfg.items():
            for sheet_name in excel.sheet_names:
                if not any(k in sheet_name for k in keywords):
                    continue
                try:
                    df = pd.read_excel(src_path, sheet_name=sheet_name, header=None)
                except Exception:
                    continue
                year_cols = detect_year_columns(df, years, scan_rows=12)
                col_map = year_cols if year_cols else fallback_col_map
                for i in range(len(df)):
                    raw_name = df.iloc[i, 0] if df.shape[1] > 0 else None
                    name = str(raw_name).strip() if raw_name is not None else ""
                    if name in ("", "nan", "数据类型", "证券代码", "证券简称", "起始年份", "截止年份"):
                        continue

                    for y in years:
                        if y not in col_map:
                            continue
                        col = col_map[y]
                        if col >= df.shape[1]:
                            continue
                        val = to_float(df.iloc[i, col])
                        if val is None:
                            continue
                        rec = SourceRecord(
                            source_type="workpaper",
                            file=str(src_path),
                            sheet=sheet_name,
                            row_idx=i + 1,
                            matched_name=name,
                            value_yiyuan=val,
                        )
                        result.setdefault((normalize_name(name), y), []).append(rec)

    return result


def load_bs_records(years: List[str], candidate_paths: Optional[List[Path]] = None) -> Dict[Tuple[str, str], List[SourceRecord]]:
    result: Dict[Tuple[str, str], List[SourceRecord]] = {}
    paths = candidate_paths if candidate_paths is not None else BS_CANDIDATES
    if not paths:
        paths = [BS_PATH]

    # fallback mapping if header years cannot be detected
    fallback_col_map = {"2022": 1, "2023": 2, "2024": 3}
    for src_path in paths:
        if not src_path.exists():
            continue
        try:
            excel = pd.ExcelFile(src_path)
        except Exception:
            continue
        target_sheets = [s for s in excel.sheet_names if "资产负债表" in s] or excel.sheet_names[:1]
        for sheet_name in target_sheets:
            try:
                df = pd.read_excel(src_path, sheet_name=sheet_name, header=None)
            except Exception:
                continue
            year_cols = detect_year_columns(df, years, scan_rows=12)
            col_map = year_cols if year_cols else fallback_col_map

            for i in range(len(df)):
                raw_name = df.iloc[i, 0] if df.shape[1] > 0 else None
                name = str(raw_name).strip() if raw_name is not None else ""
                if name in ("", "nan", "报告期", "报表类型", "证券代码", "证券简称", "起始年份", "截止年份"):
                    continue

                for y in years:
                    if y not in col_map:
                        continue
                    col = col_map[y]
                    if col >= df.shape[1]:
                        continue
                    val = to_float(df.iloc[i, col])
                    if val is None:
                        continue
                    rec = SourceRecord(
                        source_type="audit_excel",
                        file=str(src_path),
                        sheet=sheet_name,
                        row_idx=i + 1,
                        matched_name=name,
                        value_yiyuan=val,
                    )
                    result.setdefault((normalize_name(name), y), []).append(rec)

    return result


def detect_year_columns(df: pd.DataFrame, years: List[str], scan_rows: int = 20) -> Dict[str, int]:
    """Detect year columns from a real header row, avoiding meta rows like 起始年份/截止年份."""
    rows = min(scan_rows, len(df))
    candidates: List[Tuple[int, int, int, Dict[str, int]]] = []

    for r in range(rows):
        row_map: Dict[str, int] = {}
        for c in range(df.shape[1]):
            y = parse_year_token(df.iloc[r, c])
            if y and y in years and y not in row_map:
                row_map[y] = c

        # Require at least 2 years on the same row; this filters out metadata rows.
        if len(row_map) >= 2:
            cols = sorted(row_map.values())
            spread = cols[-1] - cols[0] if len(cols) >= 2 else 0
            # More years first, then wider spread, then earlier row.
            candidates.append((len(row_map), spread, -r, row_map))

    if candidates:
        candidates.sort(reverse=True)
        best = candidates[0][3]
        # Ensure each year maps to a distinct column.
        if len(set(best.values())) == len(best):
            return best
    return {}


def first_text_in_row(df: pd.DataFrame, row_idx: int, max_col: int = 3) -> str:
    cols = min(max_col, df.shape[1])
    for c in range(cols):
        v = df.iloc[row_idx, c]
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            return s
    return ""


def load_direct_ratio_records(
    years: List[str], candidate_paths: Optional[List[Path]] = None
) -> Dict[Tuple[str, str], List[RatioRecord]]:
    result: Dict[Tuple[str, str], List[RatioRecord]] = {}
    paths = candidate_paths if candidate_paths is not None else WORKPAPER_CANDIDATES
    if not paths:
        paths = [WORKPAPER_PATH]

    for src_path in paths:
        if not src_path.exists():
            continue
        try:
            excel = pd.ExcelFile(src_path)
        except Exception:
            continue

        for sheet_name in excel.sheet_names:
            try:
                df = pd.read_excel(src_path, sheet_name=sheet_name, header=None)
            except Exception:
                continue
            if df.empty:
                continue

            year_cols = detect_year_columns(df, years, scan_rows=8)
            if not year_cols:
                continue

            for i in range(1, len(df)):
                metric_name = first_text_in_row(df, i, max_col=3)
                if not metric_name:
                    continue
                metric_key = normalize_name(metric_name)
                if not metric_key:
                    continue
                for y, col in year_cols.items():
                    if col >= df.shape[1]:
                        continue
                    val = to_float(df.iloc[i, col])
                    if val is None:
                        continue
                    rec = RatioRecord(
                        source_type="direct_ratio",
                        file=str(src_path),
                        sheet=sheet_name,
                        row_idx=i + 1,
                        matched_name=metric_name,
                        value=val,
                    )
                    result.setdefault((metric_key, y), []).append(rec)

    return result


def pick_value_for_item(
    aliases: List[str],
    year: str,
    source_maps: Dict[str, Dict[Tuple[str, str], List[SourceRecord]]],
    target_report: str = "",
) -> Tuple[Optional[SourceRecord], str]:
    # returns chosen record and fill status
    alias_pairs = [(a, normalize_name(a)) for a in aliases if normalize_name(a)]

    def report_match(sheet_name: str) -> bool:
        s = str(sheet_name or "")
        if not target_report:
            return True
        if target_report == "资产负债表":
            return ("资产负债表" in s) or ("资产" in s and "现金流" not in s and "利润" not in s) or ("负债" in s and "资产负债表" in s)
        if target_report == "利润表":
            return ("利润" in s) or ("损益" in s) or ("主营构成" in s)
        if target_report == "现金流量表":
            return ("现金流" in s)
        return True

    for source_name in SOURCE_PRIORITY:
        if source_name == "rating_report":
            continue
        source_map = source_maps.get(source_name, {})
        for _, an in alias_pairs:
            records = source_map.get((an, year), [])
            if records:
                for rec in records:
                    if report_match(rec.sheet):
                        return rec, "已识别"

        # Fallback for synonym/abbreviation: normalized containment matching.
        for (src_name_norm, src_year), records in source_map.items():
            if src_year != year or not records:
                continue
            src_raw = str(records[0].matched_name or "")
            for alias_raw, an in alias_pairs:
                if an and (an in src_name_norm or src_name_norm in an):
                    if should_skip_combined_match(alias_raw, src_raw):
                        continue
                    if should_skip_qualified_subject_match(alias_raw, src_raw):
                        continue
                    for rec in records:
                        if report_match(rec.sheet):
                            return rec, "已识别"

    return None, "待补充"


def write_statement_sheet(
    wb: Workbook,
    sheet_name: str,
    items: List[dict],
    years: List[str],
    source_maps: Dict[str, Dict[Tuple[str, str], List[SourceRecord]]],
    detail_payload: Dict[Tuple[str, str, str], List[dict]],
    missing_rows: List[dict],
) -> Dict[Tuple[str, str], Optional[float]]:
    ws = wb.create_sheet(sheet_name)
    ws.append(["科目编码", "科目名称"] + [f"{y}年（万元）" for y in years] + [f"{y}年状态" for y in years])

    red_font = Font(color="FF0000")
    statement_values: Dict[Tuple[str, str], Optional[float]] = {}

    for item in items:
        code = item["code"]
        name = item["name"]
        aliases = item.get("aliases", [name])

        row_vals = [code, name]
        status_vals = []

        for y in years:
            rec, status = pick_value_for_item(aliases, y, source_maps, sheet_name)
            value_wanyuan = yiyuan_to_wanyuan(rec.value_yiyuan) if rec else None
            statement_values[(name, y)] = value_wanyuan
            row_vals.append(value_wanyuan)
            status_vals.append(status)

            detail_payload.setdefault((sheet_name, name, code), []).append(
                {
                    "报表": sheet_name,
                    "科目编码": code,
                    "科目名称": name,
                    "期间": y,
                    "值(万元)": value_wanyuan,
                    "状态": status,
                    "来源类型": rec.source_type if rec else "",
                    "来源文件": rec.file if rec else "",
                    "来源Sheet": rec.sheet if rec else "",
                    "来源行号": rec.row_idx if rec else "",
                    "匹配名称": rec.matched_name if rec else "",
                }
            )

            if status == "待补充":
                missing_rows.append(
                    {
                        "报表": sheet_name,
                        "科目编码": code,
                        "科目名称": name,
                        "期间": y,
                        "状态": status,
                        "说明": "优先来源未识别到有效数值",
                    }
                )

        ws.append(row_vals + status_vals)

        excel_row = ws.max_row
        for c in range(3, 3 + len(years)):
            cell = ws.cell(row=excel_row, column=c)
            if isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.number_format = '[Red](#,##0.00)'
                    cell.font = red_font
                else:
                    cell.number_format = '#,##0.00'

    return statement_values


def get_sheet_by_name_loose(wb: Workbook, target_name: str):
    t = str(target_name).strip()
    for ws in wb.worksheets:
        if ws.title.strip() == t:
            return ws
    return None


def load_statement_items_from_template(wb: Workbook, sheet_name: str, code_prefix: str) -> List[dict]:
    ws = get_sheet_by_name_loose(wb, sheet_name)
    if ws is None:
        return []
    items: List[dict] = []
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        code_text = str(code or "").strip()
        if not code_text.startswith(code_prefix):
            continue
        name_text = str(name or "").strip()
        if not name_text:
            continue
        items.append({"code": code_text, "name": name_text, "aliases": [name_text], "row_idx": r})
    return items


def write_statement_sheet_from_template(
    wb: Workbook,
    sheet_name: str,
    code_prefix: str,
    items: List[dict],
    years: List[str],
    source_maps: Dict[str, Dict[Tuple[str, str], List[SourceRecord]]],
    detail_payload: Dict[Tuple[str, str, str], List[dict]],
    missing_rows: List[dict],
) -> Dict[Tuple[str, str], Optional[float]]:
    ws = get_sheet_by_name_loose(wb, sheet_name)
    if ws is None:
        return {}

    # Keep template structure; only inject dynamic periods and values.
    for i, y in enumerate(years):
        ws.cell(3, 3 + i).value = f"{y}年"
        ws.cell(4, 3 + i).value = "合并报表"

    red_font = Font(color="FF0000")
    statement_values: Dict[Tuple[str, str], Optional[float]] = {}

    for item in items:
        code = item["code"]
        name = item["name"]
        aliases = item.get("aliases", [name])
        row_idx = int(item["row_idx"])

        for i, y in enumerate(years):
            rec, status = pick_value_for_item(aliases, y, source_maps, sheet_name.strip())
            value_wanyuan = yiyuan_to_wanyuan(rec.value_yiyuan) if rec else None
            statement_values[(name, y)] = value_wanyuan
            cell = ws.cell(row=row_idx, column=3 + i)
            cell.value = value_wanyuan

            if isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.number_format = '[Red](#,##0.00)'
                    cell.font = red_font
                else:
                    cell.number_format = '#,##0.00'

            detail_payload.setdefault((sheet_name, name, code), []).append(
                {
                    "报表": sheet_name,
                    "科目编码": code,
                    "科目名称": name,
                    "期间": y,
                    "值(万元)": value_wanyuan,
                    "状态": status,
                    "来源类型": rec.source_type if rec else "",
                    "来源文件": rec.file if rec else "",
                    "来源Sheet": rec.sheet if rec else "",
                    "来源行号": rec.row_idx if rec else "",
                    "匹配名称": rec.matched_name if rec else "",
                }
            )

            if status == "待补充":
                missing_rows.append(
                    {
                        "报表": sheet_name,
                        "科目编码": code,
                        "科目名称": name,
                        "期间": y,
                        "状态": status,
                        "说明": "优先来源未识别到有效数值",
                    }
                )

    # Enforce explicit BS code redirects to avoid fuzzy-match mis-hits on broad section subjects.
    if code_prefix == "BS" and FORCED_BS_CODE_REDIRECTS:
        item_by_code = {str(it.get("code", "")).strip(): it for it in items}
        for to_code, from_code in FORCED_BS_CODE_REDIRECTS.items():
            to_item = item_by_code.get(to_code)
            from_item = item_by_code.get(from_code)
            if not to_item or not from_item:
                continue
            to_row = int(to_item.get("row_idx", 0) or 0)
            from_row = int(from_item.get("row_idx", 0) or 0)
            if to_row <= 0 or from_row <= 0:
                continue
            for i, y in enumerate(years):
                val = ws.cell(row=from_row, column=3 + i).value
                dst_cell = ws.cell(row=to_row, column=3 + i)
                dst_cell.value = val
                if isinstance(val, (int, float)):
                    if float(val) < 0:
                        dst_cell.number_format = '[Red](#,##0.00)'
                        dst_cell.font = red_font
                    else:
                        dst_cell.number_format = '#,##0.00'
                statement_values[(str(to_item.get("name", "")), y)] = to_float(val)

    return statement_values


def safe_div(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def avg_or_end(curr: Optional[float], prev: Optional[float]) -> Optional[float]:
    if curr is not None and prev is not None:
        return (curr + prev) / 2.0
    return curr


def bool_text(ok: Optional[bool]) -> str:
    if ok is None:
        return "待补充"
    return "是" if ok else "否"


def load_recon_rules_from_template_sheet(ws) -> List[dict]:
    rules: List[dict] = []
    if ws is None:
        return rules
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, 1).value or "").strip() or f"R{r}"
        desc = str(ws.cell(r, 2).value or "").strip()
        formula = str(ws.cell(r, 3).value or "").strip()
        if not formula or "=" not in formula:
            continue
        rules.append({"id": rid, "description": desc, "formula": normalize_recon_formula(formula)})
    return rules


def normalize_recon_formula(formula: str) -> str:
    """Normalize template formulas and de-duplicate repeated addends in simple sum expressions."""
    text = str(formula or "").replace(" ", "")
    if "=" not in text:
        return text
    left, right = text.split("=", 1)
    return f"{normalize_sum_expr(left)}={normalize_sum_expr(right)}"


def normalize_sum_expr(expr: str) -> str:
    """
    For simple plus-only expressions like BS102=BS091+BS092+BS094+BS094,
    remove repeated codes while preserving original order.
    """
    text = str(expr or "").replace(" ", "")
    simple_sum = r"[A-Z]{2,6}\d+(?:\+[A-Z]{2,6}\d+)+"
    if not re.fullmatch(simple_sum, text):
        return text
    seen = set()
    parts = []
    for token in text.split("+"):
        key = token.upper()
        if key in seen:
            continue
        seen.add(key)
        parts.append(token)
    return "+".join(parts)


def load_recon_rules_from_template_file(template_path: Path = TEMPLATE_WORKBOOK_PATH) -> List[dict]:
    if not template_path.exists():
        return []
    wb = load_workbook(template_path, data_only=False)
    ws = get_sheet_by_name_loose(wb, "勾稽校验")
    return load_recon_rules_from_template_sheet(ws)


def split_aliases(text: str) -> List[str]:
    raw = str(text or "").strip()
    if not raw:
        return []
    parts = re.split(r"[;,，、|/]+", raw)
    return [p.strip() for p in parts if p and p.strip()]


def load_ratio_aliases_from_template_sheet(ws) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    if ws is None:
        return out
    for r in range(2, ws.max_row + 1):
        rid = str(ws.cell(r, 1).value or "").strip()
        alias_raw = str(ws.cell(r, 2).value or "").strip()
        if not rid or not alias_raw:
            continue
        vals = split_aliases(alias_raw)
        if vals:
            out[rid] = vals
    return out


def load_ratio_aliases_from_template_file(template_path: Path = TEMPLATE_WORKBOOK_PATH) -> Dict[str, List[str]]:
    if not template_path.exists():
        return {}
    wb = load_workbook(template_path, data_only=False)
    ws = get_sheet_by_name_loose(wb, "指标别名映射")
    return load_ratio_aliases_from_template_sheet(ws)


def load_recon_rules(
    cfg_path: Path = WORKBOOK_RULES_CONFIG_PATH,
    template_path: Path = TEMPLATE_WORKBOOK_PATH,
) -> List[dict]:
    """
    Preferred source: config/workbook_rules_v1.json
    Fallback source: template sheet with reconciliation rules.
    """
    cfg = load_workbook_rules_cfg(cfg_path)
    recon_cfg = cfg.get("reconciliation", {}) if isinstance(cfg, dict) else {}
    raw_rules = recon_cfg.get("rules", []) if isinstance(recon_cfg, dict) else []
    out: List[dict] = []
    for r in raw_rules:
        if not isinstance(r, dict):
            continue
        if r.get("enabled", True) is False:
            continue
        rid = str(r.get("id", "")).strip()
        formula = str(r.get("formula", "")).strip()
        if not rid or "=" not in formula:
            continue
        out.append(
            {
                "id": rid,
                "description": str(r.get("description", "")).strip(),
                "formula": normalize_recon_formula(formula),
            }
        )
    if out:
        return out
    return load_recon_rules_from_template_file(template_path)


def load_ratio_alias_map(
    cfg_path: Path = WORKBOOK_RULES_CONFIG_PATH,
    template_path: Path = TEMPLATE_WORKBOOK_PATH,
) -> Dict[str, List[str]]:
    """
    Preferred source: config/workbook_rules_v1.json
    Fallback source: template alias sheet.
    """
    cfg = load_workbook_rules_cfg(cfg_path)
    alias_cfg = cfg.get("ratio_aliases", {}) if isinstance(cfg, dict) else {}
    out: Dict[str, List[str]] = {}
    if isinstance(alias_cfg, dict):
        for rid, raw in alias_cfg.items():
            key = str(rid or "").strip()
            if not key:
                continue
            vals: List[str] = []
            if isinstance(raw, list):
                vals = [str(v).strip() for v in raw if str(v).strip()]
            else:
                vals = split_aliases(str(raw))
            if vals:
                out[key] = vals
    if out:
        return out
    return load_ratio_aliases_from_template_file(template_path)


def collect_code_values_by_year(wb: Workbook, years: List[str]) -> Dict[str, Dict[str, Optional[float]]]:
    by_year: Dict[str, Dict[str, Optional[float]]] = {y: {} for y in years}
    prefixes = ("BS", "IS", "CF", "EQ", "DEBT")

    for ws in wb.worksheets:
        for r in range(2, ws.max_row + 1):
            code = str(ws.cell(r, 1).value or "").strip().upper()
            if not any(code.startswith(p) for p in prefixes):
                continue
            for i, y in enumerate(years):
                by_year[y][code] = to_float(ws.cell(r, 3 + i).value)

    debt_ws = get_sheet_by_name_loose(wb, "有息负债明细")
    if debt_ws is not None:
        for y in years:
            for r in range(2, debt_ws.max_row + 1):
                period = str(debt_ws.cell(r, 1).value or "").strip()
                item = str(debt_ws.cell(r, 2).value or "").strip()
                if period == y and item == "有息负债合计":
                    by_year[y]["DEBT999"] = to_float(debt_ws.cell(r, 3).value)
                    break

    return by_year


def evaluate_code_expr(
    expr: str,
    year: str,
    years: List[str],
    by_year: Dict[str, Dict[str, Optional[float]]],
    code_aliases: Optional[Dict[str, str]] = None,
) -> Tuple[Optional[float], bool]:
    text = str(expr or "").strip().replace(" ", "")
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("＋", "+").replace("－", "-").replace("×", "*").replace("÷", "/")

    prev_year = None
    if year in years:
        idx = years.index(year)
        if idx > 0:
            prev_year = years[idx - 1]

    token_pat = re.compile(r"([A-Z]{2,6}\d+)(\[(t|t-1)\])?")
    missing = False
    alias_map = dict(DEFAULT_CODE_ALIASES)
    if code_aliases:
        alias_map.update({str(k).upper(): str(v).upper() for k, v in code_aliases.items() if k and v})

    def lookup_value(target_year: str, raw_code: str) -> Optional[float]:
        year_map = by_year.get(target_year, {})
        code = raw_code.upper()
        val = year_map.get(code)
        if val is not None:
            return val
        mapped = alias_map.get(code)
        if mapped:
            val = year_map.get(mapped)
            if val is not None:
                return val
        # Tolerate non-padded codes in formulas, e.g. BS27 -> BS027.
        m_code = re.fullmatch(r"([A-Z]{2,6})(\d+)", code)
        if m_code:
            prefix, digits = m_code.group(1), m_code.group(2)
            if prefix in {"BS", "IS", "CF", "EQ"} and len(digits) < 3:
                padded = f"{prefix}{int(digits):03d}"
                return year_map.get(padded)
        return None

    def repl(m):
        nonlocal missing
        code = m.group(1).upper()
        tag = m.group(3)
        target_year = year
        if tag == "t-1":
            target_year = prev_year
        if not target_year:
            missing = True
            return "0"
        val = lookup_value(target_year, code)
        if val is None:
            missing = True
            return "0"
        return str(float(val))

    parsed = token_pat.sub(repl, text)
    if not re.fullmatch(r"[0-9eE\.\+\-\*\/\(\)]+", parsed):
        return None, True
    try:
        return float(eval(parsed, {"__builtins__": {}}, {})), missing
    except Exception:
        return None, True


def build_recon_sheet_from_template_rules(
    wb: Workbook,
    years: List[str],
    recon_rules: List[dict],
    tolerance_abs: float,
    code_aliases: Optional[Dict[str, str]] = None,
) -> None:
    ws = wb.create_sheet("勾稽校验")
    ws.append(["规则ID", "规则描述", "期间", "左值", "右值", "差异", "结果(是/否)"])
    by_year = collect_code_values_by_year(wb, years)

    for rule in recon_rules:
        rid = rule.get("id", "")
        desc = rule.get("description", "")
        formula = str(rule.get("formula", "")).replace(" ", "")
        if "=" not in formula:
            continue
        left_expr, right_expr = formula.split("=", 1)
        for y in years:
            left, left_missing = evaluate_code_expr(left_expr, y, years, by_year, code_aliases=code_aliases)
            right, right_missing = evaluate_code_expr(right_expr, y, years, by_year, code_aliases=code_aliases)
            diff = None if left is None or right is None else round(left - right, 2)
            has_missing = left_missing or right_missing
            ok = None if diff is None or has_missing else abs(diff) <= tolerance_abs
            ws.append([rid, desc, y, left, right, diff, bool_text(ok)])



def load_ratio_cfg() -> Dict[str, dict]:
    cfg_excel = load_ratio_rules(PROJECT_ROOT)
    if isinstance(cfg_excel, dict) and cfg_excel.get("rules"):
        return cfg_excel
    if RATIO_RULE_CONFIG_PATH.exists():
        return json.loads(RATIO_RULE_CONFIG_PATH.read_text(encoding="utf-8"))
    return {"rules": []}


def pick_direct_ratio_value(
    aliases: List[str], year: str, ratio_map: Dict[Tuple[str, str], List[RatioRecord]]
) -> Tuple[Optional[RatioRecord], str]:
    alias_pairs = [(a, normalize_name(a)) for a in aliases if normalize_name(a)]
    for _, key in alias_pairs:
        recs = ratio_map.get((key, year), [])
        if recs:
            return recs[0], "direct_source"

    for (src_key, src_year), recs in ratio_map.items():
        if src_year != year or not recs:
            continue
        for _, key in alias_pairs:
            if key and (key in src_key or src_key in key):
                return recs[0], "direct_source"
    return None, ""


def normalize_direct_ratio_value(rule: Dict[str, Any], value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    divisor = to_float(rule.get("direct_value_divisor")) or 1.0
    if divisor not in (0, 1):
        return value / divisor
    return value


def build_ratio_sheet(
    wb: Workbook,
    years: List[str],
    all_values: Dict[str, Dict[Tuple[str, str], Optional[float]]],
    ratio_cfg: Dict[str, dict],
    direct_ratio_candidates: Optional[List[Path]] = None,
    ratio_alias_map: Optional[Dict[str, List[str]]] = None,
    runtime_cfg: Optional[Dict[str, Any]] = None,
) -> None:
    ws = wb.create_sheet("财务比率")
    ws.append(["指标ID", "指标", "分组", "期间", "数值", "取值来源", "口径说明"])

    rules = ratio_cfg.get("rules", [])
    candidates = direct_ratio_candidates if direct_ratio_candidates is not None else WORKPAPER_CANDIDATES
    direct_ratio_map = load_direct_ratio_records(years, candidates)
    by_year_codes = collect_code_values_by_year(wb, years)
    code_aliases = get_code_aliases(runtime_cfg)

    alias_map = ratio_alias_map or {}

    for y in years:
        for rule in rules:
            if not rule.get("enabled", True):
                continue
            rid = rule.get("id", "")
            name = rule.get("name", rid)
            group = rule.get("group", "")
            note = rule.get("description", "")
            aliases = [name] + alias_map.get(rid, []) + rule.get("direct_aliases", [])
            calc_formula = str(rule.get("calc_formula", "") or "").strip()
            prefer_direct = parse_bool(rule.get("prefer_direct", True), True)
            # de-duplicate while preserving order
            aliases = list(dict.fromkeys([a for a in aliases if str(a).strip()]))
            direct_rec, _ = pick_direct_ratio_value(aliases, y, direct_ratio_map)

            if direct_rec is not None and prefer_direct:
                value = normalize_direct_ratio_value(rule, direct_rec.value)
                source = "direct_source"
            else:
                value = None
                source = "pending"
                if calc_formula:
                    calc_val, calc_missing = evaluate_code_expr(
                        calc_formula,
                        y,
                        years,
                        by_year_codes,
                        code_aliases=code_aliases,
                    )
                    if calc_val is not None:
                        value = calc_val
                        source = "calculated" if not calc_missing else "calculated_zero_fill"
                if value is None and direct_rec is not None:
                    value = normalize_direct_ratio_value(rule, direct_rec.value)
                    source = "direct_source"
            ws.append([rid, name, group, y, value, source, note])
            cell = ws.cell(row=ws.max_row, column=5)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0000'


def build_missing_sheet(wb: Workbook, missing_rows: List[dict]) -> None:
    ws = wb.create_sheet("差异缺失清单")
    ws.append(["报表", "科目编码", "科目名称", "期间", "状态", "说明"])
    for row in missing_rows:
        ws.append([row["报表"], row["科目编码"], row["科目名称"], row["期间"], row["状态"], row["说明"]])


def sanitize_sheet_name(name: str) -> str:
    cleaned = str(name)
    for bad in ["\\", "/", "?", "*", "[", "]", ":"]:
        cleaned = cleaned.replace(bad, "_")
    return cleaned


def is_asset_or_liability_detail(report: str, code: str) -> bool:
    if report != "资产负债表":
        return False
    if not code.startswith("BS"):
        return False
    try:
        idx = int(code[2:])
    except ValueError:
        return False
    return 1 <= idx <= 29


def is_income_detail(report: str, subject_name: str) -> bool:
    if report != "利润表":
        return False
    return ("收入" in subject_name) or ("收益" in subject_name)


def income_category(subject_name: str) -> str:
    name = str(subject_name)
    if "投资收益" in name:
        return "投资收益"
    if ("营业收入" in name) or ("营业总收入" in name):
        return "主营收入"
    if "营业外收入" in name:
        return "其他收入"
    if "收益" in name:
        return "其他收入"
    return ""


def income_subitems(subject_name: str) -> List[str]:
    if "投资收益" in subject_name:
        return ["股权投资收益", "债权投资收益", "其他投资收益"]
    if "营业外收入" in subject_name:
        return ["政府补助", "罚没及赔偿收入", "其他营业外收入"]
    if "营业总收入" in subject_name or "营业收入" in subject_name:
        return ["产品销售收入", "工程服务收入", "其他主营业务收入"]
    return ["分项收入A", "分项收入B", "分项收入C"]


def build_detail_sheets(wb: Workbook, detail_payload: Dict[Tuple[str, str, str], List[dict]]) -> int:
    used_names = set()
    created = 0

    for (report, subject_name, code), rows in detail_payload.items():
        is_bs = is_asset_or_liability_detail(report, code)
        is_is = is_income_detail(report, subject_name)
        if not (is_bs or is_is):
            continue

        base = sanitize_sheet_name(f"明细_{subject_name}")[:31]
        ws_name = base
        suffix = 2
        while ws_name in used_names:
            tail = f"_{suffix}"
            ws_name = base[: 31 - len(tail)] + tail
            suffix += 1
        used_names.add(ws_name)

        ws = wb.create_sheet(ws_name)
        ws.append(
            [
                "报表",
                "科目编码",
                "科目名称",
                "期间",
                "对方账户",
                "分项收入/子项收入",
                "主表值(万元)",
                "明细值(万元)",
                "状态",
                "说明",
                "来源类型",
                "来源文件",
                "来源Sheet",
                "来源行号",
                "匹配名称",
            ]
        )

        for r in rows:
            note = "框架保留，请按客观资料补录明细；系统不做模拟拆分。"
            if is_bs:
                parties = ["待补充A", "待补充B", "待补充C"]
                for party in parties:
                    ws.append([
                        r["报表"],
                        r["科目编码"],
                        r["科目名称"],
                        r["期间"],
                        party,
                        "往来款项",
                        r["值(万元)"],
                        None,
                        r["状态"] if r["值(万元)"] is not None else "待补充",
                        note,
                        r.get("来源类型", ""),
                        r.get("来源文件", ""),
                        r.get("来源Sheet", ""),
                        r.get("来源行号", ""),
                        r.get("匹配名称", ""),
                    ])
            else:
                subitems = income_subitems(r["科目名称"])
                for sub in subitems:
                    ws.append([
                        r["报表"],
                        r["科目编码"],
                        r["科目名称"],
                        r["期间"],
                        "",
                        sub,
                        r["值(万元)"],
                        None,
                        r["状态"] if r["值(万元)"] is not None else "待补充",
                        note,
                        r.get("来源类型", ""),
                        r.get("来源文件", ""),
                        r.get("来源Sheet", ""),
                        r.get("来源行号", ""),
                        r.get("匹配名称", ""),
                    ])

        created += 1

    return created


def build_equity_statement_sheet(
    wb: Workbook,
    years: List[str],
    bs_values: Dict[Tuple[str, str], Optional[float]],
    is_values: Dict[Tuple[str, str], Optional[float]],
    detail_payload: Dict[Tuple[str, str, str], List[dict]],
    missing_rows: List[dict],
) -> Dict[Tuple[str, str], Optional[float]]:
    ws = wb.create_sheet("所有者权益变动表")
    ws.append(["科目编码", "科目名称"] + [f"{y}年（万元）" for y in years] + [f"{y}年状态" for y in years])

    items = [
        ("EQ001", "年初所有者权益", "所有者权益合计", "derived_prev_year"),
        ("EQ002", "本年净利润", "净利润", "from_income"),
        ("EQ003", "年末所有者权益", "所有者权益合计", "from_bs"),
    ]

    out: Dict[Tuple[str, str], Optional[float]] = {}

    for code, name, ref_name, mode in items:
        row_vals = [code, name]
        status_vals = []
        for idx, y in enumerate(years):
            val = None
            status = "待补充"

            if mode == "from_bs":
                val = bs_values.get((ref_name, y))
                status = "已识别" if val is not None else "待补充"
            elif mode == "from_income":
                val = is_values.get((ref_name, y))
                status = "已识别" if val is not None else "待补充"
            elif mode == "derived_prev_year":
                if idx > 0:
                    prev = years[idx - 1]
                    val = bs_values.get(("所有者权益合计", prev))
                status = "已推导" if val is not None else "待补充"

            out[(name, y)] = val
            row_vals.append(val)
            status_vals.append(status)

            detail_payload.setdefault(("所有者权益变动表", name, code), []).append(
                {
                    "报表": "所有者权益变动表",
                    "科目编码": code,
                    "科目名称": name,
                    "期间": y,
                    "值(万元)": val,
                    "状态": status,
                    "来源类型": "derived" if "推导" in status else "mapped",
                    "来源文件": "",
                    "来源Sheet": "",
                    "来源行号": "",
                    "匹配名称": ref_name,
                }
            )

            if val is None:
                missing_rows.append(
                    {
                        "报表": "所有者权益变动表",
                        "科目编码": code,
                        "科目名称": name,
                        "期间": y,
                        "状态": "待补充",
                        "说明": "权益表关键字段未能生成",
                    }
                )

        ws.append(row_vals + status_vals)

    return out


def build_cash_supp_sheet(
    wb: Workbook,
    years: List[str],
    detail_payload: Dict[Tuple[str, str, str], List[dict]],
    missing_rows: List[dict],
    is_values: Dict[Tuple[str, str], Optional[float]],
    cf_values: Dict[Tuple[str, str], Optional[float]],
    bs_values: Dict[Tuple[str, str], Optional[float]],
) -> Dict[Tuple[str, str], Optional[float]]:
    ws = wb.create_sheet("现金流补充资料")
    ws.append(["科目编码", "科目名称"] + [f"{y}年（万元）" for y in years] + [f"{y}年状态" for y in years])

    items = [
        ("CS001", "净利润"),
        ("CS002", "固定资产折旧"),
        ("CS003", "无形资产摊销"),
        ("CS004", "递延所得税资产减少（增加）"),
        ("CS005", "递延所得税负债增加（减少）"),
        ("CS006", "经营活动产生的现金流量净额（补充）"),
    ]

    out: Dict[Tuple[str, str], Optional[float]] = {}
    for code, name in items:
        row_vals = [code, name]
        status_vals = []
        for idx, y in enumerate(years):
            value = None
            status = "待补充"
            source_type = ""
            matched_name = name

            if name == "净利润":
                value = is_values.get(("净利润", y))
                if value is not None:
                    status = "已推导"
                    source_type = "derived_from_income_statement"
                    matched_name = "净利润"
            elif name == "经营活动产生的现金流量净额（补充）":
                value = cf_values.get(("经营活动产生的现金流量净额", y))
                if value is not None:
                    status = "已推导"
                    source_type = "derived_from_cashflow_statement"
                    matched_name = "经营活动产生的现金流量净额"

            out[(name, y)] = value
            row_vals.append(value)
            status_vals.append(status)

            if value is None:
                missing_rows.append(
                    {
                        "报表": "现金流补充资料",
                        "科目编码": code,
                        "科目名称": name,
                        "期间": y,
                        "状态": "待补充",
                        "说明": "当前结构化来源未覆盖该字段",
                    }
                )
            detail_payload.setdefault(("现金流补充资料", name, code), []).append(
                {
                    "报表": "现金流补充资料",
                    "科目编码": code,
                    "科目名称": name,
                    "期间": y,
                    "值(万元)": value,
                    "状态": status,
                    "来源类型": source_type,
                    "来源文件": "",
                    "来源Sheet": "",
                    "来源行号": "",
                    "匹配名称": matched_name,
                }
            )
        ws.append(row_vals + status_vals)

    return out


def build_interest_debt_sheet(
    wb: Workbook,
    years: List[str],
    bs_values: Dict[Tuple[str, str], Optional[float]],
    detail_payload: Dict[Tuple[str, str, str], List[dict]],
    missing_rows: List[dict],
    runtime_cfg: Optional[Dict[str, Any]] = None,
) -> Dict[Tuple[str, str], Optional[float]]:
    """Build standalone interest-bearing debt detail sheet."""
    ws = get_sheet_by_name_loose(wb, "有息负债明细")
    if ws is None:
        ws = wb.create_sheet("有息负债明细")
    else:
        ws.delete_rows(1, ws.max_row)
    ws.append(["期间", "子项", "值(万元)", "说明"])

    items = get_interest_debt_items(runtime_cfg)
    out: Dict[Tuple[str, str], Optional[float]] = {}

    for y in years:
        total = 0.0
        any_value = False
        for name in items:
            v = bs_values.get((name, y))
            if v is not None:
                total += v
                any_value = True
            ws.append([y, name, v, "有息负债构成项"])
            detail_payload.setdefault(("有息负债明细", name, "DEBT001"), []).append(
                {
                    "报表": "有息负债明细",
                    "科目编码": "DEBT001",
                    "科目名称": name,
                    "期间": y,
                    "值(万元)": v,
                    "状态": "已识别" if v is not None else "待补充",
                    "来源类型": "from_bs",
                    "来源文件": "",
                    "来源Sheet": "",
                    "来源行号": "",
                    "匹配名称": name,
                }
            )

        total_val = round(total, 2) if any_value else None
        out[("有息负债合计", y)] = total_val
        ws.append([y, "有息负债合计", total_val, "自动汇总"])
        detail_payload.setdefault(("有息负债明细", "有息负债合计", "DEBT999"), []).append(
            {
                "报表": "有息负债明细",
                "科目编码": "DEBT999",
                "科目名称": "有息负债合计",
                "期间": y,
                "值(万元)": total_val,
                "状态": "已推导" if total_val is not None else "待补充",
                "来源类型": "derived",
                "来源文件": "",
                "来源Sheet": "",
                "来源行号": "",
                "匹配名称": "有息负债合计",
            }
        )
        if total_val is None:
            missing_rows.append(
                {
                    "报表": "有息负债明细",
                    "科目编码": "DEBT999",
                    "科目名称": "有息负债合计",
                    "期间": y,
                    "状态": "待补充",
                    "说明": "有息负债构成项未识别完整，无法汇总",
                }
            )

    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Build standards-first basic financial workbook")
    parser.add_argument(
        "--project-id",
        default="default_project",
        help="Project identifier used for output isolation (default: default_project)",
    )
    parser.add_argument(
        "--input-dir",
        type=str,
        default="",
        help="Input directory override; default scans inputs/<project_id>/ then inputs/",
    )
    parser.add_argument("--workpaper-path", type=str, default="", help="Optional explicit workpaper Excel path")
    parser.add_argument("--bs-path", type=str, default="", help="Optional explicit balance-sheet Excel path")
    parser.add_argument("--output", type=str, default="", help="Optional output path override (.xlsx)")
    parser.add_argument("--skip-if-exists", action="store_true", help="Skip generation if output file already exists")
    args = parser.parse_args()
    runtime_cfg = load_runtime_cfg()
    project_id = normalize_project_id(args.project_id)
    global WORKPAPER_PATH, BS_PATH, WORKPAPER_CANDIDATES, BS_CANDIDATES
    WORKPAPER_PATH, BS_PATH, WORKPAPER_CANDIDATES, BS_CANDIDATES, discovery = discover_project_sources(
        project_id, args.input_dir, args.workpaper_path, args.bs_path
    )
    discovery_path = PROJECT_ROOT / "data" / project_id / "input_discovery.json"
    discovery_path.parent.mkdir(parents=True, exist_ok=True)
    discovery_path.write_text(json.dumps(discovery, ensure_ascii=False, indent=2), encoding="utf-8")

    if not TEMPLATE_WORKBOOK_PATH.exists():
        print(f"Template not found: {TEMPLATE_WORKBOOK_PATH}")
        return 1

    year_detect_sources: List[Path] = []
    year_detect_sources.extend(WORKPAPER_CANDIDATES or [])
    year_detect_sources.extend(BS_CANDIDATES or [])
    if not year_detect_sources:
        year_detect_sources = [WORKPAPER_PATH, BS_PATH]
    years = detect_latest_three_years(year_detect_sources)
    years = sorted(years)

    workpaper_map = load_workpaper_records(years, WORKPAPER_CANDIDATES)
    bs_map = load_bs_records(years, BS_CANDIDATES)
    source_maps = {"workpaper": workpaper_map, "audit_excel": bs_map}
    default_output = PROJECT_ROOT / "outputs" / project_id / f"{project_id}_{OUTPUT_FILENAME}"

    output_path = Path(args.output) if args.output else default_output
    if args.skip_if_exists and output_path.exists():
        print(f"Skipped: {output_path} already exists")
        print(f"Years: {', '.join(years)}")
        print("Mode: real-source")
        print(f"Workpaper source: {WORKPAPER_PATH}")
        print(f"Balance-sheet source: {BS_PATH}")
        print(f"Workpaper candidates: {len(WORKPAPER_CANDIDATES)}")
        print(f"Balance-sheet candidates: {len(BS_CANDIDATES)}")
        print(f"Discovery manifest: {discovery_path}")
        return 0

    wb = load_workbook(TEMPLATE_WORKBOOK_PATH)

    detail_payload: Dict[Tuple[str, str, str], List[dict]] = {}
    missing_rows: List[dict] = []

    all_values: Dict[str, Dict[Tuple[str, str], Optional[float]]] = {}

    statement_defs = [
        ("资产负债表", "BS"),
        ("利润表", "IS"),
        ("现金流量表", "CF"),
    ]
    for sname, code_prefix in statement_defs:
        items = load_statement_items_from_template(wb, sname, code_prefix)
        values = write_statement_sheet_from_template(
            wb=wb,
            sheet_name=sname,
            code_prefix=code_prefix,
            items=items,
            years=years,
            source_maps=source_maps,
            detail_payload=detail_payload,
            missing_rows=missing_rows,
        )
        all_values[sname] = values

    eq_values = build_equity_statement_sheet(
        wb,
        years,
        all_values.get("资产负债表", {}),
        all_values.get("利润表", {}),
        detail_payload,
        missing_rows,
    )
    all_values["所有者权益变动表"] = eq_values

    debt_values = build_interest_debt_sheet(
        wb,
        years,
        all_values.get("资产负债表", {}),
        detail_payload,
        missing_rows,
        runtime_cfg=runtime_cfg,
    )
    all_values["有息负债明细"] = debt_values

    delete_sheet_names = runtime_cfg.get("rebuild_sheets", []) + runtime_cfg.get("drop_sheets", [])
    for name in delete_sheet_names:
        ws = get_sheet_by_name_loose(wb, name)
        if ws is not None:
            del wb[ws.title]

    recon_rules = load_recon_rules()
    ratio_alias_map = load_ratio_alias_map()
    ratio_cfg = load_ratio_cfg()
    tolerance_abs = float(runtime_cfg.get("recon_tolerance_abs", 1.0))
    code_aliases = get_code_aliases(runtime_cfg)
    build_recon_sheet_from_template_rules(wb, years, recon_rules, tolerance_abs, code_aliases=code_aliases)
    build_ratio_sheet(
        wb,
        years,
        all_values,
        ratio_cfg,
        WORKPAPER_CANDIDATES,
        ratio_alias_map,
        runtime_cfg=runtime_cfg,
    )
    build_missing_sheet(wb, missing_rows)
    detail_sheet_count = 0

    # Keep key sheets first for readability.
    first_order = runtime_cfg.get("sheet_order", [])
    if first_order:
        wb._sheets.sort(key=lambda ws: first_order.index(ws.title) if ws.title in first_order else 100)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Generated: {output_path}")
    print(f"Years: {', '.join(years)}")
    print(f"Detail sheets: {detail_sheet_count}")
    print(f"Missing rows: {len(missing_rows)}")
    print("Mode: real-source")
    print(f"Workpaper source: {WORKPAPER_PATH}")
    print(f"Balance-sheet source: {BS_PATH}")
    print(f"Workpaper candidates: {len(WORKPAPER_CANDIDATES)}")
    print(f"Balance-sheet candidates: {len(BS_CANDIDATES)}")
    print(f"Discovery manifest: {discovery_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
