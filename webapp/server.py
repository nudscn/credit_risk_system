#!/usr/bin/env python3
"""Local web UI for reviewing workbook sheets and narrative edits.

Run:
  python webapp/server.py --host 127.0.0.1 --port 8787
"""

from __future__ import annotations

import argparse
import base64
import json
import re
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PROJECT_ID = "zhonggang_luonai"
DEFAULT_YEAR_COUNT = 3

SHEET_GROUPS = [
    {"id": "bs", "label": "资产负债表", "candidates": ["资产负债表", "资产负债表 "]},
    {"id": "is", "label": "利润表", "candidates": ["利润表", "利润表 "]},
    {"id": "cf", "label": "现金流量表", "candidates": ["现金流量表", "现金流量表 "]},
    {"id": "ratio", "label": "财务指标表", "candidates": ["财务比率", "财务指标", "财务指标表"]},
]

ASSET_ANALYSIS_SHEETS = {
    "scale": ["分析_规模变化_资产类"],
    "structure": ["分析_结构占比_资产类"],
}
LIABILITY_ANALYSIS_SHEETS = {
    "scale": ["分析_规模变化_负债类"],
    "structure": ["分析_结构占比_负债类"],
}
SUMMARY_ANALYSIS_ITEMS = [
    {"code": "SUM001", "name": "流动资产", "kind": "bs", "source_code": "BS000", "code_candidates": ["BS000", "BS027"]},
    {"code": "SUM002", "name": "非流动资产", "kind": "bs", "source_code": "BS028", "code_candidates": ["BS028", "BS056"]},
    {"code": "SUM003", "name": "资产总计", "kind": "bs", "source_code": "BS057", "code_candidates": ["BS057"]},
    {"code": "SUM004", "name": "流动负债", "kind": "bs", "source_code": "BS058", "code_candidates": ["BS058", "BS089"]},
    {"code": "SUM005", "name": "非流动负债", "kind": "bs", "source_code": "BS090", "code_candidates": ["BS090", "BS102"]},
    {"code": "SUM006", "name": "负债总计", "kind": "bs", "source_code": "BS103", "code_candidates": ["BS103"]},
    {"code": "SUM007", "name": "营业收入", "kind": "is", "source_code": "IS001", "code_candidates": ["IS001"], "name_keywords": ["营业总收入", "营业收入"]},
    {"code": "SUM008", "name": "经营净收益", "kind": "income_combo", "source_code": "3.1+3.2"},
    {"code": "SUM009", "name": "主营业务毛利收益", "kind": "income_node", "source_code": "2.1.1"},
    {"code": "SUM010", "name": "经常性净收益", "kind": "income_node", "source_code": "3.1"},
    {"code": "SUM011", "name": "非经常性净收益", "kind": "income_node", "source_code": "3.2"},
]
SUMMARY_TOP_COVERAGE_TARGET_PCT = 75.0
SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME = 90.0
SUMMARY_EXCLUDE_CODES = {
    "SUM001": {"BS004", "BS009", "BS015"},  # 流动资产：排除不同层级科目
    "SUM002": {"BS041", "BS044"},  # 非流动资产：排除不同层级科目
    "SUM004": {"BS062", "BS070"},  # 流动负债：排除不同层级科目
    "SUM005": {"BS094"},  # 非流动负债：排除不同层级科目
}
ANALYSIS_CODE_REDIRECTS = {
    "asset_analysis": {
        "BS028": "BS056",  # 非流动资产 <- 非流动资产合计
    },
    "liability_analysis": {
        "BS058": "BS089",  # 流动负债 <- 流动负债合计
        "BS090": "BS102",  # 非流动负债 <- 非流动负债合计
    },
}

INCOME_RULEBOOK_PATH = PROJECT_ROOT / "config" / "income_profit_rulebook_skeleton.xlsx"
INCOME_RULEBOOK_FALLBACK_PATH = PROJECT_ROOT / "config" / "rulebook.xlsx"
RATIO_RULEBOOK_PATH = PROJECT_ROOT / "config" / "ratio_analysis_rulebook.xlsx"
KEY_RATIO_RULEBOOK_PATH = PROJECT_ROOT / "config" / "key_ratio_rulebook.xlsx"
VALIDATION_RULEBOOK_PATH = PROJECT_ROOT / "config" / "validation_rulebook.xlsx"
RULE_WARNING_LOG: List[Dict[str, Any]] = []
RULE_WARNING_KEYS: set = set()
RULE_EDIT_SOURCES = [
    {
        "source_id": "rulebook_main",
        "label": "资产负债分析规则（rulebook.xlsx）",
        "path": PROJECT_ROOT / "config" / "rulebook.xlsx",
        "sheets": ["analysis_text_templates", "analysis_thresholds"],
    },
    {
        "source_id": "profit_rulebook",
        "label": "利润表规则（rulebook.xlsx）",
        "path": PROJECT_ROOT / "config" / "rulebook.xlsx",
        "sheets": ["analysis_text_templates"],
    },
    {
        "source_id": "ratio_rulebook",
        "label": "财务指标规则（ratio_analysis_rulebook.xlsx）",
        "path": PROJECT_ROOT / "config" / "ratio_analysis_rulebook.xlsx",
        "sheets": ["indicator_tree", "indicator_catalog", "trend_rules", "judgement_rules", "alert_rules", "text_templates", "display_policy"],
    },
    {
        "source_id": "income_rulebook",
        "label": "收入利润规则（income_profit_rulebook_skeleton.xlsx）",
        "path": PROJECT_ROOT / "config" / "income_profit_rulebook_skeleton.xlsx",
        "sheets": ["text_templates", "trend_thresholds", "sign_scenario_policy", "contribution_policy"],
    },
    {
        "source_id": "key_ratio_rulebook",
        "label": "重点指标规则（key_ratio_rulebook.xlsx）",
        "path": PROJECT_ROOT / "config" / "key_ratio_rulebook.xlsx",
        "sheets": ["narrative_templates", "driver_thresholds"],
    },
    {
        "source_id": "validation_rulebook",
        "label": "报表校验规则（validation_rulebook.xlsx）",
        "path": PROJECT_ROOT / "config" / "validation_rulebook.xlsx",
        "sheets": ["bs_checks", "is_checks", "cf_checks"],
    },
]


def _rulebook_template_scope_match(source_id: str, row: Dict[str, Any]) -> bool:
    key = str((row or {}).get("template_key", "") or "").strip().lower()
    if not key:
        return False
    asset_liab_prefixes = ("asset_", "liability_", "scale_phrase_", "struct_phrase_")
    is_asset_liab = key.startswith(asset_liab_prefixes)
    if source_id == "rulebook_main":
        return is_asset_liab
    if source_id == "profit_rulebook":
        return not is_asset_liab
    return True


def _detect_template_workbook_path() -> Optional[Path]:
    cfg_dir = PROJECT_ROOT / "config"
    ignore = {
        "rulebook.xlsx",
        "ratio_analysis_rulebook.xlsx",
        "income_profit_rulebook_skeleton.xlsx",
        "key_ratio_rulebook.xlsx",
        "validation_rulebook.xlsx",
    }
    candidates: List[tuple] = []
    for p in cfg_dir.glob("*.xlsx"):
        if p.name.lower() in ignore:
            continue
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
            score = len(wb.sheetnames)
            wb.close()
            candidates.append((score, p))
        except Exception:
            continue
    if not candidates:
        return None
    return sorted(candidates, key=lambda x: x[0], reverse=True)[0][1]


def _warn_rule_once(key: str, message: str, severity: str = "warn", source: str = "rules") -> None:
    k = str(key or "").strip() or f"{source}:{message}"
    if k in RULE_WARNING_KEYS:
        return
    RULE_WARNING_KEYS.add(k)
    RULE_WARNING_LOG.append(
        {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "severity": str(severity or "warn"),
            "source": str(source or "rules"),
            "key": k,
            "message": str(message or ""),
        }
    )


def _collect_code_catalog_for_preflight() -> Dict[str, set]:
    out = {"bs": set(), "is": set(), "cf": set()}
    paths: List[Path] = []
    tp = _detect_template_workbook_path()
    if tp is not None and tp.exists():
        paths.append(tp)
    try:
        wp = workbook_path(DEFAULT_PROJECT_ID)
        if wp.exists():
            paths.append(wp)
    except Exception:
        pass
    for p in paths:
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
        except Exception:
            continue
        for sid in ["bs", "is", "cf"]:
            g = next((x for x in SHEET_GROUPS if x.get("id") == sid), None)
            if not g:
                continue
            ws = get_sheet_by_loose_name(wb, g.get("candidates", []))
            if ws is None:
                continue
            pref = sid.upper()
            for r in range(2, ws.max_row + 1):
                code = str(ws.cell(r, 1).value or "").strip().upper()
                if re.match(rf"^{pref}\d{{3}}$", code):
                    out[sid].add(code)
        wb.close()
    return out


def run_rule_preflight_checks() -> None:
    # 1) rule source file / sheet existence checks
    for src in RULE_EDIT_SOURCES:
        sid = str(src.get("source_id", "")).strip()
        p = Path(src.get("path"))
        sheets = list(src.get("sheets", []) or [])
        if not p.exists():
            _warn_rule_once(f"missing_file:{sid}", f"规则文件缺失：{p}", source=sid)
            continue
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            _warn_rule_once(f"open_fail:{sid}", f"规则文件无法读取：{p} | {e}", severity="error", source=sid)
            continue
        names = set([str(x).strip() for x in wb.sheetnames])
        for s in sheets:
            if str(s).strip() not in names:
                _warn_rule_once(f"missing_sheet:{sid}:{s}", f"规则文件缺少sheet：{sid}/{s}", source=sid)
        wb.close()

    # 2) validation rulebook structural checks
    if not VALIDATION_RULEBOOK_PATH.exists():
        _warn_rule_once("validation_rulebook_missing", f"校验规则文件缺失：{VALIDATION_RULEBOOK_PATH}", source="validation")
        return
    try:
        wbv = load_workbook(VALIDATION_RULEBOOK_PATH, data_only=True)
    except Exception as e:  # noqa: BLE001
        _warn_rule_once("validation_rulebook_open_fail", f"校验规则文件无法读取：{e}", severity="error", source="validation")
        return

    catalogs = _collect_code_catalog_for_preflight()
    for gid in ["bs", "is", "cf"]:
        ws = get_sheet_by_loose_name(wbv, [f"{gid}_checks"])
        if ws is None:
            _warn_rule_once(f"validation_sheet_missing:{gid}", f"校验规则缺少sheet：{gid}_checks", source="validation")
            continue
        headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        required = ["rule_id", "left_code", "operator", "right_codes", "enabled"]
        for h in required:
            if h not in headers:
                _warn_rule_once(
                    f"validation_header_missing:{gid}:{h}",
                    f"{gid}_checks 缺少字段：{h}",
                    severity="error",
                    source="validation",
                )
        c_rule_id = headers.get("rule_id", 1)
        c_left = headers.get("left_code", 3)
        c_op = headers.get("operator", 4)
        c_right = headers.get("right_codes", 5)
        c_enabled = headers.get("enabled", 7)
        seen_ids: set = set()
        for r in range(2, ws.max_row + 1):
            enabled_cell = ws.cell(r, c_enabled).value
            enabled_raw = str("1" if enabled_cell is None else enabled_cell).strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            rid = str(ws.cell(r, c_rule_id).value or "").strip()
            left = str(ws.cell(r, c_left).value or "").strip().upper()
            op = str(ws.cell(r, c_op).value or "").strip().upper()
            right = str(ws.cell(r, c_right).value or "").strip()
            if not rid:
                _warn_rule_once(f"validation_no_id:{gid}:{r}", f"{gid}_checks 第{r}行缺少 rule_id", source="validation")
            elif rid in seen_ids:
                _warn_rule_once(f"validation_dup_id:{gid}:{rid}", f"{gid}_checks 存在重复 rule_id：{rid}", source="validation")
            seen_ids.add(rid)
            if op not in {"SUM", "EXPR"}:
                _warn_rule_once(
                    f"validation_bad_op:{gid}:{rid}",
                    f"{gid}_checks 规则{rid} 运算符非法：{op}（仅支持 SUM/EXPR）",
                    severity="error",
                    source="validation",
                )
            if not left or not re.match(rf"^{gid.upper()}\d{{3}}$", left):
                _warn_rule_once(
                    f"validation_bad_left:{gid}:{rid}",
                    f"{gid}_checks 规则{rid} 左值编码异常：{left}",
                    source="validation",
                )
            if not right:
                _warn_rule_once(
                    f"validation_empty_right:{gid}:{rid}",
                    f"{gid}_checks 规则{rid} 右值为空",
                    source="validation",
                )
                continue
            toks = [x.strip() for x in right.replace(";", ",").split(",") if x.strip()]
            for t in toks:
                code = t.lstrip("+-").strip().upper()
                if not re.match(rf"^{gid.upper()}\d{{3}}$", code):
                    _warn_rule_once(
                        f"validation_bad_code:{gid}:{rid}:{code}",
                        f"{gid}_checks 规则{rid} 引用编码异常：{code}",
                        source="validation",
                    )
                    continue
                if catalogs.get(gid) and code not in catalogs.get(gid, set()):
                    _warn_rule_once(
                        f"validation_code_not_found:{gid}:{rid}:{code}",
                        f"{gid}_checks 规则{rid} 引用编码在当前目录未找到：{code}",
                        source="validation",
                    )


def _load_bs_code_name_catalog() -> List[Dict[str, str]]:
    tpl = _detect_template_workbook_path()
    if tpl is None:
        return []
    try:
        wb = load_workbook(tpl, data_only=True, read_only=True)
    except Exception:
        return []
    ws_bs = None
    for ws in wb.worksheets:
        hit = 0
        for r in range(2, min(ws.max_row, 220) + 1):
            v = str(ws.cell(r, 1).value or "").strip()
            if re.match(r"^BS\d{3}$", v):
                hit += 1
        if hit >= 20:
            ws_bs = ws
            break
    out: List[Dict[str, str]] = []
    if ws_bs is None:
        wb.close()
        return out
    for r in range(2, ws_bs.max_row + 1):
        code = str(ws_bs.cell(r, 1).value or "").strip()
        name = str(ws_bs.cell(r, 2).value or "").strip().rstrip("：:").strip()
        if re.match(r"^BS\d{3}$", code):
            out.append({"code": code, "name": name})
    wb.close()
    out.sort(key=lambda x: x["code"])
    return out


def _read_threshold_config() -> Dict[str, Any]:
    p = PROJECT_ROOT / "config" / "rulebook.xlsx"
    global_scale_pct = 2.0
    global_struct_pp = 2.0
    summary_top_coverage_default = float(SUMMARY_TOP_COVERAGE_TARGET_PCT)
    summary_top_coverage_income = float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)
    subject_scale_map: Dict[str, float] = {}
    subject_struct_map: Dict[str, float] = {}
    if not p.exists():
        _warn_rule_once("threshold_rulebook_missing", f"阈值规则文件缺失，已使用默认阈值：{p}", source="rulebook_main")
        return {
            "global_scale_pct": global_scale_pct,
            "global_struct_pp": global_struct_pp,
            "summary_top_coverage_default": summary_top_coverage_default,
            "summary_top_coverage_income": summary_top_coverage_income,
            "subject_scale_pct": subject_scale_map,
            "subject_struct_pp": subject_struct_map,
            # backward-compatible aliases
            "global_pct": global_scale_pct,
            "subject_pct": subject_scale_map,
        }
    try:
        wb = load_workbook(p, data_only=True)
    except Exception:
        _warn_rule_once("threshold_rulebook_open_fail", f"阈值规则文件读取失败，已使用默认阈值：{p}", source="rulebook_main")
        return {
            "global_scale_pct": global_scale_pct,
            "global_struct_pp": global_struct_pp,
            "summary_top_coverage_default": summary_top_coverage_default,
            "summary_top_coverage_income": summary_top_coverage_income,
            "subject_scale_pct": subject_scale_map,
            "subject_struct_pp": subject_struct_map,
            "global_pct": global_scale_pct,
            "subject_pct": subject_scale_map,
        }
    ws = get_sheet_by_loose_name(wb, ["analysis_thresholds", "分析阈值配置"])
    if ws is None:
        _warn_rule_once("threshold_sheet_missing", "analysis_thresholds 缺失，已使用默认阈值", source="rulebook_main")
        return {
            "global_scale_pct": global_scale_pct,
            "global_struct_pp": global_struct_pp,
            "summary_top_coverage_default": summary_top_coverage_default,
            "summary_top_coverage_income": summary_top_coverage_income,
            "subject_scale_pct": subject_scale_map,
            "subject_struct_pp": subject_struct_map,
            "global_pct": global_scale_pct,
            "subject_pct": subject_scale_map,
        }
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_scope = headers.get("scope", 1)
    c_subj = headers.get("subject_code", 2)
    c_scale = headers.get("scale_stable_pct", headers.get("stable_threshold_pct", 3))
    c_struct = headers.get("struct_stable_pp", c_scale)
    c_enabled = headers.get("enabled", 4)
    for r in range(2, ws.max_row + 1):
        enabled_cell = ws.cell(r, c_enabled).value
        enabled_raw = str("1" if enabled_cell is None else enabled_cell).strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        scope = str(ws.cell(r, c_scope).value or "").strip().lower()
        subj = str(ws.cell(r, c_subj).value or "").strip().upper()
        try:
            v_scale = float(ws.cell(r, c_scale).value)
        except Exception:
            continue
        try:
            v_struct = float(ws.cell(r, c_struct).value)
        except Exception:
            v_struct = v_scale
        if scope in {"global", "全局"}:
            global_scale_pct = v_scale
            global_struct_pp = v_struct
        elif scope in {"summary", "汇总"}:
            # Reuse scale_stable_pct column as summary-top coverage threshold.
            if subj in {"SUM_ALL", "SUMMARY_DEFAULT"}:
                summary_top_coverage_default = v_scale
            elif subj in {"SUM007", "INCOME"}:
                summary_top_coverage_income = v_scale
        elif scope in {"subject", "科目"} and re.match(r"^BS\d{3}$", subj):
            subject_scale_map[subj] = v_scale
            subject_struct_map[subj] = v_struct
    return {
        "global_scale_pct": global_scale_pct,
        "global_struct_pp": global_struct_pp,
        "summary_top_coverage_default": summary_top_coverage_default,
        "summary_top_coverage_income": summary_top_coverage_income,
        "subject_scale_pct": subject_scale_map,
        "subject_struct_pp": subject_struct_map,
        "global_pct": global_scale_pct,
        "subject_pct": subject_scale_map,
    }


def _save_threshold_config(
    global_scale_pct: float,
    global_struct_pp: float,
    subject_rows: List[Dict[str, Any]],
    summary_top_coverage_default: float = SUMMARY_TOP_COVERAGE_TARGET_PCT,
    summary_top_coverage_income: float = SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME,
) -> Dict[str, Any]:
    p = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not p.exists():
        return {"ok": False, "error": f"rulebook not found: {p}"}
    try:
        wb = load_workbook(p, data_only=False)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e)}
    ws = get_sheet_by_loose_name(wb, ["analysis_thresholds", "分析阈值配置"])
    if ws is None:
        ws = wb.create_sheet("analysis_thresholds")
    headers = ["scope", "subject_code", "scale_stable_pct", "struct_stable_pp", "enabled", "notes"]
    # reset sheet
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    ws.append(["global", "", float(global_scale_pct), float(global_struct_pp), 1, "全局默认阈值"])
    ws.append(["summary", "SUM_ALL", float(summary_top_coverage_default), float(summary_top_coverage_default), 1, "分析汇总Top构成覆盖阈值(默认)"])
    ws.append(["summary", "SUM007", float(summary_top_coverage_income), float(summary_top_coverage_income), 1, "分析汇总Top构成覆盖阈值(营业收入)"])
    saved_subject = 0
    for r in subject_rows:
        code = str(r.get("code", "")).strip().upper()
        enabled = bool(r.get("enabled", False))
        if not enabled:
            continue
        if not re.match(r"^BS\d{3}$", code):
            continue
        try:
            v_scale = float(r.get("scale_threshold_pct", r.get("threshold_pct")))
        except Exception:
            continue
        try:
            v_struct = float(r.get("struct_threshold_pp", v_scale))
        except Exception:
            v_struct = v_scale
        name = str(r.get("name", "")).strip()
        ws.append(["subject", code, v_scale, v_struct, 1, name])
        saved_subject += 1
    wb.save(p)
    return {
        "ok": True,
        "global_scale_pct": float(global_scale_pct),
        "global_struct_pp": float(global_struct_pp),
        "summary_top_coverage_default": float(summary_top_coverage_default),
        "summary_top_coverage_income": float(summary_top_coverage_income),
        "saved_subject_rows": saved_subject,
        "path": str(p),
        "sheet": ws.title,
    }


def _read_analysis_thresholds_expanded(source_id: str, sheet_name: str) -> Dict[str, Any]:
    cfg = _read_threshold_config()
    catalog = _load_bs_code_name_catalog()
    subj_scale = cfg.get("subject_scale_pct", {}) if isinstance(cfg.get("subject_scale_pct"), dict) else {}
    subj_struct = cfg.get("subject_struct_pp", {}) if isinstance(cfg.get("subject_struct_pp"), dict) else {}
    headers = ["scope", "subject_code", "subject_name", "scale_stable_pct", "struct_stable_pp", "enabled", "notes"]
    rows: List[Dict[str, Any]] = []
    rows.append(
        {
            "_row": 2,
            "scope": "global",
            "subject_code": "",
            "subject_name": "全局默认",
            "scale_stable_pct": cfg.get("global_scale_pct", cfg.get("global_pct", 2.0)),
            "struct_stable_pp": cfg.get("global_struct_pp", cfg.get("global_pct", 2.0)),
            "enabled": 1,
            "notes": "全局默认阈值",
        }
    )
    rows.append(
        {
            "_row": 3,
            "scope": "summary",
            "subject_code": "SUM_ALL",
            "subject_name": "分析汇总Top构成覆盖阈值(默认)",
            "scale_stable_pct": cfg.get("summary_top_coverage_default", float(SUMMARY_TOP_COVERAGE_TARGET_PCT)),
            "struct_stable_pp": cfg.get("summary_top_coverage_default", float(SUMMARY_TOP_COVERAGE_TARGET_PCT)),
            "enabled": 1,
            "notes": "用于分析汇总页除营业收入外的Top构成披露阈值(%)",
        }
    )
    rows.append(
        {
            "_row": 4,
            "scope": "summary",
            "subject_code": "SUM007",
            "subject_name": "分析汇总Top构成覆盖阈值(营业收入)",
            "scale_stable_pct": cfg.get("summary_top_coverage_income", float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)),
            "struct_stable_pp": cfg.get("summary_top_coverage_income", float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)),
            "enabled": 1,
            "notes": "用于分析汇总页营业收入Top构成披露阈值(%)",
        }
    )
    rid = 5
    for x in catalog:
        code = str(x.get("code", "")).strip().upper()
        if not re.match(r"^BS\d{3}$", code):
            continue
        name = str(x.get("name", "")).strip()
        rows.append(
            {
                "_row": rid,
                "scope": "subject",
                "subject_code": code,
                "subject_name": name,
                "scale_stable_pct": subj_scale.get(code, cfg.get("global_scale_pct", cfg.get("global_pct", 2.0))),
                "struct_stable_pp": subj_struct.get(code, cfg.get("global_struct_pp", cfg.get("global_pct", 2.0))),
                "enabled": 1 if code in subj_scale or code in subj_struct else 0,
                "notes": name,
            }
        )
        rid += 1
    return {
        "source_id": source_id,
        "sheet_name": sheet_name,
        "path": str(PROJECT_ROOT / "config" / "rulebook.xlsx"),
        "headers": headers,
        "rows": rows,
    }


def _save_analysis_thresholds_expanded(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    global_scale_pct = 2.0
    global_struct_pp = 2.0
    summary_top_coverage_default = float(SUMMARY_TOP_COVERAGE_TARGET_PCT)
    summary_top_coverage_income = float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)
    subject_rows: List[Dict[str, Any]] = []
    for row in rows:
        scope = str(row.get("scope", "")).strip().lower()
        code = str(row.get("subject_code", "")).strip().upper()
        enabled_raw = str(row.get("enabled", "")).strip().lower()
        enabled = enabled_raw in {"1", "true", "yes", "y", "on"}
        try:
            scale_v = float(row.get("scale_stable_pct", row.get("stable_threshold_pct")))
        except Exception:
            continue
        try:
            struct_v = float(row.get("struct_stable_pp", scale_v))
        except Exception:
            struct_v = scale_v
        if scope in {"global", "全局"}:
            global_scale_pct = scale_v
            global_struct_pp = struct_v
            continue
        if scope in {"summary", "汇总"}:
            if code in {"SUM_ALL", "SUMMARY_DEFAULT"}:
                summary_top_coverage_default = scale_v
            elif code in {"SUM007", "INCOME"}:
                summary_top_coverage_income = scale_v
            continue
        if scope in {"subject", "科目"} and re.match(r"^BS\d{3}$", code):
            subject_rows.append(
                {
                    "code": code,
                    "name": str(row.get("subject_name", "")).strip() or str(row.get("notes", "")).strip(),
                    "enabled": enabled,
                    "scale_threshold_pct": scale_v,
                    "struct_threshold_pp": struct_v,
                }
            )
    return _save_threshold_config(
        global_scale_pct=global_scale_pct,
        global_struct_pp=global_struct_pp,
        subject_rows=subject_rows,
        summary_top_coverage_default=summary_top_coverage_default,
        summary_top_coverage_income=summary_top_coverage_income,
    )
DEFAULT_INCOME_SPECIAL_ITEMS = [
    {"node_id": "2.1.2.1", "label": "其他收益（政府补助等）", "code_candidates": [], "name_keywords": ["其他收益"]},
    {"node_id": "2.1.2.2", "label": "信用减值损失", "code_candidates": [], "name_keywords": ["信用减值损失"]},
    {"node_id": "2.1.2.3", "label": "资产减值损失", "code_candidates": [], "name_keywords": ["资产减值损失"]},
    {"node_id": "2.1.2.4", "label": "待判定收益项", "code_candidates": [], "name_keywords": []},
    {"node_id": "2.1.2.5", "label": "投资收益（待判定）", "code_candidates": [], "name_keywords": ["投资收益", "投资净收益"]},
    {"node_id": "2.1.2.6", "label": "公允价值变动收益（待判定）", "code_candidates": [], "name_keywords": ["公允价值变动收益", "公允价值变动净收益"]},
]
DEFAULT_RECURRING_KEYS = {"2.1.1", "2.1.2.1", "2.1.2.2", "2.1.2.3", "2.1.2.5", "2.1.2.6"}
DEFAULT_NONREC_KEYS = {"2.2.1", "2.2.2"}
DEFAULT_INCOME_FORMULAS = [
    {"node_id": "3.1", "formula": "SUM(2.1.1,2.1.2.1,2.1.2.2,2.1.2.3,2.1.2.5,2.1.2.6)"},
    {"node_id": "3.2", "formula": "SUM(2.2.1,2.2.2)"},
    {"node_id": "3.3", "formula": "SUB(3.1,3.2)"},
]
DEFAULT_INCOME_TREE = [
    {"node_id": "1", "parent_id": "", "label": "营业收入分析", "node_type": "fixed"},
    {"node_id": "1.1", "parent_id": "1", "label": "营业收入总额", "node_type": "fixed"},
    {
        "node_id": "",
        "parent_id": "1",
        "label": "",
        "node_type": "template",
        "template_name": "revenue_segment",
        "node_id_prefix": "1.",
        "start_index": 2,
        "label_suffix": "收入",
    },
    {"node_id": "2", "parent_id": "", "label": "收益贡献分析", "node_type": "fixed"},
    {"node_id": "2.1", "parent_id": "2", "label": "经常性收益分析", "node_type": "fixed"},
    {"node_id": "2.1.1", "parent_id": "2.1", "label": "主营业务毛利（自动汇总）", "node_type": "fixed"},
    {
        "node_id": "",
        "parent_id": "2.1.1",
        "label": "",
        "node_type": "template",
        "template_name": "gross_segment",
        "node_id_prefix": "2.1.1.",
        "start_index": 1,
        "label_suffix": "毛利",
    },
    {"node_id": "2.1.2", "parent_id": "2.1", "label": "其他经营收益", "node_type": "fixed"},
    {"node_id": "2.1.2.1", "parent_id": "2.1.2", "label": "其他收益（政府补助等）", "node_type": "fixed"},
    {"node_id": "2.1.2.2", "parent_id": "2.1.2", "label": "信用减值损失", "node_type": "fixed"},
    {"node_id": "2.1.2.3", "parent_id": "2.1.2", "label": "资产减值损失", "node_type": "fixed"},
    {"node_id": "2.1.2.4", "parent_id": "2.1.2", "label": "待判定收益项", "node_type": "fixed"},
    {"node_id": "2.1.2.5", "parent_id": "2.1.2", "label": "投资收益（待判定）", "node_type": "fixed"},
    {"node_id": "2.1.2.6", "parent_id": "2.1.2", "label": "公允价值变动收益（待判定）", "node_type": "fixed"},
    {"node_id": "2.2", "parent_id": "2", "label": "非经常性收益分析", "node_type": "fixed"},
    {"node_id": "2.2.1", "parent_id": "2.2", "label": "资产处置收益", "node_type": "fixed"},
    {"node_id": "2.2.2", "parent_id": "2.2", "label": "营业外收支净额（自动汇总）", "node_type": "fixed"},
    {"node_id": "3", "parent_id": "", "label": "分析输出", "node_type": "fixed"},
    {"node_id": "3.1", "parent_id": "3", "label": "经常性收益小计（自动汇总）", "node_type": "fixed"},
    {"node_id": "3.2", "parent_id": "3", "label": "非经常性收益小计（自动汇总）", "node_type": "fixed"},
    {"node_id": "3.3", "parent_id": "3", "label": "收益结构判断（自动+手动）", "node_type": "fixed"},
]
DEFAULT_RATIO_TREE = [
    {"node_id": "1", "parent_id": "", "label_zh": "偿债能力分析", "node_type": "group", "indicator_id": "", "enabled": 1, "sort_order": 10},
    {"node_id": "2", "parent_id": "", "label_zh": "盈利能力分析", "node_type": "group", "indicator_id": "", "enabled": 1, "sort_order": 20},
    {"node_id": "2.0", "parent_id": "2", "label_zh": "ROE杜邦专题", "node_type": "topic", "indicator_id": "topic_roe_dupont", "enabled": 1, "sort_order": 20.5},
    {"node_id": "2.6", "parent_id": "2", "label_zh": "毛利率贡献专题", "node_type": "topic", "indicator_id": "topic_gross_margin", "enabled": 1, "sort_order": 26},
    {"node_id": "3", "parent_id": "", "label_zh": "营运能力分析", "node_type": "group", "indicator_id": "", "enabled": 1, "sort_order": 30},
    {"node_id": "4", "parent_id": "", "label_zh": "现金流与结构分析", "node_type": "group", "indicator_id": "", "enabled": 1, "sort_order": 40},
]
KEY_RATIO_IDS = {"K1", "K2"}
DEFAULT_AMOUNT_UNIT = "元"

DEFAULT_TEXT_TEMPLATES = {
    "sheet_auto_no_latest": "{name}：最新年度暂无数值，建议补录后再分析。",
    "sheet_auto_one_year": "{name}：{latest}年为{latest_val}。",
    "sheet_auto_prev_missing": "{name}：{latest}年为{latest_val}。",
    "sheet_auto_two_year": "{name}：{latest}年为{latest_val}，较{prev}年{direction}{abs_diff}（{abs_pct}%）。",
    "income_segment_value": "{label}：{series_values}；{series_trends}。",
    "income_segment_share": "{ratio_label}：{series_values}；{series_trends}。",
    "gross_contribution_basis": "口径说明：{y1}年{b1}，{y2}年{b2}，{y3}年{b3}。",
    "gross_segment_net_attr": "净归因占比：{y1}年{n1}，{y2}年{n2}，{y3}年{n3}；{y2}较{y1}{nt21}，{y3}较{y2}{nt32}。",
    "gross_segment_dual_view": "{strength_label}：{s_latest}；{net_label}：{n_latest}。",
    "gross_segment_scenario_judgement": "场景定性（{latest_year}）：{qual_word}（{scenario}）。",
    "profit_summary_header_missing": "{title}（{latest_year}）：总额待补充。",
    "profit_summary_header_positive": "{title}（{latest_year}）：总额{total_amount}{unit}。",
    "profit_summary_header_zero": "{title}（{latest_year}）：总额为0（各分项正负对冲）。",
    "profit_summary_header_negative": "{title}（{latest_year}）：总额亏损{total_abs_amount}{unit}。",
    "profit_summary_line_positive": "{name}：贡献{amount}{unit}，{net_attr_label}{net_attr_pct}；{direction_word}。",
    "profit_summary_line_negative": "{name}：亏损{abs_amount}{unit}，{net_attr_label}{net_attr_pct}；{direction_word}。",
    "profit_summary_line_zero": "{name}：影响中性，{net_attr_label}{net_attr_pct}。",
    "ratio_indicator_value": "{name}：{series_values}。",
    "ratio_indicator_trend": "{series_trends}。",
    "key_roe_factors": "净利率：{y1}年{nm1}，{y2}年{nm2}，{y3}年{nm3}；总资产周转率：{y1}年{at1}，{y2}年{at2}，{y3}年{at3}；权益乘数：{y1}年{em1}，{y2}年{em2}，{y3}年{em3}。",
    "key_gm_top_segments": "最新年度分项贡献（按收入占比/分项毛利率）：{top_txt}。",
    "summary_abs_line": "{name}：{y1}年{v1}{unit}，{y2}年{v2}{unit}，{y3}年{v3}{unit}；{y2}较{y1}{p21}，{y3}较{y2}{p32}。",
    "summary_ratio_line": "{name}占{base_name}比例：{y1}年{r1}，{y2}年{r2}，{y3}年{r3}；{y2}较{y1}{d21}，{y3}较{y2}{d32}。",
    "summary_income_struct_intro": "{name}：分项结构如下。",
    "summary_total_not_applicable": "{name}：为汇总项，结构占比分析不适用。",
    "summary_struct_pending": "{name}：结构占比分析待补充。",
    "summary_yoy_up": "增加{value}%",
    "summary_yoy_down": "减少{value}%",
    "summary_yoy_stable": "基本稳定（变动{value}%）",
    "summary_pp_up": "上升{value}个百分点",
    "summary_pp_down": "下降{value}个百分点",
    "summary_pp_stable": "基本稳定（变动{value}个百分点）",
    "summary_top_missing": "结构迁移分析待补充。",
    "summary_top_line1": "{y3}年内部构成按占比从高到低分别为：{top_txt}，这几项占比合计{cum}%。",
    "summary_top_line2": "{y1}-{y3}年，{move_txt}；同时，{top_delta_txt}。",
    "summary_top_move_default": "内部结构整体稳定",
    "summary_top_delta_pending": "主要构成项变化待补充",
    "summary_see_income_page": "详见“收入分析”子页面。",
    "summary_sum008_struct": "{y3}年经常性净收益占比{s_rec_3}、非经常性净收益占比{s_non_3}；{y1}-{y3}年，经常性净收益占比{d_rec}；非经常性净收益占比{d_non}。",
    "summary_sum009_trend_pending": "各子业务毛利结构趋势待补充",
    "summary_sum009_struct": "{y3}年各子业务毛利影响占比及结构变化：{trend_txt}。",
    "summary_sum010_struct": "{y3}年经常性净收益中，主营业务毛利收益占比{gp_s3}，其他经常性收益占比{ot_s3}；{y1}-{y3}年，主营业务毛利收益占比{d_gp}，其他经常性收益占比{d_ot}。",
    "summary_sum011_struct": "{y3}年{name}占经营净收益比例为{r3}；{y1}-{y3}年占比{d_31}。",
}

DEFAULT_KEY_DRIVER_THRESHOLDS = {
    "significant_abs_contrib": 0.30,
    "single_driver_share": 0.40,
    "dual_driver_share_sum": 0.70,
    "dual_driver_each_min": 0.20,
    "delta_stable_pp": 2.00,
}
_KEY_RATIO_RULES_CACHE: Optional[Dict[str, Any]] = None
_MAIN_TEXT_TEMPLATE_CACHE: Optional[Dict[str, str]] = None
_MAIN_TEXT_TEMPLATE_UNIT_CACHE: Optional[Dict[str, str]] = None


def load_main_analysis_text_templates() -> Dict[str, str]:
    global _MAIN_TEXT_TEMPLATE_CACHE
    if isinstance(_MAIN_TEXT_TEMPLATE_CACHE, dict):
        return _MAIN_TEXT_TEMPLATE_CACHE
    out = {k: v for k, v in DEFAULT_TEXT_TEMPLATES.items() if k.startswith("summary_")}
    p = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not p.exists():
        _MAIN_TEXT_TEMPLATE_CACHE = out
        return out
    try:
        wb = load_workbook(p, data_only=True)
    except Exception:
        _MAIN_TEXT_TEMPLATE_CACHE = out
        return out
    ws = get_sheet_by_loose_name(wb, ["analysis_text_templates", "分析文本模板"])
    if ws is None:
        _MAIN_TEXT_TEMPLATE_CACHE = out
        return out
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_key = headers.get("template_key", 1)
    c_text = headers.get("template_text", 2)
    c_enabled = headers.get("enabled", 4)
    for r in range(2, ws.max_row + 1):
        enabled_raw = str(ws.cell(r, c_enabled).value or "1").strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        k = str(ws.cell(r, c_key).value or "").strip()
        v = str(ws.cell(r, c_text).value or "").strip()
        if not k or not _is_valid_text_cell(v):
            continue
        if k.startswith("summary_"):
            out[k] = v
    _MAIN_TEXT_TEMPLATE_CACHE = out
    return out


def load_main_analysis_template_units() -> Dict[str, str]:
    global _MAIN_TEXT_TEMPLATE_UNIT_CACHE
    if isinstance(_MAIN_TEXT_TEMPLATE_UNIT_CACHE, dict):
        return _MAIN_TEXT_TEMPLATE_UNIT_CACHE
    out: Dict[str, str] = {"summary_abs_line": DEFAULT_AMOUNT_UNIT}
    p = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not p.exists():
        _MAIN_TEXT_TEMPLATE_UNIT_CACHE = out
        return out
    try:
        wb = load_workbook(p, data_only=True)
    except Exception:
        _MAIN_TEXT_TEMPLATE_UNIT_CACHE = out
        return out
    ws = get_sheet_by_loose_name(wb, ["analysis_text_templates", "分析文本模板"])
    if ws is None:
        _MAIN_TEXT_TEMPLATE_UNIT_CACHE = out
        return out
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_key = headers.get("template_key", 1)
    c_enabled = headers.get("enabled", 4)
    c_unit = headers.get("unit")
    if not c_unit:
        _MAIN_TEXT_TEMPLATE_UNIT_CACHE = out
        return out
    for r in range(2, ws.max_row + 1):
        enabled_raw = str(ws.cell(r, c_enabled).value or "1").strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        k = str(ws.cell(r, c_key).value or "").strip()
        u = str(ws.cell(r, c_unit).value or "").strip()
        if k.startswith("summary_") and u:
            out[k] = u
    _MAIN_TEXT_TEMPLATE_UNIT_CACHE = out
    return out


def _render_template(template: str, values: Dict[str, Any]) -> str:
    txt = str(template or "")

    def repl(match: re.Match) -> str:
        key = match.group(1)
        val = values.get(key)
        if val is None:
            return "待补充"
        return str(val)

    return re.sub(r"\{([A-Za-z0-9_]+)\}", repl, txt).strip()


def normalize_project_id(project_id: str) -> str:
    text = re.sub(r"[^0-9A-Za-z._-]+", "_", (project_id or "").strip())
    return text or "default_project"


def workbook_path(project_id: str) -> Path:
    """Resolve workbook path by convention and fallback."""
    pid = normalize_project_id(project_id)
    out_dir = PROJECT_ROOT / "outputs" / pid
    preferred = out_dir / f"{pid}_项目主文件.xlsx"
    if preferred.exists():
        return preferred

    # Fallback to commonly used names.
    for name in ["基础数据_主文件.xlsx", f"{pid}_项目主文件_年份校验.xlsx"]:
        p = out_dir / name
        if p.exists():
            return p

    # Last resort: first xlsx in the project output folder.
    if out_dir.exists():
        for p in sorted(out_dir.glob("*.xlsx")):
            return p
    return preferred


def store_path(project_id: str) -> Path:
    pid = normalize_project_id(project_id)
    return PROJECT_ROOT / "data" / pid / "narrative_store.json"


def normalize_num(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text:
        return None

    # Support accounting negatives with parentheses, including full-width symbols.
    negative = False
    if (text.startswith("(") and text.endswith(")")) or (text.startswith("（") and text.endswith("）")):
        negative = True
        text = text[1:-1].strip()
    if text.startswith("-"):
        negative = True
        text = text[1:].strip()

    try:
        out = float(text)
    except Exception:
        return None
    return -out if negative else out


def get_sheet_by_loose_name(wb, candidates: List[str]):
    raw_map = {str(name).strip(): name for name in wb.sheetnames}
    for cand in candidates:
        key = str(cand).strip()
        if key in raw_map:
            return wb[raw_map[key]]
    return None


def parse_years_from_sheet(ws) -> List[str]:
    """Find years from header rows. Prefer row 3, fallback to row 1-6 scan."""
    years: List[str] = []

    def collect_from_row(row_idx: int) -> None:
        for col in range(3, ws.max_column + 1):
            text = str(ws.cell(row_idx, col).value or "")
            m = re.search(r"(20\d{2})", text)
            if not m:
                continue
            y = m.group(1)
            if y not in years:
                years.append(y)

    collect_from_row(3)
    if not years:
        for row_idx in range(1, min(ws.max_row, 6) + 1):
            collect_from_row(row_idx)
            if years:
                break
    return years[:DEFAULT_YEAR_COUNT]


def _safe_float(v: Any, default: float) -> float:
    try:
        return float(v)
    except Exception:
        return default


def _to_pct_threshold(value: float, unit: str) -> float:
    u = str(unit or "").strip().lower()
    if u == "ratio" and abs(value) <= 1:
        return value * 100.0
    return value


def _is_valid_text_cell(text: str) -> bool:
    t = str(text or "").strip()
    if not t:
        return False
    # Guard against mojibake/question-mark placeholders from bad encoding pipelines.
    if "?" in t:
        return False
    return True


def load_income_rules() -> Dict[str, Any]:
    cfg = {
        "yoy_stable_pct": 2.0,
        "share_stable_pp": 2.0,
        "trend_thresholds": {
            "rev_yoy": {"stable": 2.0, "unit": "ratio", "up_label": "增加", "down_label": "减少", "stable_label": "保持稳定"},
            "gp_value": {"stable": 2.0, "unit": "ratio", "up_label": "增加", "down_label": "减少", "stable_label": "保持稳定"},
            "rev_share": {"stable": 2.0, "unit": "pct_point", "up_label": "上升", "down_label": "下降", "stable_label": "基本稳定"},
            "gp_share_impact": {"stable": 2.0, "unit": "pct_point", "up_label": "上升", "down_label": "下降", "stable_label": "基本稳定"},
        },
        "contribution_policy": {
            "gp_share_mode": "impact_when_mixed_or_total_negative",
            "impact_ratio_label": "毛利贡献（驱动占比）",
            "algebraic_ratio_label": "毛利贡献（占总毛利比例）",
            "basis_impact_text": "按绝对值口径",
            "basis_algebraic_text": "按代数口径",
        },
        "special_items": list(DEFAULT_INCOME_SPECIAL_ITEMS),
        "recurring_keys": set(DEFAULT_RECURRING_KEYS),
        "nonrec_keys": set(DEFAULT_NONREC_KEYS),
        "formula_rules": list(DEFAULT_INCOME_FORMULAS),
        "tree_nodes": list(DEFAULT_INCOME_TREE),
        "text_templates": {
            "income_segment_value": DEFAULT_TEXT_TEMPLATES["income_segment_value"],
            "income_segment_share": DEFAULT_TEXT_TEMPLATES["income_segment_share"],
            "profit_summary_header_missing": DEFAULT_TEXT_TEMPLATES["profit_summary_header_missing"],
            "profit_summary_header_positive": DEFAULT_TEXT_TEMPLATES["profit_summary_header_positive"],
            "profit_summary_header_zero": DEFAULT_TEXT_TEMPLATES["profit_summary_header_zero"],
            "profit_summary_header_negative": DEFAULT_TEXT_TEMPLATES["profit_summary_header_negative"],
            "profit_summary_line_positive": DEFAULT_TEXT_TEMPLATES["profit_summary_line_positive"],
            "profit_summary_line_negative": DEFAULT_TEXT_TEMPLATES["profit_summary_line_negative"],
            "profit_summary_line_zero": DEFAULT_TEXT_TEMPLATES["profit_summary_line_zero"],
        },
        "text_template_units": {
            "income_segment_value": DEFAULT_AMOUNT_UNIT,
            "gross_segment_value": DEFAULT_AMOUNT_UNIT,
            "profit_summary_header_positive": DEFAULT_AMOUNT_UNIT,
            "profit_summary_header_negative": DEFAULT_AMOUNT_UNIT,
            "profit_summary_line_positive": DEFAULT_AMOUNT_UNIT,
            "profit_summary_line_negative": DEFAULT_AMOUNT_UNIT,
        },
        "sign_policies": [],
    }

    candidate_paths = [INCOME_RULEBOOK_PATH, INCOME_RULEBOOK_FALLBACK_PATH]
    loaded_any = False
    for p in candidate_paths:
        if not p.exists():
            continue
        try:
            wb = load_workbook(p, data_only=True)
        except Exception:
            _warn_rule_once(f"income_rulebook_open_fail:{p}", f"收入规则文件读取失败，已尝试回退：{p}", source="income_rulebook")
            continue
        loaded_any = True

        ws_tr = get_sheet_by_loose_name(wb, ["trend_thresholds", "趋势阈值"])
        if ws_tr is not None:
            headers = {str(ws_tr.cell(1, c).value or "").strip(): c for c in range(1, ws_tr.max_column + 1)}
            c_metric = headers.get("metric_id", 2)
            c_stable = headers.get("stable_threshold", 3)
            c_unit = headers.get("threshold_unit", 4)
            c_up = headers.get("up_label", 5)
            c_down = headers.get("down_label", 6)
            c_stable_label = headers.get("stable_label", 7)
            for r in range(2, ws_tr.max_row + 1):
                metric = str(ws_tr.cell(r, c_metric).value or "").strip()
                if not metric:
                    continue
                stable = _safe_float(ws_tr.cell(r, c_stable).value, 0.0)
                unit = str(ws_tr.cell(r, c_unit).value or "").strip()
                up_label = str(ws_tr.cell(r, c_up).value or "").strip() or cfg["trend_thresholds"].get(metric, {}).get("up_label", "")
                down_label = str(ws_tr.cell(r, c_down).value or "").strip() or cfg["trend_thresholds"].get(metric, {}).get("down_label", "")
                stable_label = (
                    str(ws_tr.cell(r, c_stable_label).value or "").strip()
                    or cfg["trend_thresholds"].get(metric, {}).get("stable_label", "")
                )
                stable_norm = _to_pct_threshold(stable, unit) if unit == "ratio" else stable
                cfg["trend_thresholds"][metric] = {
                    "stable": stable_norm,
                    "unit": unit,
                    "up_label": up_label,
                    "down_label": down_label,
                    "stable_label": stable_label,
                }
                if metric == "rev_yoy":
                    cfg["yoy_stable_pct"] = _to_pct_threshold(stable, unit)
                elif metric in {"rev_share", "gp_share_impact"}:
                    cfg["share_stable_pp"] = stable

        ws_map = get_sheet_by_loose_name(wb, ["income_special_items", "收入特殊收益映射"])
        if ws_map is not None:
            headers = {str(ws_map.cell(1, c).value or "").strip(): c for c in range(1, ws_map.max_column + 1)}
            c_node = headers.get("node_id", 1)
            c_label = headers.get("label", 2)
            c_codes = headers.get("code_candidates", 3)
            c_kw = headers.get("name_keywords", 4)
            c_enabled = headers.get("enabled", 5)
            items = []
            for r in range(2, ws_map.max_row + 1):
                node_id = str(ws_map.cell(r, c_node).value or "").strip()
                label = str(ws_map.cell(r, c_label).value or "").strip()
                enabled_raw = str(ws_map.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                if not node_id or not _is_valid_text_cell(label):
                    continue
                codes = [x.strip() for x in str(ws_map.cell(r, c_codes).value or "").replace(";", ",").split(",") if x.strip()]
                kws = [x.strip() for x in str(ws_map.cell(r, c_kw).value or "").replace(";", ",").split(",") if x.strip()]
                kws = [x for x in kws if _is_valid_text_cell(x)]
                items.append({"node_id": node_id, "label": label, "code_candidates": codes, "name_keywords": kws})
            if items:
                cfg["special_items"] = items

        ws_group = get_sheet_by_loose_name(wb, ["income_grouping", "收入汇总分组"])
        if ws_group is not None:
            headers = {str(ws_group.cell(1, c).value or "").strip(): c for c in range(1, ws_group.max_column + 1)}
            c_type = headers.get("group_type", 1)
            c_nodes = headers.get("node_ids", 2)
            c_enabled = headers.get("enabled", 3)
            rec, nonrec = set(), set()
            for r in range(2, ws_group.max_row + 1):
                gtype = str(ws_group.cell(r, c_type).value or "").strip().lower()
                nodes_raw = str(ws_group.cell(r, c_nodes).value or "").strip()
                enabled_raw = str(ws_group.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"} or not nodes_raw:
                    continue
                nodes = {x.strip() for x in nodes_raw.replace(";", ",").split(",") if x.strip()}
                if gtype in {"recurring", "经常性"}:
                    rec |= nodes
                elif gtype in {"nonrecurring", "非经常性"}:
                    nonrec |= nodes
            if rec:
                cfg["recurring_keys"] = rec
            if nonrec:
                cfg["nonrec_keys"] = nonrec

        ws_sign = get_sheet_by_loose_name(wb, ["sign_scenario_policy", "正负场景策略"])
        if ws_sign is not None:
            headers = {str(ws_sign.cell(1, c).value or "").strip(): c for c in range(1, ws_sign.max_column + 1)}
            c_pid = headers.get("policy_id", 1)
            c_scn = headers.get("scenario", 2)
            c_cond = headers.get("condition_expr", 3)
            c_metric = headers.get("primary_ratio_metric", 4)
            c_pos = headers.get("direction_label_pos", 5)
            c_neg = headers.get("direction_label_neg", 6)
            c_pri = headers.get("display_priority", 7)
            c_enabled = headers.get("enabled", 8)
            policies = []
            for r in range(2, ws_sign.max_row + 1):
                enabled_raw = str(ws_sign.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                pid = str(ws_sign.cell(r, c_pid).value or "").strip()
                if not pid:
                    continue
                policies.append(
                    {
                        "policy_id": pid,
                        "scenario": str(ws_sign.cell(r, c_scn).value or "").strip(),
                        "condition_expr": str(ws_sign.cell(r, c_cond).value or "").strip().lower(),
                        "primary_ratio_metric": str(ws_sign.cell(r, c_metric).value or "").strip(),
                        "direction_label_pos": str(ws_sign.cell(r, c_pos).value or "").strip(),
                        "direction_label_neg": str(ws_sign.cell(r, c_neg).value or "").strip(),
                        "display_priority": _safe_float(ws_sign.cell(r, c_pri).value, 9999),
                    }
                )
            if policies:
                policies.sort(key=lambda x: (x.get("display_priority", 9999), str(x.get("policy_id", ""))))
                cfg["sign_policies"] = policies

        ws_cp = get_sheet_by_loose_name(wb, ["contribution_policy", "贡献口径策略"])
        if ws_cp is not None:
            headers = {str(ws_cp.cell(1, c).value or "").strip(): c for c in range(1, ws_cp.max_column + 1)}
            c_key = headers.get("key", 1)
            c_val = headers.get("value", 2)
            c_enabled = headers.get("enabled", 3)
            cp = dict(cfg.get("contribution_policy", {}))
            for r in range(2, ws_cp.max_row + 1):
                enabled_raw = str(ws_cp.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                k = str(ws_cp.cell(r, c_key).value or "").strip()
                if not k:
                    continue
                v = str(ws_cp.cell(r, c_val).value or "").strip()
                cp[k] = v
            cfg["contribution_policy"] = cp

        ws_formula = get_sheet_by_loose_name(wb, ["income_formulas", "收入汇总公式"])
        if ws_formula is not None:
            headers = {str(ws_formula.cell(1, c).value or "").strip(): c for c in range(1, ws_formula.max_column + 1)}
            c_node = headers.get("node_id", 1)
            c_formula = headers.get("formula", 2)
            c_enabled = headers.get("enabled", 3)
            formula_rules = []
            for r in range(2, ws_formula.max_row + 1):
                enabled_raw = str(ws_formula.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                node_id = str(ws_formula.cell(r, c_node).value or "").strip()
                formula = str(ws_formula.cell(r, c_formula).value or "").strip()
                if not node_id or not formula:
                    continue
                formula_rules.append({"node_id": node_id, "formula": formula})
            if formula_rules:
                cfg["formula_rules"] = formula_rules

        ws_tree = get_sheet_by_loose_name(wb, ["income_tree", "收入分析树"])
        if ws_tree is not None:
            default_fixed_by_id = {
                str(x.get("node_id", "")).strip(): x
                for x in DEFAULT_INCOME_TREE
                if str(x.get("node_type", "fixed")).strip().lower() == "fixed" and str(x.get("node_id", "")).strip()
            }
            default_tpl_by_name = {
                str(x.get("template_name", "")).strip(): x
                for x in DEFAULT_INCOME_TREE
                if str(x.get("node_type", "")).strip().lower() == "template"
            }
            headers = {str(ws_tree.cell(1, c).value or "").strip(): c for c in range(1, ws_tree.max_column + 1)}
            c_node = headers.get("node_id", 1)
            c_parent = headers.get("parent_id", 2)
            c_label = headers.get("label", 3)
            c_type = headers.get("node_type", 4)
            c_tname = headers.get("template_name", 5)
            c_prefix = headers.get("node_id_prefix", 6)
            c_start = headers.get("start_index", 7)
            c_suffix = headers.get("label_suffix", 8)
            c_enabled = headers.get("enabled", 9)
            tree_nodes = []
            for r in range(2, ws_tree.max_row + 1):
                enabled_raw = str(ws_tree.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                node_type = str(ws_tree.cell(r, c_type).value or "fixed").strip().lower() or "fixed"
                node_id = str(ws_tree.cell(r, c_node).value or "").strip()
                parent_id = str(ws_tree.cell(r, c_parent).value or "").strip()
                label = str(ws_tree.cell(r, c_label).value or "").strip()
                template_name = str(ws_tree.cell(r, c_tname).value or "").strip()
                prefix = str(ws_tree.cell(r, c_prefix).value or "").strip()
                suffix = str(ws_tree.cell(r, c_suffix).value or "").strip()
                try:
                    start_index = int(ws_tree.cell(r, c_start).value or 1)
                except Exception:
                    start_index = 1
                if node_type == "fixed":
                    if not node_id:
                        continue
                    if not _is_valid_text_cell(label):
                        label = str(default_fixed_by_id.get(node_id, {}).get("label", "")).strip()
                    if not _is_valid_text_cell(label):
                        continue
                    if not parent_id:
                        parent_id = str(default_fixed_by_id.get(node_id, {}).get("parent_id", "")).strip()
                    tree_nodes.append(
                        {"node_id": node_id, "parent_id": parent_id, "label": label, "node_type": "fixed"}
                    )
                elif node_type == "template":
                    if not template_name:
                        continue
                    if not prefix:
                        prefix = str(default_tpl_by_name.get(template_name, {}).get("node_id_prefix", "")).strip()
                    if not prefix:
                        continue
                    if not parent_id:
                        parent_id = str(default_tpl_by_name.get(template_name, {}).get("parent_id", "")).strip()
                    if not _is_valid_text_cell(suffix):
                        suffix = str(default_tpl_by_name.get(template_name, {}).get("label_suffix", "")).strip()
                    if not start_index or start_index < 1:
                        try:
                            start_index = int(default_tpl_by_name.get(template_name, {}).get("start_index", 1))
                        except Exception:
                            start_index = 1
                    tree_nodes.append(
                        {
                            "node_id": "",
                            "parent_id": parent_id,
                            "label": "",
                            "node_type": "template",
                            "template_name": template_name,
                            "node_id_prefix": prefix,
                            "start_index": start_index,
                            "label_suffix": suffix,
                        }
                    )
            if tree_nodes:
                cfg["tree_nodes"] = tree_nodes

        ws_tpl = get_sheet_by_loose_name(wb, ["text_templates", "文本模板"])
        if ws_tpl is not None:
            headers = {str(ws_tpl.cell(1, c).value or "").strip(): c for c in range(1, ws_tpl.max_column + 1)}
            c_scene = headers.get("scene", 2)
            c_tpl = headers.get("template_text_zh", 3)
            c_enabled = headers.get("enabled", 5)
            c_unit = headers.get("unit")
            merged = dict(cfg.get("text_templates", {}))
            unit_map = dict(cfg.get("text_template_units", {}))
            for r in range(2, ws_tpl.max_row + 1):
                enabled_raw = str(ws_tpl.cell(r, c_enabled).value or "1").strip().lower()
                if enabled_raw in {"0", "false", "no"}:
                    continue
                scene = str(ws_tpl.cell(r, c_scene).value or "").strip()
                tpl = str(ws_tpl.cell(r, c_tpl).value or "").strip()
                if not scene or not _is_valid_text_cell(tpl):
                    continue
                # Keep Chinese output style by default; skip placeholder English templates.
                if not re.search(r"[\u4e00-\u9fff]", tpl):
                    continue
                merged[scene] = tpl
                if c_unit:
                    unit_val = str(ws_tpl.cell(r, c_unit).value or "").strip()
                    if unit_val:
                        unit_map[scene] = unit_val
            cfg["text_templates"] = merged
            cfg["text_template_units"] = unit_map
        break
    if not loaded_any:
        _warn_rule_once("income_rulebook_all_missing", "收入规则文件缺失，已使用代码默认规则", source="income_rulebook")
    return cfg


def _pick_income_sign_policy(policies: List[Dict[str, Any]], total_gp: Optional[float], has_pos: bool, has_neg: bool) -> Dict[str, Any]:
    if total_gp is None:
        return {}
    all_same_sign = (has_pos != has_neg) and (has_pos or has_neg)
    mixed_sign = has_pos and has_neg
    for p in policies:
        cond = str(p.get("condition_expr", "")).lower()
        if "total_gp>0" in cond and not (total_gp > 0):
            continue
        if "total_gp<0" in cond and not (total_gp < 0):
            continue
        if "total_gp==0" in cond and not (abs(total_gp) < 1e-12):
            continue
        if "all_same_sign=true" in cond and not all_same_sign:
            continue
        if "mixed_sign=true" in cond and not mixed_sign:
            continue
        return p
    return {}


def load_ratio_analysis_rules() -> Dict[str, Any]:
    cfg = {
        "tree_nodes": list(DEFAULT_RATIO_TREE),
        "catalog": {},
        "trend_threshold_pp": 2.0,
        "trend_rules": [],
        "judgement_rules": {},
        "alert_rules": [],
        "text_templates": {
            "indicator_value": DEFAULT_TEXT_TEMPLATES["ratio_indicator_value"],
            "indicator_trend": DEFAULT_TEXT_TEMPLATES["ratio_indicator_trend"],
            "indicator_alert": "风险提示：{alert_text}",
        },
    }
    if not RATIO_RULEBOOK_PATH.exists():
        return cfg
    try:
        wb = load_workbook(RATIO_RULEBOOK_PATH, data_only=True)
    except Exception:
        return cfg

    ws_tree = get_sheet_by_loose_name(wb, ["indicator_tree"])
    if ws_tree is not None:
        headers = {str(ws_tree.cell(1, c).value or "").strip(): c for c in range(1, ws_tree.max_column + 1)}
        nodes = []
        for r in range(2, ws_tree.max_row + 1):
            enabled = str(ws_tree.cell(r, headers.get("enabled", 6)).value or "1").strip().lower()
            if enabled in {"0", "false", "no"}:
                continue
            nodes.append(
                {
                    "node_id": str(ws_tree.cell(r, headers.get("node_id", 1)).value or "").strip(),
                    "parent_id": str(ws_tree.cell(r, headers.get("parent_id", 2)).value or "").strip(),
                    "label_zh": str(ws_tree.cell(r, headers.get("label_zh", 3)).value or "").strip(),
                    "node_type": str(ws_tree.cell(r, headers.get("node_type", 4)).value or "indicator").strip(),
                    "indicator_id": str(ws_tree.cell(r, headers.get("indicator_id", 5)).value or "").strip(),
                    "sort_order": _safe_float(ws_tree.cell(r, headers.get("sort_order", 7)).value, 9999),
                }
            )
        nodes = [x for x in nodes if x.get("node_id") and x.get("label_zh")]
        if nodes:
            nodes.sort(key=lambda x: (x.get("sort_order", 9999), str(x.get("node_id"))))
            cfg["tree_nodes"] = nodes

    ws_cat = get_sheet_by_loose_name(wb, ["indicator_catalog"])
    if ws_cat is not None:
        headers = {str(ws_cat.cell(1, c).value or "").strip(): c for c in range(1, ws_cat.max_column + 1)}
        for r in range(2, ws_cat.max_row + 1):
            rid = str(ws_cat.cell(r, headers.get("indicator_id", 1)).value or "").strip()
            if not rid:
                continue
            enabled = str(ws_cat.cell(r, headers.get("enabled", 8)).value or "1").strip().lower()
            if enabled in {"0", "false", "no"}:
                continue
            cfg["catalog"][rid] = {
                "name": str(ws_cat.cell(r, headers.get("indicator_name_zh", 2)).value or rid).strip(),
                "group": str(ws_cat.cell(r, headers.get("group_zh", 3)).value or "").strip(),
                "direction": str(ws_cat.cell(r, headers.get("direction", 4)).value or "higher").strip().lower(),
                "unit": str(ws_cat.cell(r, headers.get("unit", 5)).value or "").strip(),
                "value_source": str(ws_cat.cell(r, headers.get("value_source", 6)).value or "").strip(),
                "formula_expr": str(ws_cat.cell(r, headers.get("formula_expr", 7)).value or "").strip(),
            }

    ws_trend = get_sheet_by_loose_name(wb, ["trend_rules"])
    if ws_trend is not None:
        headers = {str(ws_trend.cell(1, c).value or "").strip(): c for c in range(1, ws_trend.max_column + 1)}
        c_scope = headers.get("scope", 2)
        c_type = headers.get("threshold_type", 3)
        c_stable = headers.get("stable_threshold", 4)
        c_warn = headers.get("warn_threshold", 5)
        c_crit = headers.get("critical_threshold", 6)
        c_unit = headers.get("unit", 7)
        c_enabled = headers.get("enabled", 8)
        for r in range(2, ws_trend.max_row + 1):
            enabled_raw = str(ws_trend.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            scope = str(ws_trend.cell(r, c_scope).value or "").strip().lower()
            if not scope:
                continue
            cfg["trend_rules"].append(
                {
                    "scope": scope,
                    "threshold_type": str(ws_trend.cell(r, c_type).value or "delta_abs").strip().lower(),
                    "stable_threshold": _safe_float(ws_trend.cell(r, c_stable).value, 2.0),
                    "warn_threshold": _safe_float(ws_trend.cell(r, c_warn).value, 5.0),
                    "critical_threshold": _safe_float(ws_trend.cell(r, c_crit).value, 10.0),
                    "unit": str(ws_trend.cell(r, c_unit).value or "").strip().lower(),
                }
            )

    ws_judge = get_sheet_by_loose_name(wb, ["judgement_rules"])
    if ws_judge is not None:
        headers = {str(ws_judge.cell(1, c).value or "").strip(): c for c in range(1, ws_judge.max_column + 1)}
        c_dir = headers.get("direction", 2)
        c_good = headers.get("good_label", 4)
        c_bad = headers.get("bad_label", 5)
        c_stable = headers.get("stable_label", 6)
        c_enabled = headers.get("enabled", 7)
        for r in range(2, ws_judge.max_row + 1):
            enabled_raw = str(ws_judge.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            direction = str(ws_judge.cell(r, c_dir).value or "").strip().lower()
            if direction not in {"higher", "lower", "range"}:
                continue
            cfg["judgement_rules"][direction] = {
                "good_label": str(ws_judge.cell(r, c_good).value or "改善").strip() or "改善",
                "bad_label": str(ws_judge.cell(r, c_bad).value or "弱化").strip() or "弱化",
                "stable_label": str(ws_judge.cell(r, c_stable).value or "基本稳定").strip() or "基本稳定",
            }

    ws_alert = get_sheet_by_loose_name(wb, ["alert_rules"])
    if ws_alert is not None:
        headers = {str(ws_alert.cell(1, c).value or "").strip(): c for c in range(1, ws_alert.max_column + 1)}
        c_ind = headers.get("indicator_id", 2)
        c_cond = headers.get("condition_expr", 3)
        c_sev = headers.get("severity", 4)
        c_text = headers.get("alert_text_zh", 5)
        c_enabled = headers.get("enabled", 6)
        for r in range(2, ws_alert.max_row + 1):
            enabled_raw = str(ws_alert.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            ind = str(ws_alert.cell(r, c_ind).value or "").strip().lower()
            cond = str(ws_alert.cell(r, c_cond).value or "").strip().lower()
            txt = str(ws_alert.cell(r, c_text).value or "").strip()
            if not ind or not cond or not txt:
                continue
            cfg["alert_rules"].append(
                {
                    "indicator_id": ind,
                    "condition_expr": cond,
                    "severity": str(ws_alert.cell(r, c_sev).value or "").strip().lower(),
                    "alert_text_zh": txt,
                }
            )

    ws_display = get_sheet_by_loose_name(wb, ["display_policy"])
    if ws_display is not None:
        for r in range(2, ws_display.max_row + 1):
            key = str(ws_display.cell(r, 1).value or "").strip()
            if key == "trend_threshold_default_pp":
                cfg["trend_threshold_pp"] = _safe_float(ws_display.cell(r, 2).value, 2.0)
    ws_tpl = get_sheet_by_loose_name(wb, ["text_templates", "文本模板"])
    if ws_tpl is not None:
        headers = {str(ws_tpl.cell(1, c).value or "").strip(): c for c in range(1, ws_tpl.max_column + 1)}
        c_scene = headers.get("scene", 2)
        c_tpl = headers.get("template_text_zh", 3)
        c_enabled = headers.get("enabled", 5)
        merged = dict(cfg.get("text_templates", {}))
        for r in range(2, ws_tpl.max_row + 1):
            enabled_raw = str(ws_tpl.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            scene = str(ws_tpl.cell(r, c_scene).value or "").strip()
            tpl = str(ws_tpl.cell(r, c_tpl).value or "").strip()
            if not scene or not _is_valid_text_cell(tpl):
                continue
            merged[scene] = tpl
        cfg["text_templates"] = merged
    return cfg


def _is_key_ratio_topic(node: Dict[str, Any]) -> bool:
    """Only dedicated topic rows are considered key-ratio items."""
    ntype = str(node.get("node_type", "")).strip().lower()
    if ntype != "topic":
        return False
    iid = str(node.get("indicator_id", "")).strip().lower()
    return iid in {"topic_roe_dupont", "topic_gross_margin"}


def _sanitize_ratio_tree_nodes(tree_nodes: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Guardrail for the regular ratio page:
    - keep group/indicator nodes only
    - remove any explicit key-topic rows
    """
    out: List[Dict[str, Any]] = []
    for n in tree_nodes:
        ntype = str(n.get("node_type", "")).strip().lower()
        if ntype not in {"group", "indicator"}:
            continue
        if _is_key_ratio_topic(n):
            continue
        out.append(n)
    return out


def load_key_ratio_text_rules() -> Dict[str, Any]:
    global _KEY_RATIO_RULES_CACHE
    if isinstance(_KEY_RATIO_RULES_CACHE, dict):
        return _KEY_RATIO_RULES_CACHE
    cfg: Dict[str, Any] = {
        "thresholds": {
            "roe": dict(DEFAULT_KEY_DRIVER_THRESHOLDS),
            "gross_margin": dict(DEFAULT_KEY_DRIVER_THRESHOLDS),
        },
        "templates": {
            "roe_missing_years": "杜邦分析所需年度不足，待补充。",
            "roe_headline": "{metric_name}由{y_prev}年的{v_prev}变动至{y_curr}年的{v_curr}，整体{trend_word}。",
            "trend_word_up": "上升",
            "trend_word_down": "下降",
            "trend_word_stable": "基本稳定",
            "roe_method_note": "分解方法：采用Shapley分解（乘法体系），按三因子对ROE变动的边际贡献进行公平分摊。",
            "roe_single_driver": "主要驱动因素为{driver_1}，贡献{contrib_1}个百分点，驱动占比{share_1}。",
            "roe_dual_driver": "变动由{driver_1}与{driver_2}共同驱动，分别贡献{contrib_1}/{contrib_2}个百分点，合计驱动占比{share_sum}。",
            "roe_multi_driver": "ROE变动为多因素共同作用：{driver_list}，需结合业务与资本结构综合判断。",
            "roe_contrib_detail": "三因子贡献：净利率{nm_contrib}个百分点（驱动占比{nm_share}），总资产周转率{at_contrib}个百分点（驱动占比{at_share}），权益乘数{em_contrib}个百分点（驱动占比{em_share}）。",
            "roe_reconcile": "分解校验：三因子贡献合计{contrib_sum}个百分点，对应ROE变动{roe_delta}个百分点。",
            "roe_offset": "存在显著对冲项：{offset_driver}（{offset_contrib}个百分点），抵消了{main_driver}的部分影响。",
            "gm_missing_years": "毛利率贡献分析所需年度不足，待补充。",
            "gm_missing_segments": "未识别到营业收入分项，毛利率贡献待补充。",
            "gm_headline": "{metric_name}由{y_prev}年的{v_prev}变动至{y_curr}年的{v_curr}，整体{trend_word}。",
            "gm_single_driver": "主要驱动来自{driver_1}，贡献{contrib_1}个百分点，驱动占比{share_1}。",
            "gm_dual_driver": "变动由{driver_1}与{driver_2}共同驱动，分别贡献{contrib_1}/{contrib_2}个百分点，合计驱动占比{share_sum}。",
            "gm_multi_driver": "毛利率变动为多因素共同作用：{driver_list}。",
            "gm_structure_effect": "其中结构效应{structure_effect}个百分点，价格/成本效应{price_cost_effect}个百分点，交互项{interaction_effect}个百分点。",
            "gm_negative_impact_summary": "各分项按绝对影响占比展示，不采用简单占比",
            "gm_negative_profit_case": "在总毛利为负或分项正负并存情况下，采用影响占比口径：{impact_summary}。",
            "gm_segment_header": "分项影响拆解（{y1}-{y3}）：",
            "gm_segment_line": "{name}：结构{ds}个百分点，价格/成本{dm}个百分点，交互{di}个百分点，合计{dt}个百分点；净影响占比{net_share}，影响强度占比{strength_share}。",
            "gm_segment_summary": "子项层面：正向拉动前列{top_pos}；负向拖累前列{top_neg}。",
            "gm_net_share_na": "不适用（整体变动接近0）",
            "gm_top_fallback": "待补充",
        },
    }
    if not KEY_RATIO_RULEBOOK_PATH.exists():
        _KEY_RATIO_RULES_CACHE = cfg
        return cfg
    try:
        wb = load_workbook(KEY_RATIO_RULEBOOK_PATH, data_only=True)
    except Exception:
        _KEY_RATIO_RULES_CACHE = cfg
        return cfg

    ws_th = get_sheet_by_loose_name(wb, ["driver_thresholds"])
    if ws_th is not None:
        headers = {str(ws_th.cell(1, c).value or "").strip(): c for c in range(1, ws_th.max_column + 1)}
        c_scope = headers.get("scope", 1)
        c_key = headers.get("threshold_key", 3)
        c_val = headers.get("threshold_value", 4)
        c_enabled = headers.get("enabled", 7)
        for r in range(2, ws_th.max_row + 1):
            enabled_raw = str(ws_th.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            scope = str(ws_th.cell(r, c_scope).value or "").strip().lower()
            key = str(ws_th.cell(r, c_key).value or "").strip()
            val = _safe_float(ws_th.cell(r, c_val).value, None)  # type: ignore[arg-type]
            if scope not in {"roe", "gross_margin"} or not key or val is None:
                continue
            cfg["thresholds"].setdefault(scope, dict(DEFAULT_KEY_DRIVER_THRESHOLDS))
            cfg["thresholds"][scope][key] = float(val)

    ws_tpl = get_sheet_by_loose_name(wb, ["narrative_templates", "文本模板"])
    if ws_tpl is not None:
        headers = {str(ws_tpl.cell(1, c).value or "").strip(): c for c in range(1, ws_tpl.max_column + 1)}
        c_id = headers.get("template_id", 2)
        c_tpl = headers.get("template_text", 4)
        c_enabled = headers.get("enabled", 6)
        merged = dict(cfg.get("templates", {}))
        for r in range(2, ws_tpl.max_row + 1):
            enabled_raw = str(ws_tpl.cell(r, c_enabled).value or "1").strip().lower()
            if enabled_raw in {"0", "false", "no"}:
                continue
            tid = str(ws_tpl.cell(r, c_id).value or "").strip()
            tpl = str(ws_tpl.cell(r, c_tpl).value or "").strip()
            if not tid or not _is_valid_text_cell(tpl):
                continue
            merged[tid] = tpl
        cfg["templates"] = merged
    _KEY_RATIO_RULES_CACHE = cfg
    return cfg


def _fmt_ratio_value(v: Optional[float], unit: str) -> str:
    if v is None:
        return "待补充"
    if unit == "%":
        vv = v * 100.0 if abs(v) <= 1.0 else v
        return f"{vv:.2f}%"
    return f"{v:.2f}{unit}" if unit else f"{v:.2f}"


def _fmt_ratio_number(v: Optional[float], unit: str) -> str:
    if v is None:
        return "待补充"
    if unit == "%":
        vv = v * 100.0 if abs(v) <= 1.0 else v
        return f"{vv:.2f}"
    return f"{v:.2f}"


def _fmt_ratio_trend(curr: Optional[float], prev: Optional[float], direction: str, threshold: float, unit: str) -> str:
    if curr is None or prev is None:
        return "待补充"
    c = curr
    p = prev
    if unit == "%":
        c = curr * 100.0 if abs(curr) <= 1.0 else curr
        p = prev * 100.0 if abs(prev) <= 1.0 else prev
    delta = c - p
    if abs(delta) <= threshold:
        return f"基本稳定（变动{abs(delta):.2f}{'个百分点' if unit == '%' else ''}）"
    improve = (delta > 0 and direction != "lower") or (delta < 0 and direction == "lower")
    verb = "改善" if improve else "弱化"
    return f"{verb}（变动{abs(delta):.2f}{'个百分点' if unit == '%' else ''}）"


def _ratio_trend_parts(curr: Optional[float], prev: Optional[float], direction: str, threshold: float, unit: str) -> Dict[str, str]:
    if curr is None or prev is None:
        return {"judgement": "待补充", "delta": "待补充", "unit2": "个百分点" if unit == "%" else (unit or "")}
    c = curr
    p = prev
    if unit == "%":
        c = curr * 100.0 if abs(curr) <= 1.0 else curr
        p = prev * 100.0 if abs(prev) <= 1.0 else prev
    delta = c - p
    if abs(delta) <= threshold:
        judgement = "基本稳定"
    else:
        improve = (delta > 0 and direction != "lower") or (delta < 0 and direction == "lower")
        judgement = "改善" if improve else "弱化"
    return {
        "judgement": judgement,
        "delta": f"{abs(delta):.2f}",
        "unit2": "个百分点" if unit == "%" else (unit or ""),
    }


def _ratio_scope_match(scope: str, indicator_id: str, indicator_name: str, group_name: str) -> bool:
    s = str(scope or "").strip().lower()
    iid = str(indicator_id or "").strip().lower()
    name = str(indicator_name or "").strip().lower()
    grp = str(group_name or "").strip().lower()
    if s in {"all", "*"}:
        return True
    if s in {iid, name, grp}:
        return True
    if s == "turnover":
        return ("turnover" in iid) or ("周转" in indicator_name)
    return False


def _select_ratio_trend_rule(rules: List[Dict[str, Any]], indicator_id: str, indicator_name: str, group_name: str) -> Dict[str, Any]:
    for r in rules:
        if _ratio_scope_match(r.get("scope", ""), indicator_id, indicator_name, group_name):
            return r
    return {
        "threshold_type": "delta_abs",
        "stable_threshold": 2.0,
        "warn_threshold": 5.0,
        "critical_threshold": 10.0,
        "unit": "pct_point",
    }


def _ratio_judgement_parts(
    curr: Optional[float],
    prev: Optional[float],
    direction: str,
    trend_rule: Dict[str, Any],
    labels: Dict[str, str],
    unit: str,
) -> Dict[str, str]:
    if curr is None or prev is None:
        return {"judgement": "待补充", "delta": "待补充", "unit2": "个百分点" if unit == "%" else (unit or "")}

    c = curr
    p = prev
    if unit == "%":
        c = curr * 100.0 if abs(curr) <= 1.0 else curr
        p = prev * 100.0 if abs(prev) <= 1.0 else prev
    delta = c - p
    threshold_type = str(trend_rule.get("threshold_type", "delta_abs")).strip().lower()
    stable_threshold = _safe_float(trend_rule.get("stable_threshold"), 2.0)

    if threshold_type == "delta_rate":
        if p in (None, 0):
            return {"judgement": "待补充", "delta": "待补充", "unit2": "%"}
        rel = abs((c - p) / abs(p))
        is_stable = rel <= stable_threshold
        shown_delta = rel * 100.0
        unit2 = "%"
    else:
        is_stable = abs(delta) <= stable_threshold
        shown_delta = abs(delta)
        unit2 = "个百分点" if unit == "%" else (unit or "")

    if is_stable:
        judgement = labels.get("stable_label", "基本稳定")
    else:
        improve = (delta > 0 and direction != "lower") or (delta < 0 and direction == "lower")
        judgement = labels.get("good_label", "改善") if improve else labels.get("bad_label", "弱化")
    return {"judgement": judgement, "delta": f"{shown_delta:.2f}", "unit2": unit2}


def _ratio_alert_hit(condition_expr: str, values: Dict[str, Optional[float]], years: List[str], unit: str) -> bool:
    seq_years = _year_seq(years)
    if len(seq_years) < 2:
        return False
    y_base = seq_years[0]
    y_prev = seq_years[-2]
    y_curr = seq_years[-1]
    v1, v2, v3 = values.get(y_base), values.get(y_prev), values.get(y_curr)
    if unit == "%":
        v1 = None if v1 is None else (v1 * 100.0 if abs(v1) <= 1.0 else v1)
        v2 = None if v2 is None else (v2 * 100.0 if abs(v2) <= 1.0 else v2)
        v3 = None if v3 is None else (v3 * 100.0 if abs(v3) <= 1.0 else v3)
    expr = str(condition_expr or "").strip().lower()
    if not expr:
        return False
    def _num(s: str, default: Optional[float] = None) -> Optional[float]:
        try:
            return float(str(s).strip())
        except Exception:
            return default
    def _delta() -> Optional[float]:
        if v3 is None or v2 is None:
            return None
        return v3 - v2
    def _delta_2y() -> Optional[float]:
        if len(seq_years) < 3 or v1 is None or v2 is None or v3 is None:
            return None
        return (v2 - v1) + (v3 - v2)

    # Legacy rules (3y monotonic) kept for backward compatibility.
    if expr == "down_2y":
        if len(seq_years) >= 3:
            return v1 is not None and v2 is not None and v3 is not None and (v2 < v1 and v3 < v2)
        d = _delta()
        return d is not None and d < 0
    if expr == "up_2y":
        if len(seq_years) >= 3:
            return v1 is not None and v2 is not None and v3 is not None and (v2 > v1 and v3 > v2)
        d = _delta()
        return d is not None and d > 0
    # Adaptive extensions:
    # down_last / up_last / delta_last_abs
    if expr == "down_last":
        d = _delta()
        return d is not None and d < 0
    if expr == "up_last":
        d = _delta()
        return d is not None and d > 0
    m = re.match(r"^down_last\(([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d = _delta()
        th = _num(m.group(1), 0.0) or 0.0
        return d is not None and (-d) >= th
    m = re.match(r"^up_last\(([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d = _delta()
        th = _num(m.group(1), 0.0) or 0.0
        return d is not None and d >= th
    m = re.match(r"^delta_last_abs\(([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d = _delta()
        th = _num(m.group(1), 0.0) or 0.0
        return d is not None and abs(d) >= th
    # trend_last(up|down|stable,threshold)
    m = re.match(r"^trend_last\((up|down|stable)\s*,\s*([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d = _delta()
        if d is None:
            return False
        mode = m.group(1)
        th = _num(m.group(2), 0.0) or 0.0
        if mode == "up":
            return d > th
        if mode == "down":
            return d < -th
        return abs(d) <= th
    # Optional threshold variants for 3y expressions.
    m = re.match(r"^down_2y\(([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d2 = _delta_2y()
        th = _num(m.group(1), 0.0) or 0.0
        if d2 is not None:
            return d2 <= -th
        d = _delta()
        return d is not None and d <= -th
    m = re.match(r"^up_2y\(([-+]?\d+(?:\.\d+)?)\)$", expr)
    if m:
        d2 = _delta_2y()
        th = _num(m.group(1), 0.0) or 0.0
        if d2 is not None:
            return d2 >= th
        d = _delta()
        return d is not None and d >= th
    m = re.match(r"^(curr|prev|base)\s*(<=|>=|<|>|==)\s*(-?\d+(?:\.\d+)?)$", expr)
    if not m:
        return False
    var, op, raw = m.group(1), m.group(2), float(m.group(3))
    val = v3 if var == "curr" else (v2 if var == "prev" else v1)
    if val is None:
        return False
    if op == "<":
        return val < raw
    if op == "<=":
        return val <= raw
    if op == ">":
        return val > raw
    if op == ">=":
        return val >= raw
    return abs(val - raw) < 1e-12


def _ratio_alert_text(
    indicator_id: str,
    values: Dict[str, Optional[float]],
    years: List[str],
    unit: str,
    alert_rules: List[Dict[str, Any]],
    text_templates: Dict[str, str],
) -> str:
    hits: List[str] = []
    iid = str(indicator_id or "").strip().lower()
    for r in alert_rules:
        target = str(r.get("indicator_id", "")).strip().lower()
        if target not in {"*", iid}:
            continue
        if _ratio_alert_hit(str(r.get("condition_expr", "")), values, years, unit):
            hits.append(str(r.get("alert_text_zh", "")).strip())
    if not hits:
        return ""
    tpl = text_templates.get("indicator_alert", "风险提示：{alert_text}")
    return _render_template(tpl, {"alert_text": "；".join([x for x in hits if x])})


def _pick_ratio_record(rows: List[Dict[str, Any]], code_keys: List[str], name_keys: List[str]) -> Optional[Dict[str, Any]]:
    for r in rows:
        code = str(r.get("code", "")).strip().lower()
        name = str(r.get("name", "")).strip()
        if any(k and k.lower() in code for k in code_keys):
            return r
        if any(k and k in name for k in name_keys):
            return r
    return None


def _classify_driver_case(contrib: List[tuple], thresholds: Dict[str, float]) -> Dict[str, Any]:
    usable = [(k, v) for k, v in contrib if v is not None]
    if not usable:
        return {"case": "none", "top": [], "shares": {}, "main_driver": "待补充"}
    ordered = sorted(usable, key=lambda x: abs(x[1]), reverse=True)
    total_abs = sum(abs(v) for _, v in ordered)
    shares = {k: (abs(v) / total_abs if total_abs else 0.0) for k, v in ordered}
    single = float(thresholds.get("single_driver_share", DEFAULT_KEY_DRIVER_THRESHOLDS["single_driver_share"]))
    dual_sum = float(thresholds.get("dual_driver_share_sum", DEFAULT_KEY_DRIVER_THRESHOLDS["dual_driver_share_sum"]))
    dual_each = float(thresholds.get("dual_driver_each_min", DEFAULT_KEY_DRIVER_THRESHOLDS["dual_driver_each_min"]))
    if shares.get(ordered[0][0], 0.0) >= single:
        case = "single"
    elif len(ordered) >= 2 and (shares.get(ordered[0][0], 0.0) + shares.get(ordered[1][0], 0.0) >= dual_sum) and (
        shares.get(ordered[0][0], 0.0) >= dual_each and shares.get(ordered[1][0], 0.0) >= dual_each
    ):
        case = "dual"
    else:
        case = "multi"
    return {"case": case, "top": ordered, "shares": shares, "main_driver": ordered[0][0]}


def _pct_to_ratio(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    return v if abs(v) <= 1.0 else (v / 100.0)


def _fmt_signed_pp(v: Optional[float]) -> str:
    if v is None:
        return "待补充"
    if v > 0:
        return f"+{v:.2f}"
    return f"{v:.2f}"


def _roe_shapley_contrib_pp(
    nm0: Optional[float],
    nm1: Optional[float],
    at0: Optional[float],
    at1: Optional[float],
    em0: Optional[float],
    em1: Optional[float],
) -> Dict[str, Optional[float]]:
    vals = [nm0, nm1, at0, at1, em0, em1]
    if any(v is None for v in vals):
        return {"净利率": None, "总资产周转率": None, "权益乘数": None, "sum": None}

    n0, n1 = float(nm0), float(nm1)
    a0, a1 = float(at0), float(at1)
    e0, e1 = float(em0), float(em1)

    def f(n: float, a: float, e: float) -> float:
        return n * a * e

    # Strict Shapley decomposition for three-factor multiplicative model.
    phi_n = (
        (1.0 / 3.0) * (f(n1, a0, e0) - f(n0, a0, e0))
        + (1.0 / 6.0) * (f(n1, a1, e0) - f(n0, a1, e0))
        + (1.0 / 6.0) * (f(n1, a0, e1) - f(n0, a0, e1))
        + (1.0 / 3.0) * (f(n1, a1, e1) - f(n0, a1, e1))
    )
    phi_a = (
        (1.0 / 3.0) * (f(n0, a1, e0) - f(n0, a0, e0))
        + (1.0 / 6.0) * (f(n1, a1, e0) - f(n1, a0, e0))
        + (1.0 / 6.0) * (f(n0, a1, e1) - f(n0, a0, e1))
        + (1.0 / 3.0) * (f(n1, a1, e1) - f(n1, a0, e1))
    )
    phi_e = (
        (1.0 / 3.0) * (f(n0, a0, e1) - f(n0, a0, e0))
        + (1.0 / 6.0) * (f(n1, a0, e1) - f(n1, a0, e0))
        + (1.0 / 6.0) * (f(n0, a1, e1) - f(n0, a1, e0))
        + (1.0 / 3.0) * (f(n1, a1, e1) - f(n1, a1, e0))
    )

    # Convert ratio deltas to percentage-point deltas for narrative.
    n_pp = phi_n * 100.0
    a_pp = phi_a * 100.0
    e_pp = phi_e * 100.0
    return {"净利率": n_pp, "总资产周转率": a_pp, "权益乘数": e_pp, "sum": n_pp + a_pp + e_pp}


def _dupont_topic_text(rows: List[Dict[str, Any]], years: List[str], key_rules: Dict[str, Any]) -> str:
    tpl = key_rules.get("templates", {})
    years = _year_seq(years)
    if len(years) < 2:
        return _render_template(tpl.get("roe_missing_years", "杜邦分析所需年度不足，待补充。"), {})
    y1, y2, y3 = _compat_three_periods(years)
    th = key_rules.get("thresholds", {}).get("roe", DEFAULT_KEY_DRIVER_THRESHOLDS)
    rec_roe = _pick_ratio_record(rows, ["roe"], ["净资产收益率", "ROE"])
    rec_nm = _pick_ratio_record(rows, ["net_margin", "sales_net_margin"], ["销售净利率", "净利率"])
    rec_at = _pick_ratio_record(rows, ["asset_turnover"], ["总资产周转率"])
    rec_em = _pick_ratio_record(rows, ["equity_multiplier"], ["权益乘数"])

    def gv(rec, y):
        if not rec:
            return None
        vals = rec.get("values", {}) if isinstance(rec.get("values"), dict) else {}
        return vals.get(y)

    def norm_pct(v):
        if v is None:
            return None
        return v * 100.0 if abs(v) <= 1.0 else v

    roe_pct = {y: norm_pct(gv(rec_roe, y)) for y in years}
    nm_pct = {y: norm_pct(gv(rec_nm, y)) for y in years}
    at = {y: gv(rec_at, y) for y in years}
    em = {y: gv(rec_em, y) for y in years}
    nm_ratio = {y: _pct_to_ratio(nm_pct.get(y)) for y in years}
    roe_ratio = {y: _pct_to_ratio(roe_pct.get(y)) for y in years}
    for y in years:
        if em.get(y) is None:
            n = nm_pct.get(y)
            a = at.get(y)
            r = roe_pct.get(y)
            if n not in (None, 0) and a not in (None, 0) and r is not None:
                em[y] = (r / 100.0) / ((n / 100.0) * a)
        if roe_ratio.get(y) is None and nm_ratio.get(y) is not None and at.get(y) is not None and em.get(y) is not None:
            roe_ratio[y] = nm_ratio[y] * float(at.get(y)) * float(em.get(y))
            roe_pct[y] = float(roe_ratio[y]) * 100.0

    shapley = _roe_shapley_contrib_pp(
        nm_ratio.get(y1),
        nm_ratio.get(y3),
        at.get(y1),
        at.get(y3),
        em.get(y1),
        em.get(y3),
    )
    d_roe = None if roe_pct.get(y3) is None or roe_pct.get(y1) is None else roe_pct.get(y3) - roe_pct.get(y1)
    contrib = [("净利率", shapley.get("净利率")), ("总资产周转率", shapley.get("总资产周转率")), ("权益乘数", shapley.get("权益乘数"))]
    c = _classify_driver_case(contrib, th)
    top = c["top"]
    shares = c["shares"]
    main_driver = c["main_driver"]
    headline = _render_template(
        tpl.get("roe_headline", "{metric_name}由{y_prev}年的{v_prev}变动至{y_curr}年的{v_curr}，整体{trend_word}。"),
        {
            "metric_name": "ROE",
            "y_prev": y1,
            "v_prev": _fmt_ratio_value(roe_pct.get(y1), "%"),
            "y_curr": y3,
            "v_curr": _fmt_ratio_value(roe_pct.get(y3), "%"),
            "trend_word": _trend_word(
                roe_pct.get(y3),
                roe_pct.get(y1),
                stable_pp=float(th.get("delta_stable_pp", 2.0)),
                up_label=str(tpl.get("trend_word_up", "上升")),
                down_label=str(tpl.get("trend_word_down", "下降")),
                stable_label=str(tpl.get("trend_word_stable", "基本稳定")),
            ),
        },
    )
    if len(years) >= 3:
        three_year_line = _render_template(
            tpl.get("roe_three_year_values", "ROE三年分别为：{y1}年{v1}，{y2}年{v2}，{y3}年{v3}。"),
            {
                "y1": y1,
                "y2": y2,
                "y3": y3,
                "v1": _fmt_ratio_value(roe_pct.get(y1), "%"),
                "v2": _fmt_ratio_value(roe_pct.get(y2), "%"),
                "v3": _fmt_ratio_value(roe_pct.get(y3), "%"),
            },
        )
    else:
        three_year_line = _render_template(
            tpl.get("roe_two_year_values", "ROE两年分别为：{y1}年{v1}，{y2}年{v2}。"),
            {
                "y1": y1,
                "y2": y2,
                "v1": _fmt_ratio_value(roe_pct.get(y1), "%"),
                "v2": _fmt_ratio_value(roe_pct.get(y2), "%"),
            },
        )
    method_txt = _render_template(
        tpl.get("roe_method_note", "分解方法：采用Shapley分解（乘法体系），按三因子对ROE变动的边际贡献进行公平分摊。"),
        {},
    )
    if c["case"] == "single":
        driver_txt = _render_template(
            tpl.get("roe_single_driver", "主要驱动因素为{driver_1}，贡献{contrib_1}个百分点，驱动占比{share_1}。"),
            {
                "driver_1": top[0][0],
                "contrib_1": f"{abs(top[0][1] or 0.0):.2f}",
                "share_1": f"{shares.get(top[0][0], 0.0) * 100:.2f}%",
            },
        )
    elif c["case"] == "dual":
        driver_txt = _render_template(
            tpl.get("roe_dual_driver", "变动由{driver_1}与{driver_2}共同驱动，分别贡献{contrib_1}/{contrib_2}个百分点，合计驱动占比{share_sum}。"),
            {
                "driver_1": top[0][0],
                "driver_2": top[1][0],
                "contrib_1": f"{abs(top[0][1] or 0.0):.2f}",
                "contrib_2": f"{abs(top[1][1] or 0.0):.2f}",
                "share_sum": f"{(shares.get(top[0][0], 0.0) + shares.get(top[1][0], 0.0)) * 100:.2f}%",
            },
        )
    else:
        driver_txt = _render_template(
            tpl.get("roe_multi_driver", "ROE变动为多因素共同作用：{driver_list}，需结合业务与资本结构综合判断。"),
            {"driver_list": "、".join([k for k, _ in top]) if top else "待补充"},
        )

    contrib_detail_txt = _render_template(
        tpl.get(
            "roe_contrib_detail",
            "三因子贡献：净利率{nm_contrib}个百分点（驱动占比{nm_share}），总资产周转率{at_contrib}个百分点（驱动占比{at_share}），权益乘数{em_contrib}个百分点（驱动占比{em_share}）。",
        ),
        {
            "nm_contrib": _fmt_signed_pp(shapley.get("净利率")),
            "at_contrib": _fmt_signed_pp(shapley.get("总资产周转率")),
            "em_contrib": _fmt_signed_pp(shapley.get("权益乘数")),
            "nm_share": f"{shares.get('净利率', 0.0) * 100:.2f}%",
            "at_share": f"{shares.get('总资产周转率', 0.0) * 100:.2f}%",
            "em_share": f"{shares.get('权益乘数', 0.0) * 100:.2f}%",
        },
    )
    reconcile_txt = _render_template(
        tpl.get("roe_reconcile", "分解校验：三因子贡献合计{contrib_sum}个百分点，对应ROE变动{roe_delta}个百分点。"),
        {
            "contrib_sum": _fmt_signed_pp(shapley.get("sum")),
            "roe_delta": _fmt_signed_pp(d_roe),
        },
    )

    offset_txt = ""
    if d_roe is not None:
        sig = float(th.get("significant_abs_contrib", DEFAULT_KEY_DRIVER_THRESHOLDS["significant_abs_contrib"]))
        opp = [(k, v) for k, v in top if v is not None and (v * d_roe) < 0 and abs(v) >= sig]
        if opp:
            offset_txt = _render_template(
                tpl.get("roe_offset", "存在显著对冲项：{offset_driver}（{offset_contrib}个百分点），抵消了{main_driver}的部分影响。"),
                {
                    "offset_driver": opp[0][0],
                    "offset_contrib": f"{abs(opp[0][1]):.2f}",
                    "main_driver": main_driver,
                },
            )
    if len(years) >= 3:
        factors_txt = _render_template(
            tpl.get("key_roe_factors", DEFAULT_TEXT_TEMPLATES["key_roe_factors"]),
            {
                "y1": y1,
                "y2": y2,
                "y3": y3,
                "nm1": _fmt_ratio_value(nm_pct.get(y1), "%"),
                "nm2": _fmt_ratio_value(nm_pct.get(y2), "%"),
                "nm3": _fmt_ratio_value(nm_pct.get(y3), "%"),
                "at1": _fmt_ratio_value(at.get(y1), "x"),
                "at2": _fmt_ratio_value(at.get(y2), "x"),
                "at3": _fmt_ratio_value(at.get(y3), "x"),
                "em1": _fmt_ratio_value(em.get(y1), "x"),
                "em2": _fmt_ratio_value(em.get(y2), "x"),
                "em3": _fmt_ratio_value(em.get(y3), "x"),
            },
        )
    else:
        factors_txt = _render_template(
            tpl.get("key_roe_factors_two_year", "净利率：{y1}年{nm1}，{y2}年{nm2}；总资产周转率：{y1}年{at1}，{y2}年{at2}；权益乘数：{y1}年{em1}，{y2}年{em2}。"),
            {
                "y1": y1,
                "y2": y2,
                "nm1": _fmt_ratio_value(nm_pct.get(y1), "%"),
                "nm2": _fmt_ratio_value(nm_pct.get(y2), "%"),
                "at1": _fmt_ratio_value(at.get(y1), "x"),
                "at2": _fmt_ratio_value(at.get(y2), "x"),
                "em1": _fmt_ratio_value(em.get(y1), "x"),
                "em2": _fmt_ratio_value(em.get(y2), "x"),
            },
        )
    return _append_text(
        _append_text(_append_text(_append_text(headline, three_year_line), method_txt), _append_text(driver_txt, contrib_detail_txt)),
        _append_text(_append_text(reconcile_txt, factors_txt), offset_txt),
    )


def _gross_margin_topic_text(wb, rows: List[Dict[str, Any]], years: List[str], key_rules: Dict[str, Any]) -> str:
    tpl = key_rules.get("templates", {})
    years = _year_seq(years)
    if len(years) < 2:
        return _render_template(tpl.get("gm_missing_years", "毛利率贡献分析所需年度不足，待补充。"), {})
    y1, y2, y3 = _compat_three_periods(years)
    th = key_rules.get("thresholds", {}).get("gross_margin", DEFAULT_KEY_DRIVER_THRESHOLDS)
    segs = _detect_income_segments(wb, years)
    if not segs:
        return _render_template(tpl.get("gm_missing_segments", "未识别到营业收入分项，毛利率贡献待补充。"), {})

    rec_gm = _pick_ratio_record(rows, ["gross_margin"], ["销售毛利率", "毛利率"])
    gm = {}
    for y in years:
        v = None
        if rec_gm and isinstance(rec_gm.get("values"), dict):
            v = rec_gm["values"].get(y)
        gm[y] = v * 100.0 if (v is not None and abs(v) <= 1.0) else v

    rev_total = {}
    gp_total = {}
    for y in years:
        revs = [(s.get("revenue_values", {}) or {}).get(y) for s in segs]
        gps = [(s.get("gross_values", {}) or {}).get(y) for s in segs]
        rev_total[y] = sum([v for v in revs if v is not None]) if any(v is not None for v in revs) else None
        gp_total[y] = sum([v for v in gps if v is not None]) if any(v is not None for v in gps) else None

    items = []
    for s in segs:
        name = str(s.get("label", ""))
        rv = (s.get("revenue_values", {}) or {}).get(y3)
        gv = (s.get("gross_values", {}) or {}).get(y3)
        share = None if rv is None or rev_total.get(y3) in (None, 0) else rv / rev_total[y3] * 100.0
        margin = None if rv in (None, 0) or gv is None else gv / rv * 100.0
        items.append((name, share, margin))
    items = sorted(items, key=lambda x: abs(x[1]) if x[1] is not None else -1, reverse=True)[:3]
    top_txt = "；".join([f"{n} 收入占比{_fmt_ratio_value(s, '%')}、分项毛利率{_fmt_ratio_value(m, '%')}" for n, s, m in items]) if items else "待补充"

    # Gross margin driver decomposition: structure + margin + interaction.
    seg_names = sorted({str(s.get("label", "")) for s in segs if str(s.get("label", "")).strip()})
    structure_effect = 0.0
    margin_effect = 0.0
    interaction_effect = 0.0
    has_component = False
    seg_effect_rows: List[Dict[str, Any]] = []
    for name in seg_names:
        seg = next((x for x in segs if str(x.get("label", "")) == name), None)
        if not seg:
            continue
        r1 = (seg.get("revenue_values", {}) or {}).get(y1)
        r3 = (seg.get("revenue_values", {}) or {}).get(y3)
        g1 = (seg.get("gross_values", {}) or {}).get(y1)
        g3 = (seg.get("gross_values", {}) or {}).get(y3)
        if rev_total.get(y1) in (None, 0) or rev_total.get(y3) in (None, 0) or r1 is None or r3 is None:
            continue
        s1 = r1 / rev_total[y1]
        s3 = r3 / rev_total[y3]
        m1 = None if r1 in (None, 0) or g1 is None else g1 / r1 * 100.0
        m3 = None if r3 in (None, 0) or g3 is None else g3 / r3 * 100.0
        if m1 is None or m3 is None:
            continue
        has_component = True
        ds = (s3 - s1) * m1
        dm = s1 * (m3 - m1)
        di = (s3 - s1) * (m3 - m1)
        dt = ds + dm + di
        structure_effect += ds
        margin_effect += dm
        interaction_effect += di
        seg_effect_rows.append({"name": name, "ds": ds, "dm": dm, "di": di, "dt": dt})

    contrib = [("结构效应", structure_effect if has_component else None), ("价格/成本效应", margin_effect if has_component else None), ("交互效应", interaction_effect if has_component else None)]
    c = _classify_driver_case(contrib, th)
    top = c["top"]
    shares = c["shares"]

    headline = _render_template(
        tpl.get("gm_headline", "{metric_name}由{y_prev}年的{v_prev}变动至{y_curr}年的{v_curr}，整体{trend_word}。"),
        {
            "metric_name": "毛利率",
            "y_prev": y1,
            "v_prev": _fmt_ratio_value(gm.get(y1), "%"),
            "y_curr": y3,
            "v_curr": _fmt_ratio_value(gm.get(y3), "%"),
            "trend_word": _trend_word(
                gm.get(y3),
                gm.get(y1),
                stable_pp=float(th.get("delta_stable_pp", 2.0)),
                up_label=str(tpl.get("trend_word_up", "上升")),
                down_label=str(tpl.get("trend_word_down", "下降")),
                stable_label=str(tpl.get("trend_word_stable", "基本稳定")),
            ),
        },
    )
    if len(years) >= 3:
        three_year_line = _render_template(
            tpl.get("gm_three_year_values", "毛利率三年分别为：{y1}年{v1}，{y2}年{v2}，{y3}年{v3}。"),
            {
                "y1": y1,
                "y2": y2,
                "y3": y3,
                "v1": _fmt_ratio_value(gm.get(y1), "%"),
                "v2": _fmt_ratio_value(gm.get(y2), "%"),
                "v3": _fmt_ratio_value(gm.get(y3), "%"),
            },
        )
    else:
        three_year_line = _render_template(
            tpl.get("gm_two_year_values", "毛利率两年分别为：{y1}年{v1}，{y2}年{v2}。"),
            {
                "y1": y1,
                "y2": y2,
                "v1": _fmt_ratio_value(gm.get(y1), "%"),
                "v2": _fmt_ratio_value(gm.get(y2), "%"),
            },
        )
    if c["case"] == "single" and top:
        driver_txt = _render_template(
            tpl.get("gm_single_driver", "主要驱动来自{driver_1}，贡献{contrib_1}个百分点，驱动占比{share_1}。"),
            {"driver_1": top[0][0], "contrib_1": f"{abs(top[0][1]):.2f}", "share_1": f"{shares.get(top[0][0], 0.0) * 100:.2f}%"},
        )
    elif c["case"] == "dual" and len(top) >= 2:
        driver_txt = _render_template(
            tpl.get("gm_dual_driver", "变动由{driver_1}与{driver_2}共同驱动，分别贡献{contrib_1}/{contrib_2}个百分点，合计驱动占比{share_sum}。"),
            {
                "driver_1": top[0][0],
                "driver_2": top[1][0],
                "contrib_1": f"{abs(top[0][1]):.2f}",
                "contrib_2": f"{abs(top[1][1]):.2f}",
                "share_sum": f"{(shares.get(top[0][0], 0.0) + shares.get(top[1][0], 0.0)) * 100:.2f}%",
            },
        )
    else:
        driver_txt = _render_template(
            tpl.get("gm_multi_driver", "毛利率变动为多因素共同作用：{driver_list}。"),
            {"driver_list": "、".join([k for k, _ in top]) if top else "待补充"},
        )
    effect_txt = _render_template(
        tpl.get("gm_structure_effect", "其中结构效应{structure_effect}个百分点，价格/成本效应{price_cost_effect}个百分点，交互项{interaction_effect}个百分点。"),
        {
            "structure_effect": f"{abs(structure_effect):.2f}",
            "price_cost_effect": f"{abs(margin_effect):.2f}",
            "interaction_effect": f"{abs(interaction_effect):.2f}",
        },
    )
    # Segment-level decomposition:
    # dt_i = ds_i + dm_i + di_i, and sum(dt_i) = total gross-margin change (in percentage points).
    gm_delta_pp = None if gm.get(y3) is None or gm.get(y1) is None else (gm.get(y3) - gm.get(y1))
    net_denom = gm_delta_pp if gm_delta_pp not in (None, 0) else None
    strength_denom = sum(abs(float(x.get("dt", 0.0))) for x in seg_effect_rows if x.get("dt") is not None)

    def _fmt_pp_signed(v: Optional[float]) -> str:
        if v is None:
            return "待补充"
        if v > 0:
            return f"+{v:.2f}"
        return f"{v:.2f}"

    def _fmt_pct_share(v: Optional[float]) -> str:
        if v is None:
            return "待补充"
        return f"{v:.2f}%"

    seg_header = _render_template(tpl.get("gm_segment_header", "分项影响拆解（{y1}-{y3}）："), {"y1": y1, "y3": y3})
    seg_lines: List[str] = []
    for row in sorted(seg_effect_rows, key=lambda z: abs(float(z.get("dt", 0.0))), reverse=True):
        dt = row.get("dt")
        net_share = None
        if dt is not None and net_denom not in (None, 0):
            net_share = float(dt) / float(net_denom) * 100.0
        strength_share = None
        if dt is not None and strength_denom > 1e-12:
            strength_share = abs(float(dt)) / strength_denom * 100.0
        line = _render_template(
            tpl.get(
                "gm_segment_line",
                "{name}：结构{ds}个百分点，价格/成本{dm}个百分点，交互{di}个百分点，合计{dt}个百分点；净影响占比{net_share}，影响强度占比{strength_share}。",
            ),
            {
                "name": str(row.get("name", "")),
                "ds": _fmt_pp_signed(row.get("ds")),
                "dm": _fmt_pp_signed(row.get("dm")),
                "di": _fmt_pp_signed(row.get("di")),
                "dt": _fmt_pp_signed(dt),
                "net_share": _fmt_pct_share(net_share)
                if net_denom not in (None, 0)
                else str(tpl.get("gm_net_share_na", "不适用（整体变动接近0）")),
                "strength_share": _fmt_pct_share(strength_share),
            },
        )
        seg_lines.append(line)
    pos = [x for x in seg_effect_rows if (x.get("dt") or 0) > 0]
    neg = [x for x in seg_effect_rows if (x.get("dt") or 0) < 0]
    pos.sort(key=lambda z: abs(float(z.get("dt", 0.0))), reverse=True)
    neg.sort(key=lambda z: abs(float(z.get("dt", 0.0))), reverse=True)
    top_pos = "、".join([f"{x.get('name')}（{_fmt_pp_signed(x.get('dt'))}个百分点）" for x in pos[:3]]) or str(
        tpl.get("gm_top_fallback", "待补充")
    )
    top_neg = "、".join([f"{x.get('name')}（{_fmt_pp_signed(x.get('dt'))}个百分点）" for x in neg[:3]]) or str(
        tpl.get("gm_top_fallback", "待补充")
    )
    seg_summary = _render_template(
        tpl.get("gm_segment_summary", "子项层面：正向拉动前列{top_pos}；负向拖累前列{top_neg}。"),
        {"top_pos": top_pos, "top_neg": top_neg},
    )
    seg_text = _append_text(seg_header, "\n".join(seg_lines))
    seg_text = _append_text(seg_text, seg_summary)
    top_seg_txt = _render_template(
        tpl.get("key_gm_top_segments", DEFAULT_TEXT_TEMPLATES["key_gm_top_segments"]),
        {"top_txt": top_txt},
    )
    extra = ""
    mixed_sign = any((x.get("gross_values", {}) or {}).get(y3, 0) > 0 for x in segs) and any((x.get("gross_values", {}) or {}).get(y3, 0) < 0 for x in segs)
    if gp_total.get(y3) is not None and gp_total.get(y3) < 0 or mixed_sign:
        extra = _render_template(
            tpl.get("gm_negative_profit_case", "在总毛利为负或分项正负并存情况下，采用影响占比口径：{impact_summary}。"),
            {"impact_summary": str(tpl.get("gm_negative_impact_summary", "各分项按绝对影响占比展示，不采用简单占比"))},
        )
    return _append_text(
        _append_text(_append_text(headline, three_year_line), driver_txt),
        _append_text(_append_text(_append_text(effect_txt, top_seg_txt), seg_text), extra),
    )


def build_ratio_analysis_map(wb, project_id: str) -> Dict[str, Any]:
    ws = get_sheet_by_loose_name(wb, ["财务比率", "财务指标", "财务指标表"])
    if ws is None:
        return {"sheet_title": None, "years": [], "tree": [], "nodes": []}

    ratio_data = read_ratio_rows(ws)
    years = ratio_data.get("years", [])[:3]
    if len(years) < 2:
        return {"sheet_title": ratio_data.get("sheet_title"), "years": years, "tree": [], "nodes": []}
    rows = ratio_data.get("rows", [])
    ratio_by_code = {str(r.get("code", "")).strip(): r for r in rows}
    ratio_by_name_key = {}
    for r in rows:
        nk = _norm_metric_name_key(r.get("name", ""))
        if nk and nk not in ratio_by_name_key:
            ratio_by_name_key[nk] = r

    rules = load_ratio_analysis_rules()
    tree_rules = _sanitize_ratio_tree_nodes(rules.get("tree_nodes", []))
    catalog = rules.get("catalog", {})
    fallback_threshold = float(rules.get("trend_threshold_pp", 2.0))
    trend_rules = list(rules.get("trend_rules", []))
    judgement_rules = dict(rules.get("judgement_rules", {}))
    alert_rules = list(rules.get("alert_rules", []))
    text_templates = dict(rules.get("text_templates", {}))

    tree = []
    nodes = []

    def add_node(node_id: str, parent_id: str, label: str, values: Dict[str, Optional[float]], source_code: str, source_name: str, auto_text: str):
        nodes.append(
            {
                "node_id": node_id,
                "parent_id": parent_id,
                "label": label,
                "source_code": source_code,
                "source_name": source_name,
                "values": values,
                "auto_text": auto_text,
            }
        )

    for t in tree_rules:
        nid = str(t.get("node_id", "")).strip()
        parent = str(t.get("parent_id", "")).strip()
        label = str(t.get("label_zh", "")).strip()
        ntype = str(t.get("node_type", "group")).strip().lower()
        iid = str(t.get("indicator_id", "")).strip()
        if not nid or not label:
            continue
        # Extra hard guard: regular page never accepts key page node IDs.
        if nid in KEY_RATIO_IDS:
            continue
        tree.append({"node_id": nid, "parent_id": parent, "label": label})
        if ntype != "indicator":
            continue
        cat = catalog.get(iid, {"name": label, "direction": "higher", "unit": "%"})
        rec = ratio_by_code.get(iid)
        if not rec:
            rec = ratio_by_name_key.get(_norm_metric_name_key(cat.get("name", "")))
        if not rec:
            rec = ratio_by_name_key.get(_norm_metric_name_key(label))
        rec = rec or {}
        vals = rec.get("values", {}) if isinstance(rec.get("values"), dict) else {}
        val_map = {y: vals.get(y) for y in years}
        y1, y2, y3 = _compat_three_periods(years)
        unit = cat.get("unit", "%")
        trend_rule = _select_ratio_trend_rule(
            trend_rules,
            iid,
            str(cat.get("name", "")),
            str(cat.get("group", "")),
        )
        if not trend_rule:
            trend_rule = {
                "threshold_type": "delta_abs",
                "stable_threshold": fallback_threshold,
                "warn_threshold": 5.0,
                "critical_threshold": 10.0,
                "unit": "pct_point",
            }
        labels = judgement_rules.get(cat.get("direction", "higher"), {})
        p21 = _ratio_judgement_parts(val_map.get(y2), val_map.get(y1), cat.get("direction", "higher"), trend_rule, labels, unit)
        p32 = _ratio_judgement_parts(val_map.get(y3), val_map.get(y2), cat.get("direction", "higher"), trend_rule, labels, unit)
        value_tpl = text_templates.get("indicator_value", DEFAULT_TEXT_TEMPLATES["ratio_indicator_value"])
        trend_tpl = text_templates.get("indicator_trend", DEFAULT_TEXT_TEMPLATES["ratio_indicator_trend"])
        seq_years = _year_seq(years)
        value_parts = [f"{y}年{_fmt_ratio_number(val_map.get(y), unit)}{unit}" for y in seq_years]
        trend_parts: List[str] = []
        for i in range(1, len(seq_years)):
            yp = seq_years[i - 1]
            yc = seq_years[i]
            pj = _ratio_judgement_parts(
                val_map.get(yc),
                val_map.get(yp),
                cat.get("direction", "higher"),
                trend_rule,
                labels,
                unit,
            )
            trend_parts.append(f"{yc}较{yp}{pj['judgement']}（变动{pj['delta']}{pj['unit2']}）")
        if len(seq_years) >= 3:
            y_first = seq_years[0]
            y_last = seq_years[-1]
            pj_fl = _ratio_judgement_parts(
                val_map.get(y_last),
                val_map.get(y_first),
                cat.get("direction", "higher"),
                trend_rule,
                labels,
                unit,
            )
            trend_parts.append(f"{y_last}较{y_first}{pj_fl['judgement']}（变动{pj_fl['delta']}{pj_fl['unit2']}）")
        text_ctx = {
            "name": label,
            "unit": unit,
            "y1": y1,
            "y2": y2,
            "y3": y3,
            "v1": _fmt_ratio_number(val_map.get(y1), unit),
            "v2": _fmt_ratio_number(val_map.get(y2), unit),
            "v3": _fmt_ratio_number(val_map.get(y3), unit),
            "judgement21": p21["judgement"],
            "judgement32": p32["judgement"] if len(seq_years) >= 3 else "",
            "delta21": p21["delta"],
            "delta32": p32["delta"] if len(seq_years) >= 3 else "",
            "unit2": p21["unit2"],
            "series_values": "，".join(value_parts),
            "series_trends": "；".join(trend_parts),
            "y_first": seq_years[0] if seq_years else "",
            "y_last": seq_years[-1] if seq_years else "",
            "y_prev": seq_years[-2] if len(seq_years) >= 2 else (seq_years[-1] if seq_years else ""),
        }
        if len(seq_years) == 2 and ("{y3}" in value_tpl or "{v3}" in value_tpl):
            text1 = f"{label}：{text_ctx['series_values']}。"
        else:
            text1 = _render_template(value_tpl, text_ctx)
        if len(seq_years) == 2 and ("{y3}" in trend_tpl or "{judgement32}" in trend_tpl or "{delta32}" in trend_tpl):
            text2 = f"{text_ctx['series_trends']}。"
        else:
            text2 = _render_template(trend_tpl, text_ctx)
        alert_text = _ratio_alert_text(iid, val_map, years, unit, alert_rules, text_templates)
        auto_text = _append_text(_append_text(text1, text2), alert_text)
        add_node(nid, parent, label, val_map, str(rec.get("code", "")), str(rec.get("name", "")), auto_text)

    store = load_store(project_id)
    entries = store.get("entries", {})
    for n in nodes:
        saved = entries.get(make_entry_key("ratio_analysis", n["node_id"]), {})
        manual = str(saved.get("manual_text", "") or "")
        confirmed = bool(saved.get("confirmed", False))
        n["manual_text"] = manual
        n["confirmed"] = confirmed
        n["final_text"] = manual if manual.strip() else n["auto_text"]

    return {"sheet_title": ratio_data.get("sheet_title"), "years": years, "tree": tree, "nodes": nodes}


def build_key_ratio_analysis_map(wb, project_id: str) -> Dict[str, Any]:
    ws = get_sheet_by_loose_name(wb, ["财务比率", "财务指标", "财务指标表"])
    if ws is None:
        return {"sheet_title": None, "years": [], "tree": [], "nodes": []}

    ratio_data = read_ratio_rows(ws)
    years = ratio_data.get("years", [])[:3]
    rows = ratio_data.get("rows", [])

    tree = [
        {"node_id": "K1", "parent_id": "", "label": "ROE重点分析"},
        {"node_id": "K2", "parent_id": "", "label": "毛利率重点分析"},
    ]

    key_rules = load_key_ratio_text_rules()
    roe_text = _dupont_topic_text(rows, years, key_rules)
    gm_text = _gross_margin_topic_text(wb, rows, years, key_rules)
    nodes = [
        {
            "node_id": "K1",
            "parent_id": "",
            "label": "ROE重点分析",
            "source_code": "roe",
            "source_name": "净资产收益率",
            "values": {y: None for y in years},
            "auto_text": roe_text,
        },
        {
            "node_id": "K2",
            "parent_id": "",
            "label": "毛利率重点分析",
            "source_code": "gross_margin",
            "source_name": "销售毛利率",
            "values": {y: None for y in years},
            "auto_text": gm_text,
        },
    ]
    # Hard guard: key page only renders dedicated K1/K2 nodes.
    tree = [t for t in tree if str(t.get("node_id", "")).strip() in KEY_RATIO_IDS]
    nodes = [n for n in nodes if str(n.get("node_id", "")).strip() in KEY_RATIO_IDS]

    store = load_store(project_id)
    entries = store.get("entries", {})
    for n in nodes:
        saved = entries.get(make_entry_key("key_ratio_analysis", n["node_id"]), {})
        manual = str(saved.get("manual_text", "") or "")
        confirmed = bool(saved.get("confirmed", False))
        n["manual_text"] = manual
        n["confirmed"] = confirmed
        n["final_text"] = manual if manual.strip() else n["auto_text"]

    return {"sheet_title": ratio_data.get("sheet_title"), "years": years, "tree": tree, "nodes": nodes}


def _ratio_indicator_catalog_rows() -> List[Dict[str, Any]]:
    rules = load_ratio_analysis_rules()
    out: List[Dict[str, Any]] = []
    for iid, c in (rules.get("catalog", {}) or {}).items():
        out.append(
            {
                "indicator_id": str(iid or "").strip(),
                "indicator_name_zh": str(c.get("name", "")).strip(),
                "group_zh": str(c.get("group", "")).strip(),
                "unit": str(c.get("unit", "")).strip(),
                "direction": str(c.get("direction", "")).strip(),
                "value_source": str(c.get("value_source", "")).strip(),
                "formula_expr": str(c.get("formula_expr", "")).strip(),
            }
        )
    out = [x for x in out if x.get("indicator_id")]
    out.sort(key=lambda x: x.get("indicator_id", ""))
    return out


def _norm_metric_name_key(text: Any) -> str:
    s = str(text or "").strip().lower()
    if not s:
        return ""
    s = re.sub(r"[\s\(\)（）:：,，、\.\-_/]+", "", s)
    return s


def _detect_project_years_for_ratio(project_id: str, wb=None) -> List[str]:
    own_wb = False
    if wb is None:
        wb_path = workbook_path(project_id)
        if not wb_path.exists():
            return ["2022", "2023", "2024"]
        wb = load_workbook(wb_path, data_only=False)
        own_wb = True
    try:
        ws_ratio = get_sheet_by_loose_name(wb, ["财务比率", "财务指标", "财务指标表"])
        if ws_ratio is not None:
            r = read_ratio_rows(ws_ratio)
            yrs = [str(y) for y in (r.get("years", []) or []) if str(y).strip()]
            if yrs:
                return yrs[:DEFAULT_YEAR_COUNT]
        for g in SHEET_GROUPS:
            if g.get("id") not in {"bs", "is", "cf"}:
                continue
            ws = get_sheet_by_loose_name(wb, g.get("candidates", []))
            if ws is None:
                continue
            yrs = parse_years_from_sheet(ws)
            if yrs:
                return yrs[:DEFAULT_YEAR_COUNT]
        return ["2022", "2023", "2024"]
    finally:
        if own_wb and wb is not None:
            wb.close()


def read_ratio_rows(ws) -> Dict[str, Any]:
    """Read ratio sheet in long format and pivot to wide by year."""
    # Default column positions by current template.
    idx_col, name_col, period_col, value_col = 1, 2, 4, 5

    # Try detect columns by header text in row 1.
    headers = {c: str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)}
    for c, h in headers.items():
        if h == "指标ID":
            idx_col = c
        elif h == "指标":
            name_col = c
        elif h == "期间":
            period_col = c
        elif h == "数值":
            value_col = c

    years_set = set()
    bucket: Dict[str, Dict[str, Any]] = {}

    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, idx_col).value or "").strip()
        name = str(ws.cell(r, name_col).value or "").strip()
        period_text = str(ws.cell(r, period_col).value or "").strip()
        value = normalize_num(ws.cell(r, value_col).value)
        if not code or not name:
            continue

        m = re.search(r"(20\d{2})", period_text)
        if not m:
            continue
        year = m.group(1)
        years_set.add(year)

        obj = bucket.setdefault(code, {"code": code, "name": name, "values": {}})
        obj["name"] = name
        obj["values"][year] = value

    years = sorted(years_set)[:DEFAULT_YEAR_COUNT]
    rows = list(bucket.values())
    # Keep sheet rows aligned with indicator catalog:
    # when a new indicator is added in rules, it should appear in ratio table/template automatically.
    cat_rows = _ratio_indicator_catalog_rows()
    existing_codes = {str(x.get("code", "")).strip() for x in rows}
    existing_name_keys = {_norm_metric_name_key(x.get("name", "")) for x in rows if _norm_metric_name_key(x.get("name", ""))}
    for c in cat_rows:
        iid = str(c.get("indicator_id", "")).strip()
        cname = str(c.get("indicator_name_zh", iid)).strip()
        cname_key = _norm_metric_name_key(cname)
        if not iid or iid in existing_codes:
            continue
        if cname_key and cname_key in existing_name_keys:
            continue
        rows.append({"code": iid, "name": cname, "values": {}})
        existing_codes.add(iid)
        if cname_key:
            existing_name_keys.add(cname_key)
    for row in rows:
        for y in years:
            row["values"].setdefault(y, None)

    return {"sheet_title": ws.title, "years": years, "rows": rows}


def read_sheet_rows(wb, group: Dict[str, Any], project_id: Optional[str] = None) -> Dict[str, Any]:
    ws = get_sheet_by_loose_name(wb, group["candidates"])
    if ws is None:
        return {"sheet_title": None, "years": [], "rows": []}

    if group["id"] == "ratio":
        return read_ratio_rows(ws)

    years = parse_years_from_sheet(ws)
    rows: List[Dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        name = str(ws.cell(r, 2).value or "").strip()
        if not code or not name:
            continue

        if group["id"] in {"bs", "is", "cf"} and not re.match(r"^(BS|IS|CF)\d+$", code):
            continue

        vals: Dict[str, Optional[float]] = {}
        for i, year in enumerate(years):
            vals[year] = normalize_num(ws.cell(r, 3 + i).value)

        rows.append({"code": code, "name": name, "values": vals})

    rows = apply_value_overrides_to_rows(rows, years, str(group.get("id", "")), project_id)
    return {"sheet_title": ws.title, "years": years, "rows": rows}


def _find_bs_subject_name(wb, code: str, project_id: Optional[str] = None) -> str:
    bs_group = next((g for g in SHEET_GROUPS if str(g.get("id", "")) == "bs"), None)
    if not bs_group:
        return ""
    bs_data = read_sheet_rows(wb, bs_group, project_id=project_id)
    code_u = str(code or "").strip().upper()
    for r in (bs_data.get("rows", []) or []):
        if str(r.get("code", "")).strip().upper() == code_u:
            return str(r.get("name", "")).strip()
    return ""


def _normalize_subject_key(text: Any) -> str:
    s = str(text or "").strip()
    if not s:
        return ""
    s = re.sub(r"[\s\(\)（）:：,，、\.\-_/]+", "", s)
    s = s.replace("合计", "")
    return s.lower()


def _find_bs_subject_row(wb, code: str, project_id: Optional[str] = None) -> Dict[str, Any]:
    bs_group = next((g for g in SHEET_GROUPS if str(g.get("id", "")) == "bs"), None)
    if not bs_group:
        return {}
    bs_data = read_sheet_rows(wb, bs_group, project_id=project_id)
    code_u = str(code or "").strip().upper()
    for r in (bs_data.get("rows", []) or []):
        if str(r.get("code", "")).strip().upper() == code_u:
            return r
    return {}


def _load_detail_sheet_mapping_rules() -> Dict[str, str]:
    out: Dict[str, str] = {}
    p = PROJECT_ROOT / "config" / "rulebook.xlsx"
    if not p.exists():
        return out
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception:
        return out
    try:
        ws = get_sheet_by_loose_name(wb, ["detail_sheet_mapping"])
        if ws is None:
            return out
        headers = {str(ws.cell(1, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}
        c_code = headers.get("code", 1)
        c_sheet = headers.get("sheet_name", 2)
        for r in range(2, ws.max_row + 1):
            code = str(ws.cell(r, c_code).value or "").strip().upper()
            sn = str(ws.cell(r, c_sheet).value or "").strip()
            if code and sn:
                out[code] = sn
    except Exception:
        return out
    finally:
        wb.close()
    return out


def _find_detail_sheet_for_subject(wb, subject_code: str, subject_name: str):
    code_u = str(subject_code or "").strip().upper()
    name = str(subject_name or "").strip()
    if not code_u or not name:
        return None, "none"

    # 1) explicit mapping rule: code -> sheet name
    mapping = _load_detail_sheet_mapping_rules()
    mapped_name = mapping.get(code_u)
    if mapped_name:
        ws = wb[mapped_name] if mapped_name in wb.sheetnames else None
        if ws is not None:
            return ws, "rulebook_code_mapping"

    # 2) strict by code in sheet title, e.g. 明细_BS001_货币资金
    for ws in wb.worksheets:
        title = str(ws.title or "").strip()
        if not title.startswith("明细_"):
            continue
        m = re.search(r"\b(BS\d{3})\b", title.upper())
        if m and m.group(1) == code_u:
            return ws, "sheet_title_code"

    # 3) strict exact title
    exact_title = f"明细_{name}"
    for ws in wb.worksheets:
        if str(ws.title or "").strip() == exact_title:
            return ws, "sheet_title_exact"

    # 4) strict normalized equality (no fuzzy contains)
    target_key = _normalize_subject_key(name)
    if target_key:
        for ws in wb.worksheets:
            title = str(ws.title or "").strip()
            if not title.startswith("明细_"):
                continue
            key = _normalize_subject_key(title.replace("明细_", "", 1))
            if key and key == target_key:
                return ws, "sheet_title_norm_equal"
    return None, "not_found"


def _read_detail_sheet_rows(ws) -> Dict[str, Any]:
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or "").strip()
        headers.append(v or f"列{c}")

    rows: List[List[Any]] = []
    for r in range(2, ws.max_row + 1):
        line: List[Any] = []
        has_value = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                has_value = True
            line.append(v)
        if has_value:
            rows.append(line)
    return {"headers": headers, "rows": rows}


def _canonicalize_detail_rows(headers: List[str], rows: List[List[Any]]) -> List[Dict[str, Any]]:
    hdrs = [str(h or "").strip() for h in (headers or [])]
    idx_period = 0
    idx_item = 1
    idx_value = 2
    idx_note = 3
    for i, h in enumerate(hdrs):
        if re.search(r"期间|年度|年份|日期|期末", h):
            idx_period = i
        elif re.search(r"项目|名称|对象|业务|客户|单位", h):
            idx_item = i
        elif re.search(r"明细值|金额|余额|净额|数值|账面|原值|价值|成本|收入|毛利", h):
            idx_value = i
        elif re.search(r"说明|备注|注释", h):
            idx_note = i
    out: List[Dict[str, Any]] = []
    for row in (rows or []):
        def _pick(i: int) -> str:
            if i < 0 or i >= len(row):
                return ""
            v = row[i]
            if v is None:
                return ""
            return str(v).strip()

        val_raw = _pick(idx_value)
        val = normalize_num(val_raw)
        out.append(
            {
                "period": _pick(idx_period),
                "item": _pick(idx_item),
                "value": (f"{val:.2f}" if val is not None else val_raw),
                "note": _pick(idx_note),
            }
        )
    return [x for x in out if any(str(x.get(k, "")).strip() for k in ["period", "item", "value", "note"])]


def _build_default_detail_rows(years: List[str]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for y in (years or [])[:DEFAULT_YEAR_COUNT]:
        yy = str(y).strip()
        if not yy:
            continue
        out.append({"period": f"{yy}年", "item": "", "value": "", "note": ""})
    return out


def _merge_rows_to_year_slots(src_rows: List[Dict[str, Any]], years: List[str]) -> List[Dict[str, Any]]:
    slots = _build_default_detail_rows(years)
    if not slots:
        return []
    by_year: Dict[str, Dict[str, Any]] = {}
    for r in (src_rows or []):
        if not isinstance(r, dict):
            continue
        p = str(r.get("period", "")).strip()
        m = re.search(r"(20\d{2})", p)
        if not m:
            continue
        y = m.group(1)
        if y in by_year:
            continue
        by_year[y] = {
            "period": f"{y}年",
            "item": str(r.get("item", "")).strip(),
            "value": str(r.get("value", "")).strip(),
            "note": str(r.get("note", "")).strip(),
        }
    out: List[Dict[str, Any]] = []
    for s in slots:
        m = re.search(r"(20\d{2})", str(s.get("period", "")))
        y = m.group(1) if m else ""
        out.append(by_year.get(y, s))
    return out


def _parse_detail_rows_from_xlsx_bytes(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdrs: List[str] = []
    for c in range(1, ws.max_column + 1):
        hdrs.append(str(ws.cell(1, c).value or "").strip())
    rows: List[List[Any]] = []
    for r in range(2, ws.max_row + 1):
        line: List[Any] = []
        has_value = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                has_value = True
            line.append(v)
        if has_value:
            rows.append(line)
    wb.close()
    return _canonicalize_detail_rows(hdrs, rows)


def _build_detail_auto_text(
    subject_name: str,
    detail_rows: List[Dict[str, Any]],
    max_top_items: int = 3,
) -> str:
    rows = detail_rows or []
    n_rows = len(rows)
    if n_rows == 0:
        return f"{subject_name}：未读取到明细数据，待补充。"

    year_sum: Dict[str, float] = {}
    latest_year = ""
    latest_items: List[Dict[str, Any]] = []
    for row in rows:
        val = normalize_num(row.get("value"))
        if val is None:
            continue
        period_text = str(row.get("period", ""))
        m = re.search(r"(20\d{2})", period_text)
        year = m.group(1) if m else ""
        if year:
            year_sum[year] = year_sum.get(year, 0.0) + float(val)
        if year and (not latest_year or year > latest_year):
            latest_year = year

    years = sorted(year_sum.keys())
    parts: List[str] = [f"{subject_name}：已读取明细{n_rows}条。"]
    if years:
        segs = [f"{y}年{year_sum.get(y, 0.0):.2f}" for y in years]
        parts.append("按期间汇总：" + "，".join(segs) + "。")
        if len(years) >= 2:
            y_prev, y_curr = years[-2], years[-1]
            prev = year_sum.get(y_prev, 0.0)
            curr = year_sum.get(y_curr, 0.0)
            if abs(prev) > 1e-12:
                pct = (curr - prev) / abs(prev) * 100.0
                if pct > 0:
                    parts.append(f"{y_curr}较{y_prev}增加{abs(pct):.2f}%。")
                elif pct < 0:
                    parts.append(f"{y_curr}较{y_prev}减少{abs(pct):.2f}%。")
                else:
                    parts.append(f"{y_curr}较{y_prev}基本持平。")
        if latest_year:
            for row in rows:
                if latest_year not in str(row.get("period", "") or ""):
                    continue
                val = normalize_num(row.get("value"))
                if val is None:
                    continue
                nm = str(row.get("item", "")).strip() or "未命名项"
                latest_items.append({"name": nm, "value": float(val)})
            if latest_items:
                latest_items.sort(key=lambda x: abs(float(x.get("value", 0.0))), reverse=True)
                top = latest_items[:max_top_items]
                top_txt = "、".join([f"{x['name']}（{x['value']:.2f}）" for x in top])
                parts.append(f"{latest_year}年主要项目：{top_txt}。")
    else:
        parts.append("未识别到期间字段，建议补充期间列后再生成趋势描述。")

    return "".join(parts)


def build_detail_payload(wb, project_id: str, code: str) -> Dict[str, Any]:
    code_u = str(code or "").strip().upper()
    subject_row = _find_bs_subject_row(wb, code_u, project_id=project_id)
    subject_name = str(subject_row.get("name", "")).strip()
    subject_values = subject_row.get("values", {}) if isinstance(subject_row.get("values"), dict) else {}
    years = sorted([str(y) for y in subject_values.keys() if str(y).strip()])[:DEFAULT_YEAR_COUNT]
    store = load_store(project_id)
    saved = (store.get("entries", {}) or {}).get(make_entry_key("detail_analysis", code_u), {})
    saved_rows = saved.get("detail_rows", []) if isinstance(saved.get("detail_rows", []), list) else []
    canonical_saved_rows = []
    for x in saved_rows:
        if not isinstance(x, dict):
            continue
        canonical_saved_rows.append(
            {
                "period": str(x.get("period", "")).strip(),
                "item": str(x.get("item", "")).strip(),
                "value": str(x.get("value", "")).strip(),
                "note": str(x.get("note", "")).strip(),
            }
        )
    default_rows = _build_default_detail_rows(years)
    if not subject_name:
        return {
            "code": code_u,
            "name": "",
            "detail_sheet": None,
            "headers": ["期间", "项目", "明细值（万元）", "说明"],
            "rows": [],
            "detail_rows": _merge_rows_to_year_slots(canonical_saved_rows, years) if canonical_saved_rows else default_rows,
            "auto_text": "",
            "manual_text": str(saved.get("manual_text", "") or ""),
            "confirmed": bool(saved.get("confirmed", False)),
            "years": years,
            "subject_values": {str(y): subject_values.get(str(y)) for y in years},
            "mapping_mode": "subject_not_found",
        }

    ws, mapping_mode = _find_detail_sheet_for_subject(wb, code_u, subject_name)
    if ws is None:
        return {
            "code": code_u,
            "name": subject_name,
            "detail_sheet": None,
            "headers": ["期间", "项目", "明细值（万元）", "说明"],
            "rows": [],
            "detail_rows": _merge_rows_to_year_slots(canonical_saved_rows, years) if canonical_saved_rows else default_rows,
            "auto_text": "",
            "manual_text": str(saved.get("manual_text", "") or ""),
            "confirmed": bool(saved.get("confirmed", False)),
            "years": years,
            "subject_values": {str(y): subject_values.get(str(y)) for y in years},
            "mapping_mode": mapping_mode,
        }

    detail_data = _read_detail_sheet_rows(ws)
    headers = detail_data.get("headers", []) or []
    rows = detail_data.get("rows", []) or []
    canonical_sheet_rows = _canonicalize_detail_rows(headers, rows)
    if canonical_saved_rows:
        detail_rows = _merge_rows_to_year_slots(canonical_saved_rows, years)
    elif canonical_sheet_rows:
        detail_rows = _merge_rows_to_year_slots(canonical_sheet_rows, years)
    else:
        detail_rows = default_rows
    rows_out: List[List[str]] = []
    for row in rows:
        line: List[str] = []
        for x in row:
            if x is None:
                line.append("")
            elif isinstance(x, float):
                line.append(f"{x:.2f}")
            else:
                line.append(str(x))
        rows_out.append(line)
    return {
        "code": code_u,
        "name": subject_name,
        "detail_sheet": ws.title,
        "headers": [str(h) for h in headers],
        "rows": rows_out,
        "detail_rows": detail_rows,
        "auto_text": "",
        "manual_text": str(saved.get("manual_text", "") or ""),
        "confirmed": bool(saved.get("confirmed", False)),
        "years": years,
        "subject_values": {str(y): subject_values.get(str(y)) for y in years},
        "mapping_mode": mapping_mode,
    }


def trend_text(
    curr: Optional[float],
    prev: Optional[float],
    stable_pct: float = 2.0,
    up_label: str = "增加",
    down_label: str = "减少",
    stable_label: str = "保持稳定",
) -> str:
    if curr is None or prev in (None, 0):
        return "待补充"
    pct = (curr - prev) / abs(prev) * 100.0
    if pct > stable_pct:
        return f"{up_label}{abs(pct):.2f}%"
    if pct < -stable_pct:
        return f"{down_label}{abs(pct):.2f}%"
    return f"{stable_label}（变动{abs(pct):.2f}%）"


def _empty_year_values(years: List[str]) -> Dict[str, Optional[float]]:
    return {y: None for y in years[:3]}


def _sum_values(items: List[Dict[str, Optional[float]]], years: List[str]) -> Dict[str, Optional[float]]:
    out = _empty_year_values(years)
    for y in out.keys():
        vals = [x.get(y) for x in items if x.get(y) is not None]
        out[y] = sum(vals) if vals else None
    return out


def _compat_three_periods(years: List[str]) -> tuple:
    ys = [str(y) for y in (years or []) if str(y).strip()]
    if not ys:
        return "", "", ""
    if len(ys) == 1:
        return ys[0], ys[0], ys[0]
    if len(ys) == 2:
        return ys[0], ys[1], ys[1]
    return ys[0], ys[1], ys[2]


def _year_seq(years: List[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for y in (years or []):
        ys = str(y).strip()
        if not ys or ys in seen:
            continue
        seen.add(ys)
        out.append(ys)
    return out[:3]


def _fmt_income_auto_text(
    label: str,
    values: Dict[str, Optional[float]],
    years: List[str],
    stable_pct: float = 2.0,
    trend_labels: Optional[Dict[str, str]] = None,
    text_templates: Optional[Dict[str, str]] = None,
    unit: str = DEFAULT_AMOUNT_UNIT,
) -> str:
    if len(years) < 2:
        return ""
    y1, y2, y3 = _compat_three_periods(years)
    labels = trend_labels or {}
    seq_years = _year_seq(years)
    value_parts: List[str] = []
    trend_parts: List[str] = []
    for y in seq_years:
        v = values.get(y)
        vt = f"{v:.2f}" if isinstance(v, (int, float)) and v is not None else (str(v) if v is not None else "待补充")
        value_parts.append(f"{y}年{vt}{unit or ''}")
    for i in range(1, len(seq_years)):
        yp = seq_years[i - 1]
        yc = seq_years[i]
        t = trend_text(
            values.get(yc),
            values.get(yp),
            stable_pct=stable_pct,
            up_label=str(labels.get("up_label", "增加")),
            down_label=str(labels.get("down_label", "减少")),
            stable_label=str(labels.get("stable_label", "保持稳定")),
        )
        trend_parts.append(f"{yc}较{yp}{t}")
    if len(seq_years) >= 3:
        y_first = seq_years[0]
        y_last = seq_years[-1]
        tfl = trend_text(
            values.get(y_last),
            values.get(y_first),
            stable_pct=stable_pct,
            up_label=str(labels.get("up_label", "增加")),
            down_label=str(labels.get("down_label", "减少")),
            stable_label=str(labels.get("stable_label", "保持稳定")),
        )
        trend_parts.append(f"{y_last}较{y_first}{tfl}")
    t21 = trend_parts[0].replace(f"{seq_years[1]}较{seq_years[0]}", "", 1) if len(trend_parts) >= 1 and len(seq_years) >= 2 else ""
    t32 = trend_parts[1].replace(f"{seq_years[2]}较{seq_years[1]}", "", 1) if len(trend_parts) >= 2 and len(seq_years) >= 3 else ""
    tpl = (text_templates or {}).get("income_segment_value", DEFAULT_TEXT_TEMPLATES["income_segment_value"])
    ctx = {
        "label": label,
        "y1": y1,
        "y2": y2,
        "y3": y3,
        "v1": values.get(y1) if values.get(y1) is not None else "待补充",
        "v2": values.get(y2) if values.get(y2) is not None else "待补充",
        "v3": values.get(y3) if values.get(y3) is not None else "待补充",
        "t21": t21,
        "t32": t32,
        "series_values": "，".join(value_parts),
        "series_trends": "；".join(trend_parts),
        "y_first": seq_years[0] if seq_years else "",
        "y_last": seq_years[-1] if seq_years else "",
        "y_prev": seq_years[-2] if len(seq_years) >= 2 else (seq_years[-1] if seq_years else ""),
        "unit": unit or "",
    }
    if len(seq_years) == 2 and ("{y3}" in tpl or "{v3}" in tpl or "{t32}" in tpl):
        return f"{label}：{ctx['series_values']}；{ctx['series_trends']}。"
    return _render_template(tpl, ctx)


def _fmt_pct(v: Optional[float]) -> str:
    return f"{v:.2f}%" if v is not None else "待补充"


def _calc_ratio_pct(numer: Optional[float], denom: Optional[float]) -> Optional[float]:
    if numer is None or denom in (None, 0):
        return None
    return numer / denom * 100.0


def _trend_pp_text(
    curr: Optional[float],
    prev: Optional[float],
    stable_pp: float = 2.0,
    up_label: str = "上升",
    down_label: str = "下降",
    stable_label: str = "基本稳定",
) -> str:
    if curr is None or prev is None:
        return "待补充"
    delta = curr - prev
    if delta > stable_pp:
        return f"{up_label}{abs(delta):.2f}个百分点"
    if delta < -stable_pp:
        return f"{down_label}{abs(delta):.2f}个百分点"
    return f"{stable_label}（变动{abs(delta):.2f}个百分点）"


def _trend_word(
    curr: Optional[float],
    prev: Optional[float],
    stable_pp: float = 2.0,
    up_label: str = "上升",
    down_label: str = "下降",
    stable_label: str = "基本稳定",
) -> str:
    if curr is None or prev is None:
        return "待补充"
    delta = curr - prev
    if delta > stable_pp:
        return up_label
    if delta < -stable_pp:
        return down_label
    return stable_label


def _fmt_ratio_auto_text(
    ratio_label: str,
    numer_values: Dict[str, Optional[float]],
    denom_values: Dict[str, Optional[float]],
    years: List[str],
    stable_pp: float = 2.0,
    trend_labels: Optional[Dict[str, str]] = None,
    text_templates: Optional[Dict[str, str]] = None,
) -> str:
    if len(years) < 2:
        return ""
    y1, y2, y3 = _compat_three_periods(years)
    r1 = _calc_ratio_pct(numer_values.get(y1), denom_values.get(y1))
    r2 = _calc_ratio_pct(numer_values.get(y2), denom_values.get(y2))
    r3 = _calc_ratio_pct(numer_values.get(y3), denom_values.get(y3))
    labels = trend_labels or {}
    seq_years = _year_seq(years)
    ratio_map = {y: _calc_ratio_pct(numer_values.get(y), denom_values.get(y)) for y in seq_years}
    value_parts = [f"{y}年{_fmt_pct(ratio_map.get(y))}" for y in seq_years]
    trend_parts: List[str] = []
    for i in range(1, len(seq_years)):
        yp = seq_years[i - 1]
        yc = seq_years[i]
        t = _trend_pp_text(
            ratio_map.get(yc),
            ratio_map.get(yp),
            stable_pp=stable_pp,
            up_label=str(labels.get("up_label", "上升")),
            down_label=str(labels.get("down_label", "下降")),
            stable_label=str(labels.get("stable_label", "基本稳定")),
        )
        trend_parts.append(f"{yc}较{yp}{t}")
    if len(seq_years) >= 3:
        y_first = seq_years[0]
        y_last = seq_years[-1]
        tfl = _trend_pp_text(
            ratio_map.get(y_last),
            ratio_map.get(y_first),
            stable_pp=stable_pp,
            up_label=str(labels.get("up_label", "上升")),
            down_label=str(labels.get("down_label", "下降")),
            stable_label=str(labels.get("stable_label", "基本稳定")),
        )
        trend_parts.append(f"{y_last}较{y_first}{tfl}")
    t21 = trend_parts[0].replace(f"{seq_years[1]}较{seq_years[0]}", "", 1) if len(trend_parts) >= 1 and len(seq_years) >= 2 else ""
    t32 = trend_parts[1].replace(f"{seq_years[2]}较{seq_years[1]}", "", 1) if len(trend_parts) >= 2 and len(seq_years) >= 3 else ""
    tpl = (text_templates or {}).get("income_segment_share", DEFAULT_TEXT_TEMPLATES["income_segment_share"])
    ctx = {
        "ratio_label": ratio_label,
        "y1": y1,
        "y2": y2,
        "y3": y3,
        "r1": _fmt_pct(r1),
        "r2": _fmt_pct(r2),
        "r3": _fmt_pct(r3),
        "rt21": t21,
        "rt32": t32,
        "t21": t21,
        "t32": t32,
        "i1": _fmt_pct(r1),
        "i2": _fmt_pct(r2),
        "i3": _fmt_pct(r3),
        "it21": t21,
        "it32": t32,
        "n1": _fmt_pct(r1),
        "n2": _fmt_pct(r2),
        "n3": _fmt_pct(r3),
        "nt21": t21,
        "nt32": t32,
        "series_values": "，".join(value_parts),
        "series_trends": "；".join(trend_parts),
        "y_first": seq_years[0] if seq_years else "",
        "y_last": seq_years[-1] if seq_years else "",
        "y_prev": seq_years[-2] if len(seq_years) >= 2 else (seq_years[-1] if seq_years else ""),
    }
    if len(seq_years) == 2 and ("{y3}" in tpl or "{r3}" in tpl or "{rt32}" in tpl):
        return f"{ratio_label}：{ctx['series_values']}；{ctx['series_trends']}。"
    return _render_template(tpl, ctx)


def _to_bool_like(v: Any, default: bool = False) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return default
    s = str(v).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _append_text(base: str, addon: str) -> str:
    b = str(base or "").strip()
    a = str(addon or "").strip()
    if not b:
        return a
    if not a:
        return b
    return f"{b}\n{a}"


def _eval_income_formula(
    formula: str,
    by_node: Dict[str, Dict[str, Any]],
    years: List[str],
) -> Dict[str, Optional[float]]:
    expr = str(formula or "").strip()
    out = _empty_year_values(years)
    if not expr:
        return out
    m_sum = re.match(r"^SUM\((.+)\)$", expr, re.IGNORECASE)
    if m_sum:
        ids = [x.strip() for x in m_sum.group(1).split(",") if x.strip()]
        items = [by_node.get(i, {}).get("values", _empty_year_values(years)) for i in ids if i in by_node]
        return _sum_values(items, years)
    m_sub = re.match(r"^SUB\(([^,]+),([^)]+)\)$", expr, re.IGNORECASE)
    if m_sub:
        left_id, right_id = m_sub.group(1).strip(), m_sub.group(2).strip()
        lv = by_node.get(left_id, {}).get("values", _empty_year_values(years))
        rv = by_node.get(right_id, {}).get("values", _empty_year_values(years))
        for y in out.keys():
            lval = lv.get(y)
            rval = rv.get(y)
            if lval is None and rval is None:
                out[y] = None
            else:
                out[y] = (lval or 0.0) - (rval or 0.0)
        return out
    m_ref = re.match(r"^REF\((.+)\)$", expr, re.IGNORECASE)
    if m_ref:
        rid = m_ref.group(1).strip()
        return dict(by_node.get(rid, {}).get("values", _empty_year_values(years)))
    return out


def _pick_is_row(rows: List[Dict[str, Any]], code_candidates: List[str], name_keywords: List[str]) -> Optional[Dict[str, Any]]:
    by_code = {x["code"]: x for x in rows}
    for c in code_candidates:
        if c in by_code:
            return by_code[c]
    candidates: List[Dict[str, Any]] = []
    for r in rows:
        n = str(r.get("name", ""))
        if any(k in n for k in name_keywords):
            candidates.append(r)
    if candidates:
        # Prefer top-level item over "其中：..." sub-items when multiple names match.
        candidates.sort(key=lambda x: (1 if "其中" in str(x.get("name", "")) else 0, len(str(x.get("name", "")))))
        return candidates[0]
    return None


def _find_col_by_any(headers: Dict[int, str], needles: List[str]) -> Optional[int]:
    for c, h in headers.items():
        txt = str(h or "").strip()
        if not txt:
            continue
        if any(n in txt for n in needles):
            return c
    return None


def _detect_income_segments(wb, years: List[str]) -> List[Dict[str, Any]]:
    # Try to detect real segments from detail sheet; fallback to A/B/C placeholders.
    ws = get_sheet_by_loose_name(wb, ["明细_营业总收入", "明细_营业收入"])
    if ws is None or len(_year_seq(years)) < 2:
        return [
            {"segment_id": "SEG_A", "label": "A项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
            {"segment_id": "SEG_B", "label": "B项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
            {"segment_id": "SEG_C", "label": "C项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
        ]

    headers = {c: str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)}
    year_col = _find_col_by_any(headers, ["期间", "年度"])
    seg_col = _find_col_by_any(headers, ["分项收入/子项名称", "分项收入/子项收入", "分项名称", "子项名称"])
    rev_col = _find_col_by_any(headers, ["分项收入明细值", "收入明细值", "明细值"])
    cost_col = _find_col_by_any(headers, ["分项成本明细值", "成本明细值"])
    gross_col = _find_col_by_any(headers, ["分项毛利明细值", "毛利明细值"])

    # Backward compatibility fallback by historical column positions.
    if year_col is None:
        year_col = 4
    if seg_col is None:
        seg_col = 5
    if rev_col is None:
        rev_col = 7

    revenue_bucket: Dict[str, Dict[str, Optional[float]]] = {}
    cost_bucket: Dict[str, Dict[str, Optional[float]]] = {}
    gross_bucket: Dict[str, Dict[str, Optional[float]]] = {}
    yset = set(_year_seq(years))
    for r in range(2, ws.max_row + 1):
        year_text = str(ws.cell(r, year_col).value or "").strip()
        seg_name = str(ws.cell(r, seg_col).value or "").strip()
        rev_amount = normalize_num(ws.cell(r, rev_col).value)
        cost_amount = normalize_num(ws.cell(r, cost_col).value) if cost_col else None
        gross_amount = normalize_num(ws.cell(r, gross_col).value) if gross_col else None
        m = re.search(r"(20\d{2})", year_text)
        if not seg_name or not m:
            continue
        y = m.group(1)
        if y not in yset:
            continue
        r_obj = revenue_bucket.setdefault(seg_name, _empty_year_values(years))
        c_obj = cost_bucket.setdefault(seg_name, _empty_year_values(years))
        g_obj = gross_bucket.setdefault(seg_name, _empty_year_values(years))

        if rev_amount is not None:
            r_obj[y] = (r_obj.get(y) or 0.0) + rev_amount
        if cost_amount is not None:
            c_obj[y] = (c_obj.get(y) or 0.0) + cost_amount
        if gross_amount is not None:
            g_obj[y] = (g_obj.get(y) or 0.0) + gross_amount

    all_seg_names = set(revenue_bucket.keys()) | set(cost_bucket.keys()) | set(gross_bucket.keys())
    names = []
    for n in sorted(all_seg_names):
        rv = revenue_bucket.get(n, {})
        cv = cost_bucket.get(n, {})
        gv = gross_bucket.get(n, {})
        if any(v is not None for v in rv.values()) or any(v is not None for v in cv.values()) or any(
            v is not None for v in gv.values()
        ):
            names.append(n)
    if not names:
        return [
            {"segment_id": "SEG_A", "label": "A项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
            {"segment_id": "SEG_B", "label": "B项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
            {"segment_id": "SEG_C", "label": "C项目（占位）", "revenue_values": _empty_year_values(years), "gross_values": _empty_year_values(years)},
        ]

    out = []
    for i, name in enumerate(sorted(names), start=1):
        sid = f"SEG_{i}"
        rev_vals = revenue_bucket.get(name, _empty_year_values(years))
        cost_vals = cost_bucket.get(name, _empty_year_values(years))
        gross_vals = gross_bucket.get(name, _empty_year_values(years))
        # If gross detail is absent, derive by revenue - cost when possible.
        for y in _year_seq(years):
            if gross_vals.get(y) is None and rev_vals.get(y) is not None and cost_vals.get(y) is not None:
                gross_vals[y] = (rev_vals.get(y) or 0.0) - (cost_vals.get(y) or 0.0)
        out.append(
            {
                "segment_id": sid,
                "label": name,
                "revenue_values": rev_vals,
                "cost_values": cost_vals,
                "gross_values": gross_vals,
            }
        )
    return out


def build_income_analysis_map(wb, project_id: str) -> Dict[str, Any]:
    ws = get_sheet_by_loose_name(wb, ["利润表", "利润表 "])
    if ws is None:
        return {"sheet_title": None, "years": [], "tree": [], "nodes": []}

    is_group = next((g for g in SHEET_GROUPS if g.get("id") == "is"), None)
    is_data = read_sheet_rows(wb, is_group, project_id=project_id) if is_group else {"years": [], "rows": []}
    display_years = _year_seq((is_data.get("years", []) or parse_years_from_sheet(ws))[:3])
    if len(display_years) < 2:
        return {"sheet_title": ws.title, "years": display_years, "tree": [], "nodes": []}
    years = list(display_years)
    y1, y2, y3 = _compat_three_periods(years)
    income_rules = load_income_rules()
    yoy_stable_pct = float(income_rules.get("yoy_stable_pct", 2.0))
    share_stable_pp = float(income_rules.get("share_stable_pp", 2.0))
    income_text_templates = dict(income_rules.get("text_templates", {}))
    income_text_template_units = dict(income_rules.get("text_template_units", {}))
    sign_policies = list(income_rules.get("sign_policies", []))
    trend_cfg = dict(income_rules.get("trend_thresholds", {}))
    contrib_policy = dict(income_rules.get("contribution_policy", {}))
    store = load_store(project_id)
    entries = store.get("entries", {})
    pending_class_ids = {"2.1.2.4", "2.1.2.5", "2.1.2.6"}

    def _default_bucket_for_node(nid: str, node_obj: Optional[Dict[str, Any]] = None) -> str:
        node = node_obj or {}
        pid = str(node.get("parent_id", "")).strip()
        if pid == "2.1.2":
            return "recurring"
        if pid == "2.2":
            return "nonrecurring"
        if nid in pending_class_ids:
            return "nonrecurring"
        return "nonrecurring"

    def _normalize_bucket(v: Any, default_bucket: str = "nonrecurring") -> str:
        t = str(v or "").strip().lower()
        if t in {"recurring", "经常性", "regular"}:
            return "recurring"
        if t in {"nonrecurring", "non_recurring", "非经常性", "irregular"}:
            return "nonrecurring"
        return default_bucket

    def _metric_trend(metric_id: str, fallback_stable: float, fallback_unit: str) -> Dict[str, Any]:
        raw = trend_cfg.get(metric_id, {}) if isinstance(trend_cfg, dict) else {}
        return {
            "stable": float(raw.get("stable", fallback_stable)),
            "unit": str(raw.get("unit", fallback_unit)),
            "up_label": str(raw.get("up_label", "")),
            "down_label": str(raw.get("down_label", "")),
            "stable_label": str(raw.get("stable_label", "")),
        }

    def _tpl_unit(scene: str, default: str = DEFAULT_AMOUNT_UNIT) -> str:
        u = str(income_text_template_units.get(scene, "") or "").strip()
        return u or default

    is_rows: List[Dict[str, Any]] = list(is_data.get("rows", []) or [])

    segments = _detect_income_segments(wb, years)

    def make_node(node_id: str, parent_id: str, label: str, values: Dict[str, Optional[float]], source_code: str = "", source_name: str = "") -> Dict[str, Any]:
        metric = "gp_value" if parent_id == "2.1.1" else "rev_yoy"
        mcfg = _metric_trend(metric, fallback_stable=yoy_stable_pct, fallback_unit="ratio")
        auto_text = _fmt_income_auto_text(
            label,
            values,
            years,
            stable_pct=float(mcfg["stable"]),
            trend_labels={
                "up_label": mcfg["up_label"] or "增加",
                "down_label": mcfg["down_label"] or "减少",
                "stable_label": mcfg["stable_label"] or "保持稳定",
            },
            text_templates=income_text_templates,
            unit=_tpl_unit("income_segment_value", DEFAULT_AMOUNT_UNIT),
        )
        return {
            "node_id": node_id,
            "parent_id": parent_id,
            "label": label,
            "source_code": source_code,
            "source_name": source_name,
            "values": values,
            "auto_text": auto_text,
        }

    nodes: List[Dict[str, Any]] = []
    tree: List[Dict[str, str]] = []
    tree_rules = income_rules.get("tree_nodes", DEFAULT_INCOME_TREE)
    fixed_tree = [x for x in tree_rules if str(x.get("node_type", "fixed")).strip().lower() == "fixed"]
    template_tree = [x for x in tree_rules if str(x.get("node_type", "")).strip().lower() == "template"]
    fixed_label_map = {str(x.get("node_id", "")).strip(): str(x.get("label", "")).strip() for x in fixed_tree}
    fixed_parent_map = {str(x.get("node_id", "")).strip(): str(x.get("parent_id", "")).strip() for x in fixed_tree}

    def add_tree(node_id: str, parent_id: str, label: str) -> None:
        tree.append({"node_id": node_id, "parent_id": parent_id, "label": label})

    rev_tpl = next((x for x in template_tree if str(x.get("template_name", "")).strip() == "revenue_segment"), None)
    rev_prefix = str((rev_tpl or {}).get("node_id_prefix", "1.")).strip() or "1."
    rev_parent = str((rev_tpl or {}).get("parent_id", "1")).strip() or "1"
    rev_start = int((rev_tpl or {}).get("start_index", 2) or 2)
    rev_suffix = str((rev_tpl or {}).get("label_suffix", "收入")).strip() or "收入"

    gp_tpl = next((x for x in template_tree if str(x.get("template_name", "")).strip() == "gross_segment"), None)
    gp_prefix = str((gp_tpl or {}).get("node_id_prefix", "2.1.1.")).strip() or "2.1.1."
    gp_parent = str((gp_tpl or {}).get("parent_id", "2.1.1")).strip() or "2.1.1"
    gp_start = int((gp_tpl or {}).get("start_index", 1) or 1)
    gp_suffix = str((gp_tpl or {}).get("label_suffix", "毛利")).strip() or "毛利"

    # Build left tree strictly in configured order; template nodes are expanded in place.
    for rule in tree_rules:
        ntype = str(rule.get("node_type", "fixed")).strip().lower()
        if ntype == "fixed":
            nid = str(rule.get("node_id", "")).strip()
            label = str(rule.get("label", "")).strip()
            if nid and label:
                add_tree(nid, str(rule.get("parent_id", "")).strip(), label)
            continue
        if ntype != "template":
            continue
        tname = str(rule.get("template_name", "")).strip()
        if tname == "revenue_segment":
            for i, seg in enumerate(segments, start=0):
                nid = f"{rev_prefix}{rev_start + i}"
                label = f"{seg['label']}{rev_suffix}"
                add_tree(nid, rev_parent, label)
        elif tname == "gross_segment":
            for i, seg in enumerate(segments, start=0):
                nid = f"{gp_prefix}{gp_start + i}"
                label = f"{seg['label']}{gp_suffix}"
                add_tree(nid, gp_parent, label)

    # 1 营业收入分析
    main_rev = _pick_is_row(is_rows, ["IS001"], ["营业总收入", "营业收入", "主营业务收入"])
    main_rev_vals = main_rev["values"] if main_rev else _empty_year_values(years)
    if "1.1" in fixed_label_map:
        nodes.append(
            make_node(
                "1.1",
                fixed_parent_map.get("1.1", "1"),
                fixed_label_map.get("1.1", "营业收入总额"),
                main_rev_vals,
                main_rev["code"] if main_rev else "",
                main_rev["name"] if main_rev else "",
            )
        )

    for i, seg in enumerate(segments, start=0):
        nid = f"{rev_prefix}{rev_start + i}"
        label = f"{seg['label']}{rev_suffix}"
        nodes.append(make_node(nid, rev_parent, label, seg["revenue_values"], seg["segment_id"], seg["label"]))

    gross_vals = _sum_values([x["gross_values"] for x in segments], years)
    if "2.1.1" in fixed_label_map:
        nodes.append(
            make_node(
                "2.1.1",
                fixed_parent_map.get("2.1.1", "2.1"),
                fixed_label_map.get("2.1.1", "主营业务毛利（自动汇总）"),
                gross_vals,
            )
        )

    for i, seg in enumerate(segments, start=0):
        nid = f"{gp_prefix}{gp_start + i}"
        label = f"{seg['label']}{gp_suffix}"
        nodes.append(make_node(nid, gp_parent, label, seg["gross_values"], seg["segment_id"], seg["label"]))
    # For non-main revenue items, use name-based mapping first to avoid code-shift issues
    # across different templates/projects.
    other_cfg = income_rules.get("special_items", DEFAULT_INCOME_SPECIAL_ITEMS)
    for item in other_cfg:
        nid = str(item.get("node_id", "")).strip()
        label = str(item.get("label", "")).strip()
        codes = [str(x).strip() for x in item.get("code_candidates", []) if str(x).strip()]
        kws = [str(x).strip() for x in item.get("name_keywords", []) if str(x).strip()]
        if not nid or not label:
            continue
        parent_for_node = fixed_parent_map.get(nid, "2.1.2")
        if nid == "2.1.2.4":
            nodes.append(make_node(nid, parent_for_node, label, _empty_year_values(years)))
            continue
        p = _pick_is_row(is_rows, codes, kws)
        vals = p["values"] if p else _empty_year_values(years)
        nodes.append(make_node(nid, parent_for_node, label, vals, p["code"] if p else "", p["name"] if p else ""))

    p_asset = _pick_is_row(is_rows, [], ["资产处置收益"])
    v_asset = p_asset["values"] if p_asset else _empty_year_values(years)
    if "2.2.1" in fixed_label_map:
        nodes.append(
            make_node(
                "2.2.1",
                fixed_parent_map.get("2.2.1", "2.2"),
                fixed_label_map.get("2.2.1", "资产处置收益"),
                v_asset,
                p_asset["code"] if p_asset else "",
                p_asset["name"] if p_asset else "",
            )
        )

    p_noi = _pick_is_row(is_rows, [], ["营业外收入"])
    p_noe = _pick_is_row(is_rows, [], ["营业外支出"])
    v_noi = p_noi["values"] if p_noi else _empty_year_values(years)
    v_noe = p_noe["values"] if p_noe else _empty_year_values(years)
    v_net = {y: ((v_noi.get(y) or 0.0) - (v_noe.get(y) or 0.0)) if (v_noi.get(y) is not None or v_noe.get(y) is not None) else None for y in years}
    if "2.2.2" in fixed_label_map:
        nodes.append(
            make_node(
                "2.2.2",
                fixed_parent_map.get("2.2.2", "2.2"),
                fixed_label_map.get("2.2.2", "营业外收支净额（自动汇总）"),
                v_net,
            )
        )

    # 3 分析输出（公式驱动；缺失时回退默认分组逻辑）
    by_node = {n["node_id"]: n for n in nodes}
    formula_rules = income_rules.get("formula_rules", []) or []
    formula_map = {str(x.get("node_id", "")).strip(): str(x.get("formula", "")).strip() for x in formula_rules}
    recurring_keys = set(income_rules.get("recurring_keys", DEFAULT_RECURRING_KEYS))
    nonrec_keys = set(income_rules.get("nonrec_keys", DEFAULT_NONREC_KEYS))
    editable_ids: set = set()
    for nid, n in by_node.items():
        label = str(n.get("label", ""))
        pid = str(n.get("parent_id", "")).strip()
        if nid in pending_class_ids or ("待判定" in label) or pid in {"2.1.2", "2.2"}:
            editable_ids.add(nid)
    for nid in editable_ids:
        saved = entries.get(make_entry_key("income_analysis", nid), {})
        node = by_node.get(nid)
        default_bucket = _default_bucket_for_node(nid, node)
        bucket = _normalize_bucket(
            saved.get("classification_bucket") or saved.get("profit_bucket") or saved.get("class_bucket"),
            default_bucket=default_bucket,
        )
        recurring_keys.discard(nid)
        nonrec_keys.discard(nid)
        if bucket == "recurring":
            recurring_keys.add(nid)
        else:
            nonrec_keys.add(nid)
        if node is not None:
            node["classification_editable"] = True
            node["classification_bucket"] = bucket
    if not formula_map:
        formula_map = {
            "3.1": f"SUM({','.join(sorted(recurring_keys))})",
            "3.2": f"SUM({','.join(sorted(nonrec_keys))})",
            "3.3": "SUB(3.1,3.2)",
        }
    elif editable_ids:
        # Keep rule formulas for other nodes, but 3.1/3.2 follow per-project manual classification.
        formula_map["3.1"] = f"SUM({','.join(sorted(recurring_keys))})"
        formula_map["3.2"] = f"SUM({','.join(sorted(nonrec_keys))})"
        formula_map.setdefault("3.3", "SUB(3.1,3.2)")

    output_order = [nid for nid in [x.get("node_id", "") for x in fixed_tree] if str(nid).startswith("3.")]
    if not output_order:
        output_order = ["3.1", "3.2", "3.3"]
    for nid in output_order:
        nid = str(nid).strip()
        if not nid or nid not in fixed_label_map:
            continue
        vals = _eval_income_formula(formula_map.get(nid, ""), by_node, years)
        node = make_node(
            nid,
            fixed_parent_map.get(nid, "3"),
            fixed_label_map.get(nid, nid),
            vals,
        )
        nodes.append(node)
        by_node[nid] = node

    # Enrich segment-level auto text with structure ratio and trend:
    # - Revenue segment share in total revenue
    # - Segment gross contribution share in total gross
    by_node = {n["node_id"]: n for n in nodes}
    total_revenue_values = by_node.get("1.1", {}).get("values", _empty_year_values(years))
    total_gross_values = by_node.get("2.1.1", {}).get("values", _empty_year_values(years))
    gp_nodes = [x for x in nodes if x.get("parent_id") == "2.1.1"]
    gp_abs_sum_values: Dict[str, Optional[float]] = _empty_year_values(years)
    gp_has_pos: Dict[str, bool] = {y: False for y in years}
    gp_has_neg: Dict[str, bool] = {y: False for y in years}
    for y in years:
        vals = []
        for x in gp_nodes:
            v = (x.get("values", {}) if isinstance(x.get("values"), dict) else {}).get(y)
            if v is None:
                continue
            vals.append(v)
            if v > 0:
                gp_has_pos[y] = True
            if v < 0:
                gp_has_neg[y] = True
        gp_abs_sum_values[y] = sum(abs(v) for v in vals) if vals else None

    gp_mode = str(contrib_policy.get("gp_share_mode", "impact_when_mixed_or_total_negative")).strip() or "impact_when_mixed_or_total_negative"
    gp_label_impact = str(contrib_policy.get("impact_ratio_label", "毛利贡献（驱动占比）")).strip() or "毛利贡献（驱动占比）"
    gp_label_algebraic = str(contrib_policy.get("algebraic_ratio_label", "毛利贡献（占总毛利比例）")).strip() or "毛利贡献（占总毛利比例）"
    gp_basis_impact = str(contrib_policy.get("basis_impact_text", "按绝对值口径")).strip() or "按绝对值口径"
    gp_basis_algebraic = str(contrib_policy.get("basis_algebraic_text", "按代数口径")).strip() or "按代数口径"
    net_attr_enabled = _to_bool_like(contrib_policy.get("net_attr_enabled", "0"), default=False)
    net_attr_label = (
        str(contrib_policy.get("net_attr_ratio_label", "对全部分项业务毛利影响总量的占比")).strip()
        or "对全部分项业务毛利影响总量的占比"
    )
    strength_ratio_label = str(contrib_policy.get("strength_ratio_label", "驱动占比")).strip() or "驱动占比"
    display_dual_caliber = _to_bool_like(contrib_policy.get("display_dual_caliber", "0"), default=False)
    zero_total_policy = str(contrib_policy.get("zero_total_policy", "strength_only")).strip().lower() or "strength_only"

    def _use_impact_ratio_fallback(year: str, total_gp: Optional[float]) -> bool:
        mixed = bool(gp_has_pos.get(year, False) and gp_has_neg.get(year, False))
        if gp_mode == "impact_always":
            return True
        if gp_mode == "algebraic_always":
            return False
        # default: impact when mixed sign or total gross <= 0
        return mixed or (total_gp is not None and total_gp <= 0)
    for n in nodes:
        if n.get("parent_id") == "1" and n.get("node_id") != "1.1":
            scfg = _metric_trend("rev_share", fallback_stable=share_stable_pp, fallback_unit="pct_point")
            ratio_txt = _fmt_ratio_auto_text(
                "占营业收入总额比例",
                n.get("values", {}),
                total_revenue_values,
                years,
                stable_pp=float(scfg["stable"]),
                trend_labels={
                    "up_label": scfg["up_label"] or "上升",
                    "down_label": scfg["down_label"] or "下降",
                    "stable_label": scfg["stable_label"] or "基本稳定",
                },
                text_templates=income_text_templates,
            )
            n["auto_text"] = _append_text(n.get("auto_text", ""), ratio_txt)
        if n.get("parent_id") == "2.1.1":
            gcfg = _metric_trend("gp_share_impact", fallback_stable=share_stable_pp, fallback_unit="pct_point")
            numer_values: Dict[str, Optional[float]] = _empty_year_values(years)
            denom_values: Dict[str, Optional[float]] = _empty_year_values(years)
            basis_by_year: Dict[str, str] = {y: "待补充" for y in years}
            for y in years:
                seg_gp = (n.get("values", {}) if isinstance(n.get("values"), dict) else {}).get(y)
                total_gp = total_gross_values.get(y)
                policy_y = _pick_income_sign_policy(sign_policies, total_gp, gp_has_pos.get(y, False), gp_has_neg.get(y, False))
                metric_y = str(policy_y.get("primary_ratio_metric", "")).strip().lower()
                if metric_y == "gp_share_impact":
                    use_impact = True
                elif metric_y == "gp_share_algebraic":
                    use_impact = False
                else:
                    use_impact = _use_impact_ratio_fallback(y, total_gp)
                if use_impact:
                    numer_values[y] = abs(seg_gp) if seg_gp is not None else None
                    denom_values[y] = gp_abs_sum_values.get(y)
                    basis_by_year[y] = gp_basis_impact
                else:
                    numer_values[y] = seg_gp
                    denom_values[y] = total_gp
                    basis_by_year[y] = gp_basis_algebraic
            latest_basis = basis_by_year.get(years[-1], gp_basis_impact)
            ratio_label = gp_label_impact if latest_basis == gp_basis_impact else gp_label_algebraic
            ratio_txt = _fmt_ratio_auto_text(
                ratio_label,
                numer_values,
                denom_values,
                years,
                stable_pp=float(gcfg["stable"]),
                trend_labels={
                    "up_label": gcfg["up_label"] or "上升",
                    "down_label": gcfg["down_label"] or "下降",
                    "stable_label": gcfg["stable_label"] or "基本稳定",
                },
                text_templates=income_text_templates,
            )
            n["auto_text"] = _append_text(n.get("auto_text", ""), ratio_txt)
            basis_tpl = income_text_templates.get("gross_contribution_basis", DEFAULT_TEXT_TEMPLATES["gross_contribution_basis"])
            if len(years) >= 3:
                basis_txt = _render_template(
                    basis_tpl,
                    {
                        "y1": y1,
                        "y2": y2,
                        "y3": y3,
                        "b1": basis_by_year.get(y1, "待补充"),
                        "b2": basis_by_year.get(y2, "待补充"),
                        "b3": basis_by_year.get(y3, "待补充"),
                    },
                )
            else:
                basis_txt = f"口径说明：{y1}年{basis_by_year.get(y1, '待补充')}，{y2}年{basis_by_year.get(y2, '待补充')}。"
            n["auto_text"] = _append_text(n.get("auto_text", ""), basis_txt)
            if net_attr_enabled:
                abs_total_values: Dict[str, Optional[float]] = _empty_year_values(years)
                for y in years:
                    tg = total_gross_values.get(y)
                    abs_total_values[y] = abs(tg) if tg is not None and abs(tg) > 1e-12 else None
                if zero_total_policy != "strength_only" or any(v is not None for v in abs_total_values.values()):
                    net_ratio_txt = _fmt_ratio_auto_text(
                        net_attr_label,
                        n.get("values", {}),
                        abs_total_values,
                        years,
                        stable_pp=float(gcfg["stable"]),
                        trend_labels={
                            "up_label": gcfg["up_label"] or "上升",
                            "down_label": gcfg["down_label"] or "下降",
                            "stable_label": gcfg["stable_label"] or "基本稳定",
                        },
                        text_templates={
                            "income_segment_share": income_text_templates.get(
                                "gross_segment_net_attr", DEFAULT_TEXT_TEMPLATES["gross_segment_net_attr"]
                            )
                        },
                    )
                    n["auto_text"] = _append_text(n.get("auto_text", ""), net_ratio_txt)
                    if display_dual_caliber and len(years) >= 2:
                        latest_y = years[-1]
                        seg_gp_latest = (n.get("values", {}) if isinstance(n.get("values"), dict) else {}).get(latest_y)
                        total_gp_latest = total_gross_values.get(latest_y)
                        strength_latest = _calc_ratio_pct(numer_values.get(latest_y), denom_values.get(latest_y))
                        net_latest = (
                            seg_gp_latest / abs(total_gp_latest) * 100.0
                            if seg_gp_latest is not None and total_gp_latest is not None and abs(total_gp_latest) > 1e-12
                            else None
                        )
                        dual_tpl = income_text_templates.get(
                            "gross_segment_dual_view", DEFAULT_TEXT_TEMPLATES["gross_segment_dual_view"]
                        )
                        dual_txt = _render_template(
                            dual_tpl,
                            {
                                "strength_label": strength_ratio_label,
                                "s_latest": _fmt_pct(strength_latest),
                                "net_label": net_attr_label,
                                "n_latest": _fmt_pct(net_latest),
                            },
                        )
                        n["auto_text"] = _append_text(n.get("auto_text", ""), dual_txt)
            # Scenario-based qualitative labels from sign_scenario_policy.
            latest_y = years[-1]
            latest_total_gp = total_gross_values.get(latest_y)
            latest_seg_gps = [((x.get("values", {}) if isinstance(x.get("values"), dict) else {}).get(latest_y)) for x in nodes if x.get("parent_id") == "2.1.1"]
            has_pos = any((v is not None and v > 0) for v in latest_seg_gps)
            has_neg = any((v is not None and v < 0) for v in latest_seg_gps)
            p = _pick_income_sign_policy(sign_policies, latest_total_gp, has_pos, has_neg)
            seg_gp = (n.get("values", {}) if isinstance(n.get("values"), dict) else {}).get(latest_y)
            if p:
                pos_word = str(p.get("direction_label_pos", "")).strip() or "正向项"
                neg_word = str(p.get("direction_label_neg", "")).strip() or "负向项"
                if seg_gp is None:
                    qual_word = "待补充"
                elif seg_gp > 0:
                    qual_word = pos_word
                elif seg_gp < 0:
                    qual_word = neg_word
                else:
                    qual_word = "中性项"
                scn = str(p.get("scenario", "")).strip()
                scn_tpl = income_text_templates.get(
                    "gross_segment_scenario_judgement", DEFAULT_TEXT_TEMPLATES["gross_segment_scenario_judgement"]
                )
                qual_txt = _render_template(
                    scn_tpl,
                    {"latest_year": latest_y, "qual_word": qual_word, "scenario": scn},
                )
                n["auto_text"] = _append_text(n.get("auto_text", ""), qual_txt)

    # Keep segment-level gross-profit nodes concise: only absolute value + trend.
    gp_value_cfg = _metric_trend("gp_value", fallback_stable=yoy_stable_pct, fallback_unit="ratio")
    for n in nodes:
        if n.get("parent_id") != "2.1.1":
            continue
        n["auto_text"] = _fmt_income_auto_text(
            str(n.get("label", "")),
            n.get("values", {}) if isinstance(n.get("values"), dict) else _empty_year_values(years),
            years,
            stable_pct=float(gp_value_cfg["stable"]),
            trend_labels={
                "up_label": gp_value_cfg["up_label"] or "增加",
                "down_label": gp_value_cfg["down_label"] or "减少",
                "stable_label": gp_value_cfg["stable_label"] or "保持稳定",
            },
            text_templates=income_text_templates,
            unit=_tpl_unit("income_segment_value", DEFAULT_AMOUNT_UNIT),
        )

    # Contribution analysis is shown on summary nodes only (e.g., 2.1.1 / 3.1).
    by_node = {n["node_id"]: n for n in nodes}

    def _append_summary_contrib(summary_id: str, child_ids: List[str], title: str) -> None:
        sn = by_node.get(summary_id)
        if not sn or len(years) < 2:
            return
        latest_y = years[-1]
        total_latest = (sn.get("values", {}) if isinstance(sn.get("values"), dict) else {}).get(latest_y)
        childs = [by_node[cid] for cid in child_ids if cid in by_node]
        if not childs:
            return
        seg_vals = [((x.get("values", {}) if isinstance(x.get("values"), dict) else {}).get(latest_y)) for x in childs]
        has_pos = any(v is not None and v > 0 for v in seg_vals)
        has_neg = any(v is not None and v < 0 for v in seg_vals)
        sum_abs = sum(abs(v) for v in seg_vals if v is not None)
        p = _pick_income_sign_policy(sign_policies, total_latest, has_pos, has_neg)
        pos_word = str((p or {}).get("direction_label_pos", "")).strip() or "正向项"
        neg_word = str((p or {}).get("direction_label_neg", "")).strip() or "负向项"
        lines: List[str] = []
        def _fmt_amount(v: Optional[float]) -> str:
            return f"{v:.2f}" if v is not None else "待补充"
        def _fmt_signed_pct(v: Optional[float]) -> str:
            if v is None:
                return "待补充"
            return f"{v:+.2f}%"
        hdr_missing_tpl = income_text_templates.get(
            "profit_summary_header_missing", DEFAULT_TEXT_TEMPLATES["profit_summary_header_missing"]
        )
        hdr_pos_tpl = income_text_templates.get(
            "profit_summary_header_positive", DEFAULT_TEXT_TEMPLATES["profit_summary_header_positive"]
        )
        hdr_zero_tpl = income_text_templates.get(
            "profit_summary_header_zero", DEFAULT_TEXT_TEMPLATES["profit_summary_header_zero"]
        )
        hdr_neg_tpl = income_text_templates.get(
            "profit_summary_header_negative", DEFAULT_TEXT_TEMPLATES["profit_summary_header_negative"]
        )
        line_pos_tpl = income_text_templates.get(
            "profit_summary_line_positive", DEFAULT_TEXT_TEMPLATES["profit_summary_line_positive"]
        )
        line_neg_tpl = income_text_templates.get(
            "profit_summary_line_negative", DEFAULT_TEXT_TEMPLATES["profit_summary_line_negative"]
        )
        line_zero_tpl = income_text_templates.get(
            "profit_summary_line_zero", DEFAULT_TEXT_TEMPLATES["profit_summary_line_zero"]
        )
        if total_latest is None:
            header = _render_template(
                hdr_missing_tpl,
                {"title": title, "latest_year": latest_y, "unit": _tpl_unit("profit_summary_header_missing", DEFAULT_AMOUNT_UNIT)},
            )
        elif total_latest > 0:
            header = _render_template(
                hdr_pos_tpl,
                {
                    "title": title,
                    "latest_year": latest_y,
                    "total_amount": _fmt_amount(total_latest),
                    "unit": _tpl_unit("profit_summary_header_positive", DEFAULT_AMOUNT_UNIT),
                },
            )
        elif abs(total_latest) < 1e-12:
            header = _render_template(
                hdr_zero_tpl,
                {"title": title, "latest_year": latest_y, "unit": _tpl_unit("profit_summary_header_zero", DEFAULT_AMOUNT_UNIT)},
            )
        else:
            header = _render_template(
                hdr_neg_tpl,
                {
                    "title": title,
                    "latest_year": latest_y,
                    "total_abs_amount": _fmt_amount(abs(total_latest)),
                    "unit": _tpl_unit("profit_summary_header_negative", DEFAULT_AMOUNT_UNIT),
                },
            )
        for c in childs:
            name = str(c.get("label", "")).strip()
            v = (c.get("values", {}) if isinstance(c.get("values"), dict) else {}).get(latest_y)
            if v is None:
                continue
            net_attr = (v / sum_abs * 100.0) if sum_abs > 1e-12 else None
            if v > 0:
                lines.append(
                    _render_template(
                        line_pos_tpl,
                        {
                            "name": name,
                            "amount": _fmt_amount(v),
                            "net_attr_label": net_attr_label,
                            "net_attr_pct": _fmt_signed_pct(net_attr),
                            "direction_word": pos_word,
                            "unit": _tpl_unit("profit_summary_line_positive", DEFAULT_AMOUNT_UNIT),
                        },
                    )
                )
            elif v < 0:
                lines.append(
                    _render_template(
                        line_neg_tpl,
                        {
                            "name": name,
                            "abs_amount": _fmt_amount(abs(v)),
                            "net_attr_label": net_attr_label,
                            "net_attr_pct": _fmt_signed_pct(net_attr),
                            "direction_word": neg_word,
                            "unit": _tpl_unit("profit_summary_line_negative", DEFAULT_AMOUNT_UNIT),
                        },
                    )
                )
            else:
                lines.append(
                    _render_template(
                        line_zero_tpl,
                        {
                            "name": name,
                            "net_attr_label": net_attr_label,
                            "net_attr_pct": _fmt_signed_pct(net_attr),
                        },
                    )
                )
        if not lines:
            return
        summary_text = header + "\n" + "\n".join(lines)
        sn["auto_text"] = _append_text(sn.get("auto_text", ""), summary_text)

    gp_child_ids = [str(x.get("node_id", "")) for x in nodes if x.get("parent_id") == "2.1.1"]
    _append_summary_contrib("2.1.1", gp_child_ids, "子项贡献")

    recurring_keys = set(income_rules.get("recurring_keys", DEFAULT_RECURRING_KEYS))
    recurring_child_ids = [nid for nid in recurring_keys if nid in by_node and nid != "3.1"]
    _append_summary_contrib("3.1", sorted(recurring_child_ids), "经常性收益构成贡献")

    # Merge manual store
    for n in nodes:
        saved = entries.get(make_entry_key("income_analysis", n["node_id"]), {})
        manual = str(saved.get("manual_text", "") or "")
        confirmed = bool(saved.get("confirmed", False))
        n["manual_text"] = manual
        n["confirmed"] = confirmed
        if bool(n.get("classification_editable", False)):
            default_bucket = _default_bucket_for_node(str(n.get("node_id", "")).strip(), n)
            n["classification_bucket"] = _normalize_bucket(
                saved.get("classification_bucket") or saved.get("profit_bucket") or saved.get("class_bucket"),
                default_bucket=default_bucket,
            )
        else:
            n["classification_editable"] = False
            n["classification_bucket"] = ""
        n["final_text"] = manual if manual.strip() else n["auto_text"]

    return {
        "sheet_title": ws.title,
        "years": display_years,
        "tree": tree,
        "nodes": nodes,
    }


def read_generic_sheet(ws) -> Dict[str, Any]:
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in values):
            continue
        obj: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            key = h or f"col_{i+1}"
            obj[key] = values[i]
        rows.append(obj)
    return {"sheet_title": ws.title, "headers": headers, "rows": rows}


def read_analysis_data(wb, sheet_map: Dict[str, List[str]]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, candidates in sheet_map.items():
        ws = get_sheet_by_loose_name(wb, candidates)
        if ws is None:
            out[k] = {"sheet_title": None, "headers": [], "rows": []}
            continue
        out[k] = read_generic_sheet(ws)
    return out


def build_analysis_map(wb, project_id: str, sheet_map: Dict[str, List[str]], group_id: str) -> Dict[str, Dict[str, Any]]:
    data = read_analysis_data(wb, sheet_map)
    data = apply_analysis_code_redirects(group_id, data)
    scale_rows = data.get("scale", {}).get("rows", []) or []
    struct_rows = data.get("structure", {}).get("rows", []) or []

    scale_by_code = {str(r.get("科目编码", "")).strip(): r for r in scale_rows}
    struct_by_code = {str(r.get("科目编码", "")).strip(): r for r in struct_rows}

    store = load_store(project_id)
    entries = store.get("entries", {})
    out: Dict[str, Dict[str, Any]] = {}
    th_cfg = _read_threshold_config()
    stable_scale = float(th_cfg.get("global_scale_pct", th_cfg.get("global_pct", 2.0)))
    stable_struct = float(th_cfg.get("global_struct_pp", th_cfg.get("global_pct", 2.0)))

    all_codes = set(scale_by_code.keys()) | set(struct_by_code.keys())
    for code in all_codes:
        if not code:
            continue
        scale_row = scale_by_code.get(code, {})
        struct_row = struct_by_code.get(code, {})
        raw_name = str(scale_row.get("科目名称", "") or struct_row.get("科目名称", ""))
        clean_name = raw_name.strip().rstrip("：:").strip()

        def _clean_subject_punct(text: str) -> str:
            t = str(text or "")
            if not t:
                return t
            # Remove stale punctuation artifacts from legacy generated sheets, e.g. "流动资产：，".
            t = t.replace("：，", "，").replace(":,", "，")
            if raw_name and clean_name and raw_name != clean_name:
                t = t.replace(f"{raw_name}，", f"{clean_name}，")
                t = t.replace(f"{raw_name},", f"{clean_name},")
                # Guard case: "名称："
                t = t.replace(f"{raw_name}", clean_name)
            return t

        auto_abs = _clean_subject_punct(str(scale_row.get("定量描述_绝对量", "") or ""))
        auto_rel = _clean_subject_punct(str(struct_row.get("定量描述_相对量", "") or ""))

        # Append first-last comparison (A vs C) for 3-year scenarios.
        year_keys = sorted([k for k in set(list(scale_row.keys()) + list(struct_row.keys())) if re.match(r"^20\d{2}$", str(k))])
        if len(year_keys) >= 3:
            y_first, y_last = year_keys[0], year_keys[-1]
            v_first = normalize_num(scale_row.get(y_first))
            v_last = normalize_num(scale_row.get(y_last))
            if v_first not in (None, 0) and v_last is not None:
                pct = (v_last - v_first) / abs(v_first) * 100.0
                if pct > stable_scale:
                    t_abs = f"{y_last}较{y_first}增加{abs(pct):.2f}%。"
                elif pct < -stable_scale:
                    t_abs = f"{y_last}较{y_first}减少{abs(pct):.2f}%。"
                else:
                    t_abs = f"{y_last}较{y_first}基本稳定（变动{abs(pct):.2f}%）。"
                if t_abs and t_abs not in auto_abs:
                    auto_abs = (auto_abs + ("\n" if auto_abs else "") + t_abs).strip()

            r_first = normalize_num(struct_row.get(y_first))
            r_last = normalize_num(struct_row.get(y_last))
            if r_first is not None and r_last is not None:
                d = r_last - r_first
                if d > stable_struct:
                    t_rel = f"{y_last}较{y_first}上升{abs(d):.2f}个百分点。"
                elif d < -stable_struct:
                    t_rel = f"{y_last}较{y_first}下降{abs(d):.2f}个百分点。"
                else:
                    t_rel = f"{y_last}较{y_first}基本稳定（变动{abs(d):.2f}个百分点）。"
                if t_rel and t_rel not in auto_rel:
                    auto_rel = (auto_rel + ("\n" if auto_rel else "") + t_rel).strip()

        auto_combined = auto_abs if not auto_rel else f"{auto_abs}\n{auto_rel}"

        saved = entries.get(make_entry_key(group_id, code), {})
        manual_text = str(saved.get("manual_text", "") or "")
        confirmed = bool(saved.get("confirmed", False))
        # BS主页面只展示“分析页可编辑框”的最终文本：
        # 有手工保存内容时直接展示该内容；否则回退到自动文本。
        final_text = manual_text if manual_text.strip() else auto_combined

        out[code] = {
            "code": code,
            "name": clean_name,
            "auto_abs": auto_abs,
            "auto_rel": auto_rel,
            "auto_combined": auto_combined,
            "manual_text": manual_text,
            "confirmed": confirmed,
            "final_text": final_text,
        }
    return out


def apply_analysis_code_redirects(group_id: str, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
    redirect_map = ANALYSIS_CODE_REDIRECTS.get(group_id, {})
    if not redirect_map:
        return analysis_data

    for section in ("scale", "structure"):
        part = analysis_data.get(section, {}) if isinstance(analysis_data, dict) else {}
        rows = part.get("rows", []) if isinstance(part, dict) else []
        if not isinstance(rows, list) or not rows:
            continue

        idx_by_code: Dict[str, int] = {}
        for i, row in enumerate(rows):
            if not isinstance(row, dict):
                continue
            code = str(row.get("科目编码", "")).strip()
            if code:
                idx_by_code[code] = i

        for to_code, from_code in redirect_map.items():
            i_from = idx_by_code.get(str(from_code).strip())
            if i_from is None:
                continue
            from_row = rows[i_from]
            to_code = str(to_code).strip()
            i_to = idx_by_code.get(to_code)
            to_row = rows[i_to] if i_to is not None else {}

            merged = dict(from_row)
            merged["科目编码"] = to_code
            # Keep destination subject name if present; fallback to source name.
            dst_name = str((to_row or {}).get("科目名称", "")).strip()
            if dst_name:
                merged["科目名称"] = dst_name

            if i_to is not None:
                rows[i_to] = merged
            else:
                rows.append(merged)
    return analysis_data


def build_asset_analysis_map(wb, project_id: str) -> Dict[str, Dict[str, Any]]:
    return build_analysis_map(wb, project_id, ASSET_ANALYSIS_SHEETS, "asset_analysis")


def build_liability_analysis_map(wb, project_id: str) -> Dict[str, Dict[str, Any]]:
    return build_analysis_map(wb, project_id, LIABILITY_ANALYSIS_SHEETS, "liability_analysis")


def build_summary_analysis_payload(wb, project_id: str) -> Dict[str, Any]:
    sum_tpl = load_main_analysis_text_templates()
    sum_tpl_units = load_main_analysis_template_units()
    bs_group = next((g for g in SHEET_GROUPS if g["id"] == "bs"), None)
    is_group = next((g for g in SHEET_GROUPS if g["id"] == "is"), None)
    bs_data = read_sheet_rows(wb, bs_group, project_id=project_id) if bs_group else {"years": [], "rows": []}
    is_data = read_sheet_rows(wb, is_group, project_id=project_id) if is_group else {"years": [], "rows": []}
    income_data = build_income_analysis_map(wb, project_id)

    years = (bs_data.get("years") or is_data.get("years") or income_data.get("years") or [])[:3]
    if len(years) < 2:
        years = (income_data.get("years") or bs_data.get("years") or is_data.get("years") or [])[:3]
    if len(years) < 2:
        return {
            "project_id": project_id,
            "scale": {"sheet_title": "分析汇总_规模", "headers": ["科目编码", "科目名称"], "rows": []},
            "structure": {"sheet_title": "分析汇总_结构", "headers": ["科目编码", "科目名称"], "rows": []},
            "analysis": {"rows": []},
        }
    bs_rows = bs_data.get("rows", []) or []
    is_rows = is_data.get("rows", []) or []
    income_nodes = {str(n.get("node_id", "")): n for n in (income_data.get("nodes", []) or [])}

    bs_by_code = {str(r.get("code", "")).strip().upper(): r for r in bs_rows}
    is_by_code = {str(r.get("code", "")).strip().upper(): r for r in is_rows}

    def _find_is_row(code_candidates: List[str], name_keywords: List[str]) -> Optional[Dict[str, Any]]:
        r = _pick_is_row(is_rows, code_candidates or [], name_keywords or [])
        return r

    def _pick_bs_value(item: Dict[str, Any]) -> Dict[str, Optional[float]]:
        for c in item.get("code_candidates", []):
            row = bs_by_code.get(str(c).strip().upper())
            if row:
                vals = row.get("values", {}) if isinstance(row.get("values"), dict) else {}
                return {y: vals.get(y) for y in years}
        return {y: None for y in years}

    def _pick_is_value(item: Dict[str, Any]) -> Dict[str, Optional[float]]:
        row = _find_is_row(item.get("code_candidates", []), item.get("name_keywords", []))
        vals = (row.get("values", {}) if row and isinstance(row.get("values"), dict) else {})
        return {y: vals.get(y) for y in years}

    def _pick_income_node_value(node_id: str) -> Dict[str, Optional[float]]:
        n = income_nodes.get(node_id, {})
        vals = n.get("values", {}) if isinstance(n.get("values"), dict) else {}
        return {y: vals.get(y) for y in years}

    def _fmt_amt(v: Optional[float]) -> str:
        return f"{v:.2f}" if v is not None else "待补充"

    def _yoy(curr: Optional[float], prev: Optional[float], stable: float = 2.0) -> str:
        if curr is None or prev in (None, 0):
            return "待补充"
        p = (curr - prev) / abs(prev) * 100.0
        if p > stable:
            return _render_template(sum_tpl.get("summary_yoy_up", "增加{value}%"), {"value": f"{abs(p):.2f}"})
        if p < -stable:
            return _render_template(sum_tpl.get("summary_yoy_down", "减少{value}%"), {"value": f"{abs(p):.2f}"})
        return _render_template(sum_tpl.get("summary_yoy_stable", "基本稳定（变动{value}%）"), {"value": f"{abs(p):.2f}"})

    def _ratio_pct(numer: Optional[float], denom: Optional[float]) -> Optional[float]:
        if numer is None or denom in (None, 0):
            return None
        return numer / denom * 100.0

    def _pp_change(curr: Optional[float], prev: Optional[float], stable_pp: float = 2.0) -> str:
        if curr is None or prev is None:
            return "待补充"
        d = curr - prev
        if d > stable_pp:
            return _render_template(sum_tpl.get("summary_pp_up", "上升{value}个百分点"), {"value": f"{abs(d):.2f}"})
        if d < -stable_pp:
            return _render_template(sum_tpl.get("summary_pp_down", "下降{value}个百分点"), {"value": f"{abs(d):.2f}"})
        return _render_template(sum_tpl.get("summary_pp_stable", "基本稳定（变动{value}个百分点）"), {"value": f"{abs(d):.2f}"})

    def _code_num(code: str) -> Optional[int]:
        m = re.match(r"^BS(\d{3})$", str(code or "").strip().upper())
        return int(m.group(1)) if m else None

    def _build_bs_components(num_ranges: List[tuple], excludes: Optional[set] = None) -> List[Dict[str, Any]]:
        excludes = excludes or set()
        out: List[Dict[str, Any]] = []
        for r in bs_rows:
            code = str(r.get("code", "")).strip().upper()
            n = _code_num(code)
            if n is None or code in excludes:
                continue
            hit = any(a <= n <= b for (a, b) in num_ranges)
            if not hit:
                continue
            vals = r.get("values", {}) if isinstance(r.get("values"), dict) else {}
            out.append({"code": code, "name": str(r.get("name", "")), "values": {y: vals.get(y) for y in years}})
        return out

    def _build_income_components() -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        for n in (income_data.get("nodes", []) or []):
            nid = str(n.get("node_id", "")).strip()
            pid = str(n.get("parent_id", "")).strip()
            if pid != "1" or nid == "1.1":
                continue
            vals = n.get("values", {}) if isinstance(n.get("values"), dict) else {}
            out.append({"code": nid, "name": str(n.get("label", "")), "values": {y: vals.get(y) for y in years}})
        return out

    def _build_gp_components() -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        for n in (income_data.get("nodes", []) or []):
            nid = str(n.get("node_id", "")).strip()
            pid = str(n.get("parent_id", "")).strip()
            if pid != "2.1.1":
                continue
            vals = n.get("values", {}) if isinstance(n.get("values"), dict) else {}
            out.append({"code": nid, "name": str(n.get("label", "")), "values": {y: vals.get(y) for y in years}})
        return out

    def _up_down_pp(delta: Optional[float]) -> str:
        if delta is None:
            return "待补充"
        if delta > 0:
            return _render_template(sum_tpl.get("summary_pp_up", "上升{value}个百分点"), {"value": f"{abs(delta):.2f}"})
        if delta < 0:
            return _render_template(sum_tpl.get("summary_pp_down", "下降{value}个百分点"), {"value": f"{abs(delta):.2f}"})
        return _render_template(sum_tpl.get("summary_pp_stable", "基本稳定（变动{value}个百分点）"), {"value": f"{abs(delta):.2f}"})

    def _top_migration_text(total_vals: Dict[str, Optional[float]], comps: List[Dict[str, Any]], coverage_target_pct: float = 75.0) -> str:
        if len(years) < 2:
            return _render_template(sum_tpl.get("summary_top_missing", "结构迁移分析待补充。"), {})
        y1, y_last = years[0], years[-1]
        t1 = total_vals.get(y1)
        t_last = total_vals.get(y_last)
        records: List[Dict[str, Any]] = []
        for c in comps:
            vals = c.get("values", {}) if isinstance(c.get("values"), dict) else {}
            v1 = vals.get(y1)
            v_last = vals.get(y_last)
            s1 = _ratio_pct(v1, t1)
            s_last = _ratio_pct(v_last, t_last)
            d = (s_last - s1) if (s_last is not None and s1 is not None) else None
            nm = str(c.get("name", ""))
            records.append({"name": nm, "name_norm": re.sub(r"[()（）\s]", "", nm).replace("合计", ""), "s_last": s_last, "d": d})
        latest = [x for x in records if x.get("s_last") is not None]
        if not latest:
            return _render_template(sum_tpl.get("summary_top_missing", "结构迁移分析待补充。"), {})
        latest.sort(key=lambda x: float(x.get("s_last") or 0.0), reverse=True)
        # Remove overlapping duplicates (e.g., "固定资产(合计)" vs "固定资产", "应收票据及应收账款" vs "应收账款").
        dedup_latest: List[Dict[str, Any]] = []
        for x in latest:
            n = str(x.get("name_norm", ""))
            s = float(x.get("s_last") or 0.0)
            dup = False
            for k in dedup_latest:
                nk = str(k.get("name_norm", ""))
                sk = float(k.get("s_last") or 0.0)
                if not n or not nk:
                    continue
                overlap = (n in nk) or (nk in n)
                near_share = abs(s - sk) <= 1.0
                if overlap and (near_share or ("及" in str(k.get("name", "")))):
                    dup = True
                    break
            if not dup:
                dedup_latest.append(x)
        latest = dedup_latest
        selected: List[Dict[str, Any]] = []
        cum = 0.0
        for x in latest:
            selected.append(x)
            cum += float(x.get("s3") or 0.0)
            if cum >= coverage_target_pct:
                break
        top_txt = "、".join([f"{x['name']}（{_fmt_pct(x.get('s_last'))}）" for x in selected])
        deltas = [x for x in records if x.get("d") is not None]
        up = max(deltas, key=lambda x: x["d"]) if deltas else None
        down = min(deltas, key=lambda x: x["d"]) if deltas else None
        move_bits: List[str] = []
        move_names: set = set()
        if up is not None and float(up["d"]) > 0:
            move_bits.append(f"{up['name']}占比上升{abs(float(up['d'])):.2f}个百分点")
            move_names.add(str(up["name"]))
        if down is not None and float(down["d"]) < 0:
            move_bits.append(f"{down['name']}占比下降{abs(float(down['d'])):.2f}个百分点")
            move_names.add(str(down["name"]))
        top_delta_bits = []
        for x in selected[:3]:
            if str(x["name"]) in move_names:
                continue
            top_delta_bits.append(f"{x['name']}占比{_up_down_pp(x.get('d'))}")
        move_txt = "；".join(move_bits) if move_bits else _render_template(sum_tpl.get("summary_top_move_default", "内部结构整体稳定"), {})
        top_delta_txt = (
            "；".join(top_delta_bits)
            if top_delta_bits
            else _render_template(sum_tpl.get("summary_top_delta_pending", "主要构成项变化待补充"), {})
        )
        line1 = _render_template(
            sum_tpl.get("summary_top_line1", "{y3}年内部构成按占比从高到低分别为：{top_txt}，这几项占比合计{cum}%。"),
            {"y3": y_last, "top_txt": top_txt, "cum": f"{cum:.2f}"},
        )
        line2 = _render_template(
            sum_tpl.get("summary_top_line2", "{y1}-{y3}年，{move_txt}；同时，{top_delta_txt}。"),
            {"y1": y1, "y3": y_last, "move_txt": move_txt, "top_delta_txt": top_delta_txt},
        )
        return line1 + "\n" + line2

    # Bases for structure text.
    bs_total = _pick_bs_value({"code_candidates": ["BS057"]})
    liab_total = _pick_bs_value({"code_candidates": ["BS103"]})
    income_total = _pick_is_value({"code_candidates": ["IS001"], "name_keywords": ["营业总收入", "营业收入"]})
    recurring_total = _pick_income_node_value("3.1")
    nonrec_total = _pick_income_node_value("3.2")
    op_net_total = {
        y: ((recurring_total.get(y) or 0.0) + (nonrec_total.get(y) or 0.0))
        if (recurring_total.get(y) is not None or nonrec_total.get(y) is not None)
        else None
        for y in years
    }

    scale_rows: List[Dict[str, Any]] = []
    struct_rows: List[Dict[str, Any]] = []

    summary_threshold_cfg = _read_threshold_config()
    top_coverage_default = float(
        summary_threshold_cfg.get("summary_top_coverage_default", float(SUMMARY_TOP_COVERAGE_TARGET_PCT))
    )
    top_coverage_income = float(
        summary_threshold_cfg.get("summary_top_coverage_income", float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME))
    )

    for item in SUMMARY_ANALYSIS_ITEMS:
        code = str(item["code"])
        name = str(item["name"])
        kind = str(item.get("kind", ""))
        source_code = str(item.get("source_code", ""))
        if kind == "bs":
            values = _pick_bs_value(item)
        elif kind == "is":
            values = _pick_is_value(item)
        elif kind == "income_combo":
            values = dict(op_net_total)
        elif kind == "income_node":
            values = _pick_income_node_value(source_code)
        else:
            values = {y: None for y in years}

        y1, y2, y3 = _compat_three_periods(years)
        v1, v2, v3 = values.get(y1), values.get(y2), values.get(y3)
        if len(years) == 2:
            abs_text = (
                f"{name}：{y1}年{_fmt_amt(v1)}{str(sum_tpl_units.get('summary_abs_line', DEFAULT_AMOUNT_UNIT) or DEFAULT_AMOUNT_UNIT)}，"
                f"{y2}年{_fmt_amt(v2)}{str(sum_tpl_units.get('summary_abs_line', DEFAULT_AMOUNT_UNIT) or DEFAULT_AMOUNT_UNIT)}；"
                f"{y2}较{y1}{_yoy(v2, v1)}。"
            )
        else:
            abs_text = _render_template(
                sum_tpl.get(
                    "summary_abs_line",
                    "{name}：{y1}年{v1}{unit}，{y2}年{v2}{unit}，{y3}年{v3}{unit}；{y2}较{y1}{p21}，{y3}较{y2}{p32}。",
                ),
                {
                    "name": name,
                    "y1": y1,
                    "v1": _fmt_amt(v1),
                    "y2": y2,
                    "v2": _fmt_amt(v2),
                    "y3": y3,
                    "v3": _fmt_amt(v3),
                    "p21": _yoy(v2, v1),
                    "p32": _yoy(v3, v2),
                    "unit": str(sum_tpl_units.get("summary_abs_line", DEFAULT_AMOUNT_UNIT) or DEFAULT_AMOUNT_UNIT),
                },
            )
        scale_rows.append(
            {
                "科目编码": code,
                "科目名称": name,
                y1: v1,
                y2: v2,
                y3: v3,
                "定量描述_绝对量": abs_text,
            }
        )

        # relative/structure text
        base_name = ""
        bvals: Dict[str, Optional[float]] = {y: None for y in years}
        if code in {"SUM001", "SUM002", "SUM003"}:
            base_name = "资产总计"
            bvals = bs_total
        elif code in {"SUM004", "SUM005", "SUM006"}:
            base_name = "负债总计"
            bvals = liab_total
        elif code == "SUM007":
            base_name = "营业收入"
            bvals = income_total
        elif code in {"SUM009", "SUM010", "SUM011"}:
            base_name = "经营净收益"
            bvals = dict(op_net_total)

        r1 = _ratio_pct(v1, bvals.get(y1))
        r2 = _ratio_pct(v2, bvals.get(y2))
        r3 = _ratio_pct(v3, bvals.get(y3))
        rel_text = ""
        if base_name and code not in {"SUM003", "SUM006", "SUM007", "SUM008", "SUM009", "SUM010", "SUM011"}:
            if len(years) == 2:
                rel_text = (
                    f"{name}占{base_name}比例：{y1}年{_fmt_pct(r1)}，{y2}年{_fmt_pct(r2)}；"
                    f"{y2}较{y1}{_pp_change(r2, r1)}。"
                )
            else:
                rel_text = _render_template(
                    sum_tpl.get(
                        "summary_ratio_line",
                        "{name}占{base_name}比例：{y1}年{r1}，{y2}年{r2}，{y3}年{r3}；{y2}较{y1}{d21}，{y3}较{y2}{d32}。",
                    ),
                    {
                        "name": name,
                        "base_name": base_name,
                        "y1": y1,
                        "r1": _fmt_pct(r1),
                        "y2": y2,
                        "r2": _fmt_pct(r2),
                        "y3": y3,
                        "r3": _fmt_pct(r3),
                        "d21": _pp_change(r2, r1),
                        "d32": _pp_change(r3, r2),
                    },
                )
        elif code == "SUM007":
            rel_text = _render_template(sum_tpl.get("summary_income_struct_intro", "{name}：分项结构如下。"), {"name": name})
        elif code in {"SUM003", "SUM006"}:
            rel_text = _render_template(sum_tpl.get("summary_total_not_applicable", "{name}：为汇总项，结构占比分析不适用。"), {"name": name})
        elif code == "SUM008":
            c_rec = recurring_total
            c_non = nonrec_total
            s_rec_3 = _ratio_pct(c_rec.get(y3), values.get(y3))
            s_non_3 = _ratio_pct(c_non.get(y3), values.get(y3))
            s_rec_1 = _ratio_pct(c_rec.get(y1), values.get(y1))
            s_non_1 = _ratio_pct(c_non.get(y1), values.get(y1))
            d_rec = (s_rec_3 - s_rec_1) if (s_rec_3 is not None and s_rec_1 is not None) else None
            d_non = (s_non_3 - s_non_1) if (s_non_3 is not None and s_non_1 is not None) else None
            rel_text = _render_template(
                sum_tpl.get(
                    "summary_sum008_struct",
                    "{y3}年经常性净收益占比{s_rec_3}、非经常性净收益占比{s_non_3}；{y1}-{y3}年，经常性净收益占比{d_rec}；非经常性净收益占比{d_non}。",
                ),
                {
                    "y3": y3,
                    "y1": y1,
                    "s_rec_3": _fmt_pct(s_rec_3),
                    "s_non_3": _fmt_pct(s_non_3),
                    "d_rec": _up_down_pp(d_rec),
                    "d_non": _up_down_pp(d_non),
                },
            ) + "\n" + _render_template(sum_tpl.get("summary_see_income_page", "详见“收入分析”子页面。"), {})
        elif code == "SUM009":
            # SUM009: directly reuse income-analysis 2.1.1 auto text, then add child composition trend.
            n_gp = income_nodes.get("2.1.1", {}) if isinstance(income_nodes, dict) else {}
            gp_auto = str(n_gp.get("auto_text", "") or "").strip()
            comps = _build_gp_components()
            abs_total_1 = sum(abs((c.get("values", {}) or {}).get(y1) or 0.0) for c in comps if (c.get("values", {}) or {}).get(y1) is not None)
            abs_total_3 = sum(abs((c.get("values", {}) or {}).get(y3) or 0.0) for c in comps if (c.get("values", {}) or {}).get(y3) is not None)
            bits: List[str] = []
            for c in sorted(
                comps,
                key=lambda x: abs(((x.get("values", {}) or {}).get(y3) or 0.0)),
                reverse=True,
            ):
                vals = c.get("values", {}) if isinstance(c.get("values"), dict) else {}
                v1c = vals.get(y1)
                v3c = vals.get(y3)
                s1c = (abs(v1c) / abs_total_1 * 100.0) if (v1c is not None and abs_total_1 > 1e-12) else None
                s3c = (abs(v3c) / abs_total_3 * 100.0) if (v3c is not None and abs_total_3 > 1e-12) else None
                dc = (s3c - s1c) if (s3c is not None and s1c is not None) else None
                bits.append(
                    f"{c.get('name','')}：{y3}年毛利影响占比{_fmt_pct(s3c)}（{y1}年{_fmt_pct(s1c)}，{y1}-{y3}年{_up_down_pp(dc)}）"
                )
            trend_txt = "；".join(bits) if bits else _render_template(sum_tpl.get("summary_sum009_trend_pending", "各子业务毛利结构趋势待补充"), {})
            rel_text = ((gp_auto + "\n") if gp_auto else "") + _render_template(
                sum_tpl.get("summary_sum009_struct", "{y3}年各子业务毛利影响占比及结构变化：{trend_txt}。"),
                {"y3": y3, "trend_txt": trend_txt},
            )
        elif code == "SUM010":
            # SUM010: split into 主营业务毛利收益 + 其他经常性收益 and describe structure trend.
            gp_vals = _pick_income_node_value("2.1.1")
            other_vals = {
                y: ((values.get(y) or 0.0) - (gp_vals.get(y) or 0.0))
                if (values.get(y) is not None or gp_vals.get(y) is not None)
                else None
                for y in years
            }
            gp_s1 = _ratio_pct(gp_vals.get(y1), values.get(y1))
            gp_s3 = _ratio_pct(gp_vals.get(y3), values.get(y3))
            ot_s1 = _ratio_pct(other_vals.get(y1), values.get(y1))
            ot_s3 = _ratio_pct(other_vals.get(y3), values.get(y3))
            d_gp = (gp_s3 - gp_s1) if (gp_s3 is not None and gp_s1 is not None) else None
            d_ot = (ot_s3 - ot_s1) if (ot_s3 is not None and ot_s1 is not None) else None
            rel_text = _render_template(
                sum_tpl.get(
                    "summary_sum010_struct",
                    "{y3}年经常性净收益中，主营业务毛利收益占比{gp_s3}，其他经常性收益占比{ot_s3}；{y1}-{y3}年，主营业务毛利收益占比{d_gp}，其他经常性收益占比{d_ot}。",
                ),
                {"y3": y3, "y1": y1, "gp_s3": _fmt_pct(gp_s3), "ot_s3": _fmt_pct(ot_s3), "d_gp": _up_down_pp(d_gp), "d_ot": _up_down_pp(d_ot)},
            ) + "\n" + _render_template(sum_tpl.get("summary_see_income_page", "详见“收入分析”子页面。"), {})
        elif code == "SUM011":
            rel_text = _render_template(
                sum_tpl.get("summary_sum011_struct", "{y3}年{name}占经营净收益比例为{r3}；{y1}-{y3}年占比{d_31}。"),
                {
                    "y3": y3,
                    "y1": y1,
                    "name": name,
                    "r3": _fmt_pct(r3),
                    "d_31": _up_down_pp((r3 - r1) if (r3 is not None and r1 is not None) else None),
                },
            ) + "\n" + _render_template(sum_tpl.get("summary_see_income_page", "详见“收入分析”子页面。"), {})
        else:
            rel_text = _render_template(sum_tpl.get("summary_struct_pending", "{name}：结构占比分析待补充。"), {"name": name})

        # Upgrade selected summary subjects with "Top composition + migration" narrative.
        comps: List[Dict[str, Any]] = []
        if code == "SUM001":
            comps = _build_bs_components([(1, 26)], excludes=set(SUMMARY_EXCLUDE_CODES.get("SUM001", set())))
        elif code == "SUM002":
            comps = _build_bs_components([(29, 55)], excludes=set(SUMMARY_EXCLUDE_CODES.get("SUM002", set())))
        elif code == "SUM003":
            # Asset total: use two subtotal buckets only (流动资产合计 / 非流动资产合计).
            c_cur = _pick_bs_value({"code_candidates": ["BS000", "BS027"]})
            c_non = _pick_bs_value({"code_candidates": ["BS028", "BS056"]})
            s_cur_3 = _ratio_pct(c_cur.get(y3), values.get(y3))
            s_non_3 = _ratio_pct(c_non.get(y3), values.get(y3))
            s_cur_1 = _ratio_pct(c_cur.get(y1), values.get(y1))
            s_non_1 = _ratio_pct(c_non.get(y1), values.get(y1))
            d_cur = (s_cur_3 - s_cur_1) if (s_cur_3 is not None and s_cur_1 is not None) else None
            d_non = (s_non_3 - s_non_1) if (s_non_3 is not None and s_non_1 is not None) else None
            rel_text = (
                f"{y3}年流动资产合计占比{_fmt_pct(s_cur_3)}、非流动资产合计占比{_fmt_pct(s_non_3)}；"
                f"{y1}-{y3}年，流动资产合计占比{_up_down_pp(d_cur)}；非流动资产合计占比{_up_down_pp(d_non)}。"
            )
        elif code == "SUM004":
            comps = _build_bs_components([(59, 88)], excludes=set(SUMMARY_EXCLUDE_CODES.get("SUM004", set())))
        elif code == "SUM005":
            comps = _build_bs_components([(91, 101)], excludes=set(SUMMARY_EXCLUDE_CODES.get("SUM005", set())))
        elif code == "SUM006":
            # Liability total: use two subtotal buckets only (流动负债合计 / 非流动负债合计).
            c_cur = _pick_bs_value({"code_candidates": ["BS058", "BS089"]})
            c_non = _pick_bs_value({"code_candidates": ["BS090", "BS102"]})
            s_cur_3 = _ratio_pct(c_cur.get(y3), values.get(y3))
            s_non_3 = _ratio_pct(c_non.get(y3), values.get(y3))
            s_cur_1 = _ratio_pct(c_cur.get(y1), values.get(y1))
            s_non_1 = _ratio_pct(c_non.get(y1), values.get(y1))
            d_cur = (s_cur_3 - s_cur_1) if (s_cur_3 is not None and s_cur_1 is not None) else None
            d_non = (s_non_3 - s_non_1) if (s_non_3 is not None and s_non_1 is not None) else None
            rel_text = (
                f"{y3}年流动负债合计占比{_fmt_pct(s_cur_3)}、非流动负债合计占比{_fmt_pct(s_non_3)}；"
                f"{y1}-{y3}年，流动负债合计占比{_up_down_pp(d_cur)}；非流动负债合计占比{_up_down_pp(d_non)}。"
            )
        elif code == "SUM007":
            comps = _build_income_components()
        if comps and code in {"SUM001", "SUM002", "SUM004", "SUM005", "SUM007"}:
            target_pct = top_coverage_income if code == "SUM007" else top_coverage_default
            rel_text = rel_text + "\n" + _top_migration_text(values, comps, coverage_target_pct=target_pct)

        struct_rows.append(
            {
                "科目编码": code,
                "科目名称": name,
                y1: r1,
                y2: r2,
                y3: r3,
                "定量描述_相对量": rel_text,
            }
        )

    # analysis map with manual merge
    store = load_store(project_id)
    entries = store.get("entries", {})
    analysis_rows: List[Dict[str, Any]] = []
    scale_by_code = {str(r.get("科目编码", "")): r for r in scale_rows}
    struct_by_code = {str(r.get("科目编码", "")): r for r in struct_rows}
    for item in SUMMARY_ANALYSIS_ITEMS:
        code = str(item["code"])
        name = str(item["name"])
        auto_abs = str(scale_by_code.get(code, {}).get("定量描述_绝对量", "") or "")
        auto_rel = str(struct_by_code.get(code, {}).get("定量描述_相对量", "") or "")
        auto_combined = auto_abs if not auto_rel else f"{auto_abs}\n{auto_rel}"
        saved = entries.get(make_entry_key("summary_analysis", code), {})
        manual_text = str(saved.get("manual_text", "") or "")
        confirmed = bool(saved.get("confirmed", False))
        analysis_rows.append(
            {
                "code": code,
                "name": name,
                "auto_abs": auto_abs,
                "auto_rel": auto_rel,
                "auto_combined": auto_combined,
                "manual_text": manual_text,
                "confirmed": confirmed,
                "final_text": manual_text if manual_text.strip() else auto_combined,
            }
        )

    headers_scale = ["科目编码", "科目名称"] + years + ["定量描述_绝对量"]
    headers_struct = ["科目编码", "科目名称"] + years + ["定量描述_相对量"]
    return {
        "project_id": project_id,
        "scale": {"sheet_title": "分析汇总_规模", "headers": headers_scale, "rows": scale_rows},
        "structure": {"sheet_title": "分析汇总_结构", "headers": headers_struct, "rows": struct_rows},
        "analysis": {"rows": analysis_rows},
    }


def load_store(project_id: str) -> Dict[str, Any]:
    p = store_path(project_id)
    if not p.exists():
        return {"project_id": normalize_project_id(project_id), "entries": {}}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {"project_id": normalize_project_id(project_id), "entries": {}}


def save_store(project_id: str, payload: Dict[str, Any]) -> None:
    p = store_path(project_id)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def make_entry_key(group_id: str, code: str) -> str:
    return f"{group_id}:{code}"


def make_value_override_key(group_id: str, code: str, year: str) -> str:
    return f"{group_id}:{code}:{year}"


def get_value_overrides(store: Dict[str, Any]) -> Dict[str, Any]:
    raw = store.get("value_overrides", {})
    return raw if isinstance(raw, dict) else {}


def apply_value_overrides_to_rows(
    rows: List[Dict[str, Any]],
    years: List[str],
    group_id: str,
    project_id: Optional[str],
) -> List[Dict[str, Any]]:
    if not project_id:
        return rows
    store = load_store(project_id)
    overrides = get_value_overrides(store)
    if not overrides:
        return rows
    out: List[Dict[str, Any]] = []
    for r in rows:
        code = str(r.get("code", "")).strip()
        vals = dict(r.get("values", {})) if isinstance(r.get("values"), dict) else {}
        changed_years: List[str] = []
        for y in years:
            k = make_value_override_key(group_id, code, str(y))
            obj = overrides.get(k, {})
            if isinstance(obj, dict) and obj.get("enabled", True) and obj.get("value") is not None:
                vals[str(y)] = normalize_num(obj.get("value"))
                changed_years.append(str(y))
        nr = dict(r)
        nr["values"] = vals
        nr["override_years"] = changed_years
        out.append(nr)
    return out


DEFAULT_VALIDATION_RULES = {
    "bs": [
        {
            "rule_id": "BS_CORE_001",
            "rule_name": "资产总计=流动资产合计+非流动资产合计",
            "left_code": "BS057",
            "operator": "SUM",
            "right_codes": ["BS027", "BS056"],
            "tolerance": 0.05,
            "enabled": 1,
        },
        {
            "rule_id": "BS_CORE_002",
            "rule_name": "负债总计=流动负债合计+非流动负债合计",
            "left_code": "BS103",
            "operator": "SUM",
            "right_codes": ["BS089", "BS102"],
            "tolerance": 0.05,
            "enabled": 1,
        },
    ],
    "is": [],
    "cf": [],
}


def load_validation_rules(group_id: str) -> List[Dict[str, Any]]:
    gid = str(group_id or "").strip().lower()
    defaults = list(DEFAULT_VALIDATION_RULES.get(gid, []))
    if gid not in {"bs", "is", "cf"}:
        return defaults
    if not VALIDATION_RULEBOOK_PATH.exists():
        _warn_rule_once(
            f"validation_rulebook_missing_runtime:{gid}",
            f"校验规则文件缺失，{gid} 已回退默认规则",
            source="validation",
        )
        return defaults
    try:
        wb = load_workbook(VALIDATION_RULEBOOK_PATH, data_only=True)
    except Exception:
        _warn_rule_once(
            f"validation_rulebook_open_fail_runtime:{gid}",
            f"校验规则文件读取失败，{gid} 已回退默认规则",
            source="validation",
        )
        return defaults
    ws = get_sheet_by_loose_name(wb, [f"{gid}_checks"])
    if ws is None:
        _warn_rule_once(
            f"validation_sheet_missing_runtime:{gid}",
            f"{gid}_checks 缺失，已回退默认规则",
            source="validation",
        )
        return defaults
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    c_rule_id = headers.get("rule_id", 1)
    c_rule_name = headers.get("rule_name", headers.get("rule_desc", 2))
    c_left = headers.get("left_code", 3)
    c_op = headers.get("operator", 4)
    c_right = headers.get("right_codes", headers.get("right_expr", 5))
    c_tol = headers.get("tolerance", 6)
    c_enabled = headers.get("enabled", 7)
    out: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        enabled_cell = ws.cell(r, c_enabled).value
        enabled_raw = str("1" if enabled_cell is None else enabled_cell).strip().lower()
        if enabled_raw in {"0", "false", "no"}:
            continue
        left_code = str(ws.cell(r, c_left).value or "").strip().upper()
        op = str(ws.cell(r, c_op).value or "SUM").strip().upper()
        right_raw = str(ws.cell(r, c_right).value or "").strip()
        if not left_code or not right_raw:
            continue
        right_codes = [x.strip().upper() for x in right_raw.replace(";", ",").split(",") if x.strip()]
        if not right_codes:
            continue
        try:
            tol = float(ws.cell(r, c_tol).value)
        except Exception:
            tol = 0.05
        out.append(
            {
                "rule_id": str(ws.cell(r, c_rule_id).value or "").strip() or f"{gid.upper()}_{r}",
                "rule_name": str(ws.cell(r, c_rule_name).value or "").strip() or left_code,
                "left_code": left_code,
                "operator": op,
                "right_codes": right_codes,
                "tolerance": tol,
                "enabled": 1,
            }
        )
    return out or defaults


def build_validation_map(group_id: str, rows: List[Dict[str, Any]], years: List[str], tol: float = 0.05) -> Dict[str, Dict[str, Any]]:
    by_code = {str(r.get("code", "")).strip().upper(): r for r in rows}

    def _v(code: str, year: str) -> Optional[float]:
        r = by_code.get(code, {})
        vals = r.get("values", {}) if isinstance(r.get("values"), dict) else {}
        return normalize_num(vals.get(year))

    rules = load_validation_rules(group_id)

    result: Dict[str, Dict[str, Any]] = {}
    for code in by_code.keys():
        result[code] = {"status": "未配置", "message": "该行暂未配置校验规则"}

    for rule in rules:
        left = str(rule.get("left_code", "")).strip().upper()
        op = str(rule.get("operator", "SUM")).strip().upper()
        rights = [str(x).strip().upper() for x in (rule.get("right_codes", []) or []) if str(x).strip()]
        rule_tol = float(rule.get("tolerance", tol))
        if left not in by_code:
            continue
        failed: List[str] = []
        for y in years:
            lv = _v(left, y)
            if lv is None:
                failed.append(f"{y}待补充")
                continue
            rv = None
            if op == "SUM":
                rv_codes = [x.lstrip("+-") for x in rights if x.lstrip("+-") in by_code]
                rv_list = [_v(x, y) for x in rv_codes]
                if not rv_list:
                    failed.append(f"{y}待补充")
                    continue
                rv = sum([x or 0.0 for x in rv_list])
            elif op == "EXPR":
                acc = 0.0
                ok = True
                for raw in rights:
                    sign = -1.0 if str(raw).startswith("-") else 1.0
                    code_part = str(raw)[1:] if str(raw).startswith(("+", "-")) else str(raw)
                    if code_part not in by_code:
                        ok = False
                        break
                    vv = _v(code_part, y)
                    if vv is None:
                        ok = False
                        break
                    acc += sign * vv
                rv = acc if ok else None
            if rv is None:
                continue
            diff = (lv or 0.0) - (rv or 0.0)
            if abs(diff) > rule_tol:
                failed.append(f"{y}差额{diff:.2f}")
        if failed:
            result[left] = {"status": "未通过", "message": "；".join(failed)}
        else:
            result[left] = {"status": "通过", "message": "核心勾稽已通过"}

    return result


def build_bs_validation_map(rows: List[Dict[str, Any]], years: List[str], tol: float = 0.05) -> Dict[str, Dict[str, Any]]:
    return build_validation_map("bs", rows, years, tol=tol)


def generate_auto_text(name: str, values: Dict[str, Optional[float]], years: List[str]) -> str:
    tpl = load_key_ratio_text_rules().get("templates", {})
    if not years:
        return ""
    latest = years[-1]
    latest_val = values.get(latest)
    if latest_val is None:
        return _render_template(
            tpl.get("sheet_auto_no_latest", DEFAULT_TEXT_TEMPLATES["sheet_auto_no_latest"]),
            {"name": name},
        )
    if len(years) < 2:
        return _render_template(
            tpl.get("sheet_auto_one_year", DEFAULT_TEXT_TEMPLATES["sheet_auto_one_year"]),
            {"name": name, "latest": latest, "latest_val": f"{latest_val:,.2f}"},
        )

    prev = years[-2]
    prev_val = values.get(prev)
    if prev_val in (None, 0):
        return _render_template(
            tpl.get("sheet_auto_prev_missing", DEFAULT_TEXT_TEMPLATES["sheet_auto_prev_missing"]),
            {"name": name, "latest": latest, "latest_val": f"{latest_val:,.2f}"},
        )

    diff = latest_val - prev_val
    pct = diff / abs(prev_val) * 100
    direction = "上升" if diff >= 0 else "下降"
    return _render_template(
        tpl.get("sheet_auto_two_year", DEFAULT_TEXT_TEMPLATES["sheet_auto_two_year"]),
        {
            "name": name,
            "latest": latest,
            "prev": prev,
            "latest_val": f"{latest_val:,.2f}",
            "direction": direction,
            "abs_diff": f"{abs(diff):,.2f}",
            "abs_pct": f"{abs(pct):.2f}",
        },
    )


def list_projects() -> List[str]:
    out_root = PROJECT_ROOT / "outputs"
    if not out_root.exists():
        return []
    return sorted([p.name for p in out_root.iterdir() if p.is_dir()])


def _get_rule_source(source_id: str) -> Optional[Dict[str, Any]]:
    sid = str(source_id or "").strip()
    return next((x for x in RULE_EDIT_SOURCES if str(x.get("source_id")) == sid), None)


def _id_sort_key(v: Any) -> tuple:
    s = str(v or "").strip()
    m = re.match(r"^([A-Za-z_]+)_(\d+)$", s)
    if m:
        return (m.group(1), int(m.group(2)))
    return (s, 0)


def _validation_id_prefix(sheet_name: str) -> str:
    s = str(sheet_name or "").strip().lower()
    if s == "bs_checks":
        return "BS_CHK"
    if s == "is_checks":
        return "IS_CHK"
    if s == "cf_checks":
        return "CF_CHK"
    return "CHK"


def _rule_catalog() -> List[Dict[str, Any]]:
    items = []
    for x in RULE_EDIT_SOURCES:
        p = Path(x["path"])
        items.append(
            {
                "source_id": str(x["source_id"]),
                "label": str(x["label"]),
                "path": str(p),
                "exists": p.exists(),
                "sheets": list(x.get("sheets", [])),
            }
        )
    return items


def _read_rule_sheet(source_id: str, sheet_name: str) -> Dict[str, Any]:
    if source_id == "rulebook_main" and sheet_name in {"analysis_thresholds", "分析阈值配置"}:
        return _read_analysis_thresholds_expanded(source_id, "analysis_thresholds")
    src = _get_rule_source(source_id)
    if not src:
        raise ValueError("unknown source_id")
    p = Path(src["path"])
    if not p.exists():
        raise FileNotFoundError(str(p))
    wb = load_workbook(p, data_only=False)
    ws = get_sheet_by_loose_name(wb, [sheet_name])
    if ws is None:
        raise ValueError(f"sheet not found: {sheet_name}")
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    if not any(headers):
        headers = [f"col_{c}" for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, "") for v in vals):
            continue
        obj = {"_row": r}
        for i, h in enumerate(headers):
            key = h or f"col_{i+1}"
            obj[key] = vals[i]
        rows.append(obj)
    if sheet_name in {"analysis_text_templates", "分析文本模板"} and source_id in {"rulebook_main", "profit_rulebook"}:
        rows = [x for x in rows if _rulebook_template_scope_match(source_id, x)]
    if source_id == "validation_rulebook":
        key_name = "rule_id" if "rule_id" in headers else (headers[0] if headers else "")
        if key_name:
            rows.sort(key=lambda x: _id_sort_key(x.get(key_name)))
    return {
        "source_id": source_id,
        "sheet_name": ws.title,
        "path": str(p),
        "headers": [h or f"col_{i+1}" for i, h in enumerate(headers)],
        "rows": rows,
    }


def _validate_rule_rows(sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    hset = {str(h) for h in headers}
    text_keys = [h for h in headers if "text" in h.lower() or "template" in h.lower()]
    threshold_keys = [
        h
        for h in headers
        if h.lower() in {"threshold_value", "stable_threshold", "tolerance"}
        or h.lower().endswith("_threshold")
        or h.lower().endswith("_threshold_pp")
    ]
    place_key = next((h for h in headers if h.lower() in {"placeholders", "variables"}), "")
    for i, row in enumerate(rows, start=1):
        for tk in text_keys:
            v = row.get(tk)
            if isinstance(v, str) and ("?" in v or "�" in v):
                issues.append({"row": i, "field": tk, "level": "error", "message": "疑似乱码字符"})
        for nk in threshold_keys:
            v = row.get(nk)
            if v in (None, ""):
                continue
            try:
                float(str(v).strip())
            except Exception:
                issues.append({"row": i, "field": nk, "level": "error", "message": "阈值应为数值"})
        if place_key and ("template_text" in hset or "template_text_zh" in hset):
            tpl = str(row.get("template_text", "") or row.get("template_text_zh", "") or "")
            ph_raw = str(row.get(place_key, "") or "")
            used = sorted(set(re.findall(r"\{([A-Za-z0-9_]+)\}", tpl)))
            declared = sorted(set([x.strip().strip("{}") for x in ph_raw.replace(";", ",").split(",") if x.strip()]))
            if used and declared:
                miss = [x for x in used if x not in declared]
                if miss:
                    issues.append({"row": i, "field": place_key, "level": "warn", "message": f"占位符未声明: {','.join(miss)}"})
    return issues


def _save_rule_sheet(source_id: str, sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if source_id == "rulebook_main" and sheet_name in {"analysis_thresholds", "分析阈值配置"}:
        return _save_analysis_thresholds_expanded(rows)
    src = _get_rule_source(source_id)
    if not src:
        raise ValueError("unknown source_id")
    p = Path(src["path"])
    if not p.exists():
        raise FileNotFoundError(str(p))
    wb = load_workbook(p, data_only=False)
    ws = get_sheet_by_loose_name(wb, [sheet_name])
    if ws is None:
        raise ValueError(f"sheet not found: {sheet_name}")
    issues = _validate_rule_rows(sheet_name, headers, rows)
    has_error = any(x.get("level") == "error" for x in issues)
    if has_error:
        return {"ok": False, "issues": issues, "saved_rows": 0, "path": str(p), "sheet_name": ws.title}

    rows_clean = []
    for r in rows:
        obj = dict(r)
        obj.pop("_row", None)
        rows_clean.append(obj)

    # Virtual split for rulebook.xlsx/analysis_text_templates:
    # - rulebook_main: asset/liability templates only
    # - profit_rulebook: non-asset/liability templates
    if sheet_name in {"analysis_text_templates", "分析文本模板"} and source_id in {"rulebook_main", "profit_rulebook"}:
        hdr = [str(h or "").strip() for h in headers]
        key_name = "template_key" if "template_key" in hdr else (hdr[0] if hdr else "")
        if key_name:
            # Read full rows from sheet first.
            full_rows: List[Dict[str, Any]] = []
            for rr in range(2, ws.max_row + 1):
                vals = [ws.cell(rr, c).value for c in range(1, ws.max_column + 1)]
                if all(v in (None, "") for v in vals):
                    continue
                obj = {}
                for i, h in enumerate(hdr):
                    k = h or f"col_{i+1}"
                    obj[k] = vals[i]
                full_rows.append(obj)
            kept_rows = [x for x in full_rows if not _rulebook_template_scope_match(source_id, x)]
            merged_by_key: Dict[str, Dict[str, Any]] = {}
            order: List[str] = []
            for x in kept_rows + rows_clean:
                k = str(x.get(key_name, "")).strip()
                if not k:
                    continue
                if k not in merged_by_key:
                    order.append(k)
                merged_by_key[k] = x
            rows_clean = [merged_by_key[k] for k in order]

    if source_id == "validation_rulebook":
        id_key = "rule_id" if "rule_id" in headers else (headers[0] if headers else "")
        if id_key:
            prefix = _validation_id_prefix(sheet_name)
            used_nums = set()
            for r in rows_clean:
                rid = str(r.get(id_key, "")).strip()
                m = re.match(rf"^{re.escape(prefix)}_(\d+)$", rid)
                if m:
                    used_nums.add(int(m.group(1)))
            nxt = 1
            for r in rows_clean:
                rid = str(r.get(id_key, "")).strip()
                if rid:
                    continue
                while nxt in used_nums:
                    nxt += 1
                r[id_key] = f"{prefix}_{nxt:03d}"
                used_nums.add(nxt)
                nxt += 1
            rows_clean.sort(key=lambda x: _id_sort_key(x.get(id_key)))

    # Rewrite body rows while preserving header row.
    body_start = 2
    max_cols = len(headers)
    if ws.max_row >= body_start:
        ws.delete_rows(body_start, ws.max_row - body_start + 1)
    for r in rows_clean:
        vals = [r.get(h, None) for h in headers]
        ws.append(vals)
    # Ensure header row is stable by provided headers.
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h
    wb.save(p)
    return {"ok": True, "issues": issues, "saved_rows": len(rows_clean), "path": str(p), "sheet_name": ws.title}


def render_index() -> str:
    cards = "".join(
        f'<a class="card" href="/sheet/{g["id"]}"><h3>{g["label"]}</h3><p>进入查看与编辑</p></a>' for g in SHEET_GROUPS
    )
    analysis_cards = (
        '<a class="card" href="/analysis/assets"><h3>资产分析</h3><p>查看规模变化与结构占比</p></a>'
        '<a class="card" href="/analysis/liabilities"><h3>负债分析</h3><p>查看规模变化与结构占比</p></a>'
        '<a class="card" href="/analysis/summary"><h3>分析汇总</h3><p>汇总科目综合描述与确认</p></a>'
        '<a class="card" href="/analysis/income"><h3>收入分析</h3><p>经常性/非经常性收益拆分</p></a>'
        '<a class="card" href="/analysis/ratios"><h3>财务指标分析</h3><p>指标趋势、判断与确认</p></a>'
        '<a class="card" href="/analysis/key-ratios"><h3>重要指标分析</h3><p>ROE与毛利率深度分析</p></a>'
        '<a class="card" href="/rules"><h3>规则配置</h3><p>阈值与文本模板可视化维护</p></a>'
        '<a class="card" href="/warnings"><h3>告警清单</h3><p>规则体检与降级告警</p></a>'
    )
    project_options = "".join(f'<option value="{p}">{p}</option>' for p in list_projects())
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>主文件前端</title>
  <style>
    :root {{ --bg:#f3f5f7; --card:#ffffff; --ink:#111827; --muted:#6b7280; --line:#d1d5db; --acc:#0f766e; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .wrap {{ max-width:1000px; margin:28px auto; padding:0 16px; }}
    h1 {{ margin:0 0 8px; }}
    .meta {{ color:var(--muted); margin-bottom:14px; }}
    .toolbar {{ background:#fff; border:1px solid var(--line); border-radius:10px; padding:10px; margin-bottom:14px; }}
    .grid {{ display:grid; grid-template-columns: repeat(auto-fit,minmax(220px,1fr)); gap:14px; }}
    .card {{ display:block; text-decoration:none; background:var(--card); border:1px solid var(--line); border-radius:10px; padding:18px; color:inherit; }}
    .card:hover {{ border-color:var(--acc); box-shadow:0 3px 10px rgba(0,0,0,.05); }}
    input,select,button {{ font:inherit; }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>主文件展示与确认</h1>
    <div class="meta">一级页面：基础报表 + 分析视图</div>
    <div class="toolbar">
      <label>项目ID：</label>
      <input id="pid" value="{DEFAULT_PROJECT_ID}" style="min-width:260px;padding:6px;"/>
      <select id="plist"><option value="">从 outputs 选择</option>{project_options}</select>
      <button onclick="applyProject()">应用</button>
      <button onclick="exportReportTemplate()">导出报表模板</button>
      <span id="tip" style="margin-left:8px;color:#6b7280;"></span>
    </div>
    <div class="grid" id="cards">{cards}{analysis_cards}</div>
  </div>
<script>
function getPid() {{
  return (document.getElementById('pid').value || '').trim();
}}
function rewriteLinks() {{
  const pid = getPid();
  document.querySelectorAll('#cards a.card').forEach(a => {{
    const u = new URL(a.getAttribute('href'), location.origin);
    if (pid) u.searchParams.set('project_id', pid);
    a.setAttribute('href', u.pathname + u.search);
  }});
  document.getElementById('tip').textContent = pid ? ('当前项目: ' + pid) : '未指定项目ID';
}}
function applyProject() {{
  const plist = document.getElementById('plist').value;
  if (plist) document.getElementById('pid').value = plist;
  const pid = getPid();
  const u = new URL(location.href);
  if (pid) {{
    u.searchParams.set('project_id', pid);
  }} else {{
    u.searchParams.delete('project_id');
  }}
  location.href = u.pathname + u.search;
}}
function exportReportTemplate() {{
  const pid = getPid();
  if (!pid) {{
    alert('请先输入或选择项目ID');
    return;
  }}
  const u = new URL('/api/template/report-blank-export', location.origin);
  u.searchParams.set('project_id', pid);
  location.href = u.pathname + u.search;
}}
const qpid = new URLSearchParams(location.search).get('project_id');
if (qpid) document.getElementById('pid').value = qpid;
document.getElementById('plist').addEventListener('change', () => {{
  const v = document.getElementById('plist').value;
  if (v) {{
    document.getElementById('pid').value = v;
    rewriteLinks();
  }}
}});
document.getElementById('pid').addEventListener('input', rewriteLinks);
rewriteLinks();
</script>
</body>
</html>"""


def render_thresholds_page() -> str:
    return """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>阈值配置</title>
  <style>
    :root { --bg:#f5f7fa; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; }
    body { margin:0; font-family:"Microsoft YaHei","PingFang SC",sans-serif; background:var(--bg); color:var(--ink); }
    .top { position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
    .btn { border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }
    .btn.primary { background:var(--acc); color:#fff; border-color:var(--acc); }
    .wrap { padding:12px 16px 24px; display:grid; grid-template-columns:320px 1fr; gap:12px; }
    .panel { background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }
    .panel h3 { margin:0; padding:10px 12px; border-bottom:1px solid var(--line); background:#fafafa; font-size:14px; }
    .body { padding:10px 12px; }
    .muted { color:var(--muted); font-size:12px; }
    .table-wrap { overflow:auto; border:1px solid var(--line); border-radius:8px; }
    table { width:100%; min-width:760px; border-collapse:collapse; }
    th, td { border-bottom:1px solid #eef2f7; padding:6px 8px; font-size:12px; text-align:left; }
    th { position:sticky; top:0; background:#f9fafb; }
    input[type=\"number\"] { width:120px; padding:4px 6px; }
    .ok { color:#065f46; font-weight:600; }
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/">返回主页</a>
    <strong>阈值配置（两级）</strong>
    <span class="muted" id="status">加载中...</span>
    <button class="btn" onclick="reloadAll()">刷新</button>
    <button class="btn primary" onclick="saveAll()">保存阈值</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>全局阈值</h3>
      <div class="body">
        <div class="muted">未设置科目阈值时，默认使用全局阈值。</div>
        <div style="margin-top:8px;">
          <label>稳定阈值（%）：</label>
          <input id="globalPct" type="number" step="0.01"/>
        </div>
      </div>
      <h3>说明</h3>
      <div class="body">
        <div class="muted">调用顺序：科目阈值优先；否则使用全局阈值。</div>
      </div>
    </div>
    <div class="panel">
      <h3>科目阈值（可选）</h3>
      <div class="body">
        <input id="kw" placeholder="搜索编码/科目名" style="width:260px;padding:6px;" oninput="renderTable()"/>
        <div class="table-wrap" style="margin-top:8px;">
          <table id="tbl"></table>
        </div>
      </div>
    </div>
  </div>
<script>
let rows = [];
function esc(s){ return (s ?? '').toString().replace(/[&<>"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;'}[c])); }
function setStatus(t, ok=false){ const el=document.getElementById('status'); el.textContent=t||''; el.className=ok?'ok':'muted'; }
function renderTable(){
  const kw = (document.getElementById('kw').value || '').trim().toLowerCase();
  const rs = !kw ? rows : rows.filter(r => String(r.code).toLowerCase().includes(kw) || String(r.name).toLowerCase().includes(kw));
  const head = '<tr><th>启用</th><th>科目编码</th><th>科目名称</th><th>阈值(%)</th></tr>';
  const body = rs.map((r, i) => `
    <tr data-code="${esc(r.code)}">
      <td><input type="checkbox" ${r.enabled?'checked':''} onchange="onToggle('${esc(r.code)}', this.checked)"/></td>
      <td>${esc(r.code)}</td>
      <td>${esc(r.name)}</td>
      <td><input type="number" step="0.01" value="${esc(r.threshold_pct)}" oninput="onValue('${esc(r.code)}', this.value)"/></td>
    </tr>`).join('');
  document.getElementById('tbl').innerHTML = head + body;
}
function onToggle(code, v){ const r=rows.find(x=>x.code===code); if(r) r.enabled=!!v; }
function onValue(code, v){ const r=rows.find(x=>x.code===code); if(r) r.threshold_pct=v; }
async function reloadAll(){
  setStatus('读取中...');
  const res = await fetch('/api/thresholds');
  const data = await res.json();
  if(!res.ok){ setStatus(data.error || '读取失败'); return; }
  document.getElementById('globalPct').value = data.global_pct ?? 2.0;
  rows = data.rows || [];
  renderTable();
  setStatus(`已加载 ${rows.length} 个科目`);
}
async function saveAll(){
  const gp = Number(document.getElementById('globalPct').value || '0');
  if(!Number.isFinite(gp)){ setStatus('全局阈值格式错误'); return; }
  const payload = { global_pct: gp, rows };
  setStatus('保存中...');
  const res = await fetch('/api/thresholds/save', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(payload) });
  const data = await res.json();
  if(!res.ok || !data.ok){ setStatus(data.error || '保存失败'); return; }
  setStatus(`保存成功：科目阈值 ${data.saved_subject_rows} 条`, true);
  await reloadAll();
}
reloadAll();
</script>
</body>
</html>"""


def render_warnings_page() -> str:
    return """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>告警清单</title>
  <style>
    :root { --bg:#f5f7fa; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; --warn:#92400e; --err:#991b1b; }
    body { margin:0; font-family: "Microsoft YaHei","PingFang SC",sans-serif; background:var(--bg); color:var(--ink); }
    .top { position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
    .btn { border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }
    .btn.primary { background:var(--acc); color:#fff; border-color:var(--acc); }
    .wrap { padding:12px 16px 24px; }
    .muted { color:var(--muted); font-size:12px; }
    .table-wrap { overflow:auto; border:1px solid var(--line); border-radius:8px; background:#fff; }
    table { width:100%; min-width:900px; border-collapse:collapse; }
    th, td { border-bottom:1px solid #eef2f7; padding:8px; font-size:12px; text-align:left; vertical-align:top; }
    th { position:sticky; top:0; background:#f9fafb; }
    .warn { color:var(--warn); font-weight:600; }
    .error { color:var(--err); font-weight:700; }
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/">返回主页</a>
    <strong>告警清单</strong>
    <span class="muted" id="status">加载中...</span>
    <button class="btn" onclick="reloadAll()">刷新</button>
    <button class="btn primary" onclick="clearWarnings()">清空清单</button>
  </div>
  <div class="wrap">
    <div class="table-wrap">
      <table id="tbl"></table>
    </div>
  </div>
<script>
let rows = [];
function esc(s){ return (s ?? '').toString().replace(/[&<>"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }
function render(){
  const head = '<tr><th>时间</th><th>级别</th><th>来源</th><th>键</th><th>信息</th></tr>';
  const body = (rows || []).map(r => `<tr>
    <td>${esc(r.ts || '')}</td>
    <td class="${esc((r.severity||'warn').toLowerCase())}">${esc(r.severity || '')}</td>
    <td>${esc(r.source || '')}</td>
    <td>${esc(r.key || '')}</td>
    <td>${esc(r.message || '')}</td>
  </tr>`).join('');
  document.getElementById('tbl').innerHTML = head + body;
  document.getElementById('status').textContent = `告警数: ${(rows || []).length}`;
}
async function reloadAll(){
  const res = await fetch('/api/warnings');
  const data = await res.json();
  rows = data.rows || [];
  render();
}
async function clearWarnings(){
  const res = await fetch('/api/warnings/clear', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'});
  const data = await res.json();
  if (data.ok) await reloadAll();
}
reloadAll();
</script>
</body>
</html>"""


def render_rules_page() -> str:
    return """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>规则配置</title>
  <style>
    :root { --bg:#f5f7fa; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; --warn:#92400e; --err:#991b1b; }
    body { margin:0; font-family: "Microsoft YaHei","PingFang SC",sans-serif; background:var(--bg); color:var(--ink); }
    .top { position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
    .btn { border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }
    .btn.primary { background:var(--acc); color:#fff; border-color:var(--acc); }
    .wrap { padding:12px 16px 24px; display:grid; grid-template-columns:320px 1fr; gap:12px; }
    .panel { background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }
    .panel h3 { margin:0; padding:10px 12px; border-bottom:1px solid var(--line); font-size:14px; background:#fafafa; }
    .body { padding:10px 12px; }
    .muted { color:var(--muted); font-size:12px; }
    .issues { max-height:200px; overflow:auto; border:1px solid #fde68a; background:#fffbeb; border-radius:8px; padding:8px; font-size:12px; }
    .issues .err { color:var(--err); }
    .issues .warn { color:var(--warn); }
    .helpbox { border:1px solid #bfdbfe; background:#eff6ff; border-radius:8px; padding:8px; font-size:12px; line-height:1.55; }
    .helpbox code { background:#fff; border:1px solid #dbeafe; border-radius:4px; padding:1px 4px; }
    .table-wrap { overflow:auto; border:1px solid var(--line); border-radius:8px; }
    table { width:100%; min-width:980px; border-collapse:collapse; }
    th, td { border-bottom:1px solid #eef2f7; padding:6px 8px; font-size:12px; text-align:left; vertical-align:top; }
    th { position:sticky; top:0; background:#f9fafb; }
    tr.active-row td { background:#ecfeff; }
    td input, td textarea { width:100%; box-sizing:border-box; border:1px solid #d1d5db; border-radius:6px; padding:4px 6px; font:inherit; }
    td textarea { min-height:40px; resize:vertical; }
    .row-actions { display:flex; gap:8px; margin-top:8px; }
    .ok { color:#065f46; font-weight:600; }
    .inspector { margin-top:10px; border:1px dashed var(--line); border-radius:8px; padding:10px; background:#fafafa; }
    .chip { display:inline-block; border:1px solid var(--line); border-radius:999px; padding:2px 8px; font-size:11px; margin-right:6px; background:#fff; }
    .chip.readonly { background:#f3f4f6; color:#4b5563; }
    .chip.rule { background:#ecfeff; color:#0f766e; border-color:#99f6e4; }
    .chip.tpl { background:#eff6ff; color:#1d4ed8; border-color:#bfdbfe; }
    .vars-table { width:100%; border-collapse:collapse; margin-top:8px; }
    .vars-table th,.vars-table td { border-bottom:1px solid #e5e7eb; padding:6px; font-size:12px; vertical-align:top; }
    .preview { white-space:pre-wrap; border:1px solid #e5e7eb; background:#fff; border-radius:6px; padding:8px; margin-top:8px; font-size:12px; }
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/">返回主页</a>
    <strong>规则配置（V1）</strong>
    <span class="muted" id="status">加载中...</span>
    <button class="btn" onclick="reloadSheet()">刷新</button>
    <button class="btn primary" onclick="saveSheet()">保存规则</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>规则集选择</h3>
      <div class="body">
        <div style="margin-bottom:8px;">
          <label class="muted">规则文件</label><br/>
          <select id="sourceSel" style="width:100%;padding:6px;"></select>
        </div>
        <div style="margin-bottom:8px;">
          <label class="muted" id="sheetLbl">Sheet</label><br/>
          <select id="sheetSel" style="width:100%;padding:6px;"></select>
        </div>
        <div id="ruleSheetBlock" style="margin-bottom:8px; display:none;">
          <label class="muted">具体规则</label><br/>
          <select id="ruleSheetSel" style="width:100%;padding:6px;"></select>
          <div class="muted" id="ruleSheetHint">选择具体规则类型后，再按指标维护规则。</div>
        </div>
        <div id="indicatorBlock" style="margin-bottom:8px; display:none;">
          <label class="muted">指标</label><br/>
          <select id="indicatorSel" style="width:100%;padding:6px;"></select>
          <div class="muted" id="indicatorHint">按指标筛选当前规则视图（整体维护时可忽略）。</div>
        </div>
        <div class="muted" id="sourcePath"></div>
        <div id="alertExprHelp" class="helpbox" style="margin-top:10px; display:none;"></div>
        <div class="row-actions">
          <button class="btn" onclick="addRow()">新增一行</button>
          <button class="btn" onclick="reloadSheet()">重新读取</button>
        </div>
      </div>
      <h3>校验结果</h3>
      <div class="body">
        <div class="issues" id="issuesBox">暂无</div>
      </div>
    </div>
    <div class="panel">
      <h3 id="tableTitle">规则内容</h3>
      <div class="body">
        <div class="table-wrap">
          <table id="tbl"></table>
        </div>
        <div class="inspector">
          <div><strong>变量来源说明</strong> <span class="muted" id="inspectorHint">点选一条规则后显示</span></div>
          <div class="muted" id="inspectorMeta" style="margin-top:6px;"></div>
          <div id="inspectorVars"></div>
          <div style="margin-top:8px;"><strong>实时预览</strong></div>
          <div class="preview" id="templatePreview">暂无</div>
        </div>
      </div>
    </div>
  </div>
<script>
let catalog = [];
let headers = [];
let allRows = [];
let rows = [];
let sourceId = '';
let sheetName = '';
let indicatorId = '__all';
let ratioIndicators = [];
let ratioOverallSheet = '';
let ratioRuleSheet = '';
let activeRow = -1;
let pendingLocate = null;
const RATIO_OVERALL_SHEETS = ['indicator_tree', 'indicator_catalog'];
const RATIO_RULE_SHEETS = ['trend_rules', 'judgement_rules', 'alert_rules', 'display_policy', 'text_templates'];
const DEFAULT_SHEET_BY_SOURCE = {
  ratio_rulebook: 'text_templates',
  profit_rulebook: 'analysis_text_templates',
  income_rulebook: 'trend_thresholds',
  key_ratio_rulebook: 'narrative_templates',
  validation_rulebook: 'bs_checks'
};
const ZH_NAME_MAP = {
  // 资产负债分析模板
  'asset_abs': '资产绝对量描述',
  'liability_abs': '负债绝对量描述',
  'asset_total_struct': '资产总计结构描述',
  'asset_subtotal_struct': '资产小计结构描述',
  'asset_item_struct': '资产一般科目结构描述',
  'liability_total_struct': '负债合计结构描述',
  'liability_subtotal_struct': '负债小计结构描述',
  'liability_item_struct': '负债一般科目结构描述',
  'scale_phrase_up': '绝对量变化词-增加',
  'scale_phrase_down': '绝对量变化词-减少',
  'scale_phrase_stable': '绝对量变化词-稳定',
  'struct_phrase_up': '结构变化词-上升',
  'struct_phrase_down': '结构变化词-下降',
  'struct_phrase_stable': '结构变化词-稳定',
  // 收入利润模板
  'income_segment_value': '收入分项金额描述',
  'income_segment_share': '收入分项占比描述',
  'gross_segment_value': '毛利分项金额描述',
  'gross_segment_impact': '毛利分项贡献描述',
  'gross_contribution_basis': '毛利贡献口径说明',
  'gross_segment_net_attr': '毛利净归因描述',
  'gross_segment_dual_view': '毛利双口径摘要',
  'gross_segment_scenario_judgement': '毛利场景定性描述',
  'profit_summary_header_missing': '利润汇总头句-缺失',
  'profit_summary_header_positive': '利润汇总头句-正值',
  'profit_summary_header_zero': '利润汇总头句-为零',
  'profit_summary_header_negative': '利润汇总头句-负值',
  'profit_summary_line_positive': '利润汇总分项句-正值',
  'profit_summary_line_negative': '利润汇总分项句-负值',
  'profit_summary_line_zero': '利润汇总分项句-零值',
  'gp_share_mode': '毛利贡献占比口径',
  'impact_ratio_label': '影响占比标签',
  'algebraic_ratio_label': '代数占比标签',
  'basis_impact_text': '影响口径说明词',
  'basis_algebraic_text': '代数口径说明词',
  // 常规指标模板
  'indicator_value': '指标数值描述',
  'indicator_trend': '指标趋势描述',
  // 重点指标模板
  'roe_headline': 'ROE总述',
  'roe_missing_years': 'ROE缺年度提示',
  'trend_word_up': '趋势词-上升',
  'trend_word_down': '趋势词-下降',
  'trend_word_stable': '趋势词-稳定',
  'roe_method_note': 'ROE分解方法说明',
  'roe_single_driver': 'ROE单驱动判断',
  'roe_dual_driver': 'ROE双驱动判断',
  'roe_multi_driver': 'ROE多驱动判断',
  'roe_contrib_detail': 'ROE三因子贡献明细',
  'roe_reconcile': 'ROE分解校验句',
  'roe_offset': 'ROE对冲项判断',
  'roe_l2_bridge': 'ROE二级驱动衔接',
  'gm_headline': '毛利率总述',
  'gm_missing_years': '毛利率缺年度提示',
  'gm_missing_segments': '毛利率缺分项提示',
  'gm_single_driver': '毛利率单驱动判断',
  'gm_dual_driver': '毛利率双驱动判断',
  'gm_multi_driver': '毛利率多驱动判断',
  'gm_structure_effect': '毛利率效应分解',
  'gm_negative_impact_summary': '毛利率负值场景补充词',
  'gm_negative_profit_case': '毛利率负值场景说明',
  'gm_segment_header': '毛利率分项拆解标题',
  'gm_segment_line': '毛利率分项拆解行',
  'gm_segment_summary': '毛利率分项拆解汇总',
  'gm_net_share_na': '毛利率净占比不适用提示',
  'gm_top_fallback': '毛利率Top缺失兜底词',
  'key_roe_factors': 'ROE三因子明细',
  'key_gm_top_segments': '毛利分项摘要',
  'sheet_auto_no_latest': '通用自动描述-缺最新值',
  'sheet_auto_one_year': '通用自动描述-单年',
  'sheet_auto_prev_missing': '通用自动描述-前值缺失',
  'sheet_auto_two_year': '通用自动描述-双年变化',
  'summary_abs_line': '汇总绝对量描述',
  'summary_ratio_line': '汇总结构占比描述',
  'summary_income_struct_intro': '汇总收入结构引导句',
  'summary_total_not_applicable': '汇总项结构不适用提示',
  'summary_struct_pending': '汇总结构待补充提示',
  'summary_yoy_up': '汇总同比词-上升',
  'summary_yoy_down': '汇总同比词-下降',
  'summary_yoy_stable': '汇总同比词-稳定',
  'summary_pp_up': '汇总占比词-上升',
  'summary_pp_down': '汇总占比词-下降',
  'summary_pp_stable': '汇总占比词-稳定',
  'summary_top_missing': 'Top构成缺失提示',
  'summary_top_line1': 'Top构成主句',
  'summary_top_line2': 'Top构成趋势句',
  'summary_top_move_default': 'Top构成默认迁移句',
  'summary_top_delta_pending': 'Top构成变化待补充句',
  'summary_see_income_page': '汇总页跳转说明',
  'summary_sum008_struct': '经营净收益结构描述',
  'summary_sum009_trend_pending': '主营毛利结构待补充句',
  'summary_sum009_struct': '主营毛利结构描述',
  'summary_sum010_struct': '经常性净收益结构描述',
  'summary_sum011_struct': '非经常性净收益结构描述',
  // 阈值键
  'significant_abs_contrib': '显著贡献阈值',
  'single_driver_share': '单驱动占比阈值',
  'dual_driver_share_sum': '双驱动合计阈值',
  'dual_driver_each_min': '双驱动单项最小占比',
  'delta_stable_pp': '稳定变动阈值(百分点)',
  // 常见 scene
  'summary': '总述',
  'single_driver': '单驱动',
  'dual_driver': '双驱动',
  'multi_driver': '多驱动',
  'offset': '对冲项',
  'secondary_bridge': '二级驱动衔接',
  'structure_effect': '效应分解',
  'negative_case': '负值场景'
};
const SOURCE_ZH_MAP = {
  'rulebook_main': '资产负债分析规则',
  'profit_rulebook': '利润表规则',
  'ratio_rulebook': '财务指标规则',
  'income_rulebook': '收入利润规则',
  'key_ratio_rulebook': '重点指标规则',
  'validation_rulebook': '报表校验规则'
};
const SHEET_ZH_MAP = {
  // income profit rulebook
  'scope_catalog': '范围目录',
  'metric_definitions': '指标定义',
  'trend_thresholds': '趋势阈值',
  'sign_scenario_policy': '正负场景策略',
  'contribution_policy': '贡献口径策略',
  'display_fields': '展示字段',
  'text_templates': '文本模板',
  'node_mapping': '节点映射',
  'manual_override': '人工覆盖策略',
  'income_special_items': '收入特殊项映射',
  'income_grouping': '收入分组',
  'income_tree': '收入分析树',
  'income_formulas': '收入汇总公式',
  // other common sheets
  'analysis_text_templates': '分析文本模板',
  'analysis_thresholds': '分析阈值配置',
  'bs_checks': '资产负债校验规则',
  'is_checks': '利润表校验规则',
  'cf_checks': '现金流量表校验规则',
  'display_policy': '展示策略',
  'indicator_tree': '指标树',
  'indicator_catalog': '指标目录',
  'trend_rules': '趋势规则',
  'judgement_rules': '判断词规则',
  'alert_rules': '预警规则',
  'driver_thresholds': '驱动阈值',
  'narrative_templates': '叙述模板'
};
const FIELD_ZH_MAP = {
  'tpl_id': '模板ID',
  'template_id': '模板ID',
  'scene': '场景',
  'template_key': '模板键',
  'template_text': '模板文本',
  'template_text_zh': '模板文本',
  'variables': '变量清单',
  'placeholders': '占位符声明',
  'enabled': '启用',
  'notes': '备注',
  'description': '说明',
  'rule_id': '规则ID',
  'rule_name': '规则名称',
  'alert_id': '预警ID',
  'indicator_id': '指标ID',
  'indicator_name_zh': '指标中文名',
  'group_zh': '分组',
  'node_id': '节点ID',
  'parent_id': '父节点ID',
  'label_zh': '中文标签',
  'node_type': '节点类型',
  'formula_expr': '公式表达式',
  'value_source': '取值来源',
  'warn_threshold': '预警阈值',
  'critical_threshold': '严重阈值',
  'good_label': '改善表述词',
  'bad_label': '弱化表述词',
  'alert_text_zh': '预警文本',
  'template_text': '模板文本',
  'left_code': '左值编码',
  'operator': '运算符',
  'right_codes': '右值编码集',
  'tolerance': '容差',
  'severity': '严重级别',
  'stable_threshold': '稳定阈值',
  'stable_threshold_pct': '稳定阈值(%)',
  'scale_stable_pct': '绝对量稳定阈值(%)',
  'struct_stable_pp': '结构稳定阈值(百分点)',
  'threshold_unit': '阈值单位',
  'threshold_key': '阈值键',
  'threshold_value': '阈值值',
  'up_label': '上行表述词',
  'down_label': '下行表述词',
  'stable_label': '稳定表述词',
  'direction_label_pos': '正向表述词',
  'direction_label_neg': '负向表述词',
  'metric_id': '指标ID',
  'metric_name_zh': '指标名称',
  'rule_id': '规则ID',
  'policy_id': '策略ID',
  'condition_expr': '条件表达式',
  'primary_ratio_metric': '主口径指标',
  'scope': '范围',
  'subject_code': '科目编码',
  'subject_name': '科目名称',
  'key': '键',
  'value': '值',
  'field': '字段',
  'desc_zh': '中文说明',
  'sort_order': '排序',
  'unit': '单位',
  'direction': '方向'
};
const METRIC_ZH_MAP = {
  'rev_yoy': '收入同比变动',
  'rev_share': '收入占比变动',
  'gp_value': '毛利金额变动',
  'gp_share_impact': '毛利贡献占比变动',
  'gp_share_algebraic': '毛利代数占比'
};
const RULE_ZH_MAP = {
  'TR_REV_YOY': '收入同比趋势规则',
  'TR_REV_SHARE': '收入占比趋势规则',
  'TR_GP_YOY': '毛利金额趋势规则',
  'TR_GP_SHARE_IMPACT': '毛利贡献趋势规则'
};
const POLICY_ZH_MAP = {
  'SCN1': '同号且整体盈利场景',
  'SCN2': '异号且整体盈利场景',
  'SCN3': '同号且整体亏损场景',
  'SCN4': '异号且整体亏损场景',
  'SCN5': '整体为零场景'
};
const READONLY_VARS = new Set([
  'name','label','title','base_name','latest_year','latest','prev',
  'y1','y2','y3','v1','v2','v3','r1','r2','r3','i1','i2','i3','n1','n2','n3',
  'nm1','nm2','nm3','at1','at2','at3','em1','em2','em3',
  'latest_val','total_amount','total_abs_amount','amount','abs_amount',
  'top_txt','trend_txt','gp_s3','ot_s3','d_gp','d_ot','d_rec','d_non'
]);
const RULE_TUNE_VARS = new Set([
  'unit','unit2','judgement21','judgement32','delta21','delta32','trend_word',
  't21','t32','rt21','rt32','it21','it32','nt21','nt32','p21','p32','d21','d32'
]);

function esc(s){ return (s ?? '').toString().replace(/[&<>"]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c])); }
function setStatus(t){ document.getElementById('status').textContent = t || ''; }
function currentSource(){ return catalog.find(x => x.source_id === sourceId) || null; }
function sourceLabelById(id){
  const src = catalog.find(x => x.source_id === id);
  if (!src) return id || '';
  return SOURCE_ZH_MAP[id] || src.label || id;
}
function sheetLabelByName(name){
  return SHEET_ZH_MAP[name] || name || '';
}
function keyField(){
  const candidates = ['template_key','template_id','scene','threshold_key','metric_id','key'];
  return candidates.find(k => headers.includes(k)) || '';
}
function isRatioSource(){
  return sourceId === 'ratio_rulebook';
}
function isRatioRuleSheet(name){
  return RATIO_RULE_SHEETS.includes(String(name || '').trim());
}
function activeSheetName(){
  if (!isRatioSource()) return sheetName;
  if (ratioRuleSheet) return ratioRuleSheet;
  return ratioOverallSheet || sheetName;
}
function renderAlertExprHelp(){
  const box = document.getElementById('alertExprHelp');
  if (!box) return;
  const show = isRatioSource() && activeSheetName() === 'alert_rules';
  if (!show){
    box.style.display = 'none';
    box.innerHTML = '';
    return;
  }
  box.style.display = '';
  box.innerHTML = `
    <div><strong>预警条件表达式说明（支持2年/3年自适应）</strong></div>
    <div class="muted" style="margin-top:4px;">变量口径：<code>curr</code>=最新期，<code>prev</code>=上一期，<code>base</code>=首期。</div>
    <div style="margin-top:6px;"><strong>推荐表达式</strong></div>
    <div><code>down_last</code> / <code>up_last</code>：最新期较上一期下降/上升</div>
    <div><code>down_last(2)</code> / <code>up_last(2)</code>：最新期较上一期变动幅度达到阈值</div>
    <div><code>delta_last_abs(3)</code>：最新期较上一期绝对变动值达到阈值</div>
    <div><code>trend_last(down,2)</code>：最新期趋势判断（up/down/stable）+ 阈值</div>
    <div style="margin-top:6px;"><strong>兼容表达式</strong></div>
    <div><code>down_2y</code> / <code>up_2y</code>：三年按原逻辑；两年自动降级为 last 口径</div>
    <div><code>down_2y(2)</code> / <code>up_2y(2)</code>：支持阈值，且两年自动降级</div>
    <div style="margin-top:6px;"><strong>比较表达式</strong></div>
    <div><code>curr &lt; 6</code>、<code>prev &gt;= 8</code>、<code>base == 10</code></div>
  `;
}
function indicatorApplicable(){
  return isRatioSource() && isRatioRuleSheet(activeSheetName());
}
function rowIndicatorId(row){
  const iid = String(row?.indicator_id ?? '').trim();
  if (iid) return iid;
  const scope = String(row?.scope ?? '').trim();
  if (scope && !['all','*','turnover','summary'].includes(scope.toLowerCase())) return scope;
  const mid = String(row?.metric_id ?? '').trim();
  if (mid) return mid;
  return '';
}
function indicatorMatchesRow(row, id){
  const sid = String(id || '').trim().toLowerCase();
  if (!sid || sid === '__all') return true;
  const rid = rowIndicatorId(row).toLowerCase();
  if (rid && rid === sid) return true;
  const scope = String(row?.scope ?? '').trim().toLowerCase();
  if (!scope || ['all','*','turnover','summary'].includes(scope)) return true;
  const tid = String(row?.template_id ?? '').trim().toLowerCase();
  if (tid.startsWith('trend_word_')) return true;
  return false;
}
function rebuildRows(){
  const withIndex = (allRows || []).map((r, i) => ({...r, __src_i: i}));
  if (!indicatorApplicable()){
    rows = withIndex;
    if (activeRow >= rows.length) activeRow = rows.length ? 0 : -1;
    return;
  }
  rows = withIndex.filter(r => indicatorMatchesRow(r, indicatorId));
  if (activeRow >= rows.length) activeRow = rows.length ? 0 : -1;
}
function displayHeaders(){
  const out = [...headers];
  const kf = keyField();
  if (!kf) return out;
  const idx = out.indexOf(kf);
  if (idx < 0) return out;
  out.splice(idx + 1, 0, '__zh_name');
  return out;
}
function zhNameForRow(row){
  const kf = keyField();
  const keyVal = (row && kf) ? String(row[kf] ?? '').trim() : '';
  if (keyVal && ZH_NAME_MAP[keyVal]) return ZH_NAME_MAP[keyVal];
  const rid = String(row?.rule_id ?? '').trim();
  if (rid && RULE_ZH_MAP[rid]) return RULE_ZH_MAP[rid];
  const mid = String(row?.metric_id ?? '').trim();
  if (mid && METRIC_ZH_MAP[mid]) return METRIC_ZH_MAP[mid];
  const pid = String(row?.policy_id ?? '').trim();
  if (pid && POLICY_ZH_MAP[pid]) return POLICY_ZH_MAP[pid];
  const scn = String(row?.scenario ?? '').trim();
  if (scn) return scn;
  // fallback: notes/description 可作为辅助中文说明
  const note = String((row?.notes ?? row?.description ?? '') || '').trim();
  return note || '';
}
function renderIssues(issues){
  const box = document.getElementById('issuesBox');
  if (!issues || !issues.length){ box.innerHTML = '暂无'; return; }
  box.innerHTML = issues.map(x => `<div class="${x.level==='error'?'err':'warn'}">[${esc(x.level)}] 第${esc(x.row)}行 ${esc(x.field)}：${esc(x.message)}</div>`).join('');
}
function parseVarsFromTemplate(tpl){
  const text = (tpl || '').toString();
  const ms = [...text.matchAll(/\\{([A-Za-z0-9_]+)\\}/g)].map(m => m[1]);
  return [...new Set(ms)];
}
function templateTextOfRow(row){
  if (!row) return '';
  const keys = ['template_text_zh','template_text','template','text','narrative_text'];
  for (const k of keys){
    const v = (row[k] ?? '').toString().trim();
    if (v) return v;
  }
  return '';
}
function variableTag(v){
  if (READONLY_VARS.has(v)) return {cls:'readonly', text:'只读(数据值)'};
  if (RULE_TUNE_VARS.has(v)) return {cls:'rule', text:'规则可调'};
  return {cls:'tpl', text:'模板可调'};
}
function inferVarSource(v, row){
  const lc = (v || '').toLowerCase();
  const currIndicator = rowIndicatorId(row) || indicatorId || '';
  const currKeyField = keyField();
  const currKeyValue = currKeyField ? String((row || {})[currKeyField] ?? '').trim() : '';
  const locate = {
    indicator_id: currIndicator && currIndicator !== '__all' ? currIndicator : '',
    key_field: currKeyField,
    key_value: currKeyValue,
  };
  if (lc === 'unit') {
    let targetSheet = 'analysis_text_templates';
    if (sourceId === 'income_rulebook') targetSheet = 'text_templates';
    if (sourceId === 'ratio_rulebook') targetSheet = 'text_templates';
    if (sourceId === 'key_ratio_rulebook') targetSheet = 'narrative_templates';
    return {label:'单位', field_hint:'unit', indicator_id: currIndicator, jump:{source_id:sourceId, sheet:targetSheet, locate}};
  }
  if (lc.startsWith('judgement') || lc.startsWith('delta') || lc === 'unit2' || lc === 'trend_word') {
    if (sourceId === 'ratio_rulebook' || sourceId === 'key_ratio_rulebook') {
      const fieldHint = lc.startsWith('judgement') ? 'good_label/bad_label/stable_label' : (lc.startsWith('delta') || lc==='unit2' ? 'delta_expr/threshold_unit' : 'template_text');
      return {label:'趋势/判断规则', field_hint:fieldHint, indicator_id: currIndicator, jump:{source_id:'ratio_rulebook', sheet:'judgement_rules', locate}};
    }
    return {label:'趋势阈值规则', field_hint:'up_label/down_label/stable_label/stable_threshold/threshold_unit', indicator_id: currIndicator, jump:{source_id:'income_rulebook', sheet:'trend_thresholds', locate}};
  }
  if (['t21','t32','rt21','rt32','it21','it32','nt21','nt32','p21','p32','d21','d32'].includes(lc)){
    if (sourceId === 'income_rulebook') return {label:'趋势阈值规则', field_hint:'up_label/down_label/stable_label/stable_threshold/threshold_unit', indicator_id: currIndicator, jump:{source_id:'income_rulebook', sheet:'trend_thresholds', locate}};
    if (sourceId === 'rulebook_main') return {label:'分析阈值配置', field_hint:'scale_stable_pct/struct_stable_pp', indicator_id: currIndicator, jump:{source_id:'rulebook_main', sheet:'analysis_thresholds', locate}};
  }
  return {label:'模板内占位变量', field_hint:'template_text/template_text_zh', indicator_id: currIndicator, jump:null};
}
async function jumpToRule(target){
  if (!target || !target.source_id || !target.sheet) return;
  pendingLocate = target.locate || null;
  sourceId = target.source_id;
  if (isRatioSource()){
    ratioIndicators = await loadRatioIndicators();
  } else {
    ratioIndicators = [];
    indicatorId = '__all';
  }
  renderSourceSheetOptions();
  if (sourceId === 'ratio_rulebook'){
    if (RATIO_OVERALL_SHEETS.includes(target.sheet)){
      ratioOverallSheet = target.sheet;
      ratioRuleSheet = '';
    } else {
      ratioRuleSheet = target.sheet;
      ratioOverallSheet = '';
    }
  } else {
    sheetName = target.sheet;
  }
  renderSourceSheetOptions();
  renderIndicatorOptions();
  await reloadSheet();
}
function applyPendingLocate(){
  if (!pendingLocate) return;
  const loc = pendingLocate;
  if (loc.indicator_id && isRatioSource() && indicatorApplicable()){
    indicatorId = String(loc.indicator_id);
    const sel = document.getElementById('indicatorSel');
    if (sel) sel.value = indicatorId;
    rebuildRows();
  }
  let idx = -1;
  if (loc.key_field && loc.key_value){
    idx = rows.findIndex(r => String(r?.[loc.key_field] ?? '').trim() === String(loc.key_value));
  }
  if (idx < 0 && loc.indicator_id){
    idx = rows.findIndex(r => rowIndicatorId(r) === String(loc.indicator_id));
  }
  activeRow = idx >= 0 ? idx : (rows.length ? 0 : -1);
  pendingLocate = null;
}
function renderInspector(){
  const hint = document.getElementById('inspectorHint');
  const meta = document.getElementById('inspectorMeta');
  const varsBox = document.getElementById('inspectorVars');
  const preview = document.getElementById('templatePreview');
  if (!rows.length || activeRow < 0 || activeRow >= rows.length){
    hint.textContent = '点选一条规则后显示';
    meta.textContent = '';
    varsBox.innerHTML = '';
    preview.textContent = '暂无';
    return;
  }
  const row = rows[activeRow];
  const tpl = templateTextOfRow(row);
  const kf = keyField();
  const keyVal = kf ? String(row[kf] ?? '') : '';
  hint.textContent = '可点“跳转编辑”直达对应规则';
  meta.textContent = `当前规则：${keyVal || '未命名'} | 场景：${String(row.scene || row.template_key || '').trim() || '未设置'}`;
  const vars = parseVarsFromTemplate(tpl);
  if (!tpl){
    varsBox.innerHTML = '<div class="muted" style="margin-top:6px;">当前行不是文本模板行，暂无变量。</div>';
    preview.textContent = '暂无模板文本';
    return;
  }
  const sample = {};
  vars.forEach(v => sample[v] = `‹${v}›`);
  preview.textContent = tpl.replace(/\\{([A-Za-z0-9_]+)\\}/g, (_,k)=> (sample[k] ?? `{${k}}`));
  const rowsHtml = vars.map(v => {
    const tag = variableTag(v);
    const src = inferVarSource(v, row);
    const btn = src.jump ? `<button class="btn" style="padding:2px 8px;" onclick='jumpToRule(${JSON.stringify(src.jump)})'>跳转编辑</button>` : '';
    return `<tr>
      <td><code>{${esc(v)}}</code></td>
      <td><span class="chip ${esc(tag.cls)}">${esc(tag.text)}</span></td>
      <td>${esc(src.label || '')}</td>
      <td>${esc(src.field_hint || '')}</td>
      <td>${esc(src.indicator_id || '')}</td>
      <td>${btn}</td>
    </tr>`;
  }).join('');
  varsBox.innerHTML = vars.length
    ? `<table class="vars-table"><tr><th>变量</th><th>可改性</th><th>来源</th><th>字段</th><th>当前指标ID</th><th>操作</th></tr>${rowsHtml}</table>`
    : '<div class="muted" style="margin-top:6px;">未检测到占位变量。</div>';
}
function asInput(k, v){
  const t = (v ?? '').toString();
  const isLong = /text|template|notes|description/i.test(k) || t.length > 60;
  if (isLong) {
    return `<textarea data-key="${esc(k)}" onclick="event.stopPropagation()" onmousedown="event.stopPropagation()" onmouseup="event.stopPropagation()">${esc(t)}</textarea>`;
  }
  return `<input data-key="${esc(k)}" value="${esc(t)}" onclick="event.stopPropagation()" onmousedown="event.stopPropagation()" onmouseup="event.stopPropagation()"/>`;
}
function renderTable(){
  const title = `${sourceLabelById(sourceId)} / ${sheetLabelByName(activeSheetName())}`;
  document.getElementById('tableTitle').textContent = title;
  const dHeaders = displayHeaders();
  const th = `<tr>${dHeaders.map(h => {
    if (h === '__zh_name') return '<th>中文名称</th>';
    const label = FIELD_ZH_MAP[h] || h;
    return `<th>${esc(label)}</th>`;
  }).join('')}</tr>`;
  const trs = rows.map((r, i) => {
    const cls = i === activeRow ? ' class="active-row"' : '';
    const tds = dHeaders.map(h => {
      if (h === '__zh_name') return `<td class="muted">${esc(zhNameForRow(r))}</td>`;
      return `<td>${asInput(h, r[h])}</td>`;
    }).join('');
    return `<tr${cls} data-row="${i}" data-src="${esc(r.__src_i)}" onclick="setActiveRow(${i})">${tds}</tr>`;
  }).join('');
  document.getElementById('tbl').innerHTML = th + trs;
  renderInspector();
}
function setActiveRow(i){
  const next = Number(i);
  if (next === activeRow) return;
  activeRow = next;
  renderTable();
}
function collectRows(){
  const out = [];
  document.querySelectorAll('#tbl tr[data-row]').forEach(tr => {
    const row = {};
    row.__src_i = Number(tr.dataset.src ?? -1);
    tr.querySelectorAll('[data-key]').forEach(el => {
      const k = el.getAttribute('data-key');
      row[k] = (el.value ?? '').toString();
    });
    out.push(row);
  });
  return out;
}
async function loadRatioIndicators(){
  if (!isRatioSource()) return [];
  const res = await fetch('/api/rules?source_id=ratio_rulebook&sheet=indicator_catalog');
  const data = await res.json();
  if (!res.ok) return [];
  const rs = Array.isArray(data.rows) ? data.rows : [];
  const out = rs.map(r => ({
    id: String(r.indicator_id ?? '').trim(),
    name: String(r.indicator_name_zh ?? '').trim()
  })).filter(x => x.id);
  out.sort((a,b) => a.id.localeCompare(b.id));
  return out;
}
function renderIndicatorOptions(){
  const block = document.getElementById('indicatorBlock');
  const sel = document.getElementById('indicatorSel');
  if (!block || !sel) return;
  if (!isRatioSource()){
    block.style.display = 'none';
    indicatorId = '__all';
    return;
  }
  block.style.display = '';
  sel.disabled = !indicatorApplicable();
  const opts = [{id:'__all', name:'全部指标'}, ...(ratioIndicators || [])];
  sel.innerHTML = opts.map(x => `<option value="${esc(x.id)}">${esc(x.id === '__all' ? x.name : (x.id + ' ' + (x.name || '')))}</option>`).join('');
  if (!opts.find(x => x.id === indicatorId)) indicatorId = '__all';
  sel.value = indicatorId;
}
function renderSourceSheetOptions(){
  const srcSel = document.getElementById('sourceSel');
  srcSel.innerHTML = catalog.map(x => {
    const label = sourceLabelById(x.source_id);
    return `<option value="${esc(x.source_id)}">${esc(label)}${x.exists?'':'（文件不存在）'}</option>`;
  }).join('');
  if (!sourceId && catalog.length) sourceId = catalog[0].source_id;
  srcSel.value = sourceId;
  const src = currentSource();
  const sheetSel = document.getElementById('sheetSel');
  const sheetLbl = document.getElementById('sheetLbl');
  const ruleBlock = document.getElementById('ruleSheetBlock');
  const ruleSel = document.getElementById('ruleSheetSel');
  const sheets = src ? (src.sheets || []) : [];
  if (isRatioSource()){
    const overall = sheets.filter(s => RATIO_OVERALL_SHEETS.includes(s));
    const rules = sheets.filter(s => RATIO_RULE_SHEETS.includes(s));
    if (sheetLbl) sheetLbl.textContent = '整体';
    sheetSel.innerHTML = ['<option value="">（不选择）</option>', ...overall.map(s => `<option value="${esc(s)}">${esc(sheetLabelByName(s))}</option>`)].join('');
    if (!overall.includes(ratioOverallSheet)) ratioOverallSheet = overall.includes('indicator_catalog') ? 'indicator_catalog' : (overall[0] || '');
    if (!ratioRuleSheet) {
      // Default to overall maintenance on first load.
      sheetSel.value = ratioOverallSheet || '';
    }
    if (ruleBlock) ruleBlock.style.display = '';
    if (ruleSel){
      ruleSel.innerHTML = ['<option value="">（不选择）</option>', ...rules.map(s => `<option value="${esc(s)}">${esc(sheetLabelByName(s))}</option>`)].join('');
      if (!rules.includes(ratioRuleSheet)) ratioRuleSheet = '';
      ruleSel.value = ratioRuleSheet || '';
    }
    sheetSel.value = ratioOverallSheet || '';
    sheetName = activeSheetName() || ratioOverallSheet || ratioRuleSheet || '';
  } else {
    if (sheetLbl) sheetLbl.textContent = 'Sheet';
    sheetSel.innerHTML = sheets.map(s => `<option value="${esc(s)}">${esc(sheetLabelByName(s))}</option>`).join('');
    if ((!sheetName || !sheets.includes(sheetName)) && sheets.length){
      const prefer = DEFAULT_SHEET_BY_SOURCE[sourceId];
      sheetName = (prefer && sheets.includes(prefer)) ? prefer : sheets[0];
    }
    sheetSel.value = sheetName;
    ratioOverallSheet = '';
    ratioRuleSheet = '';
    if (ruleBlock) ruleBlock.style.display = 'none';
    if (ruleSel) ruleSel.innerHTML = '';
  }
  document.getElementById('sourcePath').textContent = src ? ('规则已接入配置中心（无需手动改Excel路径）') : '';
  renderAlertExprHelp();
}
async function loadCatalog(){
  const res = await fetch('/api/rules/catalog');
  const data = await res.json();
  catalog = data.sources || [];
  renderSourceSheetOptions();
}
async function reloadSheet(){
  const currentSheet = activeSheetName();
  if (!sourceId || !currentSheet) return;
  setStatus('读取中...');
  const u = `/api/rules?source_id=${encodeURIComponent(sourceId)}&sheet=${encodeURIComponent(currentSheet)}`;
  const res = await fetch(u);
  const data = await res.json();
  if (!res.ok){ setStatus(data.error || '读取失败'); return; }
  headers = data.headers || [];
  allRows = data.rows || [];
  rebuildRows();
  applyPendingLocate();
  if (activeRow < 0 && rows.length) activeRow = 0;
  renderTable();
  renderIssues([]);
  setStatus(`已加载 ${rows.length} 行（总${allRows.length}行）`);
}
function addRow(){
  if (!headers.length) return;
  const r = {};
  headers.forEach(h => r[h] = '');
  if (indicatorApplicable() && indicatorId && indicatorId !== '__all'){
    if (headers.includes('indicator_id')) r['indicator_id'] = indicatorId;
    if (headers.includes('scope') && !r['scope']) r['scope'] = indicatorId;
  }
  allRows.push(r);
  rebuildRows();
  activeRow = rows.length ? (rows.length - 1) : -1;
  renderTable();
}
async function saveSheet(){
  const currentSheet = activeSheetName();
  if (!sourceId || !currentSheet) return;
  const edits = collectRows();
  edits.forEach(e => {
    const idx = Number(e.__src_i ?? -1);
    delete e.__src_i;
    if (Number.isInteger(idx) && idx >= 0 && idx < allRows.length) allRows[idx] = e;
    else allRows.push(e);
  });
  const payload = { source_id: sourceId, sheet_name: currentSheet, headers, rows: allRows };
  setStatus('保存中...');
  const res = await fetch('/api/rules/save', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify(payload)
  });
  const data = await res.json();
  renderIssues(data.issues || []);
  if (!res.ok || !data.ok){
    setStatus(data.error || '保存失败');
    return;
  }
  setStatus(`保存成功：${data.saved_rows} 行`);
  await reloadSheet();
}
document.getElementById('sourceSel').addEventListener('change', async (e) => {
  sourceId = e.target.value;
  sheetName = '';
  if (isRatioSource()){
    ratioIndicators = await loadRatioIndicators();
    ratioOverallSheet = '';
    ratioRuleSheet = '';
  } else {
    ratioIndicators = [];
    indicatorId = '__all';
  }
  renderSourceSheetOptions();
  renderIndicatorOptions();
  renderAlertExprHelp();
  await reloadSheet();
});
document.getElementById('sheetSel').addEventListener('change', async (e) => {
  if (isRatioSource()){
    ratioOverallSheet = String(e.target.value || '');
    if (ratioOverallSheet){
      ratioRuleSheet = '';
      const rs = document.getElementById('ruleSheetSel');
      if (rs) rs.value = '';
    }
    sheetName = activeSheetName();
    renderIndicatorOptions();
  } else {
    sheetName = e.target.value;
  }
  renderAlertExprHelp();
  await reloadSheet();
});
document.getElementById('ruleSheetSel').addEventListener('change', async (e) => {
  ratioRuleSheet = String(e.target.value || '');
  if (ratioRuleSheet){
    ratioOverallSheet = '';
    const ss = document.getElementById('sheetSel');
    if (ss) ss.value = '';
  }
  sheetName = activeSheetName();
  renderIndicatorOptions();
  renderAlertExprHelp();
  await reloadSheet();
});
document.getElementById('indicatorSel').addEventListener('change', async (e) => {
  indicatorId = String(e.target.value || '__all');
  rebuildRows();
  activeRow = rows.length ? 0 : -1;
  renderTable();
  setStatus(`已筛选 ${rows.length} 行（总${allRows.length}行）`);
});
(async () => {
  await loadCatalog();
  if (isRatioSource()) ratioIndicators = await loadRatioIndicators();
  renderIndicatorOptions();
  renderAlertExprHelp();
  await reloadSheet();
})();
</script>
</body>
</html>"""


def render_analysis_page(project_id: str, title: str, api_path: str, save_group_id: str, subject_label: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{title}</title>
  <style>
    :root {{ --bg:#f5f7fa; --card:#fff; --ink:#111827; --line:#d1d5db; --acc:#0f766e; --muted:#6b7280; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .wrap {{ display:grid; grid-template-columns:280px 1fr; gap:12px; padding:12px 16px 20px; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
    .panel h3 {{ margin:0; font-size:14px; padding:10px 12px; border-bottom:1px solid var(--line); background:#fafafa; }}
    .list {{ max-height:calc(100vh - 130px); overflow:auto; }}
    .item {{ padding:8px 10px; border-bottom:1px solid #eef2f7; cursor:pointer; }}
    .item:hover {{ background:#f8fafc; }}
    .item.active {{ background:#e6fffa; border-left:3px solid var(--acc); }}
    .cards {{ display:grid; grid-template-columns:1fr; gap:12px; }}
    .table-wrap {{ overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; }}
    th,td {{ font-size:12px; border-bottom:1px solid #eef2f7; padding:6px 8px; text-align:left; }}
    td.num {{ text-align:right; white-space:nowrap; }}
    .desc {{ white-space:pre-wrap; font-size:13px; line-height:1.5; padding:10px 12px; }}
    textarea {{ width:100%; min-height:170px; box-sizing:border-box; border:1px solid #d1d5db; border-radius:8px; padding:10px; font:inherit; }}
    .toolbar {{ padding:10px 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; border-top:1px solid #eef2f7; }}
    .ok {{ color:#065f46; font-weight:600; }}
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/?project_id={project_id}">返回主页</a>
    <strong>{title}</strong>
    <span class="muted">项目ID: {project_id}</span>
    <span class="muted" id="status"></span>
    <button class="btn primary" onclick="reloadAll()">刷新数据</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>{subject_label}列表</h3>
      <div class="list" id="subjectList"></div>
    </div>
    <div class="cards">
      <div class="panel">
        <h3>自动分析（只读）</h3>
        <div class="table-wrap"><table id="scaleTbl"></table></div>
        <div class="desc" id="scaleDesc"></div>
        <div class="desc" id="structDesc"></div>
      </div>
      <div class="panel">
        <h3>分析判断内容（可编辑）</h3>
        <div class="desc"><textarea id="manualText"></textarea></div>
        <div class="toolbar">
          <label><input type="checkbox" id="manualConfirmed"> 确认</label>
          <button class="btn primary" onclick="saveActive()">保存并确认</button>
          <span id="saveHint" class="muted"></span>
        </div>
      </div>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
const API_PATH = {json.dumps(api_path)};
const SAVE_GROUP_ID = {json.dumps(save_group_id)};
let cache = null;
let activeCode = '';

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}

function rowByCode(rows, code) {{
  return (rows || []).find(r => String((r['科目编码'] ?? r['code'] ?? '')) === code) || null;
}}

function numericCell(v) {{
  if (typeof v === 'number') return `<td class="num">${{v.toLocaleString(undefined, {{maximumFractionDigits: 4}})}}</td>`;
  return `<td>${{esc(v)}}</td>`;
}}

function normalizeNarrativeText(s) {{
  let t = (s ?? '').toString();
  if (!t) return t;
  // Normalize legacy punctuation artifacts from old generated sheets.
  t = t.replaceAll('：，', '，').replaceAll(':,', '，');
  // Also normalize accidental duplicated punctuation.
  t = t.replaceAll('，，', '，').replaceAll('：:', '：');
  return t;
}}

function normalizeSubjectName(s) {{
  return (s ?? '').toString().replace(/[：:]+$/,'').trim();
}}

function renderRowTable(elId, rowObj) {{
  const el = document.getElementById(elId);
  if (!rowObj) {{
    el.innerHTML = '<tr><td>未找到该科目数据</td></tr>';
    return;
  }}
  const row = {{...rowObj}};
  if (Object.prototype.hasOwnProperty.call(row, '科目名称')) {{
    row['科目名称'] = normalizeSubjectName(row['科目名称']);
  }}
  const keys = Object.keys(row);
  const head = `<tr>${{keys.map(k => `<th>${{esc(k)}}</th>`).join('')}}</tr>`;
  const body = `<tr>${{keys.map(k => numericCell(row[k])).join('')}}</tr>`;
  el.innerHTML = head + body;
}}

function renderSubjectList() {{
  const rows = cache.scale.rows || [];
  const html = rows.map(r => {{
    const c = String(r['科目编码'] || '');
    const n = String(r['科目名称'] || '').replace(/[：:]+$/,'').trim();
    const active = c === activeCode ? 'active' : '';
    return `<div class="item ${{active}}" data-code="${{esc(c)}}"><strong>${{esc(c)}}</strong> ${{esc(n)}}</div>`;
  }}).join('');
  const box = document.getElementById('subjectList');
  box.innerHTML = html || '<div class="item">暂无数据</div>';
  box.querySelectorAll('.item[data-code]').forEach(node => {{
    node.addEventListener('click', () => {{
      activeCode = node.dataset.code;
      renderAll();
    }});
  }});
}}

function renderAll() {{
  renderSubjectList();
  const scaleRow = rowByCode(cache.scale.rows, activeCode);
  const structRow = rowByCode(cache.structure.rows, activeCode);
  const analysisRow = rowByCode(cache.analysis.rows, activeCode) || {{}};
  renderRowTable('scaleTbl', scaleRow);
  document.getElementById('scaleDesc').textContent = normalizeNarrativeText(scaleRow?.['定量描述_绝对量'] || '');
  document.getElementById('structDesc').textContent = normalizeNarrativeText(structRow?.['定量描述_相对量'] || '');
  document.getElementById('manualText').value = analysisRow['manual_text'] || analysisRow['auto_combined'] || '';
  document.getElementById('manualConfirmed').checked = !!analysisRow['confirmed'];
  document.getElementById('saveHint').textContent = analysisRow['confirmed'] ? '当前已确认' : '';
  document.getElementById('status').textContent = `规模表行数:${{cache.scale.rows.length}} | 结构表行数:${{cache.structure.rows.length}}`;
}}

async function saveActive() {{
  if (!activeCode) return;
  const scaleRow = rowByCode(cache.scale.rows, activeCode) || {{}};
  const body = {{
    project_id: PROJECT_ID,
    group_id: SAVE_GROUP_ID,
    rows: [{{
      code: activeCode,
      name: String(scaleRow['科目名称'] || ''),
      manual_text: document.getElementById('manualText').value,
      confirmed: document.getElementById('manualConfirmed').checked
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify(body)
  }});
  const data = await res.json();
  if (data.ok) {{
    document.getElementById('saveHint').textContent = '已保存';
    await reloadAll();
  }} else {{
    document.getElementById('saveHint').textContent = '保存失败';
  }}
}}

async function reloadAll() {{
  const res = await fetch(`${{API_PATH}}?project_id=${{encodeURIComponent(PROJECT_ID)}}`);
  const data = await res.json();
  if (!res.ok) {{
    document.getElementById('status').textContent = data.error || '读取失败';
    return;
  }}
  cache = data;
  if (!activeCode && cache.scale.rows.length) {{
    activeCode = String(cache.scale.rows[0]['科目编码'] || '');
  }}
  renderAll();
}}

reloadAll();
</script>
</body>
</html>"""


def render_asset_analysis_page(project_id: str) -> str:
    return render_analysis_page(
        project_id=project_id,
        title="资产分析",
        api_path="/api/analysis/assets",
        save_group_id="asset_analysis",
        subject_label="资产科目",
    )


def render_liability_analysis_page(project_id: str) -> str:
    return render_analysis_page(
        project_id=project_id,
        title="负债分析",
        api_path="/api/analysis/liabilities",
        save_group_id="liability_analysis",
        subject_label="负债科目",
    )


def render_summary_analysis_page(project_id: str) -> str:
    return render_analysis_page(
        project_id=project_id,
        title="分析汇总",
        api_path="/api/analysis/summary",
        save_group_id="summary_analysis",
        subject_label="汇总科目",
    )


def render_income_analysis_page(project_id: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>收入分析</title>
  <style>
    :root {{ --bg:#f6f8fb; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .wrap {{ display:grid; grid-template-columns:320px 1fr; gap:12px; margin:12px auto; padding:0 16px 20px; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
    .panel h3 {{ margin:0; padding:10px 12px; border-bottom:1px solid var(--line); background:#fafafa; font-size:14px; }}
    .tree {{ max-height:calc(100vh - 130px); overflow:auto; }}
    .node {{ padding:7px 10px; border-bottom:1px solid #eef2f7; cursor:pointer; font-size:13px; }}
    .node:hover {{ background:#f8fafc; }}
    .node.active {{ background:#e6fffa; border-left:3px solid var(--acc); }}
    .cards {{ display:grid; grid-template-columns:1fr; gap:12px; }}
    .desc {{ white-space:pre-wrap; font-size:13px; line-height:1.5; padding:10px 12px; }}
    textarea {{ width:100%; min-height:180px; box-sizing:border-box; border:1px solid #d1d5db; border-radius:8px; padding:10px; font:inherit; }}
    .toolbar {{ padding:10px 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; border-top:1px solid #eef2f7; }}
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/?project_id={project_id}">返回主页</a>
    <strong>收入分析</strong>
    <span class="muted">项目ID: {project_id}</span>
    <span class="muted" id="status"></span>
    <button class="btn" onclick="exportExcel()">导出Excel计算表</button>
    <button class="btn primary" onclick="reloadData()">刷新</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>收入分析树</h3>
      <div class="tree" id="treeBox"></div>
    </div>
    <div class="cards">
      <div class="panel">
        <h3 id="nodeTitle">自动分析（只读）</h3>
        <div class="desc" id="autoText"></div>
      </div>
      <div class="panel">
        <h3>分析判断内容（可编辑）</h3>
        <div class="desc"><textarea id="manualText"></textarea></div>
        <div class="toolbar">
          <label id="classBucketWrap" style="display:inline-flex;">归类确认
            <select id="classBucket">
              <option value="recurring">经常性净收益</option>
              <option value="nonrecurring">非经常性净收益</option>
            </select>
          </label>
          <span id="classBucketHint" class="muted" style="display:inline-flex;">仅对当前项目生效</span>
          <label><input type="checkbox" id="manualConfirmed"> 确认</label>
          <button class="btn primary" onclick="saveActive()">保存并确认</button>
          <span id="saveHint" class="muted"></span>
        </div>
      </div>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
let cache = null;
let activeNodeId = '';

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}
function depthOf(id) {{
  return (id || '').split('.').length - 1;
}}
function nodeById(id) {{
  return (cache.nodes || []).find(x => String(x.node_id) === String(id)) || null;
}}
function renderTree() {{
  const tree = cache.tree || [];
  const box = document.getElementById('treeBox');
  box.innerHTML = tree.map(n => {{
    const d = depthOf(n.node_id);
    const active = n.node_id === activeNodeId ? 'active' : '';
    return `<div class="node ${{active}}" data-id="${{esc(n.node_id)}}" style="padding-left:${{10 + d * 16}}px;">${{esc(n.node_id)}} ${{esc(n.label)}}</div>`;
  }}).join('');
  box.querySelectorAll('.node[data-id]').forEach(el => {{
    el.addEventListener('click', () => {{
      activeNodeId = el.dataset.id;
      renderNode();
    }});
  }});
}}
function renderNode() {{
  renderTree();
  const n = nodeById(activeNodeId);
  if (!n) return;
  const nodeId = String(n.node_id || '');
  const canClassify = (!!n.classification_editable) || nodeId.startsWith('2.1.2.') || nodeId.startsWith('2.2.');
  const fallbackBucket = nodeId.startsWith('2.1.2.') ? 'recurring' : 'nonrecurring';
  document.getElementById('nodeTitle').textContent = `${{n.node_id}} ${{n.label}}（自动分析）`;
  document.getElementById('autoText').textContent = n.auto_text || '';
  document.getElementById('manualText').value = n.manual_text || n.auto_text || '';
  const clsSel = document.getElementById('classBucket');
  const clsHint = document.getElementById('classBucketHint');
  if (canClassify) {{
    clsSel.disabled = false;
    clsHint.textContent = '仅对当前项目生效';
    clsSel.value = (n.classification_bucket || fallbackBucket);
  }} else {{
    clsSel.disabled = true;
    clsHint.textContent = '当前节点无需归类';
    clsSel.value = fallbackBucket;
  }}
  document.getElementById('manualConfirmed').checked = !!n.confirmed;
  document.getElementById('saveHint').textContent = n.confirmed ? '当前已确认' : '';
}}
async function saveActive() {{
  const n = nodeById(activeNodeId);
  if (!n) return;
  const payload = {{
    project_id: PROJECT_ID,
    group_id: 'income_analysis',
    rows: [{{
      code: n.node_id,
      name: n.label,
      manual_text: document.getElementById('manualText').value,
      confirmed: document.getElementById('manualConfirmed').checked,
      classification_bucket: (String(n.node_id || '').startsWith('2.1.2.') || String(n.node_id || '').startsWith('2.2.') || !!n.classification_editable)
        ? document.getElementById('classBucket').value : ''
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  document.getElementById('saveHint').textContent = data.ok ? '已保存' : '保存失败';
  if (data.ok) await reloadData();
}}
async function reloadData() {{
  const res = await fetch(`/api/analysis/income?project_id=${{encodeURIComponent(PROJECT_ID)}}`);
  cache = await res.json();
  if (!res.ok) {{
    document.getElementById('status').textContent = cache.error || '读取失败';
    return;
  }}
  if (!activeNodeId && (cache.tree || []).length) {{
    activeNodeId = cache.tree[0].node_id;
  }}
  renderNode();
  document.getElementById('status').textContent = `来源Sheet: ${{cache.sheet_title || '-'}} | 年份: ${{(cache.years||[]).join(',')}}`;
}}
function exportExcel() {{
  const url = `/api/analysis/income/export?project_id=${{encodeURIComponent(PROJECT_ID)}}`;
  window.location.href = url;
}}
reloadData();
</script>
</body>
</html>"""


def render_ratio_analysis_page(project_id: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>财务指标分析</title>
  <style>
    :root {{ --bg:#f6f8fb; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .wrap {{ display:grid; grid-template-columns:320px 1fr; gap:12px; margin:12px auto; padding:0 16px 20px; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
    .panel h3 {{ margin:0; padding:10px 12px; border-bottom:1px solid var(--line); background:#fafafa; font-size:14px; }}
    .tree {{ max-height:calc(100vh - 130px); overflow:auto; }}
    .node {{ padding:7px 10px; border-bottom:1px solid #eef2f7; cursor:pointer; font-size:13px; }}
    .node:hover {{ background:#f8fafc; }}
    .node.active {{ background:#e6fffa; border-left:3px solid var(--acc); }}
    .cards {{ display:grid; grid-template-columns:1fr; gap:12px; }}
    .desc {{ white-space:pre-wrap; font-size:13px; line-height:1.5; padding:10px 12px; }}
    textarea {{ width:100%; min-height:180px; box-sizing:border-box; border:1px solid #d1d5db; border-radius:8px; padding:10px; font:inherit; }}
    .toolbar {{ padding:10px 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; border-top:1px solid #eef2f7; }}
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/?project_id={project_id}">返回主页</a>
    <strong>财务指标分析</strong>
    <span class="muted">项目ID: {project_id}</span>
    <span class="muted" id="status"></span>
    <button class="btn primary" onclick="reloadData()">刷新</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>指标分析树</h3>
      <div class="tree" id="treeBox"></div>
    </div>
    <div class="cards">
      <div class="panel">
        <h3 id="nodeTitle">自动分析（只读）</h3>
        <div class="desc" id="autoText"></div>
      </div>
      <div class="panel">
        <h3>分析判断内容（可编辑）</h3>
        <div class="desc"><textarea id="manualText"></textarea></div>
        <div class="toolbar">
          <label><input type="checkbox" id="manualConfirmed"> 确认</label>
          <button class="btn primary" onclick="saveActive()">保存并确认</button>
          <span id="saveHint" class="muted"></span>
        </div>
      </div>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
let cache = null;
let activeNodeId = '';

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}
function depthOf(id) {{
  return (id || '').split('.').length - 1;
}}
function nodeById(id) {{
  return (cache.nodes || []).find(x => String(x.node_id) === String(id)) || null;
}}
function renderTree() {{
  const tree = cache.tree || [];
  const box = document.getElementById('treeBox');
  box.innerHTML = tree.map(n => {{
    const d = depthOf(n.node_id);
    const active = n.node_id === activeNodeId ? 'active' : '';
    return `<div class="node ${{active}}" data-id="${{esc(n.node_id)}}" style="padding-left:${{10 + d * 16}}px;">${{esc(n.node_id)}} ${{esc(n.label)}}</div>`;
  }}).join('');
  box.querySelectorAll('.node[data-id]').forEach(el => {{
    el.addEventListener('click', () => {{
      activeNodeId = el.dataset.id;
      renderNode();
    }});
  }});
}}
function renderNode() {{
  renderTree();
  const n = nodeById(activeNodeId);
  if (!n) return;
  document.getElementById('nodeTitle').textContent = `${{n.node_id}} ${{n.label}}（自动分析）`;
  document.getElementById('autoText').textContent = n.auto_text || '';
  document.getElementById('manualText').value = n.manual_text || n.auto_text || '';
  document.getElementById('manualConfirmed').checked = !!n.confirmed;
  document.getElementById('saveHint').textContent = n.confirmed ? '当前已确认' : '';
}}
async function saveActive() {{
  const n = nodeById(activeNodeId);
  if (!n) return;
  const payload = {{
    project_id: PROJECT_ID,
    group_id: 'ratio_analysis',
    rows: [{{
      code: n.node_id,
      name: n.label,
      manual_text: document.getElementById('manualText').value,
      confirmed: document.getElementById('manualConfirmed').checked
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  document.getElementById('saveHint').textContent = data.ok ? '已保存' : '保存失败';
  if (data.ok) await reloadData();
}}
async function reloadData() {{
  const res = await fetch(`/api/analysis/ratios?project_id=${{encodeURIComponent(PROJECT_ID)}}`);
  cache = await res.json();
  if (!res.ok) {{
    document.getElementById('status').textContent = cache.error || '读取失败';
    return;
  }}
  if (!activeNodeId && (cache.tree || []).length) {{
    activeNodeId = cache.tree[0].node_id;
  }}
  renderNode();
  document.getElementById('status').textContent = `来源Sheet: ${{cache.sheet_title || '-'}} | 年份: ${{(cache.years||[]).join(',')}}`;
}}
reloadData();
</script>
</body>
</html>"""


def render_key_ratio_analysis_page(project_id: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>重要指标分析</title>
  <style>
    :root {{ --bg:#f6f8fb; --card:#fff; --ink:#111827; --line:#d1d5db; --muted:#6b7280; --acc:#0f766e; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; border-radius:8px; padding:6px 10px; text-decoration:none; color:inherit; cursor:pointer; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .wrap {{ display:grid; grid-template-columns:320px 1fr; gap:12px; margin:12px auto; padding:0 16px 20px; }}
    .panel {{ background:var(--card); border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
    .panel h3 {{ margin:0; padding:10px 12px; border-bottom:1px solid var(--line); background:#fafafa; font-size:14px; }}
    .tree {{ max-height:calc(100vh - 130px); overflow:auto; }}
    .node {{ padding:7px 10px; border-bottom:1px solid #eef2f7; cursor:pointer; font-size:13px; }}
    .node:hover {{ background:#f8fafc; }}
    .node.active {{ background:#e6fffa; border-left:3px solid var(--acc); }}
    .cards {{ display:grid; grid-template-columns:1fr; gap:12px; }}
    .desc {{ white-space:pre-wrap; font-size:13px; line-height:1.6; padding:10px 12px; }}
    textarea {{ width:100%; min-height:200px; box-sizing:border-box; border:1px solid #d1d5db; border-radius:8px; padding:10px; font:inherit; }}
    .toolbar {{ padding:10px 12px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; border-top:1px solid #eef2f7; }}
  </style>
</head>
<body>
  <div class="top">
    <a class="btn" href="/?project_id={project_id}">返回主页</a>
    <strong>重要指标分析</strong>
    <span class="muted">项目ID: {project_id}</span>
    <span class="muted" id="status"></span>
    <button class="btn primary" onclick="reloadData()">刷新</button>
  </div>
  <div class="wrap">
    <div class="panel">
      <h3>重点指标列表</h3>
      <div class="tree" id="treeBox"></div>
    </div>
    <div class="cards">
      <div class="panel">
        <h3 id="nodeTitle">自动分析（只读）</h3>
        <div class="desc" id="autoText"></div>
      </div>
      <div class="panel">
        <h3>分析判断内容（可编辑）</h3>
        <div class="desc"><textarea id="manualText"></textarea></div>
        <div class="toolbar">
          <label><input type="checkbox" id="manualConfirmed"> 确认</label>
          <button class="btn primary" onclick="saveActive()">保存并确认</button>
          <span id="saveHint" class="muted"></span>
        </div>
      </div>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
let cache = null;
let activeNodeId = '';

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}
function nodeById(id) {{
  return (cache.nodes || []).find(x => String(x.node_id) === String(id)) || null;
}}
function renderTree() {{
  const tree = cache.tree || [];
  const box = document.getElementById('treeBox');
  box.innerHTML = tree.map(n => {{
    const active = n.node_id === activeNodeId ? 'active' : '';
    return `<div class="node ${{active}}" data-id="${{esc(n.node_id)}}">${{esc(n.node_id)}} ${{esc(n.label)}}</div>`;
  }}).join('');
  box.querySelectorAll('.node[data-id]').forEach(el => {{
    el.addEventListener('click', () => {{
      activeNodeId = el.dataset.id;
      renderNode();
    }});
  }});
}}
function renderNode() {{
  renderTree();
  const n = nodeById(activeNodeId);
  if (!n) return;
  document.getElementById('nodeTitle').textContent = `${{n.node_id}} ${{n.label}}（自动分析）`;
  document.getElementById('autoText').textContent = n.auto_text || '';
  document.getElementById('manualText').value = n.manual_text || n.auto_text || '';
  document.getElementById('manualConfirmed').checked = !!n.confirmed;
  document.getElementById('saveHint').textContent = n.confirmed ? '当前已确认' : '';
}}
async function saveActive() {{
  const n = nodeById(activeNodeId);
  if (!n) return;
  const payload = {{
    project_id: PROJECT_ID,
    group_id: 'key_ratio_analysis',
    rows: [{{
      code: n.node_id,
      name: n.label,
      manual_text: document.getElementById('manualText').value,
      confirmed: document.getElementById('manualConfirmed').checked
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  document.getElementById('saveHint').textContent = data.ok ? '已保存' : '保存失败';
  if (data.ok) await reloadData();
}}
async function reloadData() {{
  const res = await fetch(`/api/analysis/key-ratios?project_id=${{encodeURIComponent(PROJECT_ID)}}`);
  cache = await res.json();
  if (!res.ok) {{
    document.getElementById('status').textContent = cache.error || '读取失败';
    return;
  }}
  if (!activeNodeId && (cache.tree || []).length) {{
    activeNodeId = cache.tree[0].node_id;
  }}
  renderNode();
  document.getElementById('status').textContent = `来源Sheet: ${{cache.sheet_title || '-'}} | 年份: ${{(cache.years||[]).join(',')}}`;
}}
reloadData();
</script>
</body>
</html>"""


def build_income_analysis_export_xlsx(data: Dict[str, Any], project_id: str, source_workbook: str) -> bytes:
    wb = Workbook()
    ws_nodes = wb.active
    ws_nodes.title = "收入分析_节点明细"
    years = [str(y) for y in (data.get("years", []) or [])]

    ws_nodes.append(["项目ID", project_id])
    ws_nodes.append(["来源主文件", source_workbook])
    ws_nodes.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_nodes.append([])

    headers = ["node_id", "parent_id", "label", "source_code", "source_name"] + years + [
        "auto_text",
        "manual_text",
        "final_text",
        "confirmed",
    ]
    ws_nodes.append(headers)

    node_map = {str(n.get("node_id", "")): n for n in (data.get("nodes", []) or [])}
    for t in (data.get("tree", []) or []):
        nid = str(t.get("node_id", ""))
        n = node_map.get(nid, {})
        row = [
            nid,
            str(t.get("parent_id", "")),
            str(t.get("label", "")),
            str(n.get("source_code", "")),
            str(n.get("source_name", "")),
        ]
        vals = n.get("values", {}) if isinstance(n.get("values"), dict) else {}
        for y in years:
            row.append(vals.get(y))
        row.extend(
            [
                str(n.get("auto_text", "")),
                str(n.get("manual_text", "")),
                str(n.get("final_text", "")),
                bool(n.get("confirmed", False)),
            ]
        )
        ws_nodes.append(row)

    ws_tree = wb.create_sheet("收入分析_树结构")
    ws_tree.append(["node_id", "parent_id", "label"])
    for t in (data.get("tree", []) or []):
        ws_tree.append([str(t.get("node_id", "")), str(t.get("parent_id", "")), str(t.get("label", ""))])

    ws_notes = wb.create_sheet("收入分析_汇总说明")
    ws_notes.append(["节点", "说明"])
    ws_notes.append(["3.1", "经常性收益小计（按规则公式计算）"])
    ws_notes.append(["3.2", "非经常性收益小计（按规则公式计算）"])
    ws_notes.append(["3.3", "收益结构判断（按规则公式计算）"])

    for ws in wb.worksheets:
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col[:200])
            ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_blank_report_template_xlsx(project_id: str) -> bytes:
    wb_path = workbook_path(project_id)
    if not wb_path.exists():
        raise FileNotFoundError(str(wb_path))
    wb = load_workbook(wb_path, data_only=False)

    keep_titles: List[str] = []
    for g in SHEET_GROUPS:
        if str(g.get("id", "")).strip() not in {"bs", "is", "cf"}:
            continue
        ws = get_sheet_by_loose_name(wb, g.get("candidates", []))
        if ws is not None:
            keep_titles.append(ws.title)
    keep_set = set(keep_titles)
    for title in list(wb.sheetnames):
        if title not in keep_set:
            del wb[title]

    prefix_map = {
        "bs": "BS",
        "is": "IS",
        "cf": "CF",
    }
    for g in SHEET_GROUPS:
        gid = str(g.get("id", "")).strip()
        if gid not in {"bs", "is", "cf"}:
            continue
        ws = get_sheet_by_loose_name(wb, g.get("candidates", []))
        if ws is None:
            continue
        years = parse_years_from_sheet(ws)
        n_years = max(1, len(years))
        code_prefix = prefix_map.get(gid, "")
        for r in range(2, ws.max_row + 1):
            code = str(ws.cell(r, 1).value or "").strip().upper()
            if not code or not code_prefix or not code.startswith(code_prefix):
                continue
            for c in range(3, 3 + n_years):
                ws.cell(r, c).value = None

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_ratio_indicator_template_xlsx(project_id: str) -> bytes:
    years = _detect_project_years_for_ratio(project_id)
    catalog = _ratio_indicator_catalog_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "财务指标导入模板"
    ws.append(["指标ID", "指标名称", "期间", "数值", "单位", "数据来源", "备注"])
    for c in catalog:
        iid = str(c.get("indicator_id", "")).strip()
        name = str(c.get("indicator_name_zh", "")).strip()
        unit = str(c.get("unit", "")).strip()
        source = str(c.get("value_source", "")).strip() or "external"
        for y in years:
            ws.append([iid, name, f"{y}年", None, unit, source, ""])
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col[: min(300, ws.max_row)])
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def import_ratio_indicator_template(project_id: str, xlsx_bytes: bytes) -> Dict[str, Any]:
    wb_in = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    headers = {str(ws_in.cell(1, c).value or "").strip(): c for c in range(1, ws_in.max_column + 1)}
    c_id = headers.get("指标ID", headers.get("indicator_id", 1))
    c_name = headers.get("指标名称", headers.get("indicator_name", 2))
    c_period = headers.get("期间", headers.get("period", 3))
    c_value = headers.get("数值", headers.get("value", 4))
    c_unit = headers.get("单位", headers.get("unit", 5))
    c_source = headers.get("数据来源", headers.get("source", 6))

    imported: List[Dict[str, Any]] = []
    for r in range(2, ws_in.max_row + 1):
        iid = str(ws_in.cell(r, c_id).value or "").strip()
        if not iid:
            continue
        name = str(ws_in.cell(r, c_name).value or "").strip()
        period = str(ws_in.cell(r, c_period).value or "").strip()
        m = re.search(r"(20\d{2})", period)
        if not m:
            continue
        year = m.group(1)
        val_raw = ws_in.cell(r, c_value).value
        val = normalize_num(val_raw)
        if val is None:
            continue
        imported.append(
            {
                "indicator_id": iid,
                "name": name,
                "year": year,
                "period": period or f"{year}年",
                "value": val,
                "unit": str(ws_in.cell(r, c_unit).value or "").strip(),
                "source": str(ws_in.cell(r, c_source).value or "").strip(),
            }
        )

    if not imported:
        return {"ok": False, "error": "模板中未读取到可导入数据（请检查指标ID/期间/数值）", "updated_rows": 0}

    wb_path = workbook_path(project_id)
    if not wb_path.exists():
        return {"ok": False, "error": f"workbook not found: {wb_path}", "updated_rows": 0}
    wb = load_workbook(wb_path, data_only=False)
    ws = get_sheet_by_loose_name(wb, ["财务比率", "财务指标", "财务指标表"])
    if ws is None:
        ws = wb.create_sheet("财务指标表")
        ws.append(["指标ID", "指标", "分组", "期间", "数值", "单位", "数据来源", "备注"])

    # detect or fallback columns
    h = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    idx_col = h.get("指标ID", 1)
    name_col = h.get("指标", 2)
    period_col = h.get("期间", 4)
    value_col = h.get("数值", 5)
    unit_col = h.get("单位", 6)
    source_col = h.get("数据来源", 7)

    existing_index: Dict[tuple, int] = {}
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, idx_col).value or "").strip()
        period = str(ws.cell(r, period_col).value or "").strip()
        my = re.search(r"(20\d{2})", period)
        if code and my:
            existing_index[(code, my.group(1))] = r

    updated = 0
    for rec in imported:
        key = (rec["indicator_id"], rec["year"])
        row_idx = existing_index.get(key)
        if row_idx is None:
            row_idx = ws.max_row + 1
            existing_index[key] = row_idx
        ws.cell(row_idx, idx_col).value = rec["indicator_id"]
        ws.cell(row_idx, name_col).value = rec["name"] or rec["indicator_id"]
        ws.cell(row_idx, period_col).value = rec["period"] or f"{rec['year']}年"
        ws.cell(row_idx, value_col).value = rec["value"]
        if unit_col:
            ws.cell(row_idx, unit_col).value = rec["unit"] or ws.cell(row_idx, unit_col).value
        if source_col:
            ws.cell(row_idx, source_col).value = rec["source"] or ws.cell(row_idx, source_col).value
        updated += 1

    wb.save(wb_path)
    return {"ok": True, "updated_rows": updated, "workbook": str(wb_path)}


def render_sheet_page(group: Dict[str, Any], project_id: str) -> str:
    label = group["label"]
    gid = group["id"]
    ratio_tools = ""
    if gid == "ratio":
        ratio_tools = """
    <button class=\"btn\" onclick=\"exportRatioTemplate()\">导出财务指标模板</button>
    <button class=\"btn\" onclick=\"pickRatioImport()\">导入财务指标模板</button>
    <input type=\"file\" id=\"ratioImportFile\" accept=\".xlsx\" style=\"display:none\" />
"""
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{label}</title>
  <style>
    :root {{ --bg:#f6f7f9; --ink:#111827; --line:#cfd5dc; --card:#fff; --acc:#0f766e; --muted:#6b7280; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; padding:6px 10px; border-radius:8px; cursor:pointer; text-decoration:none; color:inherit; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .wrap {{ padding:12px 16px 28px; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); background:var(--card); border-radius:10px; }}
    table {{ border-collapse:collapse; width:100%; min-width:1250px; }}
    th,td {{ border-bottom:1px solid #e5e7eb; padding:8px; vertical-align:top; font-size:13px; }}
    th {{ position:sticky; top:0; background:#f9fafb; z-index:1; text-align:left; }}
    td.num {{ text-align:right; white-space:nowrap; }}
    textarea {{ width:100%; min-height:76px; resize:vertical; font:inherit; }}
    .pill {{ display:inline-block; border:1px solid var(--line); border-radius:999px; padding:2px 8px; font-size:12px; }}
    .ok {{ background:#ecfdf5; border-color:#86efac; color:#166534; }}
    .bad {{ background:#fef2f2; border-color:#fca5a5; color:#991b1b; }}
    .na {{ background:#f8fafc; border-color:#cbd5e1; color:#475569; }}
    input.num-edit {{ width:120px; text-align:right; }}
    input.reason-edit {{ width:220px; }}
    .dirty {{ background:#fff7ed; }}
  </style>
</head>
<body>
  <div class="top">
    <a href="/?project_id={project_id}" class="btn">返回主页</a>
    <strong>{label}</strong>
    <span class="muted" id="status"></span>
    <span class="muted">项目ID: {project_id}</span>
    <button class="btn" onclick="reloadData()">刷新</button>
    {ratio_tools}
    <button class="btn primary" id="saveBtn" onclick="saveAll()">保存全部修改</button>
  </div>
  <div class="wrap">
    <div class="table-wrap">
      <table id="tbl">
        <thead id="thead"></thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
const GROUP_ID = {json.dumps(gid)};
let cache = null;
if (GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf') {{
  const btn = document.getElementById('saveBtn');
  if (btn) btn.textContent = '保存确认';
}}

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}

function markDirty(el) {{
  el.closest('tr').classList.add('dirty');
}}

async function reloadData() {{
  const res = await fetch(`/api/sheet/${{GROUP_ID}}?project_id=${{encodeURIComponent(PROJECT_ID)}}`);
  cache = await res.json();
  if (!res.ok) {{
    document.getElementById('status').textContent = cache.error || '读取失败';
    document.getElementById('thead').innerHTML = '';
    document.getElementById('tbody').innerHTML = '';
    return;
  }}
  render();
}}

function render() {{
  const years = cache.years || [];
  document.getElementById('status').textContent = `行数 ${{cache.rows.length}}` + (cache.sheet_title ? ` | 来源Sheet: ${{cache.sheet_title}}` : '');
  const head = (GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf')
    ? `<tr><th style="min-width:90px;">科目编码</th><th style="min-width:180px;">科目</th>${{years.map(y=>`<th>${{y}}</th>`).join('')}}<th style="min-width:460px;">分析判断内容</th><th style="min-width:220px;">数据校验</th><th style="min-width:220px;">修改原因</th><th style="min-width:120px;">操作</th><th>确认</th></tr>`
    : `<tr><th style="min-width:90px;">科目编码</th><th style="min-width:180px;">科目</th>${{years.map(y=>`<th>${{y}}</th>`).join('')}}<th style="min-width:280px;">自动描述</th><th style="min-width:340px;">人工确认描述</th><th>确认</th></tr>`;
  document.getElementById('thead').innerHTML = head;

  const rowsHtml = cache.rows.map((r, idx) => {{
    const vals = years.map(y => {{
      const v = r.values[y];
      if (GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf') {{
        const txt = (v == null || Number.isNaN(Number(v))) ? '' : Number(v).toFixed(2);
        return `<td class="num"><input class="num-edit" data-year="${{esc(y)}}" value="${{esc(txt)}}" oninput="markDirty(this)"/></td>`;
      }}
      return `<td class="num">${{v == null ? '' : Number(v).toLocaleString(undefined, {{minimumFractionDigits:2, maximumFractionDigits:2}})}}</td>`;
    }}).join('');
    if (GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf') {{
      const st = String(r.validation_status || '未配置');
      const cls = st === '通过' ? 'ok' : (st === '未通过' ? 'bad' : 'na');
      const analysisText = GROUP_ID === 'bs' ? (r.analysis_text || '') : (r.auto_text || '');
      const nameCell = (GROUP_ID === 'bs')
        ? `<a href="/detail?project_id=${{encodeURIComponent(PROJECT_ID)}}&code=${{encodeURIComponent(r.code)}}">${{esc(r.name)}}</a>`
        : `${{esc(r.name)}}`;
      return `<tr data-i="${{idx}}">
        <td>${{esc(r.code)}}</td>
        <td>${{nameCell}}</td>
        ${{vals}}
        <td>${{esc(analysisText)}}</td>
        <td><span class="pill ${{cls}}">${{esc(st)}}</span><div class="muted">${{esc(r.validation_message || '')}}</div></td>
        <td><input class="reason-edit" placeholder="修改原因（必填）" value="${{esc(r.override_reason || '')}}" oninput="markDirty(this)"/></td>
        <td><button class="btn" onclick="saveOneRow(${{idx}})">保存本行</button></td>
        <td><label class="pill"><input type="checkbox" ${{r.confirmed ? 'checked' : ''}} onchange="markDirty(this)">确认</label></td>
      </tr>`;
    }}
    return `<tr data-i="${{idx}}">
      <td>${{esc(r.code)}}</td>
      <td>${{esc(r.name)}}</td>
      ${{vals}}
      <td>${{esc(r.auto_text || '')}}</td>
      <td><textarea oninput="markDirty(this)">${{esc(r.manual_text || '')}}</textarea></td>
      <td><label class="pill"><input type="checkbox" ${{r.confirmed ? 'checked' : ''}} onchange="markDirty(this)">确认</label></td>
    </tr>`;
  }}).join('');
  document.getElementById('tbody').innerHTML = rowsHtml;
}}

async function saveAll() {{
  if (!cache) return;
  const years = cache.years || [];
  const rows = [...document.querySelectorAll('#tbody tr')].map(tr => {{
    const i = Number(tr.dataset.i);
    const base = cache.rows[i];
    const confirmed = tr.querySelector('input[type="checkbox"]').checked;
    if (GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf') {{
      const values = Object.assign({{}}, base.values || {{}});
      years.forEach(y => {{
        const el = tr.querySelector(`input.num-edit[data-year="${{y}}"]`);
        const raw = el ? String(el.value || '').trim() : '';
        values[y] = raw === '' ? null : Number(raw);
      }});
      const reasonEl = tr.querySelector('input.reason-edit');
      return {{
        code: base.code,
        name: base.name,
        years,
        values,
        override_reason: reasonEl ? reasonEl.value : '',
        confirmed
      }};
    }}
    const txt = tr.querySelector('textarea').value;
    return {{
      code: base.code,
      name: base.name,
      years,
      values: base.values,
      auto_text: base.auto_text,
      manual_text: txt,
      confirmed
    }};
  }});

  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify({{project_id: PROJECT_ID, group_id: GROUP_ID, rows}})
  }});
  const data = await res.json();
  document.getElementById('status').textContent = data.ok ? `已保存 ${{data.saved_rows}} 行` : (`保存失败: ` + (data.error || 'unknown'));
  if (data.ok) {{
    document.querySelectorAll('#tbody tr').forEach(x => x.classList.remove('dirty'));
  }}
}}

async function saveOneRow(idx) {{
  if (!(GROUP_ID === 'bs' || GROUP_ID === 'is' || GROUP_ID === 'cf')) return;
  if (!cache) return;
  const tr = document.querySelector(`#tbody tr[data-i="${{idx}}"]`);
  if (!tr) return;
  const base = cache.rows[idx];
  const years = cache.years || [];
  const values = Object.assign({{}}, base.values || {{}});
  years.forEach(y => {{
    const el = tr.querySelector(`input.num-edit[data-year="${{y}}"]`);
    const raw = el ? String(el.value || '').trim() : '';
    values[y] = raw === '' ? null : Number(raw);
  }});
  const reasonEl = tr.querySelector('input.reason-edit');
  const payload = {{
    project_id: PROJECT_ID,
    group_id: GROUP_ID,
    rows: [{{
      code: base.code,
      name: base.name,
      years,
      values,
      override_reason: reasonEl ? reasonEl.value : '',
      confirmed: tr.querySelector('input[type="checkbox"]').checked
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{'Content-Type':'application/json'}},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  document.getElementById('status').textContent = data.ok ? `已保存 1 行` : (`保存失败: ` + (data.error || 'unknown'));
  if (data.ok) await reloadData();
}}

function exportRatioTemplate() {{
  if (GROUP_ID !== 'ratio') return;
  const url = `/api/template/ratio-indicator-export?project_id=${{encodeURIComponent(PROJECT_ID)}}`;
  window.location.href = url;
}}

function pickRatioImport() {{
  if (GROUP_ID !== 'ratio') return;
  const el = document.getElementById('ratioImportFile');
  if (!el) return;
  el.value = '';
  el.click();
}}

async function importRatioTemplate(file) {{
  const reader = new FileReader();
  const b64 = await new Promise((resolve, reject) => {{
    reader.onload = () => {{
      try {{
        const s = String(reader.result || '');
        const parts = s.split(',');
        resolve(parts.length > 1 ? parts[1] : '');
      }} catch (e) {{
        reject(e);
      }}
    }};
    reader.onerror = reject;
    reader.readAsDataURL(file);
  }});
  const payload = {{
    project_id: PROJECT_ID,
    file_name: file.name,
    file_b64: b64
  }};
  const res = await fetch('/api/template/ratio-indicator-import', {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  if (!res.ok || !data.ok) {{
    document.getElementById('status').textContent = '导入失败: ' + (data.error || 'unknown');
    return;
  }}
  document.getElementById('status').textContent = `导入成功：${{data.updated_rows}} 行`;
  await reloadData();
}}

const importEl = document.getElementById('ratioImportFile');
if (importEl) {{
  importEl.addEventListener('change', async (e) => {{
    const f = e.target.files && e.target.files[0];
    if (!f) return;
    await importRatioTemplate(f);
  }});
}}

reloadData();
</script>
</body>
</html>"""


def render_detail_page(project_id: str, code: str) -> str:
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>科目明细</title>
  <style>
    :root {{ --bg:#f6f7f9; --ink:#111827; --line:#cfd5dc; --card:#fff; --acc:#0f766e; --muted:#6b7280; --bad:#991b1b; }}
    body {{ margin:0; font-family: "Microsoft YaHei", "PingFang SC", sans-serif; background:var(--bg); color:var(--ink); }}
    .top {{ position:sticky; top:0; background:#fff; border-bottom:1px solid var(--line); padding:10px 16px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }}
    .btn {{ border:1px solid var(--line); background:#fff; padding:6px 10px; border-radius:8px; cursor:pointer; text-decoration:none; color:inherit; }}
    .btn.primary {{ background:var(--acc); color:#fff; border-color:var(--acc); }}
    .muted {{ color:var(--muted); font-size:12px; }}
    .warn {{ color:var(--bad); font-size:12px; }}
    .wrap {{ padding:12px 16px 28px; display:grid; grid-template-columns: 1fr; gap:12px; }}
    .card {{ border:1px solid var(--line); border-radius:10px; background:var(--card); padding:10px; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:8px; }}
    table {{ border-collapse:collapse; width:100%; min-width:980px; }}
    th,td {{ border-bottom:1px solid #e5e7eb; padding:8px; font-size:13px; }}
    th {{ position:sticky; top:0; background:#f9fafb; text-align:left; }}
    textarea {{ width:100%; min-height:120px; resize:vertical; font:inherit; }}
    input.cell {{ width:100%; box-sizing:border-box; border:1px solid #d1d5db; border-radius:6px; padding:6px; font:inherit; }}
    .mini {{ display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin-bottom:8px; }}
  </style>
</head>
<body>
  <div class="top">
    <a href="/sheet/bs?project_id={project_id}" class="btn">返回资产负债表</a>
    <strong id="title">科目明细</strong>
    <span class="muted" id="status"></span>
    <span class="muted">项目ID: {project_id}</span>
    <span class="muted">编码: {code}</span>
    <button class="btn" onclick="reloadData()">刷新</button>
    <button class="btn primary" onclick="saveDetail()">保存并确认</button>
  </div>
  <div class="wrap">
    <div class="card">
      <div><strong>科目合计值（万元）</strong></div>
      <div id="mainValues" style="margin-top:6px;"></div>
      <div class="muted" id="sheetInfo" style="margin-top:6px;"></div>
      <div class="warn" id="mapInfo" style="margin-top:4px;"></div>
    </div>
    <div class="card">
      <div class="mini">
        <strong>明细录入（期间 / 项目 / 明细值（万元） / 说明）</strong>
        <button class="btn" onclick="addRow()">新增一行</button>
        <button class="btn" onclick="pickImport()">导入明细</button>
        <input type="file" id="importFile" accept=".xlsx" style="display:none"/>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th style="min-width:140px;">期间</th>
              <th style="min-width:220px;">项目</th>
              <th style="min-width:160px;">明细值（万元）</th>
              <th style="min-width:140px;">占比</th>
              <th style="min-width:260px;">说明</th>
              <th style="width:90px;">操作</th>
            </tr>
          </thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </div>
    <div class="card">
      <div><strong>分析判断内容（可编辑）</strong></div>
      <div class="muted">科目明细暂不自动生成描述，可在此手动输入。</div>
      <textarea id="manualText" placeholder="可输入人工补充判断。"></textarea>
      <label style="display:block; margin-top:6px;"><input type="checkbox" id="confirmed"/> 确认</label>
    </div>
  </div>
<script>
const PROJECT_ID = {json.dumps(project_id)};
const CODE = {json.dumps(code)};
let cache = null;
let detailRows = [];

function esc(s) {{
  return (s ?? '').toString().replace(/[&<>"]/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
}}

function normalizeRowsFromUi() {{
  const trs = [...document.querySelectorAll('#tbody tr')];
  const out = trs.map(tr => {{
    const period = (tr.querySelector('input[data-k=\"period\"]')?.value || '').trim();
    const item = (tr.querySelector('input[data-k=\"item\"]')?.value || '').trim();
    const value = (tr.querySelector('input[data-k=\"value\"]')?.value || '').trim();
    const note = (tr.querySelector('input[data-k=\"note\"]')?.value || '').trim();
    return {{ period, item, value, note }};
  }});
  return out.filter(x => x.period || x.item || x.value || x.note);
}}

function periodToYear(periodText) {{
  const m = String(periodText || '').match(/(20\\d{{2}})/);
  return m ? m[1] : '';
}}

function calcRatioText(row) {{
  if (!cache) return '';
  const y = periodToYear(row.period || '');
  if (!y) return '待补充';
  const denom = (cache.subject_values || {{}})[y];
  const num = Number(String(row.value || '').replace(/,/g, ''));
  if (denom == null || Number.isNaN(Number(denom)) || Math.abs(Number(denom)) < 1e-12) return '待补充';
  if (Number.isNaN(num)) return '待补充';
  return (num / Number(denom) * 100).toFixed(2) + '%';
}}

function onCellChange(i, key, value) {{
  if (i < 0 || i >= detailRows.length) return;
  detailRows[i][key] = value;
  renderRows();
}}

function renderRows() {{
  const html = detailRows.map((r, i) => `
    <tr data-i="${{i}}">
      <td><input class="cell" data-k="period" value="${{esc(r.period || '')}}" onchange="onCellChange(${{i}}, 'period', this.value)"/></td>
      <td><input class="cell" data-k="item" value="${{esc(r.item || '')}}" onchange="onCellChange(${{i}}, 'item', this.value)"/></td>
      <td><input class="cell" data-k="value" value="${{esc(r.value || '')}}" onchange="onCellChange(${{i}}, 'value', this.value)"/></td>
      <td>${{esc(calcRatioText(r))}}</td>
      <td><input class="cell" data-k="note" value="${{esc(r.note || '')}}" onchange="onCellChange(${{i}}, 'note', this.value)"/></td>
      <td><button class="btn" onclick="deleteRow(${{i}})">删除</button></td>
    </tr>
  `).join('');
  document.getElementById('tbody').innerHTML = html;
}}

function addRow() {{
  detailRows.push({{ period:'', item:'', value:'', note:'' }});
  renderRows();
}}

function deleteRow(i) {{
  detailRows = detailRows.filter((_, idx) => idx !== i);
  renderRows();
}}

async function reloadData() {{
  const res = await fetch(`/api/detail?project_id=${{encodeURIComponent(PROJECT_ID)}}&code=${{encodeURIComponent(CODE)}}`);
  const data = await res.json();
  cache = data;
  if (!res.ok) {{
    document.getElementById('status').textContent = data.error || '读取失败';
    return;
  }}
  render();
}}

function renderMainValues() {{
  const years = cache.years || [];
  const vals = cache.subject_values || {{}};
  const txt = years.map(y => {{
    const v = vals[y];
    if (v == null || Number.isNaN(Number(v))) return `${{y}}年待补充`;
    return `${{y}}年${{Number(v).toFixed(2)}}`;
  }}).join('；');
  document.getElementById('mainValues').textContent = txt || '待补充';
}}

function render() {{
  document.getElementById('title').textContent = `科目明细 - ${{cache.name || CODE}}`;
  document.getElementById('status').textContent = `可编辑明细行数 ${{(cache.detail_rows || []).length}}`;
  document.getElementById('manualText').value = cache.manual_text || '';
  document.getElementById('confirmed').checked = !!cache.confirmed;
  document.getElementById('sheetInfo').textContent = cache.detail_sheet ? `映射明细Sheet: ${{cache.detail_sheet}}` : '未映射到明细Sheet（可手动录入）';
  document.getElementById('mapInfo').textContent = cache.mapping_mode ? `映射方式: ${{cache.mapping_mode}}` : '';
  detailRows = (cache.detail_rows || []).map(x => ({{
    period: x.period || '',
    item: x.item || '',
    value: x.value || '',
    note: x.note || ''
  }}));
  renderMainValues();
  renderRows();
}}

function pickImport() {{
  const el = document.getElementById('importFile');
  if (!el) return;
  el.value = '';
  el.click();
}}

async function importDetail(file) {{
  const reader = new FileReader();
  const b64 = await new Promise((resolve, reject) => {{
    reader.onload = () => {{
      try {{
        const s = String(reader.result || '');
        const parts = s.split(',');
        resolve(parts.length > 1 ? parts[1] : '');
      }} catch (e) {{
        reject(e);
      }}
    }};
    reader.onerror = reject;
    reader.readAsDataURL(file);
  }});
  const payload = {{ project_id: PROJECT_ID, code: CODE, file_b64: b64 }};
  const res = await fetch('/api/detail/import', {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  if (!res.ok || !data.ok) {{
    document.getElementById('status').textContent = '导入失败: ' + (data.error || 'unknown');
    return;
  }}
  detailRows = data.detail_rows || [];
  renderRows();
  document.getElementById('status').textContent = `导入成功：${{detailRows.length}}行（尚未保存）`;
}}

async function saveDetail() {{
  if (!cache) return;
  detailRows = normalizeRowsFromUi();
  const payload = {{
    project_id: PROJECT_ID,
    group_id: 'detail_analysis',
    rows: [{{
      code: cache.code || CODE,
      name: cache.name || '',
      manual_text: document.getElementById('manualText').value || '',
      detail_rows: detailRows,
      confirmed: document.getElementById('confirmed').checked
    }}]
  }};
  const res = await fetch('/api/save', {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify(payload)
  }});
  const data = await res.json();
  document.getElementById('status').textContent = data.ok ? '已保存' : ('保存失败: ' + (data.error || 'unknown'));
  if (data.ok) {{
    await reloadData();
  }}
}}

const importEl = document.getElementById('importFile');
if (importEl) {{
  importEl.addEventListener('change', async (e) => {{
    const f = e.target.files && e.target.files[0];
    if (!f) return;
    await importDetail(f);
  }});
}}

reloadData();
</script>
</body>
</html>"""


class AppHandler(BaseHTTPRequestHandler):
    def log_message(self, format: str, *args):  # noqa: A003
        return

    def _send_json(self, obj: Dict[str, Any], status: int = 200) -> None:
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, html: str, status: int = 200) -> None:
        body = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_xlsx(self, body: bytes, filename: str, status: int = 200) -> None:
        safe_name = filename or "export.xlsx"
        self.send_response(status)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(safe_name)}")
        self.end_headers()
        self.wfile.write(body)

    def _project_id(self, qs: Dict[str, List[str]]) -> str:
        return normalize_project_id((qs.get("project_id") or [DEFAULT_PROJECT_ID])[0])

    def do_GET(self):  # noqa: N802
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)

        if path == "/":
            self._send_html(render_index())
            return

        if path.startswith("/sheet/"):
            group_id = unquote(path.split("/")[-1]).strip()
            group = next((g for g in SHEET_GROUPS if g["id"] == group_id), None)
            if not group:
                self._send_html("<h1>Unknown sheet</h1>", status=404)
                return
            self._send_html(render_sheet_page(group, self._project_id(qs)))
            return

        if path == "/detail":
            project_id = self._project_id(qs)
            code = str((qs.get("code") or [""])[0]).strip().upper()
            if not code:
                self._send_html("<h1>missing code</h1>", status=400)
                return
            self._send_html(render_detail_page(project_id, code))
            return

        if path == "/analysis/assets":
            self._send_html(render_asset_analysis_page(self._project_id(qs)))
            return

        if path == "/analysis/liabilities":
            self._send_html(render_liability_analysis_page(self._project_id(qs)))
            return

        if path == "/analysis/summary":
            self._send_html(render_summary_analysis_page(self._project_id(qs)))
            return

        if path == "/analysis/income":
            self._send_html(render_income_analysis_page(self._project_id(qs)))
            return

        if path == "/analysis/ratios":
            self._send_html(render_ratio_analysis_page(self._project_id(qs)))
            return

        if path == "/analysis/key-ratios":
            self._send_html(render_key_ratio_analysis_page(self._project_id(qs)))
            return

        if path == "/thresholds":
            self._send_html(render_thresholds_page())
            return

        if path == "/rules":
            self._send_html(render_rules_page())
            return

        if path == "/warnings":
            self._send_html(render_warnings_page())
            return

        if path.startswith("/api/sheet/"):
            group_id = unquote(path.split("/")[-1]).strip()
            project_id = self._project_id(qs)
            group = next((g for g in SHEET_GROUPS if g["id"] == group_id), None)
            if not group:
                self._send_json({"error": "unknown group"}, status=404)
                return

            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return

            wb = load_workbook(wb_path, data_only=False)
            sheet_data = read_sheet_rows(wb, group, project_id=project_id)
            years = sheet_data["years"]
            rows = sheet_data["rows"]
            asset_analysis_map = build_asset_analysis_map(wb, project_id) if group_id == "bs" else {}
            liability_analysis_map = build_liability_analysis_map(wb, project_id) if group_id == "bs" else {}

            store = load_store(project_id)
            entries = store.get("entries", {})
            value_overrides = get_value_overrides(store)
            vmap = build_validation_map(group_id, rows, years) if group_id in {"bs", "is", "cf"} else {}
            for r in rows:
                key = make_entry_key(group_id, r["code"])
                saved = entries.get(key, {})
                r["auto_text"] = generate_auto_text(r["name"], r["values"], years)
                r["manual_text"] = str(saved.get("manual_text", "") or "")
                r["confirmed"] = bool(saved.get("confirmed", False))
                if group_id == "bs":
                    code = str(r["code"])
                    a = asset_analysis_map.get(code) or liability_analysis_map.get(code) or {}
                    r["analysis_text"] = str(a.get("final_text", "") or "")
                if group_id in {"bs", "is", "cf"}:
                    code = str(r["code"])
                    chk = vmap.get(str(code).strip().upper(), {"status": "未配置", "message": "该行暂未配置校验规则"})
                    r["validation_status"] = str(chk.get("status", "未配置"))
                    r["validation_message"] = str(chk.get("message", ""))
                    reason = ""
                    for y in years:
                        ov = value_overrides.get(make_value_override_key(group_id, code, str(y)), {})
                        if isinstance(ov, dict) and str(ov.get("reason", "")).strip():
                            reason = str(ov.get("reason", "")).strip()
                            break
                    r["override_reason"] = reason

            self._send_json(
                {
                    "project_id": project_id,
                    "group_id": group_id,
                    "group_label": group["label"],
                    "sheet_title": sheet_data["sheet_title"],
                    "years": years,
                    "rows": rows,
                    "workbook": str(wb_path),
                }
            )
            return

        if path == "/api/detail":
            project_id = self._project_id(qs)
            code = str((qs.get("code") or [""])[0]).strip().upper()
            if not code:
                self._send_json({"error": "missing code"}, status=400)
                return
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            payload = build_detail_payload(wb, project_id, code)
            payload["project_id"] = project_id
            payload["workbook"] = str(wb_path)
            self._send_json(payload)
            return

        if path == "/api/analysis/assets":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            analysis = read_analysis_data(wb, ASSET_ANALYSIS_SHEETS)
            analysis = apply_analysis_code_redirects("asset_analysis", analysis)
            analysis_map = build_asset_analysis_map(wb, project_id)
            self._send_json(
                {
                    "project_id": project_id,
                    "workbook": str(wb_path),
                    "scale": analysis.get("scale", {}),
                    "structure": analysis.get("structure", {}),
                    "analysis": {
                        "rows": list(analysis_map.values()),
                    },
                }
            )
            return

        if path == "/api/analysis/liabilities":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            analysis = read_analysis_data(wb, LIABILITY_ANALYSIS_SHEETS)
            analysis = apply_analysis_code_redirects("liability_analysis", analysis)
            analysis_map = build_liability_analysis_map(wb, project_id)
            self._send_json(
                {
                    "project_id": project_id,
                    "workbook": str(wb_path),
                    "scale": analysis.get("scale", {}),
                    "structure": analysis.get("structure", {}),
                    "analysis": {
                        "rows": list(analysis_map.values()),
                    },
                }
            )
            return

        if path == "/api/analysis/summary":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            payload = build_summary_analysis_payload(wb, project_id)
            payload["workbook"] = str(wb_path)
            self._send_json(payload)
            return

        if path == "/api/analysis/income":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            data = build_income_analysis_map(wb, project_id)
            self._send_json(
                {
                    "project_id": project_id,
                    "workbook": str(wb_path),
                    "sheet_title": data.get("sheet_title"),
                    "years": data.get("years", []),
                    "tree": data.get("tree", []),
                    "nodes": data.get("nodes", []),
                }
            )
            return

        if path == "/api/analysis/ratios":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            data = build_ratio_analysis_map(wb, project_id)
            self._send_json(
                {
                    "project_id": project_id,
                    "workbook": str(wb_path),
                    "sheet_title": data.get("sheet_title"),
                    "years": data.get("years", []),
                    "tree": data.get("tree", []),
                    "nodes": data.get("nodes", []),
                }
            )
            return

        if path == "/api/analysis/key-ratios":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            data = build_key_ratio_analysis_map(wb, project_id)
            self._send_json(
                {
                    "project_id": project_id,
                    "workbook": str(wb_path),
                    "sheet_title": data.get("sheet_title"),
                    "years": data.get("years", []),
                    "tree": data.get("tree", []),
                    "nodes": data.get("nodes", []),
                }
            )
            return

        if path == "/api/analysis/income/export":
            project_id = self._project_id(qs)
            wb_path = workbook_path(project_id)
            if not wb_path.exists():
                self._send_json({"error": f"workbook not found: {wb_path}"}, status=404)
                return
            wb = load_workbook(wb_path, data_only=False)
            data = build_income_analysis_map(wb, project_id)
            body = build_income_analysis_export_xlsx(data, project_id=project_id, source_workbook=str(wb_path))
            self._send_xlsx(body, f"{project_id}_收入结构分析计算表.xlsx")
            return

        if path == "/api/template/report-blank-export":
            project_id = self._project_id(qs)
            try:
                body = build_blank_report_template_xlsx(project_id)
            except FileNotFoundError as e:
                self._send_json({"error": f"workbook not found: {e}"}, status=404)
                return
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
                return
            self._send_xlsx(body, f"{project_id}_报表录入模板.xlsx")
            return

        if path == "/api/template/ratio-indicator-export":
            project_id = self._project_id(qs)
            try:
                body = build_ratio_indicator_template_xlsx(project_id)
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
                return
            self._send_xlsx(body, f"{project_id}_财务指标录入模板.xlsx")
            return

        if path == "/api/rules/catalog":
            self._send_json({"sources": _rule_catalog()})
            return

        if path == "/api/rules":
            source_id = str((qs.get("source_id") or [""])[0]).strip()
            sheet_name = str((qs.get("sheet") or [""])[0]).strip()
            if not source_id or not sheet_name:
                self._send_json({"error": "missing source_id or sheet"}, status=400)
                return
            try:
                data = _read_rule_sheet(source_id, sheet_name)
                self._send_json(data)
            except FileNotFoundError as e:
                self._send_json({"error": f"rule file not found: {e}"}, status=404)
            except Exception as e:  # noqa: BLE001
                self._send_json({"error": str(e)}, status=400)
            return

        if path == "/api/thresholds":
            cfg = _read_threshold_config()
            catalog = _load_bs_code_name_catalog()
            subj_scale = cfg.get("subject_scale_pct", {}) if isinstance(cfg.get("subject_scale_pct"), dict) else {}
            subj_struct = cfg.get("subject_struct_pp", {}) if isinstance(cfg.get("subject_struct_pp"), dict) else {}
            rows = []
            for x in catalog:
                code = str(x.get("code", ""))
                vs = subj_scale.get(code)
                vt = subj_struct.get(code)
                rows.append(
                    {
                        "code": code,
                        "name": str(x.get("name", "")),
                        "enabled": (vs is not None) or (vt is not None),
                        "scale_threshold_pct": vs if vs is not None else cfg.get("global_scale_pct", cfg.get("global_pct", 2.0)),
                        "struct_threshold_pp": vt if vt is not None else cfg.get("global_struct_pp", cfg.get("global_pct", 2.0)),
                        # backward compatibility fields
                        "threshold_pct": vs if vs is not None else cfg.get("global_scale_pct", cfg.get("global_pct", 2.0)),
                    }
                )
            self._send_json(
                {
                    "global_scale_pct": cfg.get("global_scale_pct", cfg.get("global_pct", 2.0)),
                    "global_struct_pp": cfg.get("global_struct_pp", cfg.get("global_pct", 2.0)),
                    "summary_top_coverage_default": cfg.get("summary_top_coverage_default", float(SUMMARY_TOP_COVERAGE_TARGET_PCT)),
                    "summary_top_coverage_income": cfg.get("summary_top_coverage_income", float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)),
                    # backward compatibility field
                    "global_pct": cfg.get("global_scale_pct", cfg.get("global_pct", 2.0)),
                    "rows": rows,
                }
            )
            return

        if path == "/api/projects":
            self._send_json({"projects": list_projects()})
            return

        if path == "/api/warnings":
            self._send_json({"rows": list(RULE_WARNING_LOG)})
            return

        self._send_json({"error": "not found"}, status=404)

    def do_POST(self):  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path not in {
            "/api/save",
            "/api/rules/save",
            "/api/thresholds/save",
            "/api/warnings/clear",
            "/api/template/ratio-indicator-import",
            "/api/detail/import",
        }:
            self._send_json({"error": "not found"}, status=404)
            return

        try:
            size = int(self.headers.get("Content-Length", "0"))
            raw = self.rfile.read(size).decode("utf-8")
            body = json.loads(raw)
        except Exception:
            self._send_json({"ok": False, "error": "invalid payload"}, status=400)
            return

        if parsed.path == "/api/warnings/clear":
            RULE_WARNING_LOG.clear()
            RULE_WARNING_KEYS.clear()
            self._send_json({"ok": True, "cleared": True})
            return

        if parsed.path == "/api/template/ratio-indicator-import":
            project_id = normalize_project_id(str(body.get("project_id", DEFAULT_PROJECT_ID)))
            file_b64 = str(body.get("file_b64", "") or "").strip()
            if not file_b64:
                self._send_json({"ok": False, "error": "missing file_b64"}, status=400)
                return
            try:
                file_bytes = base64.b64decode(file_b64)
                result = import_ratio_indicator_template(project_id, file_bytes)
            except Exception as e:  # noqa: BLE001
                self._send_json({"ok": False, "error": str(e)}, status=400)
                return
            if not result.get("ok"):
                self._send_json(result, status=400)
                return
            self._send_json(result)
            return

        if parsed.path == "/api/detail/import":
            project_id = normalize_project_id(str(body.get("project_id", DEFAULT_PROJECT_ID)))
            code = str(body.get("code", "")).strip().upper()
            file_b64 = str(body.get("file_b64", "") or "").strip()
            if not code:
                self._send_json({"ok": False, "error": "missing code"}, status=400)
                return
            if not file_b64:
                self._send_json({"ok": False, "error": "missing file_b64"}, status=400)
                return
            try:
                file_bytes = base64.b64decode(file_b64)
                parsed_rows = _parse_detail_rows_from_xlsx_bytes(file_bytes)
                wb_path = workbook_path(project_id)
                years: List[str] = []
                if wb_path.exists():
                    wb = load_workbook(wb_path, data_only=False)
                    sr = _find_bs_subject_row(wb, code, project_id=project_id)
                    vals = sr.get("values", {}) if isinstance(sr.get("values"), dict) else {}
                    years = sorted([str(y) for y in vals.keys() if str(y).strip()])[:DEFAULT_YEAR_COUNT]
                    wb.close()
                detail_rows = _merge_rows_to_year_slots(parsed_rows, years) if years else parsed_rows
            except Exception as e:  # noqa: BLE001
                self._send_json({"ok": False, "error": str(e)}, status=400)
                return
            self._send_json({"ok": True, "project_id": project_id, "code": code, "detail_rows": detail_rows})
            return

        if parsed.path == "/api/rules/save":
            source_id = str(body.get("source_id", "")).strip()
            sheet_name = str(body.get("sheet_name", "")).strip()
            headers = body.get("headers", [])
            rows = body.get("rows", [])
            if not source_id or not sheet_name or not isinstance(headers, list) or not isinstance(rows, list):
                self._send_json({"ok": False, "error": "missing source_id/sheet_name/headers/rows"}, status=400)
                return
            try:
                result = _save_rule_sheet(source_id, sheet_name, [str(x) for x in headers], rows)
                if not result.get("ok"):
                    self._send_json(result, status=400)
                    return
                # invalidate runtime caches when rule sheets are edited
                global _KEY_RATIO_RULES_CACHE, _MAIN_TEXT_TEMPLATE_CACHE, _MAIN_TEXT_TEMPLATE_UNIT_CACHE
                if source_id in {"key_ratio_rulebook"}:
                    _KEY_RATIO_RULES_CACHE = None
                if source_id in {"rulebook_main"}:
                    _MAIN_TEXT_TEMPLATE_CACHE = None
                    _MAIN_TEXT_TEMPLATE_UNIT_CACHE = None
                self._send_json(result)
            except FileNotFoundError as e:
                self._send_json({"ok": False, "error": f"rule file not found: {e}"}, status=404)
            except Exception as e:  # noqa: BLE001
                self._send_json({"ok": False, "error": str(e)}, status=400)
            return

        if parsed.path == "/api/thresholds/save":
            try:
                global_scale_pct = float(body.get("global_scale_pct", body.get("global_pct")))
            except Exception:
                self._send_json({"ok": False, "error": "invalid global_scale_pct"}, status=400)
                return
            try:
                global_struct_pp = float(body.get("global_struct_pp", global_scale_pct))
            except Exception:
                global_struct_pp = global_scale_pct
            rows = body.get("rows", [])
            if not isinstance(rows, list):
                self._send_json({"ok": False, "error": "rows must be list"}, status=400)
                return
            result = _save_threshold_config(
                global_scale_pct=global_scale_pct,
                global_struct_pp=global_struct_pp,
                subject_rows=rows,
                summary_top_coverage_default=float(
                    body.get(
                        "summary_top_coverage_default",
                        _read_threshold_config().get("summary_top_coverage_default", float(SUMMARY_TOP_COVERAGE_TARGET_PCT)),
                    )
                ),
                summary_top_coverage_income=float(
                    body.get(
                        "summary_top_coverage_income",
                        _read_threshold_config().get("summary_top_coverage_income", float(SUMMARY_TOP_COVERAGE_TARGET_PCT_INCOME)),
                    )
                ),
            )
            if not result.get("ok"):
                self._send_json(result, status=400)
                return
            self._send_json(result)
            return

        project_id = normalize_project_id(str(body.get("project_id", DEFAULT_PROJECT_ID)))
        group_id = str(body.get("group_id", "")).strip()
        rows = body.get("rows", [])
        if not group_id or not isinstance(rows, list):
            self._send_json({"ok": False, "error": "missing group_id or rows"}, status=400)
            return

        store = load_store(project_id)
        entries = store.setdefault("entries", {})
        value_overrides = store.setdefault("value_overrides", {})
        if not isinstance(value_overrides, dict):
            value_overrides = {}
            store["value_overrides"] = value_overrides
        now = datetime.now().isoformat(timespec="seconds")
        saved = 0

        for r in rows:
            code = str(r.get("code", "")).strip()
            if not code:
                continue
            key = make_entry_key(group_id, code)
            raw_detail_rows = r.get("detail_rows", [])
            detail_rows = raw_detail_rows if isinstance(raw_detail_rows, list) else []
            if group_id in {"bs", "is", "cf"}:
                yrs = [str(x) for x in (r.get("years", []) if isinstance(r.get("years", []), list) else [])]
                vals = r.get("values", {}) if isinstance(r.get("values"), dict) else {}
                reason = str(r.get("override_reason", "")).strip()
                for y in yrs:
                    vk = make_value_override_key(group_id, code, y)
                    new_v = normalize_num(vals.get(y))
                    if new_v is None:
                        value_overrides.pop(vk, None)
                        continue
                    value_overrides[vk] = {
                        "group_id": group_id,
                        "code": code,
                        "year": y,
                        "value": new_v,
                        "reason": reason,
                        "enabled": True,
                        "updated_at": now,
                    }
                entries[key] = {
                    "code": code,
                    "name": str(r.get("name", "")).strip(),
                    "manual_text": str(r.get("manual_text", "")).strip(),
                    "detail_rows": detail_rows,
                    "confirmed": bool(r.get("confirmed", False)),
                    "classification_bucket": str(r.get("classification_bucket", "")).strip(),
                    "updated_at": now,
                }
            else:
                entries[key] = {
                    "code": code,
                    "name": str(r.get("name", "")).strip(),
                    "manual_text": str(r.get("manual_text", "")).strip(),
                    "detail_rows": detail_rows,
                    "confirmed": bool(r.get("confirmed", False)),
                    "classification_bucket": str(r.get("classification_bucket", "")).strip(),
                    "updated_at": now,
                }
            saved += 1

        save_store(project_id, store)
        self._send_json({"ok": True, "saved_rows": saved, "updated_at": now})


def main() -> int:
    run_rule_preflight_checks()
    parser = argparse.ArgumentParser(description="Local workbook web UI")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8787)
    args = parser.parse_args()

    server = ThreadingHTTPServer((args.host, args.port), AppHandler)
    print(f"Web UI running: http://{args.host}:{args.port}")
    print("Use query param project_id if needed, e.g. /sheet/bs?project_id=zhonggang_luonai")
    server.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
